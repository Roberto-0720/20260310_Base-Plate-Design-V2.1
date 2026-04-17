import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import os
import csv
import subprocess
import platform
import shutil
from datetime import datetime
import math
import comtypes.client

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime

import matplotlib
import matplotlib.pyplot as plt
import logging

from tomlkit import ws
logging.getLogger('matplotlib.font_manager').setLevel(logging.ERROR)
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure

class BasePlateApp:
    def __init__(self, root):
        self.root = root
        
        # Check license first
        if not self.check_license():
            return
        
        self.root.title("🏗️ BASE PLATE DESIGN")
        self.root.geometry("1500x800")
        self.root.configure(bg='#f5f5f5')

        # Initialize current file
        self.current_file = None
        self.reaction_csv_file = None
        self.coordinate_csv_file = None
        
        # Initialize SAP2000 connection
        self.SapModel = None
        self.is_sap_connected = False
        self.sap_model_file = None
        self.bpl_folder = None  # Path to "Base Plate Design" folder
        
        # Initialize material data structures
        self.material_trees = {}  # Để lưu reference các tree (sẽ set ở các hàm create sub-tab)
        self.material_status_label = None  # Sẽ tạo ở create_material_define_tab
        self.material_data = {}  # Để lưu dữ liệu material từ Data.xlsx

        # Initialize Base Plate Detail data
        self.base_plate_nodes = {}  # Store node data
        self.selected_node = None   # Currently selected node
        self.fig = None            # Matplotlib figure
        self.ax = None             # Matplotlib axes
        self.canvas = None         # Matplotlib canvas
        self.scatter = None        # Scatter plot object
        self.label_display_var = tk.StringVar(value="both")  # Display: "label", "ratio", or "both"
        self.label_texts = []      # Store text objects for labels
        
        # Create UI components
        self.create_header()
        self.create_menu_bar()
        self.create_main_content()
    
    def check_license(self):
        """Check if application license is valid"""
        from datetime import datetime
        
        expiry_date = datetime(2026, 8, 31)  # Hết tháng 8, 2026
        today = datetime.now()
        
        if today > expiry_date:
            messagebox.showerror(
                "⛔ License Expired",
                "This application has expired.\n\n"
                "Please contact Roberto to unlock the application.\n\n"
                "Application will now close."
            )
            self.root.quit()
            return False
        return True

    def create_header(self):
        """Create header section"""
        header_frame = tk.Frame(self.root, bg='#1a472a', height=70)
        header_frame.pack(fill='x', padx=0, pady=0)
        header_frame.pack_propagate(False)

        title_label = tk.Label(
            header_frame,
            text="BASE PLATE DESIGN",
            font=('Arial', 20, 'bold'),
            bg='#1a472a',
            fg='white'
        )
        title_label.pack(pady=(15, 0))

    def create_menu_bar(self):
        """Create menu bar"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # File Menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="New", command=self.new_file, accelerator="Ctrl+N")
        file_menu.add_command(label="Open", command=self.open_file, accelerator="Ctrl+O")
        file_menu.add_command(label="Save", command=self.save_file, accelerator="Ctrl+S")
        file_menu.add_command(label="Save As...", command=self.save_as_file)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)

        # Tools Menu
        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Tools", menu=tools_menu)
        tools_menu.add_command(label="Export to Excel", command=self.export_to_excel)

        # Help Menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="User Guide", command=self.show_guide)
        help_menu.add_command(label="About", command=self.show_about)

        # Keyboard shortcuts
        self.root.bind('<Control-n>', lambda e: self.new_file())
        self.root.bind('<Control-o>', lambda e: self.open_file())
        self.root.bind('<Control-s>', lambda e: self.save_file())

    def create_main_content(self):
        """Create main content area with tabs"""
        main_container = tk.Frame(self.root, bg='#f5f5f5')
        main_container.pack(fill='both', expand=True, padx=12, pady=12)

        # Create notebook for tabs
        style = ttk.Style()
        style.configure('Custom.TNotebook', background='#f5f5f5')
        style.configure('Custom.TNotebook.Tab', 
                       font=('Arial', 11, 'bold'),
                       padding=[20, 10])

        self.notebook = ttk.Notebook(main_container, style='Custom.TNotebook')
        self.notebook.pack(fill='both', expand=True)

        # Create tabs
        self.create_column_base_tab()
        self.create_material_definition_tab()
        self.create_base_plate_detail_tab()

        # Status bar at bottom
        self.create_status_bar()

        # Auto reload material data khi chuyển sang tab Material Definition
        def on_tab_changed(event):
            selected_tab = self.notebook.select()
            tab_text = self.notebook.tab(selected_tab, "text")
            if tab_text == '📋 Material Definition':
                self.reload_all_material_data()
    
        self.notebook.bind("<<NotebookTabChanged>>", on_tab_changed)        

    def create_material_definition_tab(self):
        """Create Material Definition tab with 3 sub-tabs"""
        material_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(material_frame, text='📋 Material Definition')

        # Sub-notebook cho 3 tab
        sub_style = ttk.Style()
        sub_style.configure('Material.TNotebook', background='white')
        sub_style.configure('Material.TNotebook.Tab', font=('Arial', 10, 'bold'), padding=[15, 8])
    
        material_notebook = ttk.Notebook(material_frame, style='Material.TNotebook')
        material_notebook.pack(fill='both', expand=True, padx=20, pady=(20, 0))

        # Tạo 4 sub-tab
        self.create_material_strength_tab(material_notebook)
        self.create_anchor_bolt_tab(material_notebook)
        self.create_hinge_type_tab(material_notebook)
        self.create_rebar_dev_length_tab(material_notebook)

        # Bottom frame để đặt nút Refresh ở dưới cùng, align right
        bottom_frame = tk.Frame(material_frame, bg='white')
        bottom_frame.pack(fill='x', pady=(0, 20))

        tk.Button(
            bottom_frame,
            text="🔄 Refresh",
            command=self.reload_all_material_data,
            bg='#FF9800',
            fg='white',
            font=('Arial', 11, 'bold'),
            cursor='hand2',
            relief='flat',
            padx=20,
            pady=10
        ).pack(side='right', padx=30)
    
    def create_material_strength_tab(self, parent_notebook):
            """Create Material Strength sub-tab with 4 compact tables: Anchor Bolt, Concrete, Rebar, Base Plate"""
            frame = tk.Frame(parent_notebook, bg='white')
            parent_notebook.add(frame, text='⚙️ Material Strength')

            # Main container với 4 cột đều nhau (2 hàng x 2 cột)
            container = tk.Frame(frame, bg='white')
            container.pack(fill='both', expand=True, padx=30, pady=30)

            # Configure grid
            container.grid_rowconfigure(0, weight=1)
            container.grid_rowconfigure(1, weight=1)
            container.grid_columnconfigure(0, weight=1)
            container.grid_columnconfigure(1, weight=1)

            # Hàm helper để tạo bảng chung
            def create_material_table(parent, row, col, title_text, title_icon, columns, tree_name):
                table_frame = tk.Frame(parent, bg='white', relief='groove', bd=2)
                table_frame.grid(row=row, column=col, sticky='nsew', padx=15, pady=15)

                # Title căn giữa
                title_frame = tk.Frame(table_frame, bg='#f0f0f0')
                title_frame.pack(fill='x')
                tk.Label(
                    title_frame,
                    text=f'{title_icon} {title_text}',
                    font=('Arial', 14, 'bold'),
                    bg='#f0f0f0',
                    fg='#1a472a'
                ).pack(pady=10)

                # Treeview
                tree = ttk.Treeview(table_frame, columns=columns, height=6, show='headings')
                tree.column(columns[0], anchor=tk.W, width=250)
                tree.column(columns[1], anchor=tk.CENTER, width=120)
                tree.heading(columns[0], text=columns[0])
                tree.heading(columns[1], text=columns[1])

                scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=tree.yview)
                tree.configure(yscroll=scrollbar.set)

                tree.pack(side='top', fill='both', expand=True, padx=10, pady=(0,10))
                scrollbar.pack(side='right', fill='y', padx=(0,10), pady=(0,10))

                # Buttons dưới bảng
                btn_frame = tk.Frame(table_frame, bg='white')
                btn_frame.pack(fill='x', pady=(0,10))
                tk.Button(btn_frame, text='➕ Add', command=lambda t=tree: self.add_row_to_material_tree(t, tree_name),
                          bg='#4CAF50', fg='white', font=('Arial', 9, 'bold'), width=10).pack(side='left', padx=5)
                tk.Button(btn_frame, text='❌ Delete', command=lambda t=tree: self.delete_row_from_treeview(t),
                          bg='#f44336', fg='white', font=('Arial', 9, 'bold'), width=10).pack(side='left', padx=5)
                tk.Button(btn_frame, text='💾 Save', command=self.save_material_strength_data,
                          bg='#2196F3', fg='white', font=('Arial', 9, 'bold'), width=10).pack(side='left', padx=5)

                tree.bind('<Double-1>', lambda e, t=tree: self.edit_treeview_cell(e, t))

                # Lưu tree
                self.material_trees[tree_name] = tree
                return tree

            # Tạo 4 bảng (2x2 grid)
            self.anchor_bolt_material_tree = create_material_table(
                container, 0, 0, 'ANCHOR BOLT', '🔩', 
                ('Material Type', 'futa (MPa)'), 'anchor_bolt'
            )

            self.concrete_tree = create_material_table(
                container, 0, 1, 'CONCRETE', '🏗️', 
                ('Material Type', "f'c (MPa)"), 'concrete'
            )

            self.rebar_tree = create_material_table(
                container, 1, 0, 'REBAR', '⚙️', 
                ('Material Type', 'fy (MPa)'), 'rebar'
            )

            self.base_plate_material_tree = create_material_table(
                container, 1, 1, 'BASE PLATE', '📐', 
                ('Material Type', 'fy (MPa)'), 'base_plate'
            )

            # Load data khi mở tab
            self.reload_all_material_data()
    
    def create_anchor_bolt_tab(self, parent_notebook):
        """Create Anchor Bolt Table sub-tab"""
        frame = tk.Frame(parent_notebook, bg='white')
        parent_notebook.add(frame, text='🔩 Anchor Bolt Table')
        
        # Treeview
        columns = ('db', 'Rmin', 'a', 'W', 'T', 'S', 'NutW', 'nt', 'Nut Allowance', 'Edge Min', 'Leng A1', 'Leng A2')
        tree = ttk.Treeview(frame, columns=columns, height=15)
        tree.column('#0', width=0, stretch=tk.NO)
        
        for i, col in enumerate(columns):
            width = 80 if i < 2 else 70
            tree.column(col, anchor=tk.CENTER, width=width)
            tree.heading(col, text=col, anchor=tk.CENTER)
        
        # Scrollbars
        scrollbar_y = ttk.Scrollbar(frame, orient='vertical', command=tree.yview)
        scrollbar_x = ttk.Scrollbar(frame, orient='horizontal', command=tree.xview)
        tree.configure(yscroll=scrollbar_y.set, xscroll=scrollbar_x.set)
        
        tree.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)
        scrollbar_y.grid(row=0, column=1, sticky='ns', pady=10)
        scrollbar_x.grid(row=1, column=0, sticky='ew', padx=10)
        
        # Buttons frame
        btn_frame = tk.Frame(frame, bg='white')
        btn_frame.grid(row=2, column=0, columnspan=2, sticky='ew', padx=10, pady=10)
        
        tk.Button(btn_frame, text='➕ Add Row', command=lambda: self.add_row_to_treeview(tree, 'Anchor Bolt Table', columns),
                 bg='#4CAF50', fg='white', font=('Arial', 9, 'bold'), cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_frame, text='❌ Delete Row', command=lambda: self.delete_row_from_treeview(tree),
                 bg='#f44336', fg='white', font=('Arial', 9, 'bold'), cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_frame, text='💾 Save', command=lambda: self.save_material_data(tree, 'Anchor Bolt Table'),
                 bg='#2196F3', fg='white', font=('Arial', 9, 'bold'), cursor='hand2').pack(side='left', padx=5)
        
        # Configure grid
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)
        
        # Load data from Excel
        self.load_anchor_bolt_data(tree)
        
        # Bind double-click to edit
        tree.bind('<Double-1>', lambda e: self.edit_treeview_cell(e, tree))
        
        self.material_data['Anchor Bolt Table'] = tree
        self.material_trees['Anchor Bolt Table'] = tree
    
    def create_hinge_type_tab(self, parent_notebook):
            """Create Hinge Type sub-tab"""
            frame = tk.Frame(parent_notebook, bg='white')
            parent_notebook.add(frame, text='🔩 Hinge Type')

            # Treeview - 22 cột (BỎ AB size, AB Hole, THÊM 10 cột mới)
            columns = ('Column size', 'Type', 'No.AB', 'P1', 'N', 'A', 'B', 'C', 'E', 'F', 'P2', 'Y',
                      'Np', 'Bp', 'c', 'nrb', 'drb', 'dtb', 'X-leg', 'Y-leg', 'Layer 1', 'Layer 2')
            tree = ttk.Treeview(frame, columns=columns, height=15)
            tree.column('#0', width=0, stretch=tk.NO)

            # Set column widths
            for i, col in enumerate(columns):
                if i == 0:  # Column size
                    width = 120
                else:
                    width = 70
                tree.column(col, anchor=tk.CENTER, width=width)
                tree.heading(col, text=col, anchor=tk.CENTER)

            # Scrollbars
            scrollbar_y = ttk.Scrollbar(frame, orient='vertical', command=tree.yview)
            scrollbar_x = ttk.Scrollbar(frame, orient='horizontal', command=tree.xview)
            tree.configure(yscroll=scrollbar_y.set, xscroll=scrollbar_x.set)

            tree.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)
            scrollbar_y.grid(row=0, column=1, sticky='ns', pady=10)
            scrollbar_x.grid(row=1, column=0, sticky='ew', padx=10)

            # Buttons frame
            btn_frame = tk.Frame(frame, bg='white')
            btn_frame.grid(row=2, column=0, columnspan=2, sticky='ew', padx=10, pady=10)

            tk.Button(btn_frame, text='➕ Add Row', command=lambda: self.add_row_to_treeview(tree, 'Hinge Type', columns),
                     bg='#4CAF50', fg='white', font=('Arial', 9, 'bold'), cursor='hand2').pack(side='left', padx=5)
            tk.Button(btn_frame, text='❌ Delete Row', command=lambda: self.delete_row_from_treeview(tree),
                     bg='#f44336', fg='white', font=('Arial', 9, 'bold'), cursor='hand2').pack(side='left', padx=5)
            tk.Button(btn_frame, text='💾 Save', command=lambda: self.save_material_data(tree, 'Hinge Type'),
                     bg='#2196F3', fg='white', font=('Arial', 9, 'bold'), cursor='hand2').pack(side='left', padx=5)

            # Configure grid
            frame.grid_rowconfigure(0, weight=1)
            frame.grid_columnconfigure(0, weight=1)

            # Load data from Excel
            self.load_hinge_type_data(tree)

            # Bind double-click to edit
            tree.bind('<Double-1>', lambda e: self.edit_treeview_cell(e, tree))

            self.material_data['Hinge Type'] = tree
            self.material_trees['Hinge Type'] = tree

    def create_rebar_dev_length_tab(self, parent_notebook):
            """Create Rebar Development Length sub-tab"""
            frame = tk.Frame(parent_notebook, bg='white')
            parent_notebook.add(frame, text='📏 Rebar Dev Length')

            # Treeview - 3 cột
            columns = ('Bars', 'Ld', 'Ldh')
            tree = ttk.Treeview(frame, columns=columns, height=15)
            tree.column('#0', width=0, stretch=tk.NO)

            for col in columns:
                tree.column(col, anchor=tk.CENTER, width=150)
                tree.heading(col, text=col, anchor=tk.CENTER)

            # Scrollbars
            scrollbar_y = ttk.Scrollbar(frame, orient='vertical', command=tree.yview)
            tree.configure(yscroll=scrollbar_y.set)

            tree.grid(row=0, column=0, sticky='nsew', padx=10, pady=10)
            scrollbar_y.grid(row=0, column=1, sticky='ns', pady=10)

            # Buttons frame
            btn_frame = tk.Frame(frame, bg='white')
            btn_frame.grid(row=1, column=0, columnspan=2, sticky='ew', padx=10, pady=10)

            tk.Button(btn_frame, text='➕ Add Row', command=lambda: self.add_row_to_treeview(tree, 'Rebar Development Length', columns),
                     bg='#4CAF50', fg='white', font=('Arial', 9, 'bold'), cursor='hand2').pack(side='left', padx=5)
            tk.Button(btn_frame, text='❌ Delete Row', command=lambda: self.delete_row_from_treeview(tree),
                     bg='#f44336', fg='white', font=('Arial', 9, 'bold'), cursor='hand2').pack(side='left', padx=5)
            tk.Button(btn_frame, text='💾 Save', command=lambda: self.save_material_data(tree, 'Rebar Development Length'),
                     bg='#2196F3', fg='white', font=('Arial', 9, 'bold'), cursor='hand2').pack(side='left', padx=5)

            # Configure grid
            frame.grid_rowconfigure(0, weight=1)
            frame.grid_columnconfigure(0, weight=1)

            # Load data from Excel
            self.load_rebar_dev_length_data(tree)

            # Bind double-click to edit
            tree.bind('<Double-1>', lambda e: self.edit_treeview_cell(e, tree))

            self.material_data['Rebar Development Length'] = tree
            self.material_trees['Rebar Development Length'] = tree

    def reload_all_material_data(self):
            """Reload tất cả dữ liệu từ Data.xlsx nếu bpl_folder hợp lệ"""
            if not self.bpl_folder or not os.path.exists(os.path.join(self.bpl_folder, 'Data.xlsx')):
                if self.material_status_label:
                    self.material_status_label.config(
                        text="⚠️ Please connect to SAP2000 first",
                        fg='#FF9800'
                    )
                for tree in self.material_trees.values():
                    if tree:
                        for item in tree.get_children():
                            tree.delete(item)
                return False

            # Load dữ liệu
            self.load_anchor_bolt_material_data(self.material_trees.get('anchor_bolt'))
            self.load_concrete_data(self.material_trees.get('concrete'))
            self.load_rebar_data(self.material_trees.get('rebar'))
            self.load_base_plate_material_data(self.material_trees.get('base_plate'))
            self.load_anchor_bolt_data(self.material_trees.get('Anchor Bolt Table'))
            self.load_hinge_type_data(self.material_trees.get('Hinge Type'))
            self.load_rebar_dev_length_data(self.material_trees.get('Rebar Development Length'))

            if self.material_status_label:
                sap_model_name = os.path.basename(self.sap_model_file) if self.sap_model_file else "Unknown"
                self.material_status_label.config(
                    text=f"📁 Base Plate Design folder (from {sap_model_name})",
                    fg='#4CAF50'
                )

            self.status_label.config(text="● Material data reloaded successfully", fg='#90ee90')
            return True        
    
    def load_steel_data(self, tree):
        """Load Steel data from Data.xlsx (columns A-B)"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return

        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return

        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Material Strength']

            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)

            # Read ALL Steel data from columns A-B (starting from row 2)
            for row_idx in range(2, ws.max_row + 1):
                mat_type = ws[f'A{row_idx}'].value
                strength = ws[f'B{row_idx}'].value

                # Stop when reaching empty row
                if not mat_type or mat_type == 'Material Type':
                    break
                
                tree.insert('', 'end', values=(mat_type, strength))

            wb.close()
            print(f"Debug: Loaded {len(tree.get_children())} steel materials from Data.xlsx")
        except Exception as e:
            print(f"Error loading Steel data: {e}")
    
    def load_concrete_data(self, tree):
        """Load Concrete data from Data.xlsx (columns D-E)"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return

        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return

        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Material Strength']

            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)

            # Read ALL Concrete data from columns D-E (starting from row 2)
            for row_idx in range(2, ws.max_row + 1):
                mat_type = ws[f'D{row_idx}'].value
                strength = ws[f'E{row_idx}'].value

                # Stop when reaching empty row
                if not mat_type or mat_type == 'Material Type':
                    break
                
                tree.insert('', 'end', values=(mat_type, strength))

            wb.close()
            print(f"Debug: Loaded {len(tree.get_children())} concrete materials from Data.xlsx")
        except Exception as e:
            print(f"Error loading Concrete data: {e}")

    def load_rebar_data(self, tree):
        """Load Rebar data from Data.xlsx (columns G-H)"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return
        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return
        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Material Strength']
            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)
            # Read ALL Rebar data from columns G-H (starting from row 2)
            for row_idx in range(2, ws.max_row + 1):
                mat_type = ws[f'G{row_idx}'].value
                strength = ws[f'H{row_idx}'].value
                # Stop when reaching empty row
                if not mat_type or mat_type == 'Material Type':
                    break
                
                tree.insert('', 'end', values=(mat_type, strength))
            wb.close()
            print(f"Debug: Loaded {len(tree.get_children())} rebar materials from Data.xlsx")
        except Exception as e:
            print(f"Error loading Rebar data: {e}")
    
    def save_material_strength_data(self):
            """Save all material strength data back to Data.xlsx"""
            if not self.bpl_folder or not os.path.exists(self.bpl_folder):
                messagebox.showwarning("Warning", "Base Plate Design folder not found!")
                return

            data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
            if not os.path.exists(data_file):
                messagebox.showwarning("Warning", "Data.xlsx not found!")
                return

            try:
                wb = openpyxl.load_workbook(data_file)
                ws = wb['Material Strength']

                # Clear và write Anchor Bolt (A-B)
                for row_idx in range(2, ws.max_row + 1):
                    ws[f'A{row_idx}'] = None
                    ws[f'B{row_idx}'] = None
                anchor_items = self.anchor_bolt_material_tree.get_children()
                for idx, item in enumerate(anchor_items, start=2):
                    values = self.anchor_bolt_material_tree.item(item, 'values')
                    ws[f'A{idx}'] = values[0] if values and values[0] else None
                    ws[f'B{idx}'] = values[1] if values and len(values) > 1 and values[1] else None

                # Clear và write Concrete (D-E)
                for row_idx in range(2, ws.max_row + 1):
                    ws[f'D{row_idx}'] = None
                    ws[f'E{row_idx}'] = None
                concrete_items = self.concrete_tree.get_children()
                for idx, item in enumerate(concrete_items, start=2):
                    values = self.concrete_tree.item(item, 'values')
                    ws[f'D{idx}'] = values[0] if values and values[0] else None
                    ws[f'E{idx}'] = values[1] if values and len(values) > 1 and values[1] else None

                # Clear và write Rebar (G-H)
                for row_idx in range(2, ws.max_row + 1):
                    ws[f'G{row_idx}'] = None
                    ws[f'H{row_idx}'] = None
                rebar_items = self.rebar_tree.get_children()
                for idx, item in enumerate(rebar_items, start=2):
                    values = self.rebar_tree.item(item, 'values')
                    ws[f'G{idx}'] = values[0] if values and values[0] else None
                    ws[f'H{idx}'] = values[1] if values and len(values) > 1 and values[1] else None

                # Clear và write Base Plate (J-K)
                for row_idx in range(2, ws.max_row + 1):
                    ws[f'J{row_idx}'] = None
                    ws[f'K{row_idx}'] = None
                baseplate_items = self.base_plate_material_tree.get_children()
                for idx, item in enumerate(baseplate_items, start=2):
                    values = self.base_plate_material_tree.item(item, 'values')
                    ws[f'J{idx}'] = values[0] if values and values[0] else None
                    ws[f'K{idx}'] = values[1] if values and len(values) > 1 and values[1] else None

                wb.save(data_file)
                wb.close()

                messagebox.showinfo("✅ Success",
                        f"Material Strength data saved successfully!\n\n"
                        f"Anchor Bolt: {len(anchor_items)} items\n"
                        f"Concrete: {len(concrete_items)} items\n"
                        f"Rebar: {len(rebar_items)} items\n"
                        f"Base Plate: {len(baseplate_items)} items")

                self.status_label.config(text="● Material Strength saved", fg='#90ee90')

            except Exception as e:
                messagebox.showerror("Error", f"Failed to save Material Strength:\n{str(e)}")
    
    def load_anchor_bolt_data(self, tree):
        """Load Anchor Bolt Table data from Data.xlsx"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return
        
        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return
        
        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Anchor Bolt Table']
            
            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)
            
            # Read data starting from row 2 (skip header)
            for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
                if row[0] and row[0] != 'db':
                    values = tuple('' if v is None else str(v) for v in row)
                    tree.insert('', 'end', values=values)
            
            wb.close()
        except Exception as e:
            print(f"Error loading Anchor Bolt data: {e}")
    
    def load_base_plate_dimension_data(self, tree):
        """Load Base Plate Dimension data from Data.xlsx"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return
        
        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return
        
        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Base Plate Dimension']
            
            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)
            
            # Read data starting from row 2 (skip header)
            for row in ws.iter_rows(min_row=2, max_col=14, values_only=True):
                if row[0] and row[0] != 'Column size':
                    values = tuple('' if v is None else str(v) for v in row)
                    tree.insert('', 'end', values=values)
            
            wb.close()
        except Exception as e:
            print(f"Error loading Base Plate Dimension data: {e}")
    
    def add_row_to_material_tree(self, tree, material_type):
        if material_type in ['steel', 'rebar']:
            new_row = ('', '')  # Material Type, strength
        else:  # concrete
            new_row = ('', '')
        tree.insert('', 'end', values=new_row)

    def add_row_to_treeview(self, tree, sheet_name, columns):
        """Add empty row to treeview for tables like Anchor Bolt, Hinge Type, Rebar Dev Length"""
        empty_values = tuple('' for _ in columns)
        tree.insert('', 'end', values=empty_values)
    
    def delete_row_from_treeview(self, tree):
        """Delete selected row from treeview"""
        selected = tree.selection()
        if selected:
            for item in selected:
                tree.delete(item)
    
    def edit_treeview_cell(self, event, tree):
        """Edit cell on double-click"""
        item = tree.identify('item', event.x, event.y)
        column = tree.identify('column', event.x, event.y)
        
        if not item or not column or column == '#0':
            return
        
        # Get cell position
        x, y, w, h = tree.bbox(item, column)
        
        # Create entry widget
        entry = tk.Entry(tree, width=15)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, tree.item(item, 'values')[int(column[1:]) - 1])
        entry.focus()
        
        def save_edit():
            values = list(tree.item(item, 'values'))
            values[int(column[1:]) - 1] = entry.get()
            tree.item(item, values=values)
            entry.destroy()
        
        def cancel_edit(event=None):
            entry.destroy()
        
        entry.bind('<Return>', lambda e: save_edit())
        entry.bind('<Escape>', cancel_edit)
        entry.bind('<FocusOut>', lambda e: save_edit())
    
    def save_material_data(self, tree, sheet_name):
        """Save treeview data back to Data.xlsx"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            messagebox.showwarning("Warning", "Base Plate Design folder not found!")
            return
        
        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            messagebox.showwarning("Warning", "Data.xlsx not found!")
            return
        
        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb[sheet_name]
            
            # Clear old data (keep header)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            
            # Write new data
            for idx, item in enumerate(tree.get_children(), start=2):
                values = tree.item(item, 'values')
                for col_idx, value in enumerate(values, start=1):
                    ws.cell(row=idx, column=col_idx, value=value if value else None)
            
            wb.save(data_file)
            wb.close()
            
            messagebox.showinfo("✅ Success", f"{sheet_name} data saved successfully!")
            self.status_label.config(text=f"● {sheet_name} saved", fg='#90ee90')
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save {sheet_name}:\n{str(e)}")

    def create_column_base_tab(self):
        """Create Column Base tab"""
        arrangement_frame = tk.Frame(self.notebook, bg='white')
        self.notebook.add(arrangement_frame, text='📍 Column Base')

        # Center container
        center_container = tk.Frame(arrangement_frame, bg='white')
        center_container.pack(expand=True)

        # Icon and title
        title_label = tk.Label(
            center_container,
            text="📍 COLUMN BASE DATA",
            font=('Arial', 20, 'bold'),
            bg='white',
            fg='#1a472a'
        )
        title_label.pack(pady=(50, 20))

        # Description
        desc_label = tk.Label(
            center_container,
            text="Load column base data from your structural analysis model",
            font=('Arial', 11),
            bg='white',
            fg='#666666'
        )
        desc_label.pack(pady=(0, 20))

        # Software Selection Dropdown
        software_frame = tk.Frame(center_container, bg='white')
        software_frame.pack(pady=(0, 10))
        
        software_label = tk.Label(
            software_frame,
            text="Software:",
            font=('Arial', 11, 'bold'),
            bg='white',
            fg='#1a472a'
        )
        software_label.pack(side='left', padx=(0, 10))
        
        self.coordinate_software_var = tk.StringVar(value="SAP2000")
        software_menu = ttk.OptionMenu(
            software_frame,
            self.coordinate_software_var,
            "SAP2000",
            "SAP2000",
            "STAAD PRO"
        )
        software_menu.config(width=12)
        software_menu.pack(side='left')
        
        # Include RC Pier Selection
        pier_frame = tk.Frame(center_container, bg='white')
        pier_frame.pack(pady=(0, 15))
        
        pier_label = tk.Label(
            pier_frame,
            text="Include RC Pier?",
            font=('Arial', 11, 'bold'),
            bg='white',
            fg='#1a472a'
        )
        pier_label.pack(side='left', padx=(0, 10))
        
        self.include_rc_pier_var = tk.StringVar(value="No")
        pier_menu = ttk.OptionMenu(
            pier_frame,
            self.include_rc_pier_var,
            "No",
            "No",
            "Yes"
        )
        pier_menu.config(width=12)
        pier_menu.pack(side='left')

        # Buttons frame
        buttons_frame = tk.Frame(center_container, bg='white')
        buttons_frame.pack(pady=15)

        # Get Model button (Blue)
        self.coord_get_model_btn = tk.Button(
            buttons_frame,
            text="🏗️ GET MODEL",
            command=self.get_sap_model_coordinates,
            bg='#2196F3',
            fg='white',
            font=('Arial', 11, 'bold'),
            relief='flat',
            padx=30,
            pady=12,
            cursor='hand2',
            borderwidth=0
        )
        self.coord_get_model_btn.pack(side='left', padx=5)
        self.coord_get_model_btn.bind('<Enter>', lambda e: e.widget.config(bg='#1976D2') if e.widget['state'] != 'disabled' else None)
        self.coord_get_model_btn.bind('<Leave>', lambda e: e.widget.config(bg='#2196F3') if e.widget['state'] != 'disabled' else None)

        # Disconnect button (Orange)
        self.coord_disconnect_btn = tk.Button(
            buttons_frame,
            text="🔌 DISCONNECT",
            command=self.disconnect_sap_model_coord,
            bg='#FF9800',
            fg='white',
            font=('Arial', 11, 'bold'),
            relief='flat',
            padx=30,
            pady=12,
            cursor='hand2',
            borderwidth=0,
            state='disabled'
        )
        self.coord_disconnect_btn.pack(side='right', padx=5)
        self.coord_disconnect_btn.bind('<Enter>', lambda e: e.widget.config(bg='#E68900') if e.widget['state'] != 'disabled' else None)
        self.coord_disconnect_btn.bind('<Leave>', lambda e: e.widget.config(bg='#FF9800') if e.widget['state'] != 'disabled' else None)

        # Load Auto button (Green)
        self.coord_load_auto_btn = tk.Button(
            buttons_frame,
            text="📂 LOAD (AUTO)",
            command=self.load_coordinates_auto,
            bg='#4CAF50',
            fg='white',
            font=('Arial', 11, 'bold'),
            relief='flat',
            padx=30,
            pady=12,
            cursor='hand2',
            borderwidth=0,
            state='disabled'
        )
        self.coord_load_auto_btn.pack(side='left', padx=5)
        self.coord_load_auto_btn.bind('<Enter>', lambda e: e.widget.config(bg='#45a049') if e.widget['state'] != 'disabled' else None)
        self.coord_load_auto_btn.bind('<Leave>', lambda e: e.widget.config(bg='#4CAF50') if e.widget['state'] != 'disabled' else None)

        # Status info - SAP Model Path
        self.coordinate_sap_path_label = tk.Label(
            center_container,
            text="",
            font=('Arial', 9, 'italic'),
            bg='white',
            fg='#2196F3'
        )
        self.coordinate_sap_path_label.pack(pady=(5, 0))

        # Status info - Data status
        self.coordinate_status_label = tk.Label(
            center_container,
            text="No coordinate data loaded",
            font=('Arial', 10, 'italic'),
            bg='white',
            fg='#999999'
        )
        self.coordinate_status_label.pack(pady=(10, 0))

        # Instructions
        inst_frame = tk.LabelFrame(
            center_container,
            text=" Instructions ",
            font=('Arial', 10, 'bold'),
            bg='white',
            fg='#1a472a',
            relief='solid',
            bd=1
        )
        inst_frame.pack(pady=(30, 0), padx=50, fill='x')

        instructions = """
1. Save your project first (File → Save)
2. Select software type (SAP2000 or STAAD PRO)
3. Click "GET MODEL": Connect to open SAP2000 model
4. Click "LOAD (AUTO)": Automatically extract Coordinates & Loading
  • Include RC Pier: select RC Pier + Steel Column only
  • Exclude RC Pier: select Steel Column only
  • Select Element only, Don't select Joint
5. Reaction file will be generated and opened automatically
        """
        
        inst_text = tk.Label(
            inst_frame,
            text=instructions,
            font=('Arial', 9),
            bg='white',
            fg='#424242',
            justify='left',
            anchor='w'
        )
        inst_text.pack(padx=15, pady=10, anchor='w')

    def create_base_plate_detail_tab(self):
            """Create Base Plate Detail tab with interactive plan view"""
            diagram_frame = tk.Frame(self.notebook, bg='white')
            self.notebook.add(diagram_frame, text='📊 Base Plate Detail')

            # Main container with 2 panels
            main_container = tk.Frame(diagram_frame, bg='white')
            main_container.pack(fill='both', expand=True, padx=10, pady=10)

            # LEFT PANEL (30%) - Control panel with scrollbar
            left_panel_outer = tk.Frame(main_container, bg='#f8f9fa', relief='groove', bd=2)
            left_panel_outer.pack(side='left', fill='both', padx=(0, 10), pady=0)
            left_panel_outer.pack_propagate(False)
            left_panel_outer.config(width=350)
            
            # Create Canvas and Scrollbar for left panel
            left_canvas = tk.Canvas(left_panel_outer, bg='#f8f9fa', highlightthickness=0)
            scrollbar = ttk.Scrollbar(left_panel_outer, orient='vertical', command=left_canvas.yview)
            left_panel = tk.Frame(left_canvas, bg='#f8f9fa', width=335)
            
            left_panel.bind(
                '<Configure>',
                lambda e: left_canvas.configure(scrollregion=left_canvas.bbox('all'))
            )
            
            left_canvas.create_window((0, 0), window=left_panel, anchor='nw', width=335)
            left_canvas.configure(yscrollcommand=scrollbar.set)
            
            scrollbar.pack(side='right', fill='y')
            left_canvas.pack(side='left', fill='both', expand=True)

            # RIGHT PANEL (70%) - Matplotlib canvas
            right_panel = tk.Frame(main_container, bg='white', relief='groove', bd=2)
            right_panel.pack(side='right', fill='both', expand=True)

            # ==================== LEFT PANEL CONTENT ====================

            # Title
            title_label = tk.Label(
                left_panel,
                text="📊 BASE PLATE DETAIL",
                font=('Arial', 14, 'bold'),
                bg='#f8f9fa',
                fg='#1a472a'
            )
            title_label.pack(pady=(15, 10))

            # Status section
            status_frame = tk.LabelFrame(
                left_panel,
                text=" Plan Status ",
                font=('Arial', 10, 'bold'),
                bg='#f8f9fa',
                fg='#1a472a'
            )
            status_frame.pack(fill='x', padx=15, pady=(0, 10))

            self.bpl_status_label = tk.Label(
                status_frame,
                text="No data loaded",
                font=('Arial', 10),
                bg='#f8f9fa',
                fg='#999999'
            )
            self.bpl_status_label.pack(padx=10, pady=8)

            reload_btn = tk.Button(
                status_frame,
                text="🔄 Reload Data",
                command=self.load_base_plate_plan,
                bg='#FF9800',
                fg='white',
                font=('Arial', 9, 'bold'),
                cursor='hand2',
                relief='flat',
                padx=15,
                pady=5
            )
            reload_btn.pack(pady=(0, 8))

            # Selected Node section
            node_frame = tk.LabelFrame(
                left_panel,
                text=" Selected Node ",
                font=('Arial', 10, 'bold'),
                bg='#f8f9fa',
                fg='#1a472a'
            )
            node_frame.pack(fill='both', expand=True, padx=15, pady=(0, 10))

            # Node info
            self.node_info_label = tk.Label(
                node_frame,
                text="Click a node to select",
                font=('Arial', 10, 'italic'),
                bg='#f8f9fa',
                fg='#999999',
                justify='left'
            )
            self.node_info_label.pack(padx=10, pady=10, anchor='w')

            # Dropdown frame (will be populated when node selected)
            self.dropdown_frame = tk.Frame(node_frame, bg='#f8f9fa')
            self.dropdown_frame.pack(fill='x', padx=10, pady=(0, 10))

            # Apply button
            self.apply_btn = tk.Button(
                node_frame,
                text="💾 Apply to Node",
                command=self.apply_node_settings,
                bg='#4CAF50',
                fg='white',
                font=('Arial', 10, 'bold'),
                cursor='hand2',
                relief='flat',
                padx=20,
                pady=8,
                state='disabled'
            )
            self.apply_btn.pack(pady=(5, 10))

            # ========== THÊM BUTTON MỚI ==========
            # Edit Node Detail button
            self.edit_detail_btn = tk.Button(
                node_frame,
                text="✏️ Edit Node Detail",
                command=self.edit_node_detail,
                bg='#FF9800',
                fg='white',
                font=('Arial', 10, 'bold'),
                cursor='hand2',
                relief='flat',
                padx=20,
                pady=8,
                state='disabled'
            )
            self.edit_detail_btn.pack(pady=(0, 10))

            # Copy to multiple button
            self.copy_multi_btn = tk.Button(
                node_frame,
                text="📋 Copy to Multiple Nodes",
                command=self.copy_to_multiple,
                bg='#2196F3',
                fg='white',
                font=('Arial', 9, 'bold'),
                cursor='hand2',
                relief='flat',
                padx=15,
                pady=6,
                state='disabled'
            )
            self.copy_multi_btn.pack(pady=(0, 10))

            # Run Design Check button (at bottom)
            run_frame = tk.Frame(left_panel, bg='#f8f9fa')
            run_frame.pack(side='bottom', fill='x', padx=15, pady=15)

            self.run_design_btn = tk.Button(
                run_frame,
                text="▶️ RUN DESIGN CHECK",
                command=self.run_design_check,
                bg='#1a472a',
                fg='white',
                font=('Arial', 11, 'bold'),
                cursor='hand2',
                relief='flat',
                padx=20,
                pady=12,
                state='disabled'
            )
            self.run_design_btn.pack(fill='x')

            # ==================== RIGHT PANEL - MATPLOTLIB ====================

            # Create matplotlib figure
            self.fig = Figure(figsize=(8, 6), dpi=100, facecolor='white')
            self.ax = self.fig.add_subplot(111)

            # Create canvas
            self.canvas = FigureCanvasTkAgg(self.fig, master=right_panel)
            self.canvas.draw()

            # Toolbar
            toolbar_frame = tk.Frame(right_panel, bg='white')
            toolbar_frame.pack(side='top', fill='x')
            toolbar = NavigationToolbar2Tk(self.canvas, toolbar_frame)
            toolbar.update()

            # Pack canvas
            self.canvas.get_tk_widget().pack(side='top', fill='both', expand=True)

            # Legend & Display Label frame
            bottom_frame = tk.Frame(right_panel, bg='white')
            bottom_frame.pack(side='bottom', fill='x', pady=5)
            
            # Legend
            legend_frame = tk.Frame(bottom_frame, bg='white')
            legend_frame.pack(side='left', fill='x')

            tk.Label(legend_frame, text="Legend:", font=('Arial', 9, 'bold'), bg='white').pack(side='left', padx=10)

            legends = [
                ("⬤", "#CCCCCC", "Not Assigned"),
                ("⬤", "#2196F3", "Assigned"),
                ("⬤", "#4CAF50", "Design OK"),
                ("⬤", "#f44336", "Design NG")
            ]

            for symbol, color, label in legends:
                tk.Label(legend_frame, text=symbol, fg=color, font=('Arial', 12, 'bold'), bg='white').pack(side='left', padx=(10, 2))
                tk.Label(legend_frame, text=label, font=('Arial', 9), bg='white').pack(side='left', padx=(0, 10))
            
            # Display Label options frame
            label_frame = tk.LabelFrame(
                bottom_frame,
                text=" Display Label ",
                font=('Arial', 10, 'bold'),
                bg='#f0f0f0',
                fg='#1a472a',
                relief='solid',
                bd=1
            )
            label_frame.pack(side='left', padx=20)
            
            # Radio buttons for node label/ratio display
            tk.Radiobutton(
                label_frame,
                text="Label",
                variable=self.label_display_var,
                value="label",
                font=('Arial', 9),
                bg='#f0f0f0',
                command=self.update_plot_display
            ).pack(side='left', padx=8, pady=3)
            
            tk.Radiobutton(
                label_frame,
                text="Ratio",
                variable=self.label_display_var,
                value="ratio",
                font=('Arial', 9),
                bg='#f0f0f0',
                command=self.update_plot_display
            ).pack(side='left', padx=8, pady=3)
            
            tk.Radiobutton(
                label_frame,
                text="Both",
                variable=self.label_display_var,
                value="both",
                font=('Arial', 9),
                bg='#f0f0f0',
                command=self.update_plot_display
            ).pack(side='left', padx=8, pady=3)

            # Connect click event
            self.canvas.mpl_connect('button_press_event', self.on_node_click)

            # Initial plot
            self.plot_empty_state()

            # Auto-load data when tab is created
            self.root.after(500, self.load_base_plate_plan)

    def create_status_bar(self):
        """Create status bar at bottom"""
        status_frame = tk.Frame(self.root, bg='#2d5a3d', height=35)
        status_frame.pack(fill='x', side='bottom')
        status_frame.pack_propagate(False)

        self.status_label = tk.Label(
            status_frame,
            text="● Ready",
            font=('Arial', 10, 'bold'),
            bg='#2d5a3d',
            fg='#90ee90',
            anchor='w',
            padx=15
        )
        self.status_label.pack(side='left', fill='both', expand=True)

    # ==================== File Operations ====================
    
    def new_file(self):
        """Create new file"""
        self.current_file = None
        self.reaction_csv_file = None
        self.coordinate_csv_file = None
        
        self.status_label.config(text="● New file created", fg='#90ee90')

    def save_file(self):
        """Save current data"""
        if not self.current_file:
            self.save_as_file()
        else:
            self.save_data_to_file(self.current_file)

    def save_as_file(self):
        """Save data to new file"""
        initial_dir = None
        initial_file = "Bpl.json"
        if self.bpl_folder and os.path.exists(self.bpl_folder):
            initial_dir = self.bpl_folder
        
        file_path = filedialog.asksaveasfilename(
            initialdir=initial_dir,
            initialfile=initial_file,
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save Base Plate Design Data"
        )
        if file_path:
            self.current_file = file_path
            self.save_data_to_file(file_path)

    def save_data_to_file(self, file_path):
        """Save all input data to JSON file"""
        try:
            data = {}
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4)
            
            self.status_label.config(text=f"● File saved: {os.path.basename(file_path)}", fg='#90ee90')
            messagebox.showinfo("Success", f"Data saved successfully to:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file:\n{str(e)}")
            self.status_label.config(text="● Save failed", fg='#ff4d4d')

    def open_file(self):
        """Open and load data from JSON file"""
        file_path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Open Base Plate Design Data"
        )
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                self.current_file = file_path
                self.status_label.config(text=f"● File loaded: {os.path.basename(file_path)}", fg='#90ee90')
                messagebox.showinfo("Success", f"Project loaded from:\n{file_path}")

            except Exception as e:
                messagebox.showerror("Error", f"Failed to open file:\n{str(e)}")
                self.status_label.config(text="● Load failed", fg='#ff4d4d')

    # ==================== Data Loading ====================
    
    def load_coordinate_data(self):
        """Load coordinate data from model analysis"""
        if not self.current_file:
            messagebox.showwarning("Warning", 
                "Please save your project first!\n\n"
                "The CSV file will be created in the same folder as your project file.")
            return
        
        project_dir = os.path.dirname(self.current_file)
        software = self.coordinate_software_var.get()
        
        if software == "STAAD PRO":
            coord_file = os.path.join(project_dir, "ColumnBaseCoordinate_staad.csv")
        else:
            coord_file = os.path.join(project_dir, "ColumnBaseCoordinate_sap.csv")
        
        instruction = (
            "COLUMN BASE COORDINATES LOADING INSTRUCTIONS\n\n"
            "1. A csv file has been created at:\n"
            f"   {coord_file}\n\n"
            "2. The file will open automatically\n\n"
            "3. Paste your column base coordinates (X, Y, Z) starting from cell A1\n\n"
            "4. Save the csv file and close it\n\n"
            "5. Click OK below to confirm\n\n"
            "6. System will generate column_base_data.csv\n"
        )
        
        try:
            if platform.system() == 'Windows':
                os.startfile(coord_file)
            elif platform.system() == 'Darwin':
                subprocess.call(['open', coord_file])
            else:
                subprocess.call(['xdg-open', coord_file])
        except:
            pass
        
        result = messagebox.showinfo("Instructions", instruction)
        
        if result:
            self.coordinate_csv_file = coord_file
            self.generate_column_base_data_file(coord_file, project_dir)
            
            self.coordinate_status_label.config(
                text=f"✓ Coordinates loaded: {os.path.basename(coord_file)}",
                fg='#4CAF50'
            )
            self.status_label.config(text=f"● Column base coordinates ready", fg='#90ee90')

    def generate_column_base_data_file(self, coord_file, project_dir):
        """Generate column_base_data.csv from coordinate file"""
        try:
            data_file = os.path.join(project_dir, "column_base_data.csv")
            software = self.coordinate_software_var.get()
            
            base_coords = []
            with open(coord_file, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                
                if software == "SAP2000":
                    for _ in range(3):
                        next(reader, None)
                    for row in reader:
                        if len(row) >= 5:
                            node = row[0].strip()
                            x = row[3].strip()
                            y = row[4].strip()
                            z = row[5].strip() if len(row) > 5 else "0.0"
                            base_coords.append([node, x, y, z])
                else:
                    next(reader, None)
                    for row in reader:
                        if len(row) >= 4:
                            node = row[0].strip()
                            x = row[1].strip()
                            y_staad = row[2].strip()
                            z_staad = row[3].strip()

                            y_adjusted = z_staad
                            if z_staad.strip():
                                try:
                                    z_float = float(z_staad.strip())
                                    if z_float != 0:
                                        y_adjusted = str(-z_float)
                                except:
                                    pass
                            
                            base_coords.append([node, x, y_adjusted, y_staad])
            
            with open(data_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Column Base Node', 'X (m)', 'Y (m)', 'Z (m)', 'Section Type'])
                for coord in base_coords:
                    writer.writerow(coord + [''])
            
            messagebox.showinfo("Success", 
                f"column_base_data.csv created successfully!\n\n"
                f"Location: {data_file}\n\n"
                f"Total column bases: {len(base_coords)}\n\n"
                f"Please fill in 'Section Type' column manually")
            
            if platform.system() == 'Windows':
                os.startfile(data_file)
            elif platform.system() == 'Darwin':
                subprocess.call(['open', data_file])
            else:
                subprocess.call(['xdg-open', data_file])
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate column_base_data.csv:\n{str(e)}")

    def load_reaction_data(self):
        """Load reaction data from CSV file"""
        if not self.current_file:
            messagebox.showwarning("Warning", 
                "Please save your project first!\n\n"
                "The CSV file will be created in the same folder as your project file.")
            return
        
        project_dir = os.path.dirname(self.current_file)
        software = self.reaction_software_var.get()
        
        if software == "STAAD PRO":
            csv_file = os.path.join(project_dir, "reaction_data_staadpro.csv")
            header_skip = 2
        else:
            csv_file = os.path.join(project_dir, "reaction_data_sap2000.csv")
            header_skip = 3
        
        instruction = (
            "REACTION DATA LOADING INSTRUCTIONS\n\n"
            "1. A csv file has been created at:\n"
            f"   {csv_file}\n\n"
            "2. The file will open automatically\n\n"
            "3. Paste your reaction loads starting from cell A1\n\n"
            "4. Save the csv file and close it\n\n"
            "5. Click OK below to confirm\n"
        )
        
        try:
            if platform.system() == 'Windows':
                os.startfile(csv_file)
            elif platform.system() == 'Darwin':
                subprocess.call(['open', csv_file])
            else:
                subprocess.call(['xdg-open', csv_file])
        except:
            pass
        
        result = messagebox.showinfo("Instructions", instruction)
        
        if result:
            self.reaction_csv_file = csv_file
            
            try:
                if platform.system() == 'Windows':
                    os.startfile(csv_file)
                elif platform.system() == 'Darwin':
                    subprocess.call(['open', csv_file])
                else:
                    subprocess.call(['xdg-open', csv_file])
            except:
                pass
            
            self.status_label.config(text=f"● Reaction CSV ready", fg='#90ee90')

    # ==================== SAP2000 CONNECTION METHODS ====================

    def get_sap_model_coordinates(self):
        """Connect to SAP2000 and enable auto load for coordinates"""
        if self.is_sap_connected:
            messagebox.showinfo("Info", "Already connected to SAP2000 model:\n" + self.sap_model_file)
            return
        
        self._connect_to_sap_model("coordinates")

    def get_sap_model_reaction(self):
        """Connect to SAP2000 for reaction data"""
        if self.is_sap_connected:
            messagebox.showinfo("Info", "Already connected to SAP2000 model:\n" + self.sap_model_file)
            return
        
        self._connect_to_sap_model("reaction")

    def _connect_to_sap_model(self, source_tab):
        """Internal: Connect to SAP2000 model"""
        self.coord_get_model_btn.config(state='disabled')
        self.root.update()
        
        connection_errors = []
        
        # Method 1: Try SAP2000v1.Helper (works for v21+)
        try:
            myHelper = comtypes.client.CreateObject("SAP2000v1.Helper")
            SapObject = myHelper.GetObject("CSI.SAP2000.API.SapObject")
            self.SapModel = SapObject.SapModel
            self._connection_success_sap()
            return
        except Exception as e:
            connection_errors.append(f"Method 1 (SAP2000v1.Helper): {str(e)}")
        
        # Method 2: Try GetActiveObject for older versions (v17-v20)
        try:
            SapObject = comtypes.client.GetActiveObject("CSI.SAP2000.API.SapObject")
            self.SapModel = SapObject.SapModel
            self._connection_success_sap()
            return
        except Exception as e:
            connection_errors.append(f"Method 2 (GetActiveObject): {str(e)}")
        
        # Method 3: Try with version-specific ProgID for SAP2000 v20
        try:
            SapObject = comtypes.client.GetActiveObject("SAP2000.cOAPI")
            self.SapModel = SapObject.SapModel
            self._connection_success_sap()
            return
        except Exception as e:
            connection_errors.append(f"Method 3 (SAP2000.cOAPI): {str(e)}")
        
        # All methods failed
        self.is_sap_connected = False
        
        error_details = "\n".join(connection_errors)
        messagebox.showerror("❌ Connection Error", 
                            f"Cannot connect to SAP2000.\n"
                            f"Please ensure:\n"
                            f"• SAP2000 is running (v17.2.0+)\n"
                            f"• Model is open\n"
                            f"• Python and SAP2000 are same architecture (64-bit)\n\n"
                            f"Error details:\n{error_details}")
        
        self.coord_get_model_btn.config(state='normal')
        self.status_label.config(text="● SAP2000 connection failed", fg='#ff4d4d')

    def create_or_check_data_file(self):
        """Create Data.xlsx if it doesn't exist in Base Plate Design folder"""
        try:
            data_file_path = os.path.join(self.bpl_folder, "Data.xlsx")
            
            # If Data.xlsx already exists, don't create it again
            if os.path.exists(data_file_path):
                print(f"Debug: Data.xlsx already exists at {data_file_path}")
                return
            
            # Create new Data.xlsx with 3 sheets
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # ==================== SHEET 1: Material Strength ====================
            ws1 = wb.create_sheet("Material Strength")
            
            # Anchor Bolt Table (Columns A-B) - ĐỔI TÊN TỪ STEEL
            ws1['A1'] = "Material Type"
            ws1['B1'] = "Tensile Strength futa (MPa)"
            ws1['A2'] = "ASTM F1554 Grade 36"
            ws1['B2'] = 400
            ws1['A3'] = "ASTM A193 Grade B7"
            ws1['B3'] = 800
            
            # Concrete Table (Columns D-E)
            ws1['D1'] = "Material Type"
            ws1['E1'] = "Compressive Strength f'c (MPa)"
            ws1['D2'] = "F'c = 280"
            ws1['E2'] = 28
            ws1['D3'] = "F'c = 350"
            ws1['E3'] = 35
            
            # Rebar Table (Columns G-H)
            ws1['G1'] = "Material Type"
            ws1['H1'] = "Reinforcement yield stress fy (MPa)"
            ws1['G2'] = "SD280W"
            ws1['H2'] = 280
            ws1['G3'] = "SD420W"
            ws1['H3'] = 420
            
            # ========== THÊM MỚI: Base Plate Table (Columns J-K) ==========
            ws1['J1'] = "Material Type"
            ws1['K1'] = "Yield Strength fy (MPa)"
            ws1['J2'] = "ASTM A36"
            ws1['K2'] = 245
            ws1['J3'] = "SM400"
            ws1['K3'] = 235
            ws1['J4'] = "SM490"
            ws1['K4'] = 325
            
            # ==================== SHEET 2: Anchor Bolt Table ====================
            ws2 = wb.create_sheet("Anchor Bolt Table")
            
            # Headers
            headers = ['db', 'Rmin', 'a', 'W', 'T', 'S', 'NutW', 'nt', 'Nut Allowance', 'Edge Min', 'Leng A1', 'Leng A2']
            for col, header in enumerate(headers, 1):
                ws2.cell(row=1, column=col, value=header)
            
            # Data rows
            bolt_data = [
                [16, 319, 18, 0, 0, 25, 26.16, 11, 70, 150, 420, '-'],
                [20, 340, 22, 0, 0, 30, 33, 10, 90, 150, 460, '-'],
                [24, 400, 26, 70, 22, 48, 40, 9, 130, 150, 560, '-'],
                [30, 480, 33, 80, 24, 62, 49, 7, 170, 180, 680, '-'],
                [36, 570, 39, 100, 30, 73, 58.8, 6, 190, 220, 790, '-'],
                [42, 730, 45, 110, 32, 83, 67.9, 5, 210, 255, 970, '-'],
                [48, 860, 51, 120, 34, 93, 77.6, 4.5, 240, 290, 1130, '-'],
                [56, 860, 59, 140, 40, 106, 87.2, 4.5, 270, 340, 1160, '-'],
                [64, 970, 67, 150, 42, 120, 96.8, 4, 290, 385, 1290, '-'],
            ]
            
            for row_idx, row_data in enumerate(bolt_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=value)
            
            # ==================== SHEET 3: Hinge Type ====================
            ws3 = wb.create_sheet("Hinge Type")
            
            # Headers (22 cột)
            headers = ['Column size', 'Type', 'No.AB', 'P1', 'N', 'A', 'B', 'C', 'E', 'F', 'P2', 'Y',
                      'Np', 'Bp', 'c', 'nrb', 'drb', 'dtb', 'X-leg', 'Y-leg', 'Layer 1', 'Layer 2']
            for col, header in enumerate(headers, 1):
                ws3.cell(row=1, column=col, value=header)
            
            # Data rows (18 rows mẫu)
            hinge_data = [
                ['H194X150X6X9', '-', 2, 19, 250, 120, 200, '-', 40, 125, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H244X175X7X11', '-', 2, 19, 300, 120, 250, '-', 65, 150, '-', '-', 500, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H125X125X6.5X9', '-', 2, 16, 200, 120, 200, '-', 40, 100, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H150X150X7X10', '-', 2, 19, 200, 120, 200, '-', 40, 100, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H175X175X7.5X11', '-', 2, 19, 250, 120, 250, '-', 65, 125, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H200X200X8X12', '-', 2, 25, 250, 120, 250, '-', 65, 125, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H250X250X9X14', 'B', 4, 25, 300, 120, 300, 120, 90, 90, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H294X200X8X12', 'B', 4, 25, 350, 120, 250, 120, 65, 115, '-', '-', 500, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H300X300X10X15', 'B', 4, 32, 350, 160, 350, 160, 95, 95, '-', '-', 550, 550, 50, 12, 'D19', 'D10', 3, 3, 50, 75],
                ['H340X250X9X14', 'B', 4, 32, 400, 160, 300, 160, 70, 120, '-', '-', 600, 550, 50, 12, 'D19', 'D10', 3, 3, 50, 75],
                ['H350X350X12X19', 'B', 4, 36, 400, 200, 400, 200, 100, 100, '-', '-', 650, 650, 50, 16, 'D19', 'D13', 3, 3, 50, 75],
                ['H390X300X10X16', 'B', 4, 36, 450, 160, 350, 180, 95, 135, '-', '-', 600, 550, 50, 12, 'D19', 'D13', 3, 3, 50, 75],
                ['H400X400X13X21', 'B', 4, 36, 450, 200, 450, 200, 125, 125, '-', '-', 700, 650, 50, 16, 'D19', 'D13', 3, 3, 50, 75],
                ['H488X300X11X18', 'B', 4, 32, 550, 200, 350, 240, 75, 155, '-', '-', 750, 650, 50, 18, 'D19', 'D13', 4, 3, 50, 75],
                ['H588X300X12X20', 'B', 4, 36, 650, 200, 350, 280, 75, 185, '-', '-', 750, 700, 50, 20, 'D19', 'D13', 4, 4, 50, 75],
                ['BH400X340X16X25', 'B', 4, 36, 450, 200, 400, 200, 100, 125, '-', '-', 650, 600, 50, 16, 'D19', 'D13', 3, 3, 50, 75],
                ['BH500X340X16X25', 'B', 4, 40, 550, 240, 400, 240, 80, 155, '-', '-', 750, 700, 50, 20, 'D19', 'D13', 4, 4, 50, 75],
                ['BH514X340X19X32', 'B', 4, 40, 600, 240, 400, 240, 80, 180, '-', '-', 750, 700, 50, 20, 'D19', 'D13', 4, 4, 50, 75],
            ]
            
            for row_idx, row_data in enumerate(hinge_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws3.cell(row=row_idx, column=col_idx, value=value)

            # ==================== SHEET 4: Rebar Development Length ====================
            ws4 = wb.create_sheet("Rebar Development Length")
            
            # Headers
            headers = ['Bars', 'Ld', 'Ldh']
            for col, header in enumerate(headers, 1):
                ws4.cell(row=1, column=col, value=header)
            
            # Data rows
            rebar_dev_data = [
                [10, 490, 150],
                [13, 490, 150],
                [16, 520, 160],
                [19, 950, 290],
                [22, 1110, 340],
                [25, 1270, 390],
                [32, 1600, 490],
                [36, 1760, 550],
            ]
            
            for row_idx, row_data in enumerate(rebar_dev_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws4.cell(row=row_idx, column=col_idx, value=value)
            
            # Save workbook
            wb.save(data_file_path)
            print(f"Debug: Data.xlsx created successfully at {data_file_path}")
            messagebox.showinfo("✅ Success", f"Data.xlsx created successfully in Base Plate Design folder")
            
        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to create Data.xlsx:\n{str(e)}")
            print(f"Error creating Data.xlsx: {str(e)}")

    def _copy_template_to_bpl_folder(self):
        """Copy Template.xlsx from script directory to Base Plate Design folder"""
        if not self.bpl_folder:
            return
        
        template_dest = os.path.join(self.bpl_folder, 'Template.xlsx')
        if os.path.exists(template_dest):
            return  # Already exists
        
        # Find Template.xlsx in script directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_src = os.path.join(script_dir, 'Template.xlsx')
        
        if os.path.exists(template_src):
            try:
                shutil.copy2(template_src, template_dest)
                print(f"Debug: Template.xlsx copied to {template_dest}")
            except Exception as e:
                print(f"Error copying Template.xlsx: {e}")
        else:
            print(f"Warning: Template.xlsx not found at {script_dir}")

    def _connection_success_sap(self):
        """Handle successful SAP2000 connection"""
        try:
            model_file = self.SapModel.GetModelFilename()
            
            if not model_file:
                messagebox.showwarning("⚠️ Warning", "Please open a SAP2000 model file first!")
                self.coord_get_model_btn.config(state='normal')
                return
            
            self.is_sap_connected = True
            self.sap_model_file = model_file
            
            # Create "Base Plate Design" folder in the same directory as SAP model
            sap_dir = os.path.dirname(model_file)
            self.bpl_folder = os.path.join(sap_dir, "Base Plate Design")
            try:
                folder_exists = os.path.exists(self.bpl_folder)
                
                if not folder_exists:
                    os.makedirs(self.bpl_folder)
                    # Folder is newly created, so create Data.xlsx
                    self.create_or_check_data_file()
                else:
                    # Folder already exists, check if Data.xlsx exists
                    self.create_or_check_data_file()
                
                # Copy Template.xlsx from script directory if not exists
                self._copy_template_to_bpl_folder()
                    
            except Exception as e:
                messagebox.showwarning("⚠️ Warning", f"Could not create Base Plate Design folder:\n{str(e)}")
            
            # Update UI
            sap_path_text = f"📁 SAP Model: {os.path.basename(model_file)}"
            self.coordinate_sap_path_label.config(text=sap_path_text)
            
            self.coord_load_auto_btn.config(state='normal')
            self.coord_get_model_btn.config(state='disabled')
            self.coord_disconnect_btn.config(state='normal')
            
            messagebox.showinfo("✅ Connected to SAP2000", 
                              f"Successfully connected to SAP2000!\n\n"
                              f"Model: {os.path.basename(model_file)}")
            
            self.status_label.config(text=f"● Connected to SAP2000: {os.path.basename(model_file)}", fg='#90ee90')
            self.reload_all_material_data()
        
        finally:
            self.root.update()

    def disconnect_sap_model(self):
        """Disconnect from SAP2000 and reset state"""
        self.SapModel = None
        self.is_sap_connected = False
        self.sap_model_file = None
        self.bpl_folder = None
        
        self.coordinate_sap_path_label.config(text="")
        self.coord_load_auto_btn.config(state='disabled')
        self.coord_get_model_btn.config(state='normal')
        self.coord_disconnect_btn.config(state='disabled')
        
        self.status_label.config(text="● Disconnected from SAP2000", fg='#90ee90')
        messagebox.showinfo("Disconnected", "Disconnected from SAP2000 model")
    
    def disconnect_sap_model_coord(self):
        """Wrapper for disconnect from Column Base tab"""
        self.disconnect_sap_model()
    
    # ==================== AUTO LOAD METHODS ====================
    
    def load_coordinates_auto(self):
        """Load column base coordinates and reaction data from selected beam elements"""
        if not self.is_sap_connected or not self.SapModel:
            messagebox.showerror("Error", "Not connected to SAP2000 model!")
            return

        if not self.current_file:
            messagebox.showwarning("Warning", "Please save your project first!")
            return

        self.load_coordinates_and_reaction_auto()

    def load_coordinates_and_reaction_auto(self):
        """Load column base coordinates from SELECTED BEAM ELEMENTS and create reaction data files from start point nodes"""
        if not self.is_sap_connected or not self.SapModel:
            messagebox.showerror("Error", "Not connected to SAP2000 model!")
            return

        if not self.current_file:
            messagebox.showwarning("Warning", "Please save your project first!")
            return

        try:
            project_dir = os.path.dirname(self.current_file)

            # Get SELECTED FRAME ELEMENTS
            selected_frame_elements = []
            selected_joints = []

            try:
                ret = self.SapModel.SelectObj.GetSelected()
                num_sel = ret[0]

                if num_sel > 0:
                    obj_types = ret[1]
                    obj_names = ret[2]

                    for i in range(num_sel):
                        obj_type = obj_types[i]
                        obj_name = obj_names[i]
                        
                        print(f"Debug: Selected object: {obj_name}, Type: {obj_type}")
                        
                        try:
                            ret_prop = self.SapModel.FrameObj.GetPoints(obj_name)
                            print(f"Debug: GetPoints result for {obj_name}: {ret_prop}")
                            if ret_prop[2] == 0:
                                start_node = ret_prop[0]
                                
                                section_name = "Unknown"
                                try:
                                    ret_section = self.SapModel.FrameObj.GetSection(obj_name)
                                    print(f"Debug: GetSection result for {obj_name}: {ret_section}")
                                    if len(ret_section) >= 3 and ret_section[2] == 0:
                                        section_name = ret_section[0] if ret_section[0] else "Unknown"
                                        print(f"Debug: Got section name from GetSection: {section_name}")
                                    elif len(ret_section) >= 1 and ret_section[0]:
                                        section_name = ret_section[0]
                                        print(f"Debug: Using section name from index 0: {section_name}")
                                except Exception as e:
                                    print(f"Debug: GetSection failed: {e}")
                                    section_name = "Unknown"
                                
                                print(f"Debug: Final section name for {obj_name}: {section_name}")
                                
                                selected_joints.append(start_node)
                                print(f"Debug: Added frame element {obj_name} with start node {start_node}, section {section_name}")
                            else:
                                print(f"Debug: GetPoints failed for {obj_name}, return code: {ret_prop[2]}")
                        except Exception as e:
                            print(f"Debug: {obj_name} error: {e}")
                            continue
            except Exception as e:
                print(f"SelectObj.GetSelected failed: {e}")

            # Cấu trúc đúng — toàn bộ nằm trong for loop:
            for i in range(num_sel):
                obj_type = obj_types[i]
                obj_name = obj_names[i]

                try:
                    ret_prop = self.SapModel.FrameObj.GetPoints(obj_name)
                    if ret_prop[2] == 0:
                        start_node = ret_prop[0]

                        # Get section name
                        section_name = "Unknown"
                        try:
                            ret_section = self.SapModel.FrameObj.GetSection(obj_name)
                            if len(ret_section) >= 3 and ret_section[2] == 0:
                                section_name = ret_section[0] if ret_section[0] else "Unknown"
                            elif len(ret_section) >= 1 and ret_section[0]:
                                section_name = ret_section[0]
                        except Exception as e:
                            print(f"Debug: GetSection failed: {e}")
                            section_name = "Unknown"

                        # ✅ Get beta — TRONG loop, SAU section_name, TRƯỚC append
                        beta = 0
                        try:
                            ret_axes = self.SapModel.FrameObj.GetLocalAxes(obj_name)
                            if ret_axes[2] == 0:
                                raw_angle = ret_axes[0]
                                beta = 0 if abs(raw_angle % 180) < 45 else 90
                                print(f"Debug: {obj_name} beta = {beta} (raw={raw_angle})")
                        except Exception as e:
                            print(f"Debug: GetLocalAxes failed for {obj_name}: {e}")
                            beta = 0

                        # ✅ Append với beta — TRONG loop
                        selected_frame_elements.append((obj_name, start_node, section_name, beta))
                        selected_joints.append(start_node)
                        print(f"Debug: Added {obj_name}, node={start_node}, section={section_name}, beta={beta}")
                    else:
                        print(f"Debug: GetPoints failed for {obj_name}")
                except Exception as e:
                    print(f"Debug: {obj_name} error: {e}")
                    continue

            if not selected_frame_elements:
                messagebox.showwarning("⚠️ No Beam Elements Selected", 
                    "No beam/column elements detected in selection!\n\n"
                    "Please:\n"
                    "1. Select beam/column elements in SAP2000\n"
                    "2. Do NOT select nodes directly\n"
                    "3. Ensure selection is visible (highlighted)\n"
                    "4. Click 'Load (Auto)' again")
                return

            # Branch for Include RC Pier? = Yes
            if self.include_rc_pier_var.get() == "Yes":
                self._load_with_rc_pier(selected_frame_elements, project_dir)
                return

            selected_joints = list(dict.fromkeys(selected_joints))

            print(f"Debug: Found {len(selected_frame_elements)} selected frame elements")
            print(f"Debug: Extracted {len(selected_joints)} unique start point nodes")

            # Get coordinates for SELECTED joints
            node_to_section = {}
            node_to_beta = {}
            for elem_name, start_node, section_name, beta in selected_frame_elements:
                if start_node not in node_to_section:
                    node_to_section[start_node] = section_name
                    node_to_beta[start_node] = beta
            
            base_coords = []
            for joint_name in selected_joints:
                ret = self.SapModel.PointObj.GetCoordCartesian(joint_name)
                x, y, z = ret[0], ret[1], ret[2]
                section_name = node_to_section.get(joint_name, "Unknown")
                beta = node_to_beta.get(joint_name, 0)
                base_coords.append([joint_name, str(x), str(y), str(z), section_name, str(beta)])

            # Sort by name
            def sort_key(base_name):
                import re
                parts = re.split(r'(\d+)', base_name[0])
                result = []
                for i, part in enumerate(parts):
                    if i % 2 == 0:
                        result.append((0, part))
                    else:
                        result.append((1, int(part)))
                return result
            
            base_coords.sort(key=sort_key)

            # Create bpl_coordinate.csv in Base Plate Design folder
            if not self.bpl_folder:
                self.bpl_folder = os.path.join(project_dir, "Base Plate Design")
                if not os.path.exists(self.bpl_folder):
                    os.makedirs(self.bpl_folder)
            
            coord_file = os.path.join(self.bpl_folder, "bpl_coordinate.csv")
            with open(coord_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Column Base', 'X (m)', 'Y (m)', 'Z (m)', 'Section', 'Beta'])
                for coord in base_coords:
                    writer.writerow(coord)

            messagebox.showinfo("✅ Success", 
                f"Column base coordinates loaded from SAP2000!\n\n"
                f"Selected joints: {len(base_coords)}")

            self.coordinate_status_label.config(
                text=f"✓ Auto-loaded: {len(base_coords)} SELECTED joints",
                fg='#4CAF50'
            )
            self.status_label.config(text=f"● {len(base_coords)} selected column bases loaded", fg='#90ee90')
            
            # Load reaction data for same selected joints
            self._load_reaction_data_for_selected_joints(selected_joints, project_dir)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load column base coordinates:\n{str(e)}")
            self.status_label.config(text="● Auto-load failed", fg='#ff4d4d')

    def _load_reaction_data_for_selected_joints(self, selected_joints, project_dir):
        """Load reaction data for the selected joints"""
        try:
            # Show Load Case Selection Dialog
            selected_lcs = self.show_load_case_selection_dialog()
            if not selected_lcs:
                self.status_label.config(text="● Load case selection cancelled", fg='#ff9800')
                return

            # Confirm re-run analysis
            confirm = messagebox.askyesno("Confirm Analysis",
                f"To retrieve results for the {len(selected_lcs)} selected cases/combos,\n"
                "the output selection needs to be updated and the analysis re-run in SAP2000.\n\n"
                "This may take some time depending on the model size.\n"
                "Continue?")
            if not confirm:
                self.status_label.config(text="● Cancelled by user", fg='#ff9800')
                return

            # Create Progress Dialog
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Loading Reaction Data")
            progress_window.geometry("500x150")
            progress_window.transient(self.root)
            progress_window.grab_set()
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (progress_window.winfo_width() // 2)
            y = (progress_window.winfo_screenheight() // 2) - (progress_window.winfo_height() // 2)
            progress_window.geometry(f"+{x}+{y}")
            tk.Label(progress_window, text="Extracting Reaction Data from SAP2000...", font=('Arial', 12, 'bold')).pack(pady=(20, 10))
            progress_label = tk.Label(progress_window, text="Initializing...", font=('Arial', 10))
            progress_label.pack(pady=5)
            progress_bar = ttk.Progressbar(progress_window, orient='horizontal', length=400, mode='determinate')
            progress_bar.pack(pady=10)
            progress_percent = tk.Label(progress_window, text="0%", font=('Arial', 10, 'bold'), fg='#2196F3')
            progress_percent.pack(pady=5)
            
            def update_progress(value, label_text):
                progress_bar['value'] = value
                progress_label.config(text=label_text)
                progress_percent.config(text=f"{int(value)}%")
                progress_window.update()

            update_progress(5, "Getting load case/combo lists...")
            
            ret_cases = self.SapModel.LoadCases.GetNameList()
            num_cases = ret_cases[0]
            case_names = ret_cases[1] if num_cases > 0 else []

            ret_combos = self.SapModel.RespCombo.GetNameList()
            num_combos = ret_combos[0]
            combo_names = ret_combos[1] if num_combos > 0 else []

            ret_cases = self.SapModel.LoadCases.GetNameList()
            case_set = set(ret_cases[1]) if ret_cases[0] > 0 else set()
            ret_combos = self.SapModel.RespCombo.GetNameList()
            combo_set = set(ret_combos[1]) if ret_combos[0] > 0 else set()
            
            update_progress(10, "Deselecting all cases/combos...")
            self.SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()

            update_progress(15, "Setting selected cases/combos...")
            for lc_name in selected_lcs:
                if lc_name in case_set:
                    self.SapModel.Results.Setup.SetCaseSelectedForOutput(lc_name)
                elif lc_name in combo_set:
                    self.SapModel.Results.Setup.SetComboSelectedForOutput(lc_name)
            
            update_progress(20, "Running analysis (may take time)...")
            ret_analyze = self.SapModel.Analyze.RunAnalysis()
            if ret_analyze != 0:
                progress_window.after(500, progress_window.destroy)
                messagebox.showerror("Error", "Analysis failed in SAP2000! Check model.")
                self.status_label.config(text="● Analysis failed", fg='#ff4d4d')
                return
            
            update_progress(40, "Extracting reaction data...")
            
            reaction_data = []
            reaction_data.append(['TABLE: Joint Reactions'])
            reaction_data.append(['Joint', 'LoadCase', 'CaseType', 'F1', 'F2', 'F3', 'M1', 'M2', 'M3'])
            reaction_data.append(['Text', 'Text', 'Text', 'KN', 'KN', 'KN', 'KN-m', 'KN-m', 'KN-m'])
            
            for node in selected_joints:
                try:
                    ret = self.SapModel.Results.JointReact(node, 0)
                    num_results = ret[0]
                    if num_results > 0:
                        load_cases = ret[3]
                        f1_vals = ret[6]
                        f2_vals = ret[7]
                        f3_vals = ret[8]
                        m1_vals = ret[9]
                        m2_vals = ret[10]
                        m3_vals = ret[11]
                        
                        for i in range(num_results):
                            lc_name = load_cases[i].strip()
                            if lc_name in selected_lcs:
                                case_type_str = "Combination"
                                reaction_data.append([
                                    node,
                                    lc_name,
                                    case_type_str,
                                    f"{f1_vals[i]:.6g}",
                                    f"{f2_vals[i]:.6g}",
                                    f"{f3_vals[i]:.6g}",
                                    f"{m1_vals[i]:.6g}",
                                    f"{m2_vals[i]:.6g}",
                                    f"{m3_vals[i]:.6g}"
                                ])
                                print(f"Debug: Got reaction for {node}, {lc_name}")
                except Exception as e:
                    print(f"Error for joint {node}: {e}")
            
            if len(reaction_data) <= 3:
                progress_window.after(500, progress_window.destroy)
                messagebox.showwarning("No Data",
                    f"No reaction data extracted!\n\n"
                    f"Check:\n- Model analyzed successfully\n- Selected cases have results")
                return
            
            update_progress(70, "Writing CSV file...")
            
            if not self.bpl_folder:
                self.bpl_folder = os.path.join(project_dir, "Base Plate Design")
                if not os.path.exists(self.bpl_folder):
                    os.makedirs(self.bpl_folder)
            
            reaction_csv_file = os.path.join(self.bpl_folder, "reaction_data_sap2000.csv")
            with open(reaction_csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in reaction_data:
                    writer.writerow(row)
            
            self.reaction_csv_file = reaction_csv_file
            
            update_progress(100, "Done!")
            progress_window.after(500, progress_window.destroy)
            
            messagebox.showinfo("✅ Reaction Data Loaded", 
                f"Reaction data extracted successfully!\n\n"
                f"Nodes: {len(selected_joints)}\n"
                f"Load Cases/Combos: {len(selected_lcs)}\n"
                f"Data points: {len(reaction_data)}")
            
            try:
                if platform.system() == 'Windows':
                    os.startfile(reaction_csv_file)
                elif platform.system() == 'Darwin':
                    subprocess.call(['open', reaction_csv_file])
                else:
                    subprocess.call(['xdg-open', reaction_csv_file])
            except:
                pass
            
            self.status_label.config(text=f"● Reaction data loaded: {len(reaction_data)} data points", fg='#90ee90')
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load reaction data:\n{str(e)}")
            self.status_label.config(text="● Reaction load failed", fg='#ff4d4d')

    def _load_with_rc_pier(self, selected_frame_elements, project_dir):
        """Load coordinates and element joint forces when Include RC Pier = Yes.
        
        Algorithm:
        1. Get Start/End node coordinates for all selected elements
        2. Pair elements by shared node (Base Plate Node)
        3. Distinguish: element with lower-Z other node = RC Pier, higher-Z = Steel Column
        4. Write bpl_coordinate.csv using shared node coords + Steel Column section
        5. Extract Element Joint Forces at End node of each RC Pier
        """
        try:
            import re

            # Step 1: Get full node info for all selected elements
            element_info = []  # [(elem_name, start_node, end_node, start_z, end_z, section_name)]
            
            for elem_name, start_node, section_name, beta in selected_frame_elements:
                try:
                    ret_points = self.SapModel.FrameObj.GetPoints(elem_name)
                    if ret_points[2] == 0:
                        s_node = ret_points[0]
                        e_node = ret_points[1]
                        
                        ret_s = self.SapModel.PointObj.GetCoordCartesian(s_node)
                        s_x, s_y, s_z = ret_s[0], ret_s[1], ret_s[2]
                        
                        ret_e = self.SapModel.PointObj.GetCoordCartesian(e_node)
                        e_x, e_y, e_z = ret_e[0], ret_e[1], ret_e[2]

                        ret_axes = self.SapModel.FrameObj.GetLocalAxes(elem_name)
                        raw_angle = ret_axes[0] if ret_axes[2] == 0 else 0
                        beta = 0 if abs(raw_angle % 180) < 45 else 90
                        
                        element_info.append({
                            'name': elem_name,
                            'start_node': s_node,
                            'end_node': e_node,
                            'start_xyz': (s_x, s_y, s_z),
                            'end_xyz': (e_x, e_y, e_z),
                            'section': section_name,
                            'beta': beta
                        })
                        print(f"Debug RC Pier: Element {elem_name}: Start={s_node}(z={s_z:.3f}), End={e_node}(z={e_z:.3f}), Section={section_name}")
                except Exception as e:
                    print(f"Debug RC Pier: Error getting info for {elem_name}: {e}")
                    continue

            if len(element_info) < 2:
                messagebox.showwarning("⚠️ Insufficient Elements", 
                    "Need at least 2 frame elements (RC Pier + Steel Column pair)!\n\n"
                    "Please select pairs of elements:\n"
                    "- RC Pier (below base plate)\n"
                    "- Steel Column (above base plate)")
                return

            # Step 2: Build node-to-elements map for pairing
            node_to_elements = {}  # node_name -> [element_info, ...]
            for elem in element_info:
                for node in [elem['start_node'], elem['end_node']]:
                    if node not in node_to_elements:
                        node_to_elements[node] = []
                    node_to_elements[node].append(elem)

            # Step 3: Find pairs sharing a common node (Base Plate Node)
            pairs = []  # [(pier_elem, column_elem, shared_node)]
            used_elements = set()
            
            for node, elems in node_to_elements.items():
                if len(elems) == 2 and elems[0]['name'] != elems[1]['name']:
                    e1, e2 = elems[0], elems[1]
                    
                    if e1['name'] in used_elements or e2['name'] in used_elements:
                        continue
                    
                    # Find the "other" node (non-shared) Z for each element
                    def get_other_z(elem, shared_node):
                        if elem['start_node'] == shared_node:
                            return elem['end_xyz'][2]  # Z of end node
                        else:
                            return elem['start_xyz'][2]  # Z of start node
                    
                    z1_other = get_other_z(e1, node)
                    z2_other = get_other_z(e2, node)
                    
                    # Element with lower-Z other node = RC Pier
                    if z1_other < z2_other:
                        pier_elem, column_elem = e1, e2
                    else:
                        pier_elem, column_elem = e2, e1
                    
                    pairs.append((pier_elem, column_elem, node))
                    used_elements.add(e1['name'])
                    used_elements.add(e2['name'])
                    
                    print(f"Debug RC Pier: Pair found at node {node}: "
                          f"Pier={pier_elem['name']}(section={pier_elem['section']}), "
                          f"Column={column_elem['name']}(section={column_elem['section']})")

            if not pairs:
                messagebox.showwarning("⚠️ No Pairs Found", 
                    "Could not find RC Pier + Steel Column pairs!\n\n"
                    "Make sure:\n"
                    "1. Each pair shares a common node\n"
                    "2. RC Pier is below, Steel Column is above\n"
                    "3. Select both elements of each pair\n"
                    "4. Don't select node, select element only")
                return

            print(f"Debug RC Pier: Found {len(pairs)} pier-column pairs")

            # Step 4: Create bpl_coordinate.csv using shared node coords + Steel Column section
            base_coords = []
            pier_elements_for_force = []  # [(pier_name, shared_node)]
            
            for pier_elem, column_elem, shared_node in pairs:
                # Get shared node coordinates
                ret_coord = self.SapModel.PointObj.GetCoordCartesian(shared_node)
                x, y, z = ret_coord[0], ret_coord[1], ret_coord[2]
                
                # Section name from Steel Column
                section_name = column_elem['section']
                beta = column_elem.get('beta', 0)
                
                base_coords.append([shared_node, str(x), str(y), str(z), section_name, str(beta)])
                pier_elements_for_force.append((pier_elem['name'], shared_node))
                
                print(f"Debug RC Pier: Base plate at node {shared_node}: "
                      f"({x:.3f}, {y:.3f}, {z:.3f}), Section={section_name}")

            # Sort by name
            def sort_key(base_name):
                parts = re.split(r'(\d+)', base_name[0])
                result = []
                for i, part in enumerate(parts):
                    if i % 2 == 0:
                        result.append((0, part))
                    else:
                        result.append((1, int(part)))
                return result
            
            base_coords.sort(key=sort_key)

            # Write bpl_coordinate.csv
            if not self.bpl_folder:
                self.bpl_folder = os.path.join(project_dir, "Base Plate Design")
                if not os.path.exists(self.bpl_folder):
                    os.makedirs(self.bpl_folder)
            
            coord_file = os.path.join(self.bpl_folder, "bpl_coordinate.csv")
            with open(coord_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['Column Base', 'X (m)', 'Y (m)', 'Z (m)', 'Section', 'Beta'])
                for coord in base_coords:
                    writer.writerow(coord)

            messagebox.showinfo("✅ Success", 
                f"Column base coordinates loaded (RC Pier mode)!\n\n"
                f"Pairs found: {len(pairs)}\n"
                f"File: bpl_coordinate.csv")

            self.coordinate_status_label.config(
                text=f"✓ RC Pier mode: {len(pairs)} pairs loaded",
                fg='#4CAF50'
            )
            self.status_label.config(text=f"● {len(pairs)} pier-column pairs loaded", fg='#90ee90')

            # Step 5: Load Element Joint Forces for RC Pier end nodes
            self._load_elejointforce_for_piers(pier_elements_for_force, project_dir)

        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to load with RC Pier:\n{str(e)}")
            self.status_label.config(text="● RC Pier load failed", fg='#ff4d4d')

    def _load_elejointforce_for_piers(self, pier_elements, project_dir):
        """Load Element Joint Forces at End node of each RC Pier and write to elejointforce.csv.
        
        Args:
            pier_elements: list of (pier_name, shared_node) tuples
            project_dir: project directory path
        """
        try:
            # Show Load Case Selection Dialog
            selected_lcs = self.show_load_case_selection_dialog()
            if not selected_lcs:
                self.status_label.config(text="● Load case selection cancelled", fg='#ff9800')
                return

            # Confirm re-run analysis
            confirm = messagebox.askyesno("Confirm Analysis",
                f"To retrieve Element Joint Forces for {len(pier_elements)} RC Pier elements\n"
                f"and {len(selected_lcs)} selected cases/combos,\n"
                "the analysis needs to be re-run in SAP2000.\n\n"
                "Continue?")
            if not confirm:
                self.status_label.config(text="● Cancelled by user", fg='#ff9800')
                return

            # Create Progress Dialog
            progress_window = tk.Toplevel(self.root)
            progress_window.title("Loading Element Joint Forces")
            progress_window.geometry("500x150")
            progress_window.transient(self.root)
            progress_window.grab_set()
            progress_window.update_idletasks()
            x = (progress_window.winfo_screenwidth() // 2) - (progress_window.winfo_width() // 2)
            y = (progress_window.winfo_screenheight() // 2) - (progress_window.winfo_height() // 2)
            progress_window.geometry(f"+{x}+{y}")
            tk.Label(progress_window, text="Extracting Element Joint Forces from SAP2000...", font=('Arial', 12, 'bold')).pack(pady=(20, 10))
            progress_label = tk.Label(progress_window, text="Initializing...", font=('Arial', 10))
            progress_label.pack(pady=5)
            progress_bar = ttk.Progressbar(progress_window, orient='horizontal', length=400, mode='determinate')
            progress_bar.pack(pady=10)
            progress_percent = tk.Label(progress_window, text="0%", font=('Arial', 10, 'bold'), fg='#2196F3')
            progress_percent.pack(pady=5)
            
            def update_progress(value, label_text):
                progress_bar['value'] = value
                progress_label.config(text=label_text)
                progress_percent.config(text=f"{int(value)}%")
                progress_window.update()

            update_progress(5, "Getting load case/combo lists...")
            
            ret_cases = self.SapModel.LoadCases.GetNameList()
            case_set = set(ret_cases[1]) if ret_cases[0] > 0 else set()
            ret_combos = self.SapModel.RespCombo.GetNameList()
            combo_set = set(ret_combos[1]) if ret_combos[0] > 0 else set()
            
            update_progress(10, "Deselecting all cases/combos...")
            self.SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()

            update_progress(15, "Setting selected cases/combos...")
            for lc_name in selected_lcs:
                if lc_name in case_set:
                    self.SapModel.Results.Setup.SetCaseSelectedForOutput(lc_name)
                elif lc_name in combo_set:
                    self.SapModel.Results.Setup.SetComboSelectedForOutput(lc_name)
            
            update_progress(20, "Running analysis (may take time)...")
            ret_analyze = self.SapModel.Analyze.RunAnalysis()
            if ret_analyze != 0:
                progress_window.after(500, progress_window.destroy)
                messagebox.showerror("Error", "Analysis failed in SAP2000! Check model.")
                self.status_label.config(text="● Analysis failed", fg='#ff4d4d')
                return
            
            update_progress(40, "Extracting element joint forces...")
            
            # Build CSV data
            force_data = []
            force_data.append(['TABLE: Element Joint Forces - Pier End'])
            force_data.append(['Joint', 'PierElement', 'LoadCase', 'CaseType', 'F1', 'F2', 'F3', 'M1', 'M2', 'M3'])
            force_data.append(['Text', 'Text', 'Text', 'Text', 'KN', 'KN', 'KN', 'KN-m', 'KN-m', 'KN-m'])
            
            total_piers = len(pier_elements)
            for idx, (pier_name, shared_node) in enumerate(pier_elements):
                progress_pct = 40 + (idx / total_piers) * 50
                update_progress(progress_pct, f"Processing pier {pier_name} ({idx+1}/{total_piers})...")
                
                try:
                    # FrameJointForce returns forces at element joints (2 per element: start & end)
                    # Parameters: (Name, ItemTypeElm) where 0 = ObjectElm
                    ret = self.SapModel.Results.FrameJointForce(pier_name, 0)
                    num_results = ret[0]
                    
                    print(f"Debug RC Pier: FrameJointForce for {pier_name}: {num_results} results")
                    
                    if num_results > 0:
                        obj_names = ret[1]   # Object names
                        elm_names = ret[2]   # Element names  
                        point_elms = ret[3]  # Point element names (joint names)
                        load_cases = ret[4]  # Load case names
                        step_types = ret[5]  # Step types
                        step_nums = ret[6]   # Step numbers
                        f1_vals = ret[7]     # F1 values
                        f2_vals = ret[8]     # F2 values
                        f3_vals = ret[9]     # F3 values
                        m1_vals = ret[10]    # M1 values
                        m2_vals = ret[11]    # M2 values
                        m3_vals = ret[12]    # M3 values
                        
                        for i in range(num_results):
                            lc_name = load_cases[i].strip()
                            joint_name = point_elms[i].strip() if point_elms[i] else ""
                            
                            # Only keep results at END node (shared node = base plate node)
                            if joint_name == shared_node and lc_name in selected_lcs:
                                case_type_str = "Combination" if lc_name in combo_set else "LinStatic"
                                force_data.append([
                                    shared_node,
                                    pier_name,
                                    lc_name,
                                    case_type_str,
                                    f"{f1_vals[i]:.6g}",
                                    f"{f2_vals[i]:.6g}",
                                    f"{f3_vals[i]:.6g}",
                                    f"{m1_vals[i]:.6g}",
                                    f"{m2_vals[i]:.6g}",
                                    f"{m3_vals[i]:.6g}"
                                ])
                                print(f"Debug RC Pier: Got force for {pier_name} at {shared_node}, LC={lc_name}")
                                
                except Exception as e:
                    print(f"Error for pier {pier_name}: {e}")
                    import traceback
                    traceback.print_exc()
            
            if len(force_data) <= 3:
                progress_window.after(500, progress_window.destroy)
                messagebox.showwarning("No Data",
                    f"No element joint force data extracted!\n\n"
                    f"Check:\n- Model analyzed successfully\n- Selected cases have results\n"
                    f"- RC Pier elements are correctly identified")
                return
            
            update_progress(95, "Writing CSV file...")
            
            if not self.bpl_folder:
                self.bpl_folder = os.path.join(project_dir, "Base Plate Design")
                if not os.path.exists(self.bpl_folder):
                    os.makedirs(self.bpl_folder)
            
            force_csv_file = os.path.join(self.bpl_folder, "elejointforce.csv")
            with open(force_csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in force_data:
                    writer.writerow(row)
            
            # Map elejointforce → reaction_data_sap2000.csv (drop PierElement, negate forces)
            reaction_data = []
            reaction_data.append(['TABLE: Joint Reactions'])
            reaction_data.append(['Joint', 'LoadCase', 'CaseType', 'F1', 'F2', 'F3', 'M1', 'M2', 'M3'])
            reaction_data.append(['Text', 'Text', 'Text', 'KN', 'KN', 'KN', 'KN-m', 'KN-m', 'KN-m'])
            
            for row in force_data[3:]:  # Skip 3 header rows
                # row = [Joint, PierElement, LoadCase, CaseType, F1, F2, F3, M1, M2, M3]
                joint = row[0]
                load_case = row[2]   # Col C → Col B
                case_type = row[3]   # Col D → Col C
                # Negate forces: Col E-J → Col D-I
                f1 = f"{-float(row[4]):.6g}"
                f2 = f"{-float(row[5]):.6g}"
                f3 = f"{-float(row[6]):.6g}"
                m1 = f"{-float(row[7]):.6g}"
                m2 = f"{-float(row[8]):.6g}"
                m3 = f"{-float(row[9]):.6g}"
                reaction_data.append([joint, load_case, case_type, f1, f2, f3, m1, m2, m3])
            
            reaction_csv_file = os.path.join(self.bpl_folder, "reaction_data_sap2000.csv")
            with open(reaction_csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in reaction_data:
                    writer.writerow(row)
            
            self.reaction_csv_file = reaction_csv_file
            
            update_progress(100, "Done!")
            progress_window.after(500, progress_window.destroy)
            
            data_point_count = len(force_data) - 3  # Exclude header rows
            messagebox.showinfo("✅ Element Joint Forces Loaded", 
                f"Element Joint Forces extracted successfully!\n\n"
                f"RC Pier elements: {len(pier_elements)}\n"
                f"Load Cases/Combos: {len(selected_lcs)}\n"
                f"Data points: {data_point_count}")
                #f"Files created in Base Plate Design folder:\n"
                #f"- elejointforce.csv (raw data, reference)\n"
                #f"- reaction_data_sap2000.csv (mapped, sign-reversed)")
            
            try:
                if platform.system() == 'Windows':
                    os.startfile(reaction_csv_file)
                elif platform.system() == 'Darwin':
                    subprocess.call(['open', reaction_csv_file])
                else:
                    subprocess.call(['xdg-open', reaction_csv_file])
            except:
                pass
            
            self.status_label.config(text=f"● Reaction data loaded (RC Pier): {data_point_count} data points", fg='#90ee90')
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to load element joint forces:\n{str(e)}")
            self.status_label.config(text="● Element joint force load failed", fg='#ff4d4d')

    def show_load_case_selection_dialog(self):
        """Show dialog to select Load Cases/Combinations from SAP2000 model"""
        if not self.SapModel:
            return None
        try:
            ret_cases = self.SapModel.LoadCases.GetNameList()
            num_cases = ret_cases[0]
            case_names = ret_cases[1] if num_cases > 0 else []

            ret_combos = self.SapModel.RespCombo.GetNameList()
            num_combos = ret_combos[0]
            combo_names = ret_combos[1] if num_combos > 0 else []

            if num_cases == 0 and num_combos == 0:
                messagebox.showwarning("No Load Cases", "No Load Cases or Combinations found in model!")
                return None

            import re
            def sort_key(name):
                num_part = re.search(r'\d+', name)
                return (int(num_part.group()) if num_part else float('inf'), name)

            case_names = sorted(case_names, key=sort_key)
            combo_names = sorted(combo_names, key=sort_key)

            dialog = tk.Toplevel(self.root)
            dialog.title("Select Load Cases/Combinations")
            dialog.geometry("900x550")
            dialog.transient(self.root)
            dialog.grab_set()
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
            y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
            dialog.geometry(f"+{x}+{y}")

            tk.Label(dialog, text="Select Load Cases/Combinations for Analysis", font=('Arial', 12, 'bold'), fg='#1a472a').pack(pady=(10, 3))

            case_vars = {}
            combo_vars = {}

            # Top: Deselect All
            btn_top_frame = tk.Frame(dialog)
            btn_top_frame.pack(pady=3)
            def deselect_all():
                for var in case_vars.values():
                    var.set(False)
                for var in combo_vars.values():
                    var.set(False)
            tk.Button(btn_top_frame, text="✗ Deselect All", command=deselect_all, bg='#f44336', fg='white', font=('Arial', 9, 'bold'), cursor='hand2').pack()

            # Helper functions
            def get_prefix_digit(name):
                num_part = re.search(r'\d+', name)
                if num_part and num_part.group():
                    return num_part.group()[0]
                return None

            def parse_prefix_input(prefix_input):
                if not prefix_input.strip():
                    return set()
                ranges = []
                for part in prefix_input.strip().split(','):
                    part = part.strip()
                    if '-' in part:
                        try:
                            start, end = map(int, part.split('-'))
                            ranges.extend(range(start, end + 1))
                        except:
                            pass
                    else:
                        try:
                            ranges.append(int(part))
                        except:
                            pass
                return set(str(d) for d in ranges)

            selection_state = {'last_clicked_index': None, 'last_type': None}

            # ==================== TWO-COLUMN LAYOUT ====================
            columns_frame = tk.Frame(dialog)
            columns_frame.pack(fill='both', expand=True, padx=10, pady=5)
            columns_frame.columnconfigure(0, weight=3)
            columns_frame.columnconfigure(1, weight=2)

            def create_lc_column(parent, col_idx, title, title_bg, title_fg, names, var_dict, item_type):
                col_frame = tk.Frame(parent, bg='white', bd=1, relief='groove')
                col_frame.grid(row=0, column=col_idx, sticky='nsew', padx=(0 if col_idx == 0 else 5, 0))
                
                header = tk.Frame(col_frame, bg=title_bg)
                header.pack(fill='x')
                tk.Label(header, text=title, font=('Arial', 10, 'bold'), 
                        bg=title_bg, fg=title_fg).pack(side='left', padx=5, pady=4)
                
                ctrl = tk.Frame(col_frame, bg='#f5f5f5')
                ctrl.pack(fill='x', padx=3, pady=3)
                
                tk.Label(ctrl, text="Prefix:", font=('Arial', 8), bg='#f5f5f5').pack(side='left', padx=2)
                prefix_entry = tk.Entry(ctrl, width=8, font=('Arial', 9))
                prefix_entry.pack(side='left', padx=2)
                
                def apply_prefix():
                    prefix_set = parse_prefix_input(prefix_entry.get())
                    if prefix_set:
                        for name, var in var_dict.items():
                            if get_prefix_digit(name) in prefix_set:
                                var.set(True)
                def select_all():
                    for var in var_dict.values():
                        var.set(True)
                def select_none():
                    for var in var_dict.values():
                        var.set(False)
                
                tk.Button(ctrl, text="Apply", command=apply_prefix, bg='#2196F3', fg='white', 
                         font=('Arial', 8, 'bold'), cursor='hand2').pack(side='left', padx=2)
                tk.Button(ctrl, text="✓ All", command=select_all, bg='#4CAF50', fg='white', 
                         font=('Arial', 8, 'bold'), cursor='hand2').pack(side='left', padx=2)
                tk.Button(ctrl, text="✗ None", command=select_none, bg='#FF9800', fg='white', 
                         font=('Arial', 8, 'bold'), cursor='hand2').pack(side='left', padx=2)
                tk.Label(ctrl, text=f"({len(names)})", font=('Arial', 8), bg='#f5f5f5', fg='#999').pack(side='left', padx=3)
                
                scroll_container = tk.Frame(col_frame, bg='white')
                scroll_container.pack(fill='both', expand=True, padx=2, pady=2)
                
                canvas = tk.Canvas(scroll_container, bg='white', highlightthickness=0)
                scrollbar = ttk.Scrollbar(scroll_container, orient='vertical', command=canvas.yview)
                scrollable = tk.Frame(canvas, bg='white')
                
                scrollable.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
                canvas.create_window((0, 0), window=scrollable, anchor='nw')
                canvas.configure(yscrollcommand=scrollbar.set)
                
                def _on_mw(event):
                    canvas.yview_scroll(int(-1*(event.delta/120)), "units")
                def _bind_mw(event):
                    canvas.bind_all("<MouseWheel>", _on_mw)
                def _unbind_mw(event):
                    canvas.unbind_all("<MouseWheel>")
                canvas.bind("<Enter>", _bind_mw)
                canvas.bind("<Leave>", _unbind_mw)
                
                canvas.pack(side='left', fill='both', expand=True)
                scrollbar.pack(side='right', fill='y')
                
                for idx, name in enumerate(names):
                    var = tk.BooleanVar(value=False)
                    var_dict[name] = var
                    cb = tk.Checkbutton(scrollable, text=f" {name}", variable=var, 
                                       font=('Arial', 10), bg='white', anchor='w')
                    
                    def on_click(event, index=idx, item_name=name, itype=item_type):
                        if event.state & 0x1:
                            last_idx = selection_state['last_clicked_index']
                            if last_idx is not None and selection_state['last_type'] == itype:
                                start, end = min(last_idx, index), max(last_idx, index)
                                for i in range(start, end + 1):
                                    var_dict[names[i]].set(True)
                            else:
                                var_dict[item_name].set(True)
                        else:
                            var_dict[item_name].set(not var_dict[item_name].get())
                        selection_state['last_clicked_index'] = index
                        selection_state['last_type'] = itype
                        return 'break'
                    
                    cb.pack(fill='x', padx=5, pady=1)
                    cb.bind('<Button-1>', on_click)
                
                if not names:
                    tk.Label(scrollable, text="(None)", font=('Arial', 9, 'italic'), 
                            bg='white', fg='#ccc').pack(pady=20)

            # LEFT: Load Combinations (more commonly used)
            create_lc_column(columns_frame, 0, 
                           "🔗 Load Combinations", '#FFF3E0', '#E65100',
                           combo_names, combo_vars, 'combo')
            
            # RIGHT: Load Cases
            create_lc_column(columns_frame, 1, 
                           "📋 Load Cases", '#E8F5E9', '#2E7D32',
                           case_names, case_vars, 'case')

            tk.Label(dialog, text=f"Total: {num_cases} Load Cases, {num_combos} Combinations | Shift+Click for range", 
                    font=('Arial', 9, 'italic'), fg='#666666').pack(pady=3)

            result = {'selected': None}
            def on_ok():
                selected = []
                for name, var in case_vars.items():
                    if var.get():
                        selected.append(name)
                for name, var in combo_vars.items():
                    if var.get():
                        selected.append(name)
                if not selected:
                    messagebox.showwarning("No Selection", "Please select at least one Load Case or Combination!")
                    return
                result['selected'] = selected
                dialog.destroy()

            def on_cancel():
                dialog.destroy()

            btn_frame = tk.Frame(dialog)
            btn_frame.pack(pady=10)
            tk.Button(btn_frame, text="OK", command=on_ok, bg='#4CAF50', fg='white', font=('Arial', 10, 'bold'), width=12, cursor='hand2').pack(side='left', padx=10)
            tk.Button(btn_frame, text="Cancel", command=on_cancel, bg='#f44336', fg='white', font=('Arial', 10, 'bold'), width=12, cursor='hand2').pack(side='left', padx=10)

            dialog.wait_window()
            return result['selected']
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get Load Cases:\n{str(e)}")
            return None


    def export_to_excel(self):
        """Export calculation reports for checked nodes to Calculation Report.xlsx"""
        # Check if any nodes have been checked
        checked_nodes = [name for name, data in self.base_plate_nodes.items() 
                        if data['design_status'] in ['OK', 'NG']]
        
        if not checked_nodes:
            messagebox.showwarning("No Data", 
                "Please run design check first!\n\n"
                "Steps:\n"
                "1. Select nodes in Base Plate Detail\n"
                "2. Click 'RUN DESIGN CHECK'\n"
                "3. Then export")
            return
        
        # Create selection dialog for nodes
        self._show_export_dialog(checked_nodes)
    
    def _show_export_dialog(self, checked_nodes):
        """Show dialog to select nodes for export"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Export to Excel")
        dialog.geometry("500x650")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Header
        header_frame = tk.Frame(dialog, bg='#f0f0f0')
        header_frame.pack(fill='x', padx=15, pady=(15, 10))
        tk.Label(header_frame, text="Select nodes to export:", 
                font=('Arial', 12, 'bold'), bg='#f0f0f0', fg='#1a472a').pack(side='left')
        tk.Label(header_frame, text=f"({len(checked_nodes)} nodes available)", 
                font=('Arial', 10), bg='#f0f0f0', fg='#666').pack(side='left', padx=(15, 0))
        
        # Scrollable frame for checkboxes
        scroll_container = tk.Frame(dialog, bg='white')
        scroll_container.pack(fill='both', expand=True, padx=15, pady=10)
        
        canvas = tk.Canvas(scroll_container, bg='white', highlightthickness=0, height=350)
        scrollbar = ttk.Scrollbar(scroll_container, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind('<Configure>', 
                             lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Mouse wheel scroll
        def _on_mousewheel(event):
            if canvas.winfo_containing(event.x_root, event.y_root) == canvas:
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind("<MouseWheel>", _on_mousewheel)
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Checkboxes
        node_vars = {}
        for node_name in sorted(checked_nodes):
            var = tk.BooleanVar(value=True)  # Default: all selected
            node_vars[node_name] = var
            
            # Node info
            node_data = self.base_plate_nodes[node_name]
            status_color = '#4CAF50' if node_data['design_status'] == 'OK' else '#f44336'
            status_text = f"[{node_data['design_status']}]"
            ratio_text = f"(R: {node_data.get('max_ratio', '?'):.2f})" if node_data.get('max_ratio') else ""
            
            cb_frame = tk.Frame(scrollable_frame, bg='white')
            cb_frame.pack(fill='x', padx=10, pady=4)
            
            cb = tk.Checkbutton(cb_frame, text=node_name, variable=var, 
                               font=('Arial', 10), bg='white', anchor='w')
            cb.pack(side='left')
            
            status_label = tk.Label(cb_frame, text=status_text, 
                                   font=('Arial', 9, 'bold'), bg='white', fg=status_color)
            status_label.pack(side='left', padx=(10, 5))
            
            ratio_label = tk.Label(cb_frame, text=ratio_text, 
                                  font=('Arial', 9), bg='white', fg='#666')
            ratio_label.pack(side='left')
        
        # Buttons
        def select_all():
            for var in node_vars.values():
                var.set(True)
        
        def deselect_all():
            for var in node_vars.values():
                var.set(False)
        
        def export_selected():
            selected = [name for name, var in node_vars.items() if var.get()]
            
            if not selected:
                messagebox.showwarning("No Selection", "Please select at least one node")
                return
            
            dialog.destroy()
            self._create_calculation_report(selected)
        
        # Button frame
        btn_frame = tk.Frame(dialog, bg='white')
        btn_frame.pack(fill='x', pady=(10, 15), padx=15)
        
        # Top buttons
        select_frame = tk.Frame(btn_frame, bg='white')
        select_frame.pack(fill='x', pady=(0, 10))
        
        tk.Button(select_frame, text="✓ All", command=select_all,
                 bg='#2196F3', fg='white', font=('Arial', 9, 'bold'),
                 width=12, cursor='hand2').pack(side='left', padx=5)
        tk.Button(select_frame, text="✗ None", command=deselect_all,
                 bg='#FF9800', fg='white', font=('Arial', 9, 'bold'),
                 width=12, cursor='hand2').pack(side='left', padx=5)
        
        # Bottom buttons
        bottom_frame = tk.Frame(btn_frame, bg='white')
        bottom_frame.pack(fill='x')
        
        tk.Button(bottom_frame, text="✓ Export", command=export_selected,
                 bg='#4CAF50', fg='white', font=('Arial', 10, 'bold'),
                 width=15, cursor='hand2').pack(side='left', padx=5)
        tk.Button(bottom_frame, text="✗ Cancel", command=dialog.destroy,
                 bg='#f44336', fg='white', font=('Arial', 10, 'bold'),
                 width=15, cursor='hand2').pack(side='left', padx=5)
    
    def _create_calculation_report(self, selected_nodes):
        """Create Calculation Report.xlsx using Template.xlsx"""
        try:
            # 1. Find Template.xlsx in bpl_folder or script directory
            template_path = os.path.join(self.bpl_folder, 'Template.xlsx')
            if not os.path.exists(template_path):
                script_dir = os.path.dirname(os.path.abspath(__file__))
                template_src = os.path.join(script_dir, 'Template.xlsx')
                if os.path.exists(template_src):
                    shutil.copy2(template_src, template_path)
                else:
                    messagebox.showerror("Error", 
                        "Template.xlsx not found!\n\n"
                        "Please place Template.xlsx in:\n"
                        f"1. {self.bpl_folder}\n"
                        f"2. {script_dir}")
                    return
            
            # 2. Copy Template.xlsx → Calculation Report.xlsx
            report_file = os.path.join(self.bpl_folder, 'Calculation Report.xlsx')
            shutil.copy2(template_path, report_file)
            
            # 3. Open workbooks
            report_wb = openpyxl.load_workbook(report_file)
            hinge_file = os.path.join(self.bpl_folder, 'Hinge Type.xlsx')
            if not os.path.exists(hinge_file):
                messagebox.showerror("Error", "Hinge Type.xlsx not found!")
                report_wb.close()
                return
            hinge_wb = openpyxl.load_workbook(hinge_file)
            
            # 4. Get template sheet (first sheet)
            template_ws = report_wb.worksheets[0]
            
            # 5. Create copies of template sheet for additional nodes
            for i in range(1, len(selected_nodes)):
                new_ws = report_wb.copy_worksheet(template_ws)
                
                # Copy drawing (images) from template sheet to new sheet
                try:
                   if hasattr(template_ws, '_drawing') and template_ws._drawing:
                       # Deep copy the drawing object
                       import copy
                       new_ws._drawing = copy.deepcopy(template_ws._drawing)
                       if new_ws._drawing:
                           new_ws._drawing.anchor = new_ws
                except Exception as e:
                   print(f"Warning: Could not copy images: {e}")
            
            # 6. Rename sheets to node names
            for i, node_name in enumerate(selected_nodes):
                report_wb.worksheets[i].title = node_name
            
            # 7. Map data for each node from Hinge Type.xlsx
            for i, node_name in enumerate(selected_nodes):
                ws = report_wb.worksheets[i]
                hinge_sheet_name = f"Node {node_name}"
                
                if hinge_sheet_name not in hinge_wb.sheetnames:
                    print(f"Warning: Sheet '{hinge_sheet_name}' not found in Hinge Type.xlsx")
                    continue
                
                hs = hinge_wb[hinge_sheet_name]
                
                # Map Page 1 data (rows 1-44)
                self._map_page1_data(ws, hs)
                
                # Map Page 2 data (rows 46-85)
                self._map_page2_data(ws, hs)
                
                # Map Page 3 data (rows 87-134)
                self._map_page3_data(ws, hs)
                
                # Map Page 4 data (rows 135-167)
                self._map_page4_data(ws, hs)
                
                # Map Page 5 data (rows 168-212)
                self._map_page5_data(ws, hs)
                
                # Map Page 6 data (rows 213-247)
                self._map_page6_data(ws, hs)
                
                # Apply formatting: Page Break Preview + Hide Gridlines
                ws.sheet_view.view = 'pageBreakPreview'  # Enable Page Break Preview
                ws.sheet_view.showGridLines = False  # Hide gridlines
                
                print(f"Exported: Node {node_name}")
            
            hinge_wb.close()
            
            # 8. Create Summary sheet at position 0
            self._create_summary_sheet(report_wb, selected_nodes)
            
            # 9. Save
            report_wb.save(report_file)
            report_wb.close()
            
            # Show success message
            messagebox.showinfo("Export Success!", 
                f"Calculation Report created!\n\n"
                f"File: Calculation Report.xlsx\n"
                f"Location: {self.bpl_folder}\n\n"
                f"Nodes exported: {len(selected_nodes)}\n"
                f"Total sheets: {len(selected_nodes) + 1} (Summary + {len(selected_nodes)} nodes)")
            
            # Open file
            try:
                if platform.system() == 'Windows':
                    os.startfile(report_file)
                elif platform.system() == 'Darwin':
                    subprocess.call(['open', report_file])
                else:
                    subprocess.call(['xdg-open', report_file])
            except:
                pass
            
            self.status_label.config(text=f"● Export complete: {len(selected_nodes)} nodes", fg='#90ee90')
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to create report:\n{str(e)}")
            print(f"Export error: {e}")
            import traceback
            traceback.print_exc()
    
    def _map_page1_data(self, ws, hs):
        """Map Page 1 data (rows 1-44) from Hinge Type sheet to Template"""
        
        def sv(value):
            """Safe value - convert to number if possible, return '' for None"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, 6)
            except:
                return value
        
        # ===== 1. Column Section (Row 6-10) =====
        ws['J6'] = hs['A1'].value or ''      # Section name e.g. "H340X250X9X14"
        ws['K7'] = sv(hs['A3'].value)         # d (depth)
        ws['K8'] = sv(hs['B3'].value)         # bf (width)
        ws['K9'] = sv(hs['C3'].value)         # tw (web)
        ws['K10'] = sv(hs['D3'].value)        # tf (flange)
        
        # ===== 2. Base Plate Detail (Row 12-19) =====
        n_val = sv(hs['F3'].value)
        b_val = sv(hs['G3'].value)
        p1_val = sv(hs['H3'].value)
        ws['J12'] = f"PL-{n_val}X{b_val}X{p1_val}" if all([n_val != '', b_val != '', p1_val != '']) else ''
        ws['K13'] = n_val                     # N (length)
        ws['K14'] = b_val                     # B (width)
        ws['K15'] = p1_val                    # P1 (thickness)
        ws['K16'] = sv(hs['I3'].value)        # C (bolt spacing)
        ws['K17'] = sv(hs['J3'].value)        # A (bolt spacing)
        ws['K18'] = sv(hs['L3'].value)        # F (edge distance)
        ws['K19'] = sv(hs['K3'].value)        # E (edge distance)
        
        # ===== 3. Anchor Bolt Detail (Row 21-28) =====
        nb = sv(hs['F7'].value)
        db = sv(hs['G7'].value)
        lo = sv(hs['J7'].value)
        ws['J21'] = f"{nb}-M{db}, Lo={lo} mm" if all([nb != '', db != '', lo != '']) else ''
        ws['K22'] = nb                        # nb
        ws['K23'] = db                        # db
        ws['K24'] = sv(hs['H7'].value)        # Ase
        ws['K25'] = sv(hs['I7'].value)        # futa
        ws['K26'] = lo                        # Lo (A1)
        ws['K27'] = sv(hs['K7'].value)        # proj
        ws['K28'] = sv(hs['L7'].value)        # hef
        
        # ===== 4. Material (Row 31-35) =====
        ws['K31'] = sv(hs['A7'].value)        # fy (steel)
        ws['K32'] = sv(hs['B7'].value)        # f'c
        ws['K33'] = sv(hs['F15'].value)       # μ (friction) from Other Coefficient
        ws['K34'] = sv(hs['C7'].value)        # fy (mainbar)
        ws['K35'] = sv(hs['D7'].value)        # fy (tiebar)
        
        # ===== 5. Pier Detail (Row 31-34 right side) =====
        ws['AB31'] = sv(hs['A11'].value)      # Np
        ws['AB32'] = sv(hs['B11'].value)      # Bp
        ws['AB33'] = sv(hs['C11'].value)      # TG
        ws['AB34'] = sv(hs['D11'].value)      # Cover c
        
        # ===== 6. Control Ratios - Base Plate Design =====
        rat1 = sv(hs['B28'].value)
        rat2 = sv(hs['B38'].value)
        rat3 = sv(hs['B51'].value)
        rat6 = sv(hs['B106'].value)
        
        ws['J40'] = 'NA' if rat1 == 0 else hs['O2'].value or ''       # L/C Compression
        ws['P40'] = rat1        # rat1 (Concrete Bearing)
        ws['J41'] = 'NA' if rat2 == 0 else hs['O2'].value or ''       # L/C Compression (same case)
        ws['P41'] = rat2        # rat2 (Plate Yielding C)
        ws['J42'] = 'NA' if rat3 == 0 else hs['O3'].value or ''       # L/C Tensile
        ws['P42'] = rat3        # rat3 (Plate Yielding T)
        ws['C44'] = hs['B107'].value or ''     # Control type text
        ws['J44'] = 'NA' if rat6 == 0 else hs['O4'].value or ''       # L/C Anchor Bolt
        ws['P44'] = rat6        # rat6 (Anchor Bolt interaction)
        
        # ===== 7. Control Ratios - Pier Reinforcement =====
        rat7 = sv(hs['B121'].value)
        rat8 = sv(hs['B125'].value)
        rat9 = sv(hs['B146'].value)
        rat10 = sv(hs['B162'].value)
        
        ws['Y40'] = 'NA' if rat7 == 0 else hs['O3'].value or ''       # L/C Main Bar (Tensile case)
        ws['AE40'] = rat7       # rat7 (Main Bar Area)
        ws['Y41'] = 'NA' if rat8 == 0 else hs['O3'].value or ''       # L/C Dev Length (Tensile case)
        ws['AE41'] = rat8       # rat8 (Development Length)
        ws['Y43'] = 'NA' if rat9 == 0 else hs['O5'].value or ''       # L/C Tie Bar X-Dir
        ws['AE43'] = rat9       # rat9 (Tie Bar Area X)
        ws['Y44'] = 'NA' if rat10 == 0 else hs['O6'].value or ''      # L/C Tie Bar Y-Dir
        ws['AE44'] = rat10      # rat10 (Tie Bar Area Y)
    
    def _map_page2_data(self, ws, hs):
        """Map Page 2 data (rows 46-85): Base Plate Design - Bearing & Yielding"""
        
        def sv(value, dp=4):
            """Safe value - convert to number if possible"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, dp)
            except:
                return value
        
        # ===== 1.1 Design Requirements for Bearing =====
        # Row 49: Control case
        ws['I49'] = hs['O2'].value or ''         # L/C (Compression control case)
        
        # Row 50: Vertical load
        ws['I50'] = sv(hs['B15'].value)          # Vertical load (from Hinge row 15)
        
        # Row 57-58: A1, A2
        ws['M57'] = sv(hs['B18'].value)          # A1 = N x B (M57)
        ws['M58'] = sv(hs['B19'].value)          # A2 = Np x Bp (M58)
        
        # Row 60-61: Pp (AISC)
        ws['M60'] = sv(hs['B20'].value)          # Pp (AISC) (M60)
        ws['M61'] = sv(hs['B21'].value)          # 1.7f'cA1 (M61)
        ws['D61'] = '=IF(M60<=M61,"≤",">")'     # Comparison: Pp <= 1.7f\'cA1
        
        # Row 65: Pp (ACI)
        ws['E65'] = sv(hs['B24'].value)          # Pp (ACI) (E65)
        
        # Row 70-71: Pn, PhiPn
        ws['K70'] = sv(hs['B26'].value)          # Pn
        ws['K71'] = sv(hs['B27'].value)          # PhiPn
        ws['O71'] = '=IF(K71<P71,"<",">")'      # Comparison: PhiPn < Pu
        ws['P71'] = sv(hs['B15'].value)          # Pu (P71 for comparison)
        
        # Row 71: OK/NG check - based on Ratio 1
        try:
            rat1 = float(hs['B28'].value) if hs['B28'].value else 0
            ws['U71'] = 'OK' if rat1 < 1.0 else 'NG'
        except:
            ws['U71'] = ''
        
        # Row 71: Ratio1
        ws['AE71'] = 'Ratio1 ='
        ws['AF71'] = sv(hs['B28'].value)         # rat1
        
        # ===== Base Plate Yielding Limit =====
        # Row 75-77: m, n, n' - set to 2 decimal places
        ws['K75'] = sv(hs['B32'].value, dp=2)    # m
        ws['K76'] = sv(hs['B33'].value, dp=2)    # n
        ws['K77'] = sv(hs['B34'].value, dp=2)    # n'
        
        # Row 79: l = max(m, n, n')
        ws['K79'] = sv(hs['B35'].value, dp=2)    # l
        
        # Row 84: tmin, P1, OK/NG, Ratio2
        ws['K84'] = sv(hs['B37'].value, dp=2)    # tmin
        ws['S84'] = sv(hs['H3'].value)           # P1 (plate thickness)
        ws['P84'] = '=IF(K84<=S84,"≤",">")'     # Comparison: tmin <= P1
        
        # OK/NG check - based on Ratio 2
        try:
            rat2 = float(hs['B38'].value) if hs['B38'].value else 0
            ws['X84'] = 'OK' if rat2 < 1.0 else 'NG'
        except:
            ws['X84'] = ''
        
        ws['AE84'] = 'Ratio2 ='
        ws['AF84'] = sv(hs['B38'].value)         # rat2
    
    def _map_page3_data(self, ws, hs):
        """Map Page 3 data (rows 87-134): Tensile Loading & AB Design (Tension)"""
        
        def sv(value, dp=4):
            """Safe value - convert to number if possible"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, dp)
            except:
                return value
        
        # ===== 1.2 Design Requirements for tensile loading =====
        ws['I88'] = hs['O3'].value or ''         # Control Case
        ws['I89'] = sv(hs['B43'].value)          # Vertical Load (F3)
        ws['L93'] = sv(hs['B46'].value)          # Nua = Vert.Load / nb
        ws['L94'] = sv(hs['B47'].value)          # Mu
        ws['L97'] = sv(hs['B48'].value)          # beff
        ws['J99'] = 0.9                         # Resistance Factor (J99-K99)
        ws['K102'] = sv(hs['B50'].value, dp=2)   # treq'd
        ws['S102'] = sv(hs['H3'].value)          # P1
        ws['P102'] = '=IF(K102<=S102,"≤",">")'  # Comparison: treq'd <= P1
        
        # OK/NG check for Ratio 3
        try:
            rat3 = float(hs['B51'].value) if hs['B51'].value else 0
            ws['X102'] = 'OK' if rat3 < 1.0 else 'NG'
        except:
            ws['X102'] = ''
        ws['AF102'] = sv(hs['B51'].value)        # Ratio3
        
        # ===== 2. Anchor Bolt Design (Tensile resistance) =====
        ws['I106'] = hs['O4'].value or ''        # Control case
        ws['I107'] = sv(hs['B56'].value)         # Vertical load
        ws['I108'] = sv(hs['B57'].value)         # Horizontal load (X Dir.)
        ws['I109'] = sv(hs['B58'].value)         # Horizontal load (Y Dir.)
        
        # 2.1 Design Requirements for tensile loading
        ws['I114'] = sv(hs['B62'].value)         # Nsa (I114-K114)
        ws['F118'] = sv(hs['B65'].value)         # Abrg (F118-H118)
        ws['M119'] = sv(hs['B66'].value)         # Nap (M119-O119)
        ws['E121'] = 1.4                         # Psi_c,P (E121-F121)
        ws['M122'] = sv(hs['B68'].value)         # Npn (M122-O122)
        
        # Side-face Blowout
        ws['E125'] = sv(hs['B71'].value)         # Ca1 (E125-G125)
        ws['K125'] = sv(hs['B72'].value)         # 2.5Ca1 (K125-M125)
        ws['Q125'] = sv(hs['B73'].value, dp=1)   # hef (Q125-S125)
        ws['T125'] = hs['B77'].value or ''       # Blowout message
        
        ws['E126'] = sv(hs['B74'].value)         # Ca2 (E126-G126)
        ws['K126'] = sv(hs['B75'].value)         # Ca2/Ca1 (K126-M126)
        ws['Q126'] = sv(hs['B76'].value)         # f1 (Q126-S126)
        ws['T126'] = hs['B78'].value or ''       # Corner message
        
        ws['Q128'] = hs['B79'].value or ''       # Nsb (Q128-S128)
        
        ws['E130'] = sv(hs['B80'].value)         # S1 (E130-G130)
        ws['L130'] = sv(hs['B81'].value)         # 6 x Ca1 (L130-N130)
        ws['Q130'] = sv(hs['B82'].value)         # f2 (Q130-S130)
        ws['T130'] = hs['B83'].value or ''       # Spacing message
        
        ws['Q133'] = hs['B84'].value or ''       # Nsbg (Q133-S133)
    
    def _map_page4_data(self, ws, hs):
        """Map Page 4 data (rows 135-167): Tensile/Shear Summaries & Interaction"""
        
        def sv(value, dp=4):
            """Safe value - convert to number if possible"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, dp)
            except:
                return value
        
        # ===== Tensile Design Summary =====
        ws['J137'] = sv(hs['B87'].value)         # Phi = 0.75 (J137-K137)
        ws['M138'] = sv(hs['B88'].value)         # Nn (M138-O138)
        ws['M139'] = sv(hs['B89'].value)         # PhiNn (M139-O139)
        
        ws['F142'] = sv(hs['B90'].value, dp=2)   # Nua (F142-H142)
        ws['M142'] = sv(hs['B89'].value)         # PhiNn (M142-O142)
        
        try:
            rat4 = float(hs['B91'].value) if hs['B91'].value else 0
            ws['R142'] = 'OK' if rat4 < 1.0 else 'NG'
        except:
            ws['R142'] = ''
        
        ws['AF143'] = sv(hs['B91'].value)        # Ratio4 (AF143-AH143)
        
        # ===== 2.2 Design Requirements for shear loading =====
        ws['M147'] = sv(hs['B95'].value)         # Vsa (M147-O147)
        
        # ===== Shear Design Summary =====
        ws['J150'] = sv(hs['B98'].value)         # Phi = 0.65 (J150-K150)
        ws['J151'] = sv(hs['B99'].value)         # PhiVn (J151-L151)
        
        ws['E154'] = sv(hs['B100'].value, dp=2)  # Vua (E154-G154)
        ws['L154'] = sv(hs['B99'].value)         # PhiVn (L154-N154)
        
        try:
            rat5 = float(hs['B101'].value) if hs['B101'].value else 0
            ws['Q154'] = 'OK' if rat5 < 1.0 else 'NG'
        except:
            ws['Q154'] = ''
        
        ws['AF155'] = sv(hs['B101'].value)       # Ratio5 (AF155-AH155)
        
        # ===== 2.3 Interaction of Tensile and Shear Forces =====
        ws['I157'] = sv(hs['B104'].value, dp=2)  # Vua / PhiVn (I157-K157)
        ws['I158'] = sv(hs['B105'].value, dp=2)  # Nua / PhiNn (I158-K158)
        
        # Get values from Hinge Type C104 & C105 for L157 & L158
        c104_value = sv(hs['C104'].value, dp=2) if hs['C104'].value else ''
        c105_value = sv(hs['C105'].value, dp=2) if hs['C105'].value else ''
        ws['L157'] = c104_value    # Result from Hinge Type C104
        ws['L158'] = c105_value    # Result from Hinge Type C105
        
        ws['C159'] = hs['B107'].value or ''      # Interaction control text (moved to C159)
        try:
            rat6 = float(hs['B106'].value) if hs['B106'].value else 0
            ws['I159'] = 'OK' if rat6 < 1.0 else 'NG'
        except:
            ws['I159'] = ''
        
        ws['AF159'] = sv(hs['B106'].value)       # Ratio6 (AF159-AH159)
    
    def _map_page5_data(self, ws, hs):
        """Map Page 5 data (rows 168-212): Pier Reinforcement - Vertical Rebars"""
        
        def sv(value, dp=4):
            """Safe value - convert to number if possible"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, dp)
            except:
                return value
        
        # ===== 3.1 Transfer of Anchor Load to Vertical Rebars =====
        ws['I171'] = hs['O3'].value or ''         # Control case (Tensile)
        ws['I172'] = sv(hs['B43'].value)          # Vertical load (F3)
        
        # Row 175: Rebar text e.g. "For 14 - D22 bars,"
        main_qty = sv(hs['F11'].value)
        main_size = hs['G11'].value or ''
        ws['C175'] = f"For {main_qty} - {main_size} bars,"
        ws['K175'] = sv(hs['B121'].value)         # As prov/req? Actually Hinge B121 is rat7. 
        # Wait, looking at template Page 5, J175:L175 is As. 
        # In create_or_update_hinge_fixed_xlsx, B120 is nreqd, B121 is rat7.
        # Let's map As from Hinge sheet calculation if available.
        # Actually Template says "As = 380 mm2" (area of one bar).
        import math
        try:
            size_num = int(str(main_size).replace('D', '').strip())
            ws['K175'] = sv(math.pi * (size_num ** 2) / 4, dp=0)
        except:
            ws['K175'] = ''

        # Row 179: dmax
        ws['K179'] = sv(hs['B115'].value, dp=0)   # dmax
        ws['T179'] = sv(hs['B116'].value, dp=0)   # 0.5hef
        ws['Y179'] = hs['C116'].value or ''       # OK/NG (Y179-AA179)
        
        # Row 182: dactual
        ws['S182'] = sv(hs['B117'].value, dp=0)   # dact (S182-U182)
        ws['AA182'] = sv(hs['B115'].value, dp=0)  # dmax (AA182-AC182)
        ws['AF182'] = hs['C117'].value or ''      # OK/NG
        
        # Row 185-186: Ca checks
        ws['G185'] = sv(hs['B118'].value, dp=0)   # Ca1-c-db/2 (G185-I185)
        ws['O185'] = sv(float(hs['I3'].value)/2 if hs['I3'].value else 0, dp=0) # S1/2 (O185-Q185)
        ws['T185'] = hs['C118'].value or ''       # OK/NG
        
        ws['G186'] = sv(hs['B119'].value, dp=0)   # Ca2-c-db/2 (G186-I186)
        ws['O186'] = sv(float(hs['J3'].value)/2 if hs['J3'].value else 0, dp=0) # S2/2 (O186-Q186)
        ws['T186'] = hs['C119'].value or ''       # OK/NG
        
        # Check for development length
        ws['J189'] = 0.75                         # Resistance Factor (J189-K189)
        ws['L190'] = sv(hs['B120'].value, dp=2)   # nreqd (L190-M190)
        ws['O190'] = sv(hs['F11'].value, dp=0)    # nprov (O190-P190)
        try:
            rat7 = float(hs['B121'].value) if hs['B121'].value else 0
            ws['Q190'] = 'OK' if rat7 < 1.0 else 'NG'
        except:
            ws['Q190'] = ''
        ws['AF191'] = sv(hs['B121'].value, dp=2)  # Ratio7 (AF191-AH191)
        
        ws['F193'] = sv(hs['B122'].value, dp=0)   # ldprov (F193-H193)
        ws['N195'] = hs['G11'].value or ''        # Main bar size from G11 (e.g. D22)
        ws['R195'] = sv(hs['B123'].value, dp=2)   # ld min (R195-T195)
        ws['R196'] = sv(hs['B124'].value, dp=2)   # ldreqd (R196-T196)
        ws['R197'] = sv(hs['B122'].value, dp=2)   # ldprov (R197-T197)
        try:
            rat8 = float(hs['B125'].value) if hs['B125'].value else 0
            ws['W197'] = 'OK' if rat8 < 1.0 else 'NG'
        except:
            ws['W197'] = ''
        ws['AF198'] = sv(hs['B125'].value, dp=2)  # Ratio8
        
        # 3.2 Horizontal Rebars
        ws['K202'] = sv(hs['B130'].value, dp=0)   # ldh
        ws['H204'] = 1                            # Psi_e
        ws['L204'] = 1                            # Lambda
        
        # X-Direction
        ws['I207'] = hs['O5'].value or ''         # Control Case X
        ws['I208'] = sv(hs['B134'].value)         # Vertical Load X
        ws['I209'] = sv(hs['B135'].value)         # Horizontal Load X
        ws['N211'] = sv(hs['B136'].value, dp=0)   # Vx
        
    def _map_page6_data(self, ws, hs):
        """Map Page 6 data (rows 213-247): Pier Reinforcement - Horizontal Rebars (X & Z)"""
        
        def sv(value, dp=4):
            """Safe value - convert to number if possible"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, dp)
            except:
                return value
        
        # ===== X-Direction Continued =====
        # Extract diameter number from H11 (e.g. D10 -> 10)
        h11_val = hs['H11'].value or ''
        tie_diameter = str(h11_val).replace('D', '').replace('d', '') if h11_val else ''
        ws['H214'] = tie_diameter  # Tie diameter number only
        ws['N214'] = sv(hs['B137'].value, dp=1)   # Atie (N214-P214)
        ws['F215'] = sv(hs['B138'].value, dp=1)   # lda_A_L (F215-H215)
        ws['N215'] = sv(hs['B139'].value, dp=1)   # lda_A_R (N215-P215)
        ws['F216'] = sv(hs['B140'].value, dp=1)   # lda_B_L (F216-H216)
        ws['N216'] = sv(hs['B141'].value, dp=1)   # lda_B_R (N216-P216)
        ws['M218'] = sv(hs['B142'].value, dp=1)   # fs (M218-O218)
        ws['J220'] = 0.75                         # Resistance Factor (J220-K220)
        ws['K222'] = sv(hs['B144'].value, dp=3)  # Atie_reqd (K222-M222)
        ws['F224'] = sv(hs['I11'].value, dp=3)   # Legs
        ws['J224'] = hs['H11'].value or ''        # Tie bar size from H11 (e.g. D10)
        ws['F225'] = sv(hs['B145'].value, dp=3)   # Atie_prov (F225-H225)
        ws['P225'] = sv(hs['B144'].value, dp=3)   # Atie_reqd (P225-R225)
        try:
            rat9 = float(hs['B146'].value) if hs['B146'].value else 0
            ws['U225'] = 'OK' if rat9 < 1.0 else 'NG'
        except:
            ws['U225'] = ''
        ws['AF226'] = sv(hs['B146'].value, dp=2)  # Ratio9 (AF226-AH226)
        
        # ===== Y-Direction =====
        ws['I229'] = hs['O6'].value or ''         # Control Case Y (I229-K229)
        ws['I230'] = sv(hs['B150'].value)         # Vertical Load Y (I230-K230)
        
        # Check if vertical load is tension
        try:
            vy_load = float(hs['B150'].value) if hs['B150'].value else 0
            if vy_load < 0: ws['M230'] = '( Tension )'
        except: pass

        ws['I231'] = sv(hs['B151'].value)         # Horizontal Load Y (I231-K231)
        ws['N233'] = sv(hs['B152'].value, dp=2)   # Vy (N233-P233)
        
        # Extract diameter number from H11 (e.g. D10 -> 10)
        ws['H235'] = tie_diameter  # Tie diameter number only
        ws['N235'] = sv(hs['B153'].value, dp=1)   # Atie (N235-P235)
        ws['F236'] = sv(hs['B154'].value, dp=1)   # lda_A_L (F236-H236)
        ws['N236'] = sv(hs['B155'].value, dp=1)   # lda_A_R (N236-P236)
        ws['F237'] = sv(hs['B156'].value, dp=1)   # lda_B_L (F237-H237)
        ws['N237'] = sv(hs['B157'].value, dp=1)   # lda_B_R (N237-P237)
        
        # fs for Y
        ws['K239'] = sv(hs['B158'].value, dp=1)   # fs (K239-M239)
        
        ws['J241'] = 0.75                         # Resistance Factor (J241-K241)
        ws['K243'] = sv(hs['B160'].value, dp=3)   # Atie_reqd (K243-M243)
        ws['F245'] = sv(hs['J11'].value, dp=3)    # Legs
        ws['J245'] = hs['H11'].value or ''        # Tie bar size from H11 (e.g. D10)
        ws['F246'] = sv(hs['B161'].value, dp=3)   # Atie_prov (F246-H246)
        ws['P246'] = sv(hs['B160'].value, dp=3)   # Atie_reqd (P246-R246)
        try:
            rat10 = float(hs['B162'].value) if hs['B162'].value else 0
            ws['U246'] = 'OK' if rat10 < 1.0 else 'NG'
        except:
            ws['U246'] = ''
        ws['AF247'] = sv(hs['B162'].value, dp=2)  # Ratio10 (AF247-AH247)
    
    def _create_summary_sheet(self, workbook, selected_nodes):
        """Create summary sheet with node list and status"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        ws = workbook.create_sheet("Summary", 0)
        
        # Header
        ws['A1'] = "Calculation Report Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Info
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Total Nodes: {len(selected_nodes)}"
        
        # Table headers
        headers = ['Node', 'Status', 'Max Ratio', 'Design', 'L/C Control']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1a472a', end_color='1a472a', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_idx, node_name in enumerate(sorted(selected_nodes), 7):
            node_data = self.base_plate_nodes[node_name]
            
            # Node name
            ws.cell(row=row_idx, column=1).value = node_name
            
            # Status
            status = node_data.get('design_status', 'N/A')
            status_cell = ws.cell(row=row_idx, column=2)
            status_cell.value = status
            status_cell.fill = PatternFill(
                start_color='4CAF50' if status == 'OK' else 'f44336',
                end_color='4CAF50' if status == 'OK' else 'f44336',
                fill_type='solid'
            )
            status_cell.font = Font(color='FFFFFF', bold=True)
            status_cell.alignment = Alignment(horizontal='center')
            
            # Max Ratio
            max_ratio = node_data.get('max_ratio')
            ratio_cell = ws.cell(row=row_idx, column=3)
            ratio_cell.value = f"{max_ratio:.2f}" if max_ratio else "N/A"
            ratio_cell.alignment = Alignment(horizontal='center')
            
            # Design check
            design_cell = ws.cell(row=row_idx, column=4)
            design_cell.value = "✓ Checked" if status != 'Not Checked' else "✗ Not Checked"
            design_cell.alignment = Alignment(horizontal='center')
            
            # L/C Control (placeholder - would need to read from sheet)
            lc_cell = ws.cell(row=row_idx, column=5)
            lc_cell.value = "See detail"
            lc_cell.alignment = Alignment(horizontal='center')
            
            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = thin_border
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

    # ==================== Help ====================

    def show_guide(self):
        """Show user guide"""
        guide_window = tk.Toplevel(self.root)
        guide_window.title("Instructions & Workflow")
        guide_window.geometry("550x650")
        guide_window.configure(bg='white')
        guide_window.transient(self.root)
        
        # Title
        title_frame = tk.Frame(guide_window, bg='white')
        title_frame.pack(fill='x', pady=(15, 5))
        tk.Label(title_frame, text="📋 HOW TO USE THIS TOOL", 
                font=('Arial', 14, 'bold'), bg='white').pack()
        
        # Scrollable text area
        text_frame = tk.Frame(guide_window, bg='white')
        text_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        text_widget = tk.Text(text_frame, wrap='word', font=('Consolas', 10), 
                             bg='#fafafa', fg='#333', relief='groove', bd=1,
                             padx=15, pady=15, spacing2=3)
        scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        text_widget.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        guide_text = """🏗  BASE PLATE DESIGN TOOL

📋 PREPARATION STEPS:
  • Have your SAP2000 model file (.sdb) ready
  • Unit system: kN, m

⚡ WORKFLOW:
  1  □ "Get Model" to connect to SAP2000
  2  □ Select frame elements in SAP2000 model
  3  □ "Load (Auto)" to extract coordinates & reactions
  4  □ Select load cases/combinations for analysis
  5  □ Define Material Properties
  6  □ Assign Base Plate details for each node
  7  □ "Run Analysis" to calculate all ratios
  8  □ "Export to Excel" to generate Calculation Report

📤 EXPORT & RESULTS:
  • Calculation Report.xlsx - Detailed calculation sheets
  • Summary sheet with all node ratios
  • Color-coded pass/fail indication

⚠ IMPORTANT NOTES:
  • Select ELEMENTS only for Load (Auto), DON'T select any nodes
  • "Fixed Type" is still under development
  • "STAAD Pro" software option is under development

💡 TIPS:
  • Add/Modify Material to suit your project 
  • Click Reload Data to update Base Plate Plan View
  • Click a Node on Base Plate Plan View to assign Parameters
  • Use "Copy to Multiple Nodes" for same-section nodes

📞 SUPPORT: Contact Roberto for technical assistance"""
        
        text_widget.insert('1.0', guide_text)
        text_widget.config(state='disabled')

    def show_about(self):
        """Show about information"""
        about_window = tk.Toplevel(self.root)
        about_window.title("About")
        about_window.geometry("450x350")
        about_window.configure(bg='white')
        about_window.transient(self.root)
        about_window.resizable(False, False)
        
        # Icon
        tk.Label(about_window, text="ℹ️", font=('Arial', 28), bg='white').pack(pady=(20, 5))
        
        # Tool name
        tk.Label(about_window, text="Base Plate Design Tool", 
                font=('Arial', 13, 'bold'), bg='white', fg='#1a472a').pack()
        
        # Version info
        info_frame = tk.Frame(about_window, bg='white')
        info_frame.pack(pady=10)
        tk.Label(info_frame, text="Version: 2.0\nAnalysis Type: Base Plate Design", 
                font=('Arial', 10), bg='white', fg='#555', justify='center').pack()
        
        # Features
        features_frame = tk.Frame(about_window, bg='white')
        features_frame.pack(pady=5)
        tk.Label(features_frame, text="Features:\n"
                "- Load combination analysis\n"
                "- Compression & Tensile checks\n"
                "- Shear & Interaction checks\n"
                "- Visual capacity ratio diagrams\n"
                "- Color-coded failure indication",
                font=('Arial', 9), bg='white', fg='#666', justify='left').pack(anchor='w', padx=40)
        
        # Credits
        tk.Label(about_window, text="Civil & Structural Engineering Analysis\nRoberto Inhouse", 
                font=('Arial', 9, 'italic'), bg='white', fg='#888').pack(pady=(10, 2))
        tk.Label(about_window, text="© 2026 All rights reserved", 
                font=('Arial', 9), bg='white', fg='#aaa').pack()
        
        # OK button
        tk.Button(about_window, text="OK", command=about_window.destroy,
                 width=10, font=('Arial', 10)).pack(pady=(10, 15))

# ========================================================================
    def load_anchor_bolt_material_data(self, tree):
            """Load Anchor Bolt Material data from Data.xlsx (columns A-B)"""
            if not self.bpl_folder or not os.path.exists(self.bpl_folder):
                return
            data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
            if not os.path.exists(data_file):
                return
            try:
                wb = openpyxl.load_workbook(data_file)
                ws = wb['Material Strength']
                for item in tree.get_children():
                    tree.delete(item)
                for row_idx in range(2, ws.max_row + 1):
                    mat_type = ws[f'A{row_idx}'].value
                    strength = ws[f'B{row_idx}'].value
                    if not mat_type or mat_type == 'Material Type':
                        break
                    tree.insert('', 'end', values=(mat_type, strength))
                wb.close()
            except Exception as e:
                print(f"Error loading Anchor Bolt Material data: {e}")

    def load_base_plate_material_data(self, tree):
        """Load Base Plate Material data from Data.xlsx (columns J-K)"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return
        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return
        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Material Strength']
            for item in tree.get_children():
                tree.delete(item)
            for row_idx in range(2, ws.max_row + 1):
                mat_type = ws[f'J{row_idx}'].value
                strength = ws[f'K{row_idx}'].value
                if not mat_type or mat_type == 'Material Type':
                    break
                tree.insert('', 'end', values=(mat_type, strength))
            wb.close()
        except Exception as e:
            print(f"Error loading Base Plate Material data: {e}")

    def load_hinge_type_data(self, tree):
            """Load Hinge Type data from Data.xlsx"""
            if not self.bpl_folder or not os.path.exists(self.bpl_folder):
                return
            data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
            if not os.path.exists(data_file):
                return
            try:
                wb = openpyxl.load_workbook(data_file)
                ws = wb['Hinge Type']
                for item in tree.get_children():
                    tree.delete(item)
                # Đọc 22 cột
                for row in ws.iter_rows(min_row=2, max_col=22, values_only=True):
                    if row[0] and row[0] != 'Column size':
                        values = tuple('' if v is None else str(v) for v in row)
                        tree.insert('', 'end', values=values)
                wb.close()
            except Exception as e:
                print(f"Error loading Hinge Type data: {e}")

    def load_rebar_dev_length_data(self, tree):
        """Load Rebar Development Length data from Data.xlsx"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return
        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return
        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Rebar Development Length']
            for item in tree.get_children():
                tree.delete(item)
            # Read data starting from row 2 (skip header)
            for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                if row[0] and row[0] != 'Bars':
                    values = tuple('' if v is None else str(v) for v in row)
                    tree.insert('', 'end', values=values)
            wb.close()
        except Exception as e:
            print(f"Error loading Rebar Development Length data: {e}")

    def lookup_rebar_dev_length(self, bar_size, column_name):
            """Lookup Ld or Ldh from Rebar Development Length sheet

            Args:
                bar_size: int (e.g., 19 for D19)
                column_name: 'Ld' or 'Ldh'

            Returns:
                float value or 0
            """
            if not self.bpl_folder:
                return 0

            data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
            if not os.path.exists(data_file):
                return 0

            try:
                wb = openpyxl.load_workbook(data_file)
                ws = wb['Rebar Development Length']

                # Find column index
                col_idx = 2 if column_name == 'Ld' else 3  # Ld = col B, Ldh = col C

                # Search for bar size
                for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                    if row[0] and int(float(row[0])) == bar_size:
                        result = row[col_idx - 1] if len(row) >= col_idx else 0
                        wb.close()
                        return float(result) if result else 0

                wb.close()
                return 0
            except Exception as e:
                print(f"Error looking up rebar dev length: {e}")
                return 0

    # ==================== BASE PLATE DETAIL METHODS ====================
    
    def plot_empty_state(self):
        """Plot empty state message"""
        self.ax.clear()
        self.ax.text(0.5, 0.5, 'No base plate data loaded\n\nClick "Reload Data" to load from bpl_coordinate.csv',
                    ha='center', va='center', fontsize=12, color='#999999',
                    transform=self.ax.transAxes)
        self.ax.set_xlim(0, 1)
        self.ax.set_ylim(0, 1)
        self.ax.axis('off')
        self.canvas.draw()
    
    def load_base_plate_plan(self):
        """Load base plate coordinates from CSV and plot"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            messagebox.showwarning("Warning", 
                "Base Plate Design folder not found!\n\n"
                "Please connect to SAP2000 model first (Column Base tab)")
            return
        
        csv_file = os.path.join(self.bpl_folder, "bpl_coordinate.csv")
        if not os.path.exists(csv_file):
            messagebox.showwarning("Warning", 
                f"bpl_coordinate.csv not found!\n\n"
                f"Please load column base data first (Column Base tab)")
            return
        
        try:
            # Read CSV
            self.base_plate_nodes = {}
            with open(csv_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    node_name = row['Column Base'].strip()
                    x = float(row['X (m)'].strip())
                    y = float(row['Y (m)'].strip())
                    section = row.get('Section', '').strip()
                    beta = int(float(row.get('Beta', '0').strip() or '0'))
                    
                    self.base_plate_nodes[node_name] = {
                        'x': x,
                        'y': y,
                        'section': section,
                        'beta': beta,
                        'material_anchor_bolt': None,
                        'material_base_plate': None,
                        'material_concrete': None,
                        'material_rebar': None,
                        'bolt_size': None,
                        'hinge_fixed_type': None,
                        'detail_type': None,
                        'design_status': 'Not Checked'
                    }
            
            if not self.base_plate_nodes:
                messagebox.showwarning("Warning", "No valid data found in bpl_coordinate.csv!")
                return
            
            # Update status
            self.bpl_status_label.config(
                text=f"✓ {len(self.base_plate_nodes)} nodes loaded",
                fg='#4CAF50'
            )
            
            # Enable run button
            self.run_design_btn.config(state='normal')
            
            # Plot
            self.plot_base_plate_plan()
            
            self.status_label.config(text=f"● Base plate plan loaded: {len(self.base_plate_nodes)} nodes", fg='#90ee90')
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load base plate data:\n{str(e)}")
    
    def plot_base_plate_plan(self):
        """Plot all base plate nodes"""
        self.ax.clear()
        
        if not self.base_plate_nodes:
            self.plot_empty_state()
            return
        
        # Extract coordinates and colors
        x_coords = []
        y_coords = []
        colors = []
        
        for node_name, data in self.base_plate_nodes.items():
            x_coords.append(data['x'])
            y_coords.append(data['y'])
            
            # Color based on status
            if data['design_status'] == 'OK':
                colors.append('#4CAF50')  # Green
            elif data['design_status'] == 'NG':
                colors.append('#f44336')  # Red
            elif data['bolt_size'] is not None:  # Has assignment
                colors.append('#2196F3')  # Blue
            else:
                colors.append('#CCCCCC')  # Gray
        
        # Plot scatter
        self.scatter = self.ax.scatter(x_coords, y_coords, c=colors, s=150, 
                                      edgecolors='black', linewidths=1.5,
                                      picker=True, pickradius=5)
        
        # Clear previous labels
        self.label_texts = []
        
        # Add labels based on display option
        display_mode = self.label_display_var.get()
        
        if display_mode != "none":  # If not "none", display something
            for node_name, data in self.base_plate_nodes.items():
                text_label = ""
                
                # Build label based on display mode
                if display_mode == "label":
                    text_label = node_name
                elif display_mode == "ratio":
                    if data.get('max_ratio') is not None:
                        text_label = f"{data['max_ratio']:.2f}"
                    else:
                        text_label = "N/A"
                elif display_mode == "both":
                    text_label = node_name
                    if data.get('max_ratio') is not None:
                        text_label += f"\n{data['max_ratio']:.2f}"
                
                # Use annotate for better font styling (no bbox border)
                if text_label:
                    txt = self.ax.annotate(text_label, 
                                           (data['x'], data['y']), 
                                           xytext=(5, 5), 
                                           textcoords='offset points', 
                                           fontsize=8, 
                                           fontfamily='Aptos',
                                           alpha=0.85)
                    self.label_texts.append(txt)
        
        # Styling
        self.ax.set_xlabel('X (m)', fontsize=10, fontweight='bold')
        self.ax.set_ylabel('Y (m)', fontsize=10, fontweight='bold')
        self.ax.set_title('Base Plate Plan View', fontsize=12, fontweight='bold', pad=15)
        self.ax.grid(True, alpha=0.3, linestyle='--')
        self.ax.set_aspect('equal', adjustable='datalim')
        
        # Auto-adjust limits with margin
        if x_coords and y_coords:
            x_margin = (max(x_coords) - min(x_coords)) * 0.1 or 1
            y_margin = (max(y_coords) - min(y_coords)) * 0.1 or 1
            self.ax.set_xlim(min(x_coords) - x_margin, max(x_coords) + x_margin)
            self.ax.set_ylim(min(y_coords) - y_margin, max(y_coords) + y_margin)
        
        self.canvas.draw()
    
    def on_node_click(self, event):
        """Handle node click event"""
        if event.inaxes != self.ax or not self.base_plate_nodes:
            return
        
        # Find closest node
        min_dist = float('inf')
        closest_node = None
        
        for node_name, data in self.base_plate_nodes.items():
            dist = math.sqrt((event.xdata - data['x'])**2 + (event.ydata - data['y'])**2)
            if dist < min_dist:
                min_dist = dist
                closest_node = node_name
        
        # Select if close enough (within 2 units)
        if min_dist < 2.0:
            self.select_node(closest_node)
    
    def select_node(self, node_name):
        """Select a node and show its details"""
        self.selected_node = node_name
        node_data = self.base_plate_nodes[node_name]
        
        # Update info label
        info_text = (
            f"Node: {node_name}\n"
            f"Section: {node_data['section'] or 'N/A'}\n"
            f"Position: ({node_data['x']:.2f}, {node_data['y']:.2f})"
        )
        self.node_info_label.config(text=info_text, fg='#1a472a')
        
        # If design check has been run, show control load table
        if node_data.get('design_status') and node_data['design_status'] != 'Not Checked':
            self.show_control_load_table(node_name)
        
        # Clear and rebuild dropdown frame
        for widget in self.dropdown_frame.winfo_children():
            widget.destroy()
        
        # Get material lists from Data.xlsx
        anchor_bolt_list = self.get_material_list('anchor_bolt')
        base_plate_list = self.get_material_list('base_plate')
        concrete_list = self.get_material_list('concrete')
        rebar_list = self.get_material_list('rebar')
        bolt_list = self.get_bolt_size_list()
        detail_list = self.get_detail_type_list(node_data['section'])
        
        # Helper function to get default value (first non-empty item or stored value)
        def get_default(stored_value, option_list):
            if stored_value:
                return stored_value
            # Find first non-empty item in list
            for item in option_list:
                if item and item.strip():  # Skip empty strings
                    return item
            return ''
        
        # Anchor Bolt Material
        tk.Label(self.dropdown_frame, text="Anchor Bolt Material:", font=('Arial', 9, 'bold'), 
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(5, 2))
        self.anchor_bolt_var = tk.StringVar(value=get_default(node_data.get('material_anchor_bolt', ''), anchor_bolt_list))
        anchor_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.anchor_bolt_var, 
                                   values=anchor_bolt_list, state='readonly', width=25)
        anchor_menu.pack(fill='x', pady=(0, 8))
        
        # Base Plate Material
        tk.Label(self.dropdown_frame, text="Base Plate Material:", font=('Arial', 9, 'bold'), 
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.base_plate_var = tk.StringVar(value=get_default(node_data.get('material_base_plate', ''), base_plate_list))
        baseplate_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.base_plate_var, 
                                      values=base_plate_list, state='readonly', width=25)
        baseplate_menu.pack(fill='x', pady=(0, 8))
        
        # Concrete Material
        tk.Label(self.dropdown_frame, text="Concrete Material:", font=('Arial', 9, 'bold'), 
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.concrete_var = tk.StringVar(value=get_default(node_data.get('material_concrete', ''), concrete_list))
        concrete_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.concrete_var, 
                                     values=concrete_list, state='readonly', width=25)
        concrete_menu.pack(fill='x', pady=(0, 8))
        
        # Mainbar Material
        tk.Label(self.dropdown_frame, text="Mainbar Material:", font=('Arial', 9, 'bold'), 
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.mainbar_var = tk.StringVar(value=get_default(node_data.get('material_mainbar', ''), rebar_list))
        mainbar_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.mainbar_var, 
                                  values=rebar_list, state='readonly', width=25)
        mainbar_menu.pack(fill='x', pady=(0, 8))
        
        # Tiebar Material
        tk.Label(self.dropdown_frame, text="Tiebar Material:", font=('Arial', 9, 'bold'), 
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.tiebar_var = tk.StringVar(value=get_default(node_data.get('material_tiebar', ''), rebar_list))
        tiebar_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.tiebar_var, 
                                  values=rebar_list, state='readonly', width=25)
        tiebar_menu.pack(fill='x', pady=(0, 8))
        
        # Anchor Bolt Size
        tk.Label(self.dropdown_frame, text="Anchor Bolt Size:", font=('Arial', 9, 'bold'), 
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.bolt_var = tk.StringVar(value=get_default(node_data.get('bolt_size', ''), bolt_list))
        bolt_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.bolt_var, 
                                values=bolt_list, state='readonly', width=25)
        bolt_menu.pack(fill='x', pady=(0, 8))
        
        # Hinge/Fixed Type
        tk.Label(self.dropdown_frame, text="Hinge/Fixed Type:", font=('Arial', 9, 'bold'), 
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        hinge_list = ['Hinge Type', 'Fixed Type']
        self.hinge_fixed_var = tk.StringVar(value=get_default(node_data.get('hinge_fixed_type', ''), hinge_list))
        hinge_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.hinge_fixed_var, 
                                  values=hinge_list, state='readonly', width=25)
        hinge_menu.pack(fill='x', pady=(0, 8))
        
        # Detail Type
        tk.Label(self.dropdown_frame, text="Detail Type:", font=('Arial', 9, 'bold'), 
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.detail_var = tk.StringVar(value=get_default(node_data.get('detail_type', ''), detail_list))
        detail_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.detail_var, 
                                   values=detail_list, state='readonly', width=25)
        detail_menu.pack(fill='x', pady=(0, 8))
        
        # Block mouse wheel on all Comboboxes to prevent accidental value changes while scrolling
        def _block_mousewheel(event):
            return "break"
        for combo in [anchor_menu, baseplate_menu, concrete_menu, mainbar_menu, 
                      tiebar_menu, bolt_menu, hinge_menu, detail_menu]:
            combo.bind("<MouseWheel>", _block_mousewheel)
        
        # Enable buttons
        self.apply_btn.config(state='normal')
        self.edit_detail_btn.config(state='normal')  # THÊM DÒNG NÀY
        self.copy_multi_btn.config(state='normal')
        
        # Highlight selected node in plot
        self.highlight_selected_node()
    
    def highlight_selected_node(self):
        """Highlight the selected node on plot"""
        if not self.selected_node:
            return
        
        # Redraw plot with highlight
        self.plot_base_plate_plan()
        
        # Add highlight circle
        node_data = self.base_plate_nodes[self.selected_node]
        circle = plt.Circle((node_data['x'], node_data['y']), 1.5, 
                           color='yellow', fill=False, linewidth=3, 
                           linestyle='--', alpha=0.8)
        self.ax.add_patch(circle)
        
        self.canvas.draw()
    
    def get_material_list(self, material_type):
        """Get material list from Material Strength tree"""
        tree = self.material_trees.get(material_type)
        if not tree:
            return ['']
        
        materials = ['']  # Empty option
        for item in tree.get_children():
            values = tree.item(item, 'values')
            if values and values[0]:
                materials.append(values[0])
        
        return materials
    
    def get_bolt_size_list(self):
        """Get bolt sizes from Anchor Bolt Table"""
        tree = self.material_trees.get('Anchor Bolt Table')
        if not tree:
            return ['']
        
        bolts = ['']
        for item in tree.get_children():
            values = tree.item(item, 'values')
            if values and values[0]:  # db column
                bolts.append(f"M{values[0]}")
        
        return bolts
    
    def get_detail_type_list(self, section):
            """Get detail types from Hinge Type that match section"""
            tree = self.material_trees.get('Hinge Type')
            if not tree:
                return ['']
            
            details = ['']
            for item in tree.get_children():
                values = tree.item(item, 'values')
                if values and values[0]:
                    if section and section.upper() in values[0].upper():
                        detail_type = values[1] if len(values) > 1 and values[1] else 'Standard'
                        detail_str = f"{values[0]} - Type {detail_type}"
                        if detail_str not in details:
                            details.append(detail_str)
            
            if len(details) == 1:
                for item in tree.get_children():
                    values = tree.item(item, 'values')
                    if values and values[0]:
                        detail_type = values[1] if len(values) > 1 and values[1] else 'Standard'
                        details.append(f"{values[0]} - Type {detail_type}")
            
            return details
    
    def apply_node_settings(self):
            """Apply settings to selected node and create/update Hinge Type.xlsx"""
            if not self.selected_node:
                return

            node_data = self.base_plate_nodes[self.selected_node]

            # Validate required fields
            if not node_data.get('section'):
                messagebox.showwarning("Warning", "Section is required!")
                return

            # Update node data
            node_data['material_anchor_bolt'] = self.anchor_bolt_var.get() if self.anchor_bolt_var.get() else None
            node_data['material_base_plate'] = self.base_plate_var.get() if self.base_plate_var.get() else None
            node_data['material_concrete'] = self.concrete_var.get() if self.concrete_var.get() else None
            node_data['material_mainbar'] = self.mainbar_var.get() if self.mainbar_var.get() else None
            node_data['material_tiebar'] = self.tiebar_var.get() if self.tiebar_var.get() else None
            node_data['bolt_size'] = self.bolt_var.get() if self.bolt_var.get() else None
            node_data['hinge_fixed_type'] = self.hinge_fixed_var.get() if self.hinge_fixed_var.get() else None
            node_data['detail_type'] = self.detail_var.get() if self.detail_var.get() else None

            # Reset design status
            node_data['design_status'] = 'Not Checked'

            # ✅ THÊM: Clear tất cả edited_* khi Apply lại từ dropdown
            for key in ['edited_section','edited_pier_detail', 'edited_base_plate_detail', 
                        'edited_material', 'edited_anchor_bolt', 
                        'edited_main_bar', 'edited_other_coeff']:
                node_data.pop(key, None)
            
            # Check if Fixed Type is selected (not yet supported)
            if node_data.get('hinge_fixed_type') == 'Fixed Type':
                messagebox.showwarning("⚠️ Not Supported", 
                    "Fixed Type is still under development.\n\n"
                    "Please select 'Hinge Type' instead.")
                return

            # Create/Update Hinge Type.xlsx
            try:
                filepath = self.create_or_update_hinge_fixed_xlsx(self.selected_node, node_data)

                # Refresh plot
                self.plot_base_plate_plan()
                self.highlight_selected_node()

                self.status_label.config(text=f"● Settings applied to {self.selected_node}, sheet created in {os.path.basename(filepath)}", fg='#90ee90')

                messagebox.showinfo("✅ Success", 
                    f"Settings applied to node {self.selected_node}\n\n"
                    #f"Sheet created/updated:\n{filepath}\n\n"
                    f"Sheet name: Node {self.selected_node}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create sheet:\n{str(e)}")
                print(f"Error details: {e}")
    
    def copy_to_multiple(self):
        """Copy current settings to multiple nodes"""
        if not self.selected_node:
            return
        
        # Get source data
        source_data = self.base_plate_nodes[self.selected_node]
        source_section = source_data.get('section', '')
        
        # Split nodes into same section vs different section
        same_section_nodes = []
        diff_section_nodes = []
        for node_name in sorted(self.base_plate_nodes.keys()):
            if node_name != self.selected_node:
                node_section = self.base_plate_nodes[node_name].get('section', '')
                if node_section and source_section and node_section.upper() == source_section.upper():
                    same_section_nodes.append(node_name)
                else:
                    diff_section_nodes.append(node_name)
        
        # Create selection dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Copy to Multiple Nodes")
        dialog.geometry("950x600")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Header
        header_frame = tk.Frame(dialog, bg='#f0f0f0')
        header_frame.pack(fill='x', padx=15, pady=(15, 5))
        tk.Label(header_frame, text=f"Copy settings from: {self.selected_node}", 
                font=('Arial', 11, 'bold'), bg='#f0f0f0', fg='#1a472a').pack(side='left')
        tk.Label(header_frame, text=f"Section: {source_section or 'N/A'}", 
                font=('Arial', 10, 'italic'), bg='#f0f0f0', fg='#666').pack(side='left', padx=(20, 0))
        
        # Two-column container
        columns_frame = tk.Frame(dialog, bg='white')
        columns_frame.pack(fill='both', expand=True, padx=15, pady=10)
        columns_frame.columnconfigure(0, weight=1)
        columns_frame.columnconfigure(1, weight=1)
        
        node_vars = {}
        
        def create_column(parent, col_idx, title, subtitle, nodes, title_bg, title_fg):
            """Create a scrollable column with checkboxes"""
            col_frame = tk.Frame(parent, bg='white', bd=1, relief='groove')
            col_frame.grid(row=0, column=col_idx, sticky='nsew', padx=(0 if col_idx == 0 else 5, 0))
            
            # Column header
            header = tk.Frame(col_frame, bg=title_bg)
            header.pack(fill='x')
            tk.Label(header, text=title, font=('Arial', 10, 'bold'), 
                    bg=title_bg, fg=title_fg).pack(side='left', padx=10, pady=5)
            tk.Label(header, text=f"({len(nodes)})", font=('Arial', 9), 
                    bg=title_bg, fg=title_fg).pack(side='left')
            
            # Subtitle
            tk.Label(col_frame, text=subtitle, font=('Arial', 8, 'italic'), 
                    bg='white', fg='#999').pack(anchor='w', padx=10, pady=(3, 0))
            
            # Select All / None buttons
            btn_row = tk.Frame(col_frame, bg='white')
            btn_row.pack(fill='x', padx=5, pady=3)
            
            def select_all():
                for nn in nodes:
                    if nn in node_vars:
                        node_vars[nn].set(True)
            def select_none():
                for nn in nodes:
                    if nn in node_vars:
                        node_vars[nn].set(False)
            
            tk.Button(btn_row, text="✓ All", command=select_all, bg='#4CAF50', fg='white', 
                     font=('Arial', 8, 'bold'), cursor='hand2', width=6).pack(side='left', padx=2)
            tk.Button(btn_row, text="✗ None", command=select_none, bg='#FF9800', fg='white', 
                     font=('Arial', 8, 'bold'), cursor='hand2', width=6).pack(side='left', padx=2)
            
            # Scrollable checkbox area
            scroll_container = tk.Frame(col_frame, bg='white')
            scroll_container.pack(fill='both', expand=True, padx=2, pady=2)
            
            canvas = tk.Canvas(scroll_container, bg='white', highlightthickness=0)
            scrollbar = ttk.Scrollbar(scroll_container, orient='vertical', command=canvas.yview)
            scrollable = tk.Frame(canvas, bg='white')
            
            scrollable.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
            canvas.create_window((0, 0), window=scrollable, anchor='nw')
            canvas.configure(yscrollcommand=scrollbar.set)
            
            def _on_mousewheel(event):
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            canvas.bind("<MouseWheel>", _on_mousewheel)
            
            canvas.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            
            # Add checkboxes
            for nn in nodes:
                var = tk.BooleanVar(value=False)
                node_vars[nn] = var
                cb = tk.Checkbutton(scrollable, text=f" {nn}", variable=var, 
                                   font=('Arial', 10), bg='white', anchor='w')
                cb.pack(fill='x', padx=5, pady=2)
            
            if not nodes:
                tk.Label(scrollable, text="(No nodes)", font=('Arial', 9, 'italic'), 
                        bg='white', fg='#ccc').pack(pady=20)
        
        # Left column: same section
        create_column(columns_frame, 0, 
                     f"✅ Same Section", f"{source_section}",
                     same_section_nodes, '#E8F5E9', '#2E7D32')
        
        # Right column: different section
        create_column(columns_frame, 1, 
                     "⚠️ Different Section", "Other sections",
                     diff_section_nodes, '#FFF3E0', '#E65100')
        
        def apply_to_selected():
            count = 0
            failed_nodes = []
            
            for node_name, var in node_vars.items():
                if var.get():
                    try:
                        # Update node data (basic dropdowns/materials)
                        self.base_plate_nodes[node_name]['material_anchor_bolt'] = source_data.get('material_anchor_bolt')
                        self.base_plate_nodes[node_name]['material_base_plate'] = source_data.get('material_base_plate')
                        self.base_plate_nodes[node_name]['material_concrete'] = source_data.get('material_concrete')
                        self.base_plate_nodes[node_name]['material_mainbar'] = source_data.get('material_mainbar')
                        self.base_plate_nodes[node_name]['material_tiebar'] = source_data.get('material_tiebar')
                        self.base_plate_nodes[node_name]['bolt_size'] = source_data.get('bolt_size')
                        self.base_plate_nodes[node_name]['hinge_fixed_type'] = source_data.get('hinge_fixed_type')
                        self.base_plate_nodes[node_name]['detail_type'] = source_data.get('detail_type')
                        self.base_plate_nodes[node_name]['design_status'] = 'Not Checked'
                        
                        # Copy edited values if source node has them (from Edit Node Detail)
                        if 'edited_pier_detail' in source_data:
                            self.base_plate_nodes[node_name]['edited_pier_detail'] = source_data.get('edited_pier_detail')
                        if 'edited_base_plate_detail' in source_data:
                            self.base_plate_nodes[node_name]['edited_base_plate_detail'] = source_data.get('edited_base_plate_detail')
                        if 'edited_material' in source_data:
                            self.base_plate_nodes[node_name]['edited_material'] = source_data.get('edited_material')
                        if 'edited_anchor_bolt' in source_data:
                            self.base_plate_nodes[node_name]['edited_anchor_bolt'] = source_data.get('edited_anchor_bolt')
                        if 'edited_main_bar' in source_data:
                            self.base_plate_nodes[node_name]['edited_main_bar'] = source_data.get('edited_main_bar')
                        if 'edited_other_coeff' in source_data:
                            self.base_plate_nodes[node_name]['edited_other_coeff'] = source_data.get('edited_other_coeff')
                        
                        # Create/update Hinge Type.xlsx sheet for this node
                        self.create_or_update_hinge_fixed_xlsx(node_name, self.base_plate_nodes[node_name])
                        count += 1
                    except Exception as e:
                        failed_nodes.append(f"{node_name}: {str(e)}")
                        print(f"Error applying to {node_name}: {e}")
            
            if count > 0:
                self.plot_base_plate_plan()
                self.highlight_selected_node()
                
                msg = f"Settings copied to {count} node(s)"
                if failed_nodes:
                    msg += f"\n\nFailed ({len(failed_nodes)}):\n" + "\n".join(failed_nodes[:5])
                    if len(failed_nodes) > 5:
                        msg += f"\n... and {len(failed_nodes) - 5} more"
                    messagebox.showwarning("Partial Success", msg)
                else:
                    messagebox.showinfo("Success", msg)
            else:
                messagebox.showwarning("No Selection", "Please select at least one node")
            
            dialog.destroy()
        
        # Buttons
        btn_frame = tk.Frame(dialog, bg='white')
        btn_frame.pack(side='bottom', fill='x', pady=15, padx=15)
        
        tk.Button(btn_frame, text="✓ Apply", command=apply_to_selected, 
                 bg='#4CAF50', fg='white', font=('Arial', 10, 'bold'), 
                 width=15, cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_frame, text="✗ Cancel", command=dialog.destroy, 
                 bg='#f44336', fg='white', font=('Arial', 10, 'bold'), 
                 width=15, cursor='hand2').pack(side='left', padx=5)
    
    def update_plot_display(self):
        """Update plot when display mode changes (Label, Ratio, or Both)"""
        # Refresh plot
        self.plot_base_plate_plan()
        if self.selected_node:
            self.highlight_selected_node()
        
        display_mode = self.label_display_var.get()
        self.status_label.config(
            text=f"● Display: {display_mode.title()}", 
            fg='#FF9800'
        )
    
    def run_design_check(self):
        """Run design check for all assigned nodes"""
        if not self.base_plate_nodes:
            return
        
        # Count assigned nodes
        assigned_count = sum(1 for data in self.base_plate_nodes.values() 
                           if data['bolt_size'] is not None)
        
        if assigned_count == 0:
            messagebox.showwarning("Warning", "No nodes have been assigned!\n\nPlease assign materials and bolt sizes first.")
            return
        
        # Confirm
        confirm = messagebox.askyesno("Confirm Design Check", 
                                     f"Run design check for {assigned_count} assigned node(s)?")
        if not confirm:
            return
        
        # Run design check - Read ratios from Hinge Type.xlsx sheets
        checked_count = 0
        for node_name, data in self.base_plate_nodes.items():
            if data['bolt_size'] is not None:
                # Try to read max ratio from node's Hinge Type sheet
                max_ratio = self._get_max_ratio_from_sheet(node_name)
                
                if max_ratio is not None:
                    data['max_ratio'] = max_ratio
                    # Determine status based on ratio
                    if max_ratio <= 1.0:
                        data['design_status'] = 'OK'
                    else:
                        data['design_status'] = 'NG'
                else:
                    # If can't read sheet, set to "Not Checked"
                    data['design_status'] = 'Not Checked'
                    data['max_ratio'] = None
                
                checked_count += 1
        
        # Refresh plot
        self.plot_base_plate_plan()
        if self.selected_node:
            self.highlight_selected_node()
        
        messagebox.showinfo("Design Check Complete", 
                          f"Design check completed for {checked_count} node(s)!\n\n"
                          f"Check the plan view for results:\n"
                          f"🟢 Green = Design OK (Ratio ≤ 1.0)\n"
                          f"🔴 Red = Design NG (Ratio > 1.0)\n"
                          f"⚪ Gray = Not Checked")
        
        self.status_label.config(text=f"● Design check complete: {checked_count} nodes", fg='#90ee90')
    
    def show_control_load_table(self, node_name):
        """Display Control Load Table (N1:S6) from node's Hinge Type sheet in a dialog"""
        try:
            if not self.bpl_folder:
                return
            
            # Open Hinge Type.xlsx
            filepath = os.path.join(self.bpl_folder, "Hinge Type.xlsx")
            if not os.path.exists(filepath):
                return
            
            wb = openpyxl.load_workbook(filepath)
            sheet_name = f"Node {node_name}"
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                return
            
            ws = wb[sheet_name]
            
            # Read Control Load Table (N1:S6)
            table_data = []
            headers = ['Control Load', 'L/C', 'F1', 'F2', 'F3', 'Ratio']
            
            # Row 1: Headers
            table_data.append(headers)
            
            # Rows 2-6: Data
            control_checks = ['Compression', 'Tensile', 'Anchor Bolt', 'X-Shear check', 'Y-Shear check']
            for row in range(2, 7):
                row_data = []
                row_data.append(ws[f'N{row}'].value or '')
                row_data.append(ws[f'O{row}'].value or '')
                
                # Format forces and ratio with 2 decimal places
                try:
                    f1 = ws[f'P{row}'].value
                    f1_str = f"{float(f1):.2f}" if f1 is not None else ''
                except:
                    f1_str = ws[f'P{row}'].value or ''
                row_data.append(f1_str)
                
                try:
                    f2 = ws[f'Q{row}'].value
                    f2_str = f"{float(f2):.2f}" if f2 is not None else ''
                except:
                    f2_str = ws[f'Q{row}'].value or ''
                row_data.append(f2_str)
                
                try:
                    f3 = ws[f'R{row}'].value
                    f3_str = f"{float(f3):.2f}" if f3 is not None else ''
                except:
                    f3_str = ws[f'R{row}'].value or ''
                row_data.append(f3_str)
                
                # Ratio - format with 2 decimal places
                try:
                    ratio = ws[f'S{row}'].value
                    ratio_str = f"{float(ratio):.2f}" if ratio is not None else ''
                except:
                    ratio_str = ws[f'S{row}'].value or ''
                row_data.append(ratio_str)
                
                table_data.append(row_data)
            
            wb.close()
            
            # Create dialog to display table
            dialog = tk.Toplevel(self.root)
            dialog.title(f"Control Load Table - Node {node_name}")
            dialog.geometry("700x300")
            dialog.transient(self.root)
            
            # Add title
            tk.Label(dialog, text=f"Control Load Table - Node {node_name}", 
                    font=('Arial', 12, 'bold'), fg='#1a472a').pack(pady=10)
            
            # Create treeview for table
            tree_frame = tk.Frame(dialog)
            tree_frame.pack(fill='both', expand=True, padx=10, pady=10)
            
            # Columns
            columns = ('Control Load', 'L/C', 'F1', 'F2', 'F3', 'Ratio')
            tree = ttk.Treeview(tree_frame, columns=columns, height=6, show='headings')
            
            # Define column widths and headings
            widths = [120, 100, 80, 80, 80, 80]
            for i, (col, width) in enumerate(zip(columns, widths)):
                tree.column(col, width=width, anchor='center')
                tree.heading(col, text=col)
            
            # Add data rows
            for row in table_data[1:]:  # Skip header
                tree.insert('', 'end', values=row)
            
            # Add scrollbar
            scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
            tree.configure(yscroll=scrollbar.set)
            
            tree.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            
            # Add close button
            tk.Button(dialog, text="Close", command=dialog.destroy,
                     bg='#1a472a', fg='white', font=('Arial', 10, 'bold'),
                     padx=20, pady=8).pack(pady=10)
            
        except Exception as e:
            print(f"Error displaying control load table: {e}")
    
    def _get_max_ratio_from_sheet(self, node_name):
        """Get max ratio (from S2-S6) from node's Hinge Type sheet"""
        try:
            if not self.bpl_folder:
                return None
            
            # Try to open Hinge Type.xlsx
            filepath = os.path.join(self.bpl_folder, "Hinge Type.xlsx")
            if not os.path.exists(filepath):
                return None
            
            wb = openpyxl.load_workbook(filepath)
            sheet_name = f"Node {node_name}"
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                return None
            
            ws = wb[sheet_name]
            
            # Get ratios from S2-S6
            ratios = []
            for row in range(2, 7):  # Rows 2-6
                try:
                    ratio_val = ws[f'S{row}'].value
                    if ratio_val is not None:
                        ratios.append(float(ratio_val))
                except:
                    pass
            
            wb.close()
            
            # Return max ratio
            if ratios:
                return max(ratios)
            else:
                return None
                
        except Exception as e:
            print(f"Error reading ratio from {node_name}: {e}")
            return None

# ==================== CALCULATE CONTROL RATIOS ====================
    
    def _calculate_control_ratios(self, f1, f2, f3, ws, node_data):
        """
        Calculate 5 control ratios for a given load combination (f1, f2, f3)
        
        Returns: dict with keys 'compression', 'tensile', 'anchor_bolt', 'x_shear', 'y_shear'
        """
        import math
        
        def safe_float(value, default=0):
            """Safely convert value to float"""
            try:
                return float(value) if value else default
            except:
                return default
        
        # Get values from sheet (same as original code)
        n_val = safe_float(ws['F3'].value)
        b_val = safe_float(ws['G3'].value)
        p1_val = safe_float(ws['H3'].value)
        d_val = safe_float(ws['A3'].value)
        bf_val = safe_float(ws['B3'].value)
        
        np_val = safe_float(ws['A11'].value)
        bp_val = safe_float(ws['B11'].value)
        
        fy_steel = safe_float(ws['A7'].value)
        fc_val = safe_float(ws['B7'].value)
        fy_mainbar = safe_float(ws['C7'].value)  # fy(Mainbar)
        fy_tiebar = safe_float(ws['D7'].value)  # fy(Tiebar)
        friction_mu = safe_float(ws['F15'].value)  # Friction μ from Other Coefficient
        
        # Main bar
        main_qty_raw = ws['F11'].value
        main_size_raw = ws['G11'].value
        
        # Calculate Ast
        try:
            main_qty_num = int(safe_float(main_qty_raw))
            main_size = str(main_size_raw) if main_size_raw else ''
            size_num = int(main_size.replace('D', '').strip()) if main_size else 0
            area_one_bar = math.pi * (size_num ** 2) / 4
            ast = main_qty_num * area_one_bar
        except:
            ast = 0
        
        # Tie bar
        tie_size_raw = ws['H11'].value
        try:
            tie_size_str = str(tie_size_raw) if tie_size_raw else ''
            d_tiebar = int(tie_size_str.replace('D', '').strip()) if tie_size_str else 0
        except:
            d_tiebar = 0
        
        ratios = {}
        
        # ==================== 1. COMPRESSION RATIO ====================
        # Vertical load (F3)
        vertical_load = f3
        if vertical_load < 0:
            pu = 0
        else:
            pu = vertical_load
        
        # A1 and A2
        a1 = n_val * b_val
        a2 = np_val * bp_val
        
        # Pp (AISC)
        pp = 0.85 * fc_val * a1 * math.sqrt(a2 / a1) * 0.001 if a1 != 0 else 0
        val_1_7 = 1.7 * fc_val * a1 * 0.001
        
        # Pp (ACI) - using fy(Mainbar)
        pp_aci = 0.8 * (0.85 * fc_val * (a2 - ast) + fy_mainbar * ast) / 1000
        
        # Pn and φPn
        pn = min(min(pp, val_1_7), pp_aci)
        phi_pn = 0.65 * pn
        
        # Rat1 (Compression)
        rat1 = abs(vertical_load) / phi_pn if phi_pn != 0 else 0
        
        # m, n, n' for plate yielding
        m_val = (n_val - 0.95 * d_val) / 2
        n_calc = (b_val - 0.8 * bf_val) / 2
        n_prime = (d_val * bf_val) ** 0.5 / 4
        l_val = max(m_val, n_calc, n_prime)
        
        # tmin
        if fy_steel != 0 and b_val != 0 and n_val != 0:
            tmin = l_val * math.sqrt(2 * pu * 1000 / (0.9 * fy_steel * b_val * n_val))
        else:
            tmin = 0
        
        # Rat2 (Plate yielding)
        rat2 = tmin / p1_val if p1_val != 0 else 0
        
        # Max compression ratio
        ratios['compression'] = max(rat1, rat2)
        
        # ==================== 2. TENSILE RATIO ====================
        vertical_load_tensile = f3
        
        # Nua
        nb_val = safe_float(ws['F7'].value)
        if vertical_load_tensile < 0:
            nua = abs(vertical_load_tensile / nb_val) if nb_val != 0 else 0
        else:
            nua = 0
        
        # Mu
        a_val = safe_float(ws['J3'].value)
        tw_val = safe_float(ws['C3'].value)
        mu = nua * (a_val / 2 - tw_val / 2)
        
        # beff
        beff = (b_val / 2 - tw_val / 2) * 2
        
        # t_reqd
        if fy_steel != 0 and beff != 0:
            t_reqd = math.sqrt(4 * mu * 1000 / (0.9 * fy_steel * beff))
        else:
            t_reqd = 0
        
        # Rat3 (Tensile)
        rat3 = t_reqd / p1_val if p1_val != 0 else 0
        
        ratios['tensile'] = rat3
        
        # ==================== 3. ANCHOR BOLT RATIO ====================
        vertical_load_ab = f3
        h_load_x = f1
        h_load_y = f2
        
        # Steel tension
        ase_val = safe_float(ws['H7'].value)
        futa_val = safe_float(ws['I7'].value)
        nsa = ase_val * futa_val / 1000
        
        # Pullout resistance
        db_num = safe_float(ws['G7'].value)
        bolt_data = self.get_anchor_bolt_data(f"M{int(db_num)}")
        w_val = safe_float(bolt_data.get('W', 0)) if bolt_data else 0
        nutw_val = safe_float(bolt_data.get('NutW', 0)) if bolt_data else 0
        
        if w_val > 0:
            abrg = w_val ** 2 - math.pi / 4 * (db_num ** 2)
        else:
            abrg = 0.866 * (nutw_val ** 2)
        
        nap = 8 * abrg * fc_val / 1000
        npn = 1.4 * nap
        
        # Side-face blowout
        c_val = safe_float(ws['I3'].value)
        a_val = safe_float(ws['J3'].value)
        ca1 = min((bp_val - a_val) / 2, (np_val - c_val) / 2)
        ca2 = max((bp_val - a_val) / 2, (np_val - c_val) / 2)
        hef_val = safe_float(ws['L7'].value)
        
        ca2_ca1_ratio = ca2 / ca1 if ca1 != 0 else 0
        
        if ca2_ca1_ratio < 1:
            f1_factor = 1
        elif ca2_ca1_ratio > 3:
            f1_factor = 1 + (ca1 / ca2) / 4
        else:
            f1_factor = (1 + ca2_ca1_ratio) / 4
        
        if hef_val > 2.5 * ca1:
            nsb = f1_factor * npn
        else:
            nsb = npn
        
        # Pryout
        np_pier = safe_float(ws['A11'].value)
        bp_pier = safe_float(ws['B11'].value)
        
        ca1_p = min((bp_pier - a_val) / 2, (np_pier - c_val) / 2)
        np_effect = min(2 * ca1_p, hef_val / 2)
        ncp = 16 * np_effect * fc_val / 1000
        
        # Tension capacity
        nna = min(nsa, nsb, ncp)
        phi_nn = 0.65 * nna
        
        # Rat4 (Tension)
        if vertical_load_ab < 0:
            nua_ab = abs(vertical_load_ab)
        else:
            nua_ab = 0
        
        rat4 = nua_ab / phi_nn if phi_nn != 0 else 0
        
        # Shear
        vsa = (f1_factor * nsa if hef_val > 2.5 * ca1 else nsa) if abs(h_load_x) > 0 or abs(h_load_y) > 0 else nsa
        phi_vn = 0.65 * vsa
        
        # Vua
        h_resultant = math.sqrt(h_load_x ** 2 + h_load_y ** 2)
        if vertical_load_ab < 0:
            pu_shear = 0
        else:
            pu_shear = vertical_load_ab
        
        nb_val = safe_float(ws['F7'].value)
        vua_calc = (h_resultant - friction_mu * pu_shear) / nb_val if nb_val != 0 else 0
        vua = max(vua_calc, 0)
        
        # Rat5 (Shear)
        rat5 = vua / phi_vn if phi_vn != 0 else 0
        
        # Rat6 (Interaction)
        d104_val = 1 if rat5 > 0.2 else 0
        d105_val = 1 if rat4 > 0.2 else 0
    
        if d104_val * d105_val == 1:
            rat6 = rat5 + rat4
        else:
            rat6 = max(rat5, rat4)
    
        ratios['anchor_bolt'] = rat6
        
        # ==================== SETUP PIER/SPACING BASED ON BETA ====================
        beta_val = node_data.get('beta', 90)

        if beta_val == 0:
            pier_x  = np_val
            space_x = safe_float(ws['I3'].value)   # C
            pier_y  = bp_val
            space_y = safe_float(ws['J3'].value)   # A
        else:
            pier_x  = bp_val
            space_x = safe_float(ws['J3'].value)   # A
            pier_y  = np_val
            space_y = safe_float(ws['I3'].value)   # C

        # ==================== 4. X-SHEAR CHECK ====================
        vertical_load_x = f3
        h_load_x_pier = f1

        if vertical_load_x < 0:
            pu_x = 0
        else:
            pu_x = vertical_load_x

        vx_calc = abs(h_load_x_pier) - friction_mu * pu_x
        vx = max(vx_calc, 0)

        atie = math.pi * (d_tiebar ** 2) / 4 if d_tiebar > 0 else 1

        e_bp = safe_float(ws['K3'].value)
        f_bp = safe_float(ws['L3'].value)
        layer1 = safe_float(ws['K11'].value)
        layer2 = safe_float(ws['L11'].value)
        cover_c = safe_float(ws['D11'].value)

        ldh_value = 0.24 * fy_tiebar * d_tiebar / (fc_val ** 0.5) if fc_val > 0 and d_tiebar > 0 else 1

        # ✅ dùng pier_x, space_x
        tan35 = math.tan(35 * math.pi / 180)
        lda_a_lx = (pier_x + space_x) / 2 - db_num - cover_c - layer1 * tan35
        lda_a_rx = (pier_x - 2 * cover_c) - lda_a_lx
        lda_b_lx = (pier_x + space_x) / 2 - db_num - cover_c - (layer1 + layer2) * tan35
        lda_b_rx = (pier_x - 2 * cover_c) - lda_b_lx

        fsx = min(lda_a_lx, lda_a_rx, lda_b_lx, lda_b_rx) / ldh_value * fy_tiebar if ldh_value != 0 else 0

        x_legs = safe_float(ws['I11'].value)
        atie_reqd_x = vx / 2 * 1000 / (0.75 * fsx) if fsx != 0 else 0
        atie_prov_x = x_legs * atie

        rat9 = atie_reqd_x / atie_prov_x if atie_prov_x != 0 else 0
        ratios['x_shear'] = rat9

        # ==================== 5. Y-SHEAR CHECK ====================
        vertical_load_y = f3
        h_load_y_pier = f2

        if vertical_load_y < 0:
            pu_y = 0
        else:
            pu_y = vertical_load_y

        vy_calc = abs(h_load_y_pier) - friction_mu * pu_y
        vy = max(vy_calc, 0)

        # ✅ dùng pier_y, space_y
        lda_a_ly = (pier_y + space_y) / 2 - db_num - cover_c - layer1 * tan35
        lda_a_ry = (pier_y - 2 * cover_c) - lda_a_ly
        lda_b_ly = (pier_y + space_y) / 2 - db_num - cover_c - (layer1 + layer2) * tan35
        lda_b_ry = (pier_y - 2 * cover_c) - lda_b_ly

        fsy = min(lda_a_ly, lda_a_ry, lda_b_ly, lda_b_ry) / ldh_value * fy_tiebar if ldh_value != 0 else 0

        y_legs = safe_float(ws['J11'].value)
        atie_reqd_y = vy / 2 * 1000 / (0.75 * fsy) if fsy != 0 else 0
        atie_prov_y = y_legs * atie

        rat10 = atie_reqd_y / atie_prov_y if atie_prov_y != 0 else 0
        ratios['y_shear'] = rat10
        
        return ratios

# ==================== HINGE TYPE XLSX GENERATION ====================
    
    def parse_section_name(self, section):
        """Parse section name like 'H340X250X9X14' to get d, bf, tw, tf"""
        import re
        # Remove 'H' or 'BH' prefix
        section_clean = section.upper().replace('BH', '').replace('H', '')
        # Split by 'X'
        parts = section_clean.split('X')
        if len(parts) >= 4:
            try:
                return {
                    'd': float(parts[0]),
                    'bf': float(parts[1]),
                    'tw': float(parts[2]),
                    'tf': float(parts[3])
                }
            except:
                pass
        return {'d': '', 'bf': '', 'tw': '', 'tf': ''}
    
    def get_hinge_type_row_data(self, section):
        """Get row data from Hinge Type sheet matching section"""
        tree = self.material_trees.get('Hinge Type')
        if not tree:
            return None
        
        for item in tree.get_children():
            values = tree.item(item, 'values')
            if values and values[0]:
                if section.upper() in values[0].upper():
                    # Return dict with all column values
                    cols = ['Column size', 'Type', 'No.AB', 'P1', 'N', 'A', 'B', 'C', 'E', 'F', 'P2', 'Y',
                           'Np', 'Bp', 'c', 'nrb', 'drb', 'dtb', 'X-leg', 'Y-leg', 'Layer 1', 'Layer 2']
                    return {cols[i]: values[i] if i < len(values) else '' for i in range(len(cols))}
        return None
    
    def get_material_strength(self, material_name, material_type):
        """Get strength value from Material Strength sheet"""
        tree_map = {
            'anchor_bolt': 'anchor_bolt',
            'base_plate': 'base_plate',
            'concrete': 'concrete',
            'rebar': 'rebar'
        }
        
        tree = self.material_trees.get(tree_map.get(material_type))
        if not tree or not material_name:
            return ''
        
        for item in tree.get_children():
            values = tree.item(item, 'values')
            if values and len(values) >= 2 and values[0] == material_name:
                return values[1]
        return ''
    
    def get_anchor_bolt_data(self, db_value):
        """Get anchor bolt data from Anchor Bolt Table"""
        tree = self.material_trees.get('Anchor Bolt Table')
        if not tree:
            return None
        
        # Parse db value (e.g., "M30" -> 30)
        try:
            db_num = int(db_value.replace('M', '').strip())
        except:
            return None
        
        for item in tree.get_children():
            values = tree.item(item, 'values')
            if values and values[0]:
                try:
                    if int(float(values[0])) == db_num:
                        # Return dict: db, Rmin, a, W, T, S, NutW, nt, Nut Allowance, Edge Min, Leng A1, Leng A2
                        return {
                            'db': values[0] if len(values) > 0 else '',
                            'Rmin': values[1] if len(values) > 1 else '',
                            'a': values[2] if len(values) > 2 else '',
                            'W': values[3] if len(values) > 3 else '',
                            'T': values[4] if len(values) > 4 else '',
                            'S': values[5] if len(values) > 5 else '',
                            'NutW': values[6] if len(values) > 6 else '',
                            'nt': values[7] if len(values) > 7 else '',
                            'Nut_Allowance': values[8] if len(values) > 8 else '',
                            'Edge_Min': values[9] if len(values) > 9 else '',
                            'Leng_A1': values[10] if len(values) > 10 else '',
                            'Leng_A2': values[11] if len(values) > 11 else ''
                        }
                except:
                    continue
        return None
    
    def calculate_anchor_bolt_ase(self, db_value):
        """Calculate Ase = PI()/4*(db-25.4*0.9743/nt)^2"""
        import math
        
        bolt_data = self.get_anchor_bolt_data(db_value)
        if not bolt_data:
            return ''
        
        try:
            db = float(bolt_data['db'])
            nt = float(bolt_data['nt'])
            
            ase = (math.pi / 4) * ((db - 25.4 * 0.9743 / nt) ** 2)
            return round(ase, 1)
        except:
            return ''
    
    def calculate_anchor_bolt_values(self, db_value, bolt_data, futa):
            """Calculate A1, Proj, heff for anchor bolt"""
            try:
                # A1 = Leng A1 (từ cột K trong Anchor Bolt Table)
                a1 = float(bolt_data.get('Leng_A1', 0)) if bolt_data.get('Leng_A1') else ''

                # Proj = Nut Allowance (từ cột I trong Anchor Bolt Table)
                proj = float(bolt_data.get('Nut_Allowance', 0)) if bolt_data.get('Nut_Allowance') else ''

                # heff = A1 - Proj - 30 - S
                s_val = float(bolt_data.get('S', 0)) if bolt_data.get('S') else 0
                if a1 != '' and proj != '':
                    heff = a1 - proj - 30 - s_val
                else:
                    heff = ''

                return a1, proj, heff
            except Exception as e:
                print(f"Error calculating anchor bolt values: {e}")
                return '', '', ''
    
    def calculate_main_bar_ase(self, qty, size):
        """Calculate Main Bar Ase based on qty and rebar size"""
        import math
        try:
            qty_val = int(qty) if qty else 0
            # Parse size: "D19" -> 19
            size_num = int(size.replace('D', '').strip()) if size else 0
            
            # Ase = qty * π * d² / 4
            ase = qty_val * math.pi * (size_num ** 2) / 4
            return round(ase, 1)
        except:
            return ''
    
    def create_or_update_hinge_fixed_xlsx(self, node_name, node_data):
        """Create or update Hinge Type.xlsx / Fixed Type.xlsx with node sheet"""
        if not self.bpl_folder:
            messagebox.showerror("Error", "Base Plate Design folder not found!")
            return
        
        # Determine file name based on hinge/fixed type
        hinge_fixed = node_data.get('hinge_fixed_type', 'Hinge Type')
        if 'Fixed' in hinge_fixed:
            filename = 'Fixed Type.xlsx'
        else:
            filename = 'Hinge Type.xlsx'
        
        filepath = os.path.join(self.bpl_folder, filename)
        
        # Load or create workbook
        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # Create or get sheet for this node
        sheet_name = f"Node {node_name}"
        sheet_exists = sheet_name in wb.sheetnames
        
        # Always delete and recreate sheet to reflect latest dropdown selections
        if sheet_exists:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        is_update_mode = False  # Always write full input data
        
        # Get all required data
        section = node_data.get('section', '')
        section_parts = self.parse_section_name(section)
        hinge_data = self.get_hinge_type_row_data(section)
        
        # Materials
        fy_steel = self.get_material_strength(node_data.get('material_base_plate'), 'base_plate')
        fc = self.get_material_strength(node_data.get('material_concrete'), 'concrete')
        fy_mainbar = self.get_material_strength(node_data.get('material_mainbar'), 'rebar')
        fy_tiebar = self.get_material_strength(node_data.get('material_tiebar'), 'rebar')
        futa = self.get_material_strength(node_data.get('material_anchor_bolt'), 'anchor_bolt')
        
        # Anchor Bolt
        bolt_size = node_data.get('bolt_size', '')
        bolt_data = self.get_anchor_bolt_data(bolt_size)
        nb = hinge_data.get('No.AB', '') if hinge_data else ''
        db = bolt_size.replace('M', '') if bolt_size else ''
        ase_bolt = self.calculate_anchor_bolt_ase(bolt_size)
        a1, proj, heff = self.calculate_anchor_bolt_values(bolt_size, bolt_data, futa) if bolt_data else ('', '', '')
        
        # Main Bar
        main_qty = hinge_data.get('nrb', '') if hinge_data else ''
        main_size = hinge_data.get('drb', '') if hinge_data else ''
        main_ase = self.calculate_main_bar_ase(main_qty, main_size)
        
        # ==================== INPUT DATA SECTION (only for new sheets) ====================
        if not is_update_mode:
            # ==================== ROW 1: Base Plate Detail Title ====================
            ws.merge_cells('A1:D1')
            ws['A1'] = section
            ws['A1'].font = Font(bold=True, size=12)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            ws.merge_cells('F1:L1')
            ws['F1'] = 'Base Plate Detail'
            ws['F1'].font = Font(bold=True, size=12)
            ws['F1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # ==================== ROW 2: Headers ====================
            ws['A2'] = 'd'
            ws['B2'] = 'bf'
            ws['C2'] = 'tw'
            ws['D2'] = 'tf'
            
            ws['F2'] = 'N'
            ws['G2'] = 'B'
            ws['H2'] = 'P1'
            ws['I2'] = 'C'
            ws['J2'] = 'A'
            ws['K2'] = 'E'
            ws['L2'] = 'F'
            
            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}2'].font = Font(bold=True)
                ws[f'{col}2'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}2'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            
            # ==================== ROW 3: Values ====================
            if 'edited_section' in node_data and node_data['edited_section']:
                ws['A3'] = node_data['edited_section'].get('d', section_parts.get('d', ''))
                ws['B3'] = node_data['edited_section'].get('bf', section_parts.get('bf', ''))
                ws['C3'] = node_data['edited_section'].get('tw', section_parts.get('tw', ''))
                ws['D3'] = node_data['edited_section'].get('tf', section_parts.get('tf', ''))
            else:
                ws['A3'] = section_parts.get('d', '')
                ws['B3'] = section_parts.get('bf', '')
                ws['C3'] = section_parts.get('tw', '')
                ws['D3'] = section_parts.get('tf', '')
            
            # Check if edited values exist for Base Plate Detail
            if 'edited_base_plate_detail' in node_data and node_data['edited_base_plate_detail']:
                ws['F3'] = node_data['edited_base_plate_detail'].get('N', hinge_data.get('N', '') if hinge_data else '')
                ws['G3'] = node_data['edited_base_plate_detail'].get('B', hinge_data.get('B', '') if hinge_data else '')
                ws['H3'] = node_data['edited_base_plate_detail'].get('P1', hinge_data.get('P1', '') if hinge_data else '')
                ws['I3'] = node_data['edited_base_plate_detail'].get('C', hinge_data.get('C', '') if hinge_data else '')
                ws['J3'] = node_data['edited_base_plate_detail'].get('A', hinge_data.get('A', '') if hinge_data else '')
                ws['K3'] = node_data['edited_base_plate_detail'].get('E', hinge_data.get('E', '') if hinge_data else '')
                ws['L3'] = node_data['edited_base_plate_detail'].get('F', hinge_data.get('F', '') if hinge_data else '')
            else:
                if hinge_data:
                    ws['F3'] = hinge_data.get('N', '')
                    ws['G3'] = hinge_data.get('B', '')
                    ws['H3'] = hinge_data.get('P1', '')
                    ws['I3'] = hinge_data.get('C', '')
                    ws['J3'] = hinge_data.get('A', '')
                    ws['K3'] = hinge_data.get('E', '')
                    ws['L3'] = hinge_data.get('F', '')
            
            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')
            
            # ==================== ROW 5: Material Title ====================
            ws.merge_cells('A5:D5')
            ws['A5'] = 'Material'
            ws['A5'].font = Font(bold=True, size=12)
            ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A5'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            ws.merge_cells('F5:L5')
            ws['F5'] = 'Anchor Bolt'
            ws['F5'].font = Font(bold=True, size=12)
            ws['F5'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F5'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # ==================== ROW 6: Headers ====================
            ws['A6'] = 'fy(Steel)'
            ws['B6'] = "f'c"
            ws['C6'] = 'fy(Mainbar)'
            ws['D6'] = 'fy(Tiebar)'
            
            ws['F6'] = 'nb'
            ws['G6'] = 'db'
            ws['H6'] = 'Ase'
            ws['I6'] = 'futa'
            ws['J6'] = 'A1'
            ws['K6'] = 'Proj'
            ws['L6'] = 'heff'
            
            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}6'].font = Font(bold=True)
                ws[f'{col}6'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}6'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            
            # ==================== ROW 7: Values ====================
            # Check if edited material values exist
            if 'edited_material' in node_data and node_data['edited_material']:
                ws['A7'] = node_data['edited_material'].get('fy_steel', fy_steel)
                ws['B7'] = node_data['edited_material'].get('fc', fc)
                ws['C7'] = node_data['edited_material'].get('fy_mainbar', fy_mainbar)
                ws['D7'] = node_data['edited_material'].get('fy_tiebar', fy_tiebar)
            else:
                ws['A7'] = fy_steel
                ws['B7'] = fc
                ws['C7'] = fy_mainbar
                ws['D7'] = fy_tiebar
            
            # Check if edited anchor bolt values exist
            if 'edited_anchor_bolt' in node_data and node_data['edited_anchor_bolt']:
                ws['F7'] = node_data['edited_anchor_bolt'].get('nb', nb)
                ws['G7'] = node_data['edited_anchor_bolt'].get('db', db)
                ws['H7'] = node_data['edited_anchor_bolt'].get('Ase', ase_bolt)
                ws['I7'] = node_data['edited_anchor_bolt'].get('futa', futa)
                ws['J7'] = node_data['edited_anchor_bolt'].get('A1', a1)
                ws['K7'] = node_data['edited_anchor_bolt'].get('Proj', proj)
                ws['L7'] = node_data['edited_anchor_bolt'].get('heff', heff)
            else:
                ws['F7'] = nb
                ws['G7'] = db
                ws['H7'] = ase_bolt
                ws['I7'] = futa
                ws['J7'] = a1
                ws['K7'] = proj
                ws['L7'] = heff
            
            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}7'].alignment = Alignment(horizontal='center', vertical='center')
            
            # ==================== ROW 9: Pier Detail Title ====================
            ws.merge_cells('A9:D9')
            ws['A9'] = 'Pier Detail'
            ws['A9'].font = Font(bold=True, size=12)
            ws['A9'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A9'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            ws.merge_cells('F9:L9')
            ws['F9'] = 'Main Bar                                Tie Bar'
            ws['F9'].font = Font(bold=True, size=12)
            ws['F9'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F9'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # ==================== ROW 10: Headers ====================
            ws['A10'] = 'Np'
            ws['B10'] = 'Bp'
            ws['C10'] = 'TG'
            ws['D10'] = 'Cover c'
            
            ws['F10'] = 'Qty'
            ws['G10'] = 'Size'
            ws['H10'] = 'Size'
            ws['I10'] = 'X-legs'
            ws['J10'] = 'Y-legs'
            ws['K10'] = 'Layer 1'
            ws['L10'] = 'Layer 2'
            
            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}10'].font = Font(bold=True)
                ws[f'{col}10'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}10'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            
            # ==================== ROW 11: Values ====================
            # Check if edited values exist, use them; otherwise use hinge_data
            if 'edited_pier_detail' in node_data and node_data['edited_pier_detail']:
                ws['A11'] = node_data['edited_pier_detail'].get('Np', hinge_data.get('Np', '') if hinge_data else '')
                ws['B11'] = node_data['edited_pier_detail'].get('Bp', hinge_data.get('Bp', '') if hinge_data else '')
                ws['C11'] = node_data['edited_pier_detail'].get('TG', 25)  # Default 25 if not set
                ws['D11'] = node_data['edited_pier_detail'].get('c', hinge_data.get('c', '') if hinge_data else '')
            else:
                if hinge_data:
                    ws['A11'] = hinge_data.get('Np', '')
                    ws['B11'] = hinge_data.get('Bp', '')
                ws['C11'] = 25  # TG fixed value
                if hinge_data:
                    ws['D11'] = hinge_data.get('c', '')
            
            # Check if edited main bar values exist
            if 'edited_main_bar' in node_data and node_data['edited_main_bar']:
                ws['F11'] = node_data['edited_main_bar'].get('Qty', main_qty)
                ws['G11'] = node_data['edited_main_bar'].get('Size', main_size)
            else:
                ws['F11'] = main_qty
                ws['G11'] = main_size
            
            if hinge_data:
                ws['H11'] = hinge_data.get('dtb', '')
                ws['I11'] = hinge_data.get('X-leg', '')
                ws['J11'] = hinge_data.get('Y-leg', '')
                ws['K11'] = hinge_data.get('Layer 1', '')
                ws['L11'] = hinge_data.get('Layer 2', '')
            
            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}11'].alignment = Alignment(horizontal='center', vertical='center')
            
            # ==================== ROW 13: Other Coefficient Title ====================
            ws.merge_cells('F13:I13')
            ws['F13'] = 'Other Coefficient'
            ws['F13'].font = Font(bold=True, size=11)
            ws['F13'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F13'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            
            # ==================== ROW 14: Headers ====================
            ws['F14'] = 'Friction μ'
            ws['G14'] = 'Ψc,P'
            ws['H14'] = 'Ψe'
            ws['I14'] = 'λ'
            
            for col in ['F', 'G', 'H', 'I']:
                ws[f'{col}14'].font = Font(bold=True)
                ws[f'{col}14'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}14'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            
            # ==================== ROW 15: Values ====================
            # Check if edited other coefficient values exist
            if 'edited_other_coeff' in node_data and node_data['edited_other_coeff']:
                ws['F15'] = node_data['edited_other_coeff'].get('friction_mu', 0.4)
                ws['G15'] = node_data['edited_other_coeff'].get('psi_c_p', 1.4)
                ws['H15'] = node_data['edited_other_coeff'].get('psi_e', 1)
                ws['I15'] = node_data['edited_other_coeff'].get('lambda', 1)
            else:
                ws['F15'] = 0.4  # Friction μ default value
                ws['G15'] = 1.4  # Ψc,P default value
                ws['H15'] = 1    # Ψe default value
                ws['I15'] = 1    # λ default value
            
            for col in ['F', 'G', 'H', 'I']:
                ws[f'{col}15'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 3
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 10
            ws.column_dimensions['K'].width = 10
            ws.column_dimensions['L'].width = 10
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, 16):
                for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                    if row <= 11 or (row >= 13 and row <= 15):
                        if (row <= 11 and col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']) or \
                           (row >= 13 and col in ['F', 'G', 'H', 'I']):
                            ws[f'{col}{row}'].border = thin_border

# ==================== CONTROL LOAD TABLE (N1:S6) ====================
        
        # Read reaction data from CSV
        reaction_file = os.path.join(self.bpl_folder, 'reaction_data_sap2000.csv')
        
        if os.path.exists(reaction_file):
            try:
                # Read CSV and find data for this node
                with open(reaction_file, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                
                # Skip header lines (first 3 lines)
                node_reactions = []
                for line in lines[3:]:
                    parts = line.strip().split(',')
                    if len(parts) >= 9 and parts[0].strip() == node_name:
                        # Found reaction for this node
                        node_reactions.append({
                            'node': parts[0].strip(),
                            'loadcase': parts[1].strip(),
                            'casetype': parts[2].strip(),
                            'f1': parts[3].strip(),
                            'f2': parts[4].strip(),
                            'f3': parts[5].strip(),
                            'm1': parts[6].strip() if len(parts) > 6 else '',
                            'm2': parts[7].strip() if len(parts) > 7 else '',
                            'm3': parts[8].strip() if len(parts) > 8 else ''
                        })
                
                if node_reactions:
                    # ==================== INITIALIZE TRACKING FOR MAX RATIOS ====================
                    max_ratios = {
                        'compression': {'ratio': -1, 'lc': None, 'f1': 0, 'f2': 0, 'f3': 0},
                        'tensile': {'ratio': -1, 'lc': None, 'f1': 0, 'f2': 0, 'f3': 0},
                        'anchor_bolt': {'ratio': -1, 'lc': None, 'f1': 0, 'f2': 0, 'f3': 0},
                        'x_shear': {'ratio': -1, 'lc': None, 'f1': 0, 'f2': 0, 'f3': 0},
                        'y_shear': {'ratio': -1, 'lc': None, 'f1': 0, 'f2': 0, 'f3': 0}
                    }
                    
                    # ==================== LOOP THROUGH ALL COMBOS ====================
                    print(f"Debug: Processing {len(node_reactions)} load combinations for node {node_name}")
                    
                    for combo in node_reactions:
                        lc_name = combo['loadcase']
                        try:
                            f1_val = float(combo['f1'])
                            f2_val = float(combo['f2'])
                            f3_val = float(combo['f3'])
                        except:
                            continue
                        
                        # Calculate 5 ratios for this combo
                        ratios = self._calculate_control_ratios(
                            f1=f1_val, f2=f2_val, f3=f3_val,
                            ws=ws, node_data=node_data
                        )
                        
                        # Update max ratios if this combo is better
                        for check_type in ['compression', 'tensile', 'anchor_bolt', 'x_shear', 'y_shear']:
                            current_ratio = ratios.get(check_type, 0)
                            if current_ratio > max_ratios[check_type]['ratio']:
                                max_ratios[check_type]['ratio'] = current_ratio
                                max_ratios[check_type]['lc'] = lc_name
                                max_ratios[check_type]['f1'] = f1_val
                                max_ratios[check_type]['f2'] = f2_val
                                max_ratios[check_type]['f3'] = f3_val
                        
                        print(f"Debug: LC {lc_name}: Compress={ratios.get('compression', 0):.3f}, "
                              f"Tensile={ratios.get('tensile', 0):.3f}, "
                              f"AB={ratios.get('anchor_bolt', 0):.3f}, "
                              f"X={ratios.get('x_shear', 0):.3f}, "
                              f"Y={ratios.get('y_shear', 0):.3f}")
                    
                    # ==================== FILL CONTROL LOAD TABLE WITH MAX RATIOS ====================
                    
                    # ROW 1: Headers
                    ws['N1'] = 'Control Load'
                    ws['O1'] = 'L/C'
                    ws['P1'] = 'F1'
                    ws['Q1'] = 'F2'
                    ws['R1'] = 'F3'
                    ws['S1'] = 'Ratio'
                    
                    for col in ['N', 'O', 'P', 'Q', 'R', 'S']:
                        ws[f'{col}1'].font = Font(bold=True, size=11)
                        ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
                        ws[f'{col}1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                    
                    # ROW 2-6: Data with best combos
                    control_loads = [
                        ('N2', 'Compression', 'compression'),
                        ('N3', 'Tensile', 'tensile'),
                        ('N4', 'Anchor Bolt', 'anchor_bolt'),
                        ('N5', 'X-Shear check', 'x_shear'),
                        ('N6', 'Y-Shear check', 'y_shear')
                    ]
                    
                    for cell_ref, label, check_type in control_loads:
                        row = cell_ref[1]  # Get row number
                        best_combo = max_ratios[check_type]
                        
                        # Control Load label
                        ws[cell_ref] = label
                        ws[cell_ref].alignment = Alignment(horizontal='left', vertical='center')
                        
                        # L/C (best combo for this check type)
                        ws[f'O{row}'] = best_combo['lc'] if best_combo['lc'] else ''
                        ws[f'O{row}'].alignment = Alignment(horizontal='center', vertical='center')
                        
                        # F1, F2, F3
                        ws[f'P{row}'] = best_combo['f1']
                        ws[f'P{row}'].alignment = Alignment(horizontal='center', vertical='center')
                        
                        ws[f'Q{row}'] = best_combo['f2']
                        ws[f'Q{row}'].alignment = Alignment(horizontal='center', vertical='center')
                        
                        ws[f'R{row}'] = best_combo['f3']
                        ws[f'R{row}'].alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Ratio (will be filled during calculation)
                        ws[f'S{row}'] = ''
                        ws[f'S{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Set column widths for Control Load table
                    ws.column_dimensions['N'].width = 15
                    ws.column_dimensions['O'].width = 12
                    ws.column_dimensions['P'].width = 12
                    ws.column_dimensions['Q'].width = 12
                    ws.column_dimensions['R'].width = 12
                    ws.column_dimensions['S'].width = 12
                    
                    # Add borders to Control Load table
                    for row in range(1, 7):
                        for col in ['N', 'O', 'P', 'Q', 'R', 'S']:
                            ws[f'{col}{row}'].border = thin_border
                    
                    print(f"Control Load table added for Node {node_name}")
                    print(f"Best LCs: Compress={max_ratios['compression']['lc']} ({max_ratios['compression']['ratio']:.3f}), "
                          f"Tensile={max_ratios['tensile']['lc']} ({max_ratios['tensile']['ratio']:.3f}), "
                          f"AB={max_ratios['anchor_bolt']['lc']} ({max_ratios['anchor_bolt']['ratio']:.3f}), "
                          f"X={max_ratios['x_shear']['lc']} ({max_ratios['x_shear']['ratio']:.3f}), "
                          f"Y={max_ratios['y_shear']['lc']} ({max_ratios['y_shear']['ratio']:.3f})")
                
                else:
                    print(f"No reaction data found for Node {node_name}")
                    
            except Exception as e:
                print(f"Error reading reaction data: {e}")
        else:
            print(f"Reaction file not found: {reaction_file}")

# ==================== COMPRESSION DESIGN CALCULATION (Row 13-39) ====================
        
        # Get values from sheet and convert to float
        try:
            def safe_float(value, default=0):
                """Safely convert value to float"""
                try:
                    return float(value) if value else default
                except:
                    return default
            
            n_val = safe_float(ws['F3'].value)
            b_val = safe_float(ws['G3'].value)
            p1_val = safe_float(ws['H3'].value)  # Base plate thickness
            d_val = safe_float(ws['A3'].value)
            bf_val = safe_float(ws['B3'].value)
            
            np_val = safe_float(ws['A11'].value)
            bp_val = safe_float(ws['B11'].value)
            
            fy_steel = safe_float(ws['A7'].value)
            fc_val = safe_float(ws['B7'].value)
            fy_mainbar = safe_float(ws['C7'].value)  # fy(Mainbar)
            fy_tiebar = safe_float(ws['D7'].value)  # fy(Tiebar)
            
            # Main bar
            main_qty_raw = ws['F11'].value
            main_size_raw = ws['G11'].value
            
            # Calculate Ast
            try:
                main_qty_num = int(safe_float(main_qty_raw))
                # Parse D19 -> 19
                main_size = str(main_size_raw) if main_size_raw else ''
                size_num = int(main_size.replace('D', '').strip()) if main_size else 0
                # Area = π × d² / 4
                import math
                area_one_bar = math.pi * (size_num ** 2) / 4
                ast = main_qty_num * area_one_bar
            except Exception as e:
                print(f"Error calculating Ast: {e}")
                ast = 0
            
            vertical_load = safe_float(ws['R2'].value)  # F3 from Control Load
            
            # Calculate Pu
            if vertical_load < 0:
                pu = 0
            else:
                pu = vertical_load
            
            print(f"Debug values: N={n_val}, B={b_val}, P1={p1_val}, d={d_val}, bf={bf_val}")
            print(f"Debug values: Np={np_val}, Bp={bp_val}, fy={fy_steel}, fc={fc_val}")
            print(f"Debug values: Ast={ast}, Vertical Load={vertical_load}")
            
            # ==================== ROW 13-15: Title & Control ====================
            ws['A13'] = '1. Base Plate Design'
            ws['A13'].font = Font(bold=True, size=11)
            
            ws['A14'] = 'Control Case'
            ws['B14'] = ws['O2'].value if ws['O2'].value else ''  # L/C from Control Load
            
            ws['A15'] = 'Vertical Load'
            ws['B15'] = vertical_load
            ws['C15'] = 'kN'
            
            # ==================== ROW 16-28: Concrete Bearing Limit ====================
            ws['A16'] = '* Concrete Bearing Limit'
            ws['A16'].font = Font(italic=True, bold=True)
            
            # Row 17: As per AISC 360-16
            ws['A17'] = 'As per AISC 360-16:'
            
            # Row 18: A1
            ws['A18'] = 'A1='
            a1 = n_val * b_val
            ws['B18'] = a1
            ws['C18'] = 'mm2'
            
            print(f"Debug A1 calculation: {n_val} * {b_val} = {a1}")
            
            # Row 19: A2
            ws['A19'] = 'A2='
            a2 = np_val * bp_val
            ws['B19'] = a2
            ws['C19'] = 'mm2'
            
            # Row 20: Pp
            ws['A20'] = 'Pp='
            pp = 0.85 * fc_val * a1 * math.sqrt(a2 / a1) * 0.001 if a1 != 0 else 0
            ws['B20'] = pp
            ws['C20'] = 'kN'
            
            # Row 21: 1.7f'cA1
            ws['A21'] = "1.7f'cA1="
            val_1_7 = 1.7 * fc_val * a1 * 0.001
            ws['B21'] = val_1_7
            ws['C21'] = 'kN'
            
            # Row 23: As per ACI 318M-14
            ws['A23'] = 'As per ACI 318M-14:'
            
            # Row 24: Pp (ACI) - using fy(Mainbar)
            ws['A24'] = 'Pp='
            pp_aci = 0.8 * (0.85 * fc_val * (a2 - ast) + fy_mainbar * ast) / 1000  # Convert to kN
            ws['B24'] = pp_aci
            ws['C24'] = 'kN'
            
            # Row 25: Resistant factor
            ws['A25'] = 'Resitant factor φ ='
            ws['B25'] = 0.65
            
            # Row 26: Pn
            ws['A26'] = 'Pn ='
            pn = min(min(pp, val_1_7), pp_aci)
            ws['B26'] = pn
            ws['C26'] = 'kN'
            
            # Row 27: φPn
            ws['A27'] = 'φPn'
            phi_pn = 0.65 * pn
            ws['B27'] = phi_pn
            ws['C27'] = 'kN'
            
            # Row 28: Rat1
            ws['A28'] = 'Rat 1 ='
            rat1 = abs(vertical_load) / phi_pn if phi_pn != 0 else 0
            ws['B28'] = rat1
            #ws['C28'] = '(NG or OK: so sánh B28 với 1)'
            
            # ==================== ROW 30-38: Base Plate Yielding Limit ====================
            ws['A30'] = '* Base Plate Yielding Limit ( For concrete bearing )'
            ws['A30'].font = Font(italic=True, bold=True)
            
            # Row 31
            ws['A31'] = 'For the critical base plate cantilever dimension, l'
            
            # Row 32: m
            ws['A32'] = 'm='
            m_val = (n_val - 0.95 * d_val) / 2
            ws['B32'] = m_val
            ws['C32'] = 'mm'
            
            # Row 33: n
            ws['A33'] = 'n='
            n_calc = (b_val - 0.8 * bf_val) / 2
            ws['B33'] = n_calc
            ws['C33'] = 'mm'
            
            # Row 34: n'
            ws['A34'] = "n'="
            n_prime = (d_val * bf_val) ** 0.5 / 4
            ws['B34'] = n_prime
            ws['C34'] = 'mm'
            
            # Row 35: l
            ws['A35'] = 'l='
            l_val = max(m_val, n_calc, n_prime)
            ws['B35'] = l_val
            ws['C35'] = 'mm'
            
            # Row 36: Resistant factor
            ws['A36'] = 'Resitant factor φ ='
            ws['B36'] = 0.9
            
            # Row 37: tmin
            ws['A37'] = 'tmin='
            if fy_steel != 0 and b_val != 0 and n_val != 0:
                tmin = l_val * math.sqrt(2 * pu * 1000 / (0.9 * fy_steel * b_val * n_val))
            else:
                tmin = 0
            ws['B37'] = tmin
            ws['C37'] = 'mm'
            
            # Row 38: Rat2
            ws['A38'] = 'Rat 2 ='
            rat2 = tmin / p1_val if p1_val != 0 else 0
            ws['B38'] = rat2
            #ws['C38'] = '(NG or OK: so sánh B38 với 1)'
            
            # ==================== UPDATE S2: Final Ratio ====================
            final_ratio = max(rat1, rat2)
            ws['S2'] = final_ratio
            ws['S2'].alignment = Alignment(horizontal='center', vertical='center')

# ==================== ROW 41-51: TENSILE LOADING ====================
            ws['A41'] = '1.2 Design Requirements for tensile loading'
            ws['A41'].font = Font(bold=True, size=11)
            
            # Row 42: Control Case
            ws['A42'] = 'Control Case'
            ws['B42'] = ws['O3'].value if ws['O3'].value else ''  # L/C from Tensile row
            
            # Row 43: Vertical Load
            ws['A43'] = 'Vertical Load'
            vertical_load_tensile = safe_float(ws['R3'].value)  # F3 from Tensile row
            ws['B43'] = vertical_load_tensile
            ws['C43'] = 'kN'
            
            # Row 45: Title
            ws['A45'] = '* Base Plate Yielding Limit ( For bolt tension )'
            ws['A45'].font = Font(italic=True, bold=True)
            
            # Row 46: Nua
            ws['A46'] = 'Nua='
            nb_val = safe_float(ws['F7'].value)  # nb from Anchor Bolt section
            if vertical_load_tensile < 0:
                nua = abs(vertical_load_tensile / nb_val) if nb_val != 0 else 0
            else:
                nua = 0
            ws['B46'] = nua
            ws['C46'] = 'kN'
            
            # Row 47: Mu
            ws['A47'] = 'Mu='
            a_val = safe_float(ws['J3'].value)  # A from Base Plate Detail
            tw_val = safe_float(ws['C3'].value)  # tw
            mu = nua * (a_val / 2 - tw_val / 2)
            ws['B47'] = mu
            ws['C47'] = 'kN-mm'
            
            # Row 48: beff
            ws['A48'] = 'beff='
            beff = (b_val / 2 - tw_val / 2) * 2
            ws['B48'] = beff
            ws['C48'] = 'mm'
            
            # Row 49: Resistant factor
            ws['A49'] = 'Resitant factor φ ='
            ws['B49'] = 0.9
            
            # Row 50: t_reqd
            ws['A50'] = 't_reqd='
            if fy_steel != 0 and beff != 0:
                t_reqd = math.sqrt(4 * mu *1000 / (0.9 * fy_steel * beff))
            else:
                t_reqd = 0
            ws['B50'] = t_reqd
            ws['C50'] = 'mm'
            
            # Row 51: Rat 3
            ws['A51'] = 'Rat 3 ='
            rat3 = t_reqd / p1_val if p1_val != 0 else 0
            ws['B51'] = rat3
            
            # ==================== UPDATE S3: Rat3 ====================
            ws['S3'] = rat3
            ws['S3'].alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"Tensile calculation added: Nua={nua:.3f}, Mu={mu:.3f}, t_reqd={t_reqd:.3f}, Rat3={rat3:.3f}")

            # ==================== ROW 54-58: ANCHOR BOLT DESIGN HEADER ====================
            ws['A54'] = '2. Anchor Bolt Design'
            ws['A54'].font = Font(bold=True, size=11)
            
            ws['A55'] = 'Control Case'
            ws['B55'] = ws['O4'].value if ws['O4'].value else ''  # L/C from Anchor Bolt row
            
            ws['A56'] = 'Vertical Load'
            vertical_load_ab = safe_float(ws['R4'].value)  # F3
            ws['B56'] = vertical_load_ab
            ws['C56'] = 'kN'
            
            ws['A57'] = 'Horizontal load'
            h_load_x = safe_float(ws['P4'].value)  # F1 (X-Dir)
            ws['B57'] = h_load_x
            ws['C57'] = 'kN (X-Dir)'
            
            ws['A58'] = 'Horizontal load'
            h_load_y = safe_float(ws['Q4'].value)  # F2 (Y-Dir)
            ws['B58'] = h_load_y
            ws['C58'] = 'kN (Y-Dir)'
            
            # ==================== ROW 60-91: TENSILE LOADING ====================
            ws['A60'] = '2.1 Design Requirements for tensile loading'
            ws['A60'].font = Font(bold=True, italic=True, size=10)
            
            # Row 61: Steel Strength
            ws['A61'] = '* Steel Strength of Anchor in Tension'
            ws['A61'].font = Font(italic=True, bold=True)
            
            # Row 62: Nsa
            ws['A62'] = 'Nsa='
            ase_val = safe_float(ws['H7'].value)  # Ase from Anchor Bolt section
            futa_val = safe_float(ws['I7'].value)  # futa
            nsa = ase_val * futa_val / 1000  # Convert to kN
            ws['B62'] = nsa
            ws['C62'] = 'kN'
            
            # Row 64: Pullout Resistance
            ws['A64'] = '* Pullout Resistance of Anchor in Tension'
            ws['A64'].font = Font(italic=True, bold=True)
            
            # Row 65: Abrg
            ws['A65'] = 'Abrg='
            db_num = safe_float(ws['G7'].value)  # db
            
            # Get W and NutW from Anchor Bolt Table
            bolt_data = self.get_anchor_bolt_data(f"M{int(db_num)}")
            w_val = safe_float(bolt_data.get('W', 0)) if bolt_data else 0
            nutw_val = safe_float(bolt_data.get('NutW', 0)) if bolt_data else 0
            
            if w_val > 0:
                abrg = w_val ** 2 - math.pi / 4 * (db_num ** 2)
            else:
                abrg = 0.866 * (nutw_val ** 2)
            ws['B65'] = abrg
            ws['C65'] = 'mm2'
            
            # Row 66: Nap
            ws['A66'] = 'Nap='
            nap = 8 * abrg * fc_val / 1000  # Convert to kN
            ws['B66'] = nap
            ws['C66'] = 'kN'
            
            # Row 67: Ψc,P (from Other Coefficient G15)
            ws['A67'] = 'Ψc,P ='
            ws['B67'] = f'=G15'
            
            # Row 68: Npn
            ws['A68'] = 'Npn'
            npn = 1.4 * nap
            ws['B68'] = npn
            ws['C68'] = 'kN'
            
            # Row 70: Side-face Blowout
            ws['A70'] = '* Side-face Blowout Resistance of Anchor in Tension'
            ws['A70'].font = Font(italic=True, bold=True)
            
            # Row 71: Ca1
            ws['A71'] = 'Ca1='
            bp_val = safe_float(ws['B11'].value)  # Bp from Pier Detail
            np_pier = safe_float(ws['A11'].value)  # Np from Pier Detail
            a_val = safe_float(ws['J3'].value)  # A from Base Plate Detail (khoảng cách bolt)
            c_val = safe_float(ws['I3'].value)  # C from Base Plate Detail (khoảng cách bolt)
            
            ca1 = min((bp_val - a_val) / 2, (np_pier - c_val) / 2)
            ws['B71'] = ca1
            ws['C71'] = 'mm'
            
            # Row 72: 2.5Ca1
            ws['A72'] = '2.5Ca1='
            val_2_5ca1 = 2.5 * ca1
            ws['B72'] = val_2_5ca1
            ws['C72'] = 'mm'
            
            # Row 73: hef
            ws['A73'] = 'hef='
            hef_val = safe_float(ws['L7'].value)  # heff from Anchor Bolt
            ws['B73'] = hef_val
            ws['C73'] = 'mm'
            
            # Row 74: Ca2
            ws['A74'] = 'Ca2='
            ca2 = max((bp_val - a_val) / 2, (np_pier - c_val) / 2)
            ws['B74'] = ca2
            ws['C74'] = 'mm'
            
            # Row 75: ca2/ca1
            ws['A75'] = 'ca2/Ca1='
            ca2_ca1_ratio = ca2 / ca1 if ca1 != 0 else 0
            ws['B75'] = ca2_ca1_ratio
            
            # Row 76: f1
            ws['A76'] = 'f1='
            if ca2_ca1_ratio < 1:
                f1 = 1
            elif ca2_ca1_ratio > 3:
                f1 = 1 + (ca1 / ca2) / 4
            else:
                f1 = (1 + ca2_ca1_ratio) / 4
            ws['B76'] = f1
            
            # Row 77: Check if hef > 2.5Ca1 (side-face blowout applies)
            #ws['A77'] = 'IF(hef>2.5Ca1,f1*( Consider side-face blowout )",*( Side-face blowout strength does not apply )")'
            if hef_val > val_2_5ca1:
                ws['B77'] = "Consider side-face blowout"
            else:
                ws['B77'] = "Side-face blowout strength does not apply"
            ws['B77'].font = Font(italic=True, color='0000FF')
            
            # Row 78: Check corner effect
            #ws['A78'] = 'IF(AND(Ca2/Ca1<1,Ca2/Ca1>3),"( Corner effect does not apply )","( Consider corner effect )")'
            if ca2_ca1_ratio < 1 or ca2_ca1_ratio > 3:
                ws['B78'] = "Corner effect does not apply"
            else:
                ws['B78'] = "Consider corner effect"
            ws['B78'].font = Font(italic=True, color='0000FF')
            
            # Row 79: Nsb
            ws['A79'] = 'Nsb='
            if hef_val > val_2_5ca1:
                nsb = f1 * (13 * ca1 * abrg * 0.5) * (fc_val ** 0.5) / 1000
            else:
                nsb = 0  # NA
            ws['B79'] = nsb if nsb != 0 else 'NA'
            ws['C79'] = 'kN'
            
            # Row 80: S1
            ws['A80'] = 'S1'
            s1 = min(a_val, c_val)
            ws['B80'] = s1
            ws['C80'] = 'mm'
            
            # Row 81: 6 x Ca1
            ws['A81'] = '6 x Ca1'
            val_6ca1 = 6 * ca1
            ws['B81'] = val_6ca1
            ws['C81'] = 'mm'
            
            # Row 82: f2
            ws['A82'] = 'f2'
            if s1 < val_6ca1:
                f2 = 1 + s1 / (6 * ca1) if ca1 != 0 else 0
            else:
                f2 = 1
            ws['B82'] = f2
            
            # Row 83: Check close spacing effect
            #ws['A83'] = 'IF(S1<6Ca1,"( Consider close spacing effect )","( Close spacing effect does not apply )")'
            if s1 < val_6ca1:
                ws['B83'] = "Consider close spacing effect"
            else:
                ws['B83'] = "Close spacing effect does not apply"
            ws['B83'].font = Font(italic=True, color='0000FF')
            
            # Row 84: NSbg
            ws['A84'] = 'NSbg='
            if hef_val > val_2_5ca1 and nsb != 'NA':
                nsbg = f2 * nsb / nb_val * 2 if nb_val != 0 else 0
            else:
                nsbg = 'NA'
            ws['B84'] = nsbg if nsbg != 'NA' else 'NA'
            ws['C84'] = 'kN'
            
            # Row 86: Tensile Design Summary
            ws['A86'] = '* Tensile Design Summary'
            ws['A86'].font = Font(italic=True, bold=True)
            
            # Row 87: Resistant factor
            ws['A87'] = 'Resitant factor φ ='
            ws['B87'] = 0.75
            
            # Row 88: Nn
            ws['A88'] = 'Nn='
            # IF(Nsbg="NA", MIN(Nsa,Npn), MIN(Nsa,Npn,Nsbg))
            if nsbg == 'NA':
                nn = min(nsa, npn)
            else:
                nn = min(nsa, npn, nsbg)
            ws['B88'] = nn
            ws['C88'] = 'kN'
            
            # Row 89: φNn
            ws['A89'] = 'φNn='
            phi_nn = 0.75 * nn
            ws['B89'] = phi_nn
            ws['C89'] = 'kN'
            
            # Row 90: Nua
            ws['A90'] = 'Nua='
            if vertical_load_ab < 0:
                nua_ab = abs(vertical_load_ab / nb_val) if nb_val != 0 else 0
            else:
                nua_ab = 0
            ws['B90'] = nua_ab
            ws['C90'] = 'kN'
            
            # Row 91: Rat 4
            ws['A91'] = 'Rat 4 ='
            rat4 = nua_ab / phi_nn if phi_nn != 0 else 0
            ws['B91'] = rat4
            
            # ==================== ROW 93-101: SHEAR LOADING ====================
            ws['A93'] = '2.2 Design Requirements for shear loading'
            ws['A93'].font = Font(bold=True, italic=True, size=10)
            
            # Row 94: Steel Strength
            ws['A94'] = '* Steel Strength of Anchor in Shear'
            ws['A94'].font = Font(italic=True, bold=True)
            
            # Row 95: Vsa
            ws['A95'] = 'Vsa='
            vsa = 0.6 * ase_val * futa_val * 0.8 / 1000
            ws['B95'] = vsa
            ws['C95'] = 'kN'
            
            # Row 97: Shear Design Summary
            ws['A97'] = '* Shear Design Summary'
            ws['A97'].font = Font(italic=True, bold=True)
            
            # Row 98: Resistant factor
            ws['A98'] = 'Resitant factor φ ='
            ws['B98'] = 0.65
            
            # Row 99: φVn
            ws['A99'] = 'φVn='
            phi_vn = 0.65 * vsa
            ws['B99'] = phi_vn
            ws['C99'] = 'kN'
            
            # Row 100: Vua
            ws['A100'] = 'Vua='
            friction_mu = safe_float(ws['C7'].value)  # Friction μ
            h_resultant = math.sqrt(h_load_x ** 2 + h_load_y ** 2)
            if vertical_load_ab < 0:
                pu_shear = 0
            else:
                pu_shear = vertical_load_ab
            
            vua_calc = (h_resultant - friction_mu * pu_shear) / nb_val
            vua = max(vua_calc, 0)
            ws['B100'] = vua
            ws['C100'] = 'kN'
            
            # Row 101: Rat 5
            ws['A101'] = 'Rat 5 ='
            rat5 = vua / phi_vn if phi_vn != 0 else 0
            ws['B101'] = rat5
            
            # ==================== ROW 103-107: INTERACTION ====================
            ws['A103'] = '2.3 Interaction of Tensile and Shear Forces'
            ws['A103'].font = Font(bold=True, italic=True, size=10)
            
            # Row 104: Vua / φVn
            ws['A104'] = 'Vua / φVn ='
            ws['B104'] = rat5
            ws['C104'] = '≤0.2' if rat5 <= 0.2 else '>0.2'
            
            # Row 105: Nua / φNn
            ws['A105'] = 'Nua / φNn ='
            ws['B105'] = rat4
            ws['C105'] = '≤0.2' if rat4 <= 0.2 else '>0.2'
            
            # Row 106: Rat 6
            ws['A106'] = 'Rat 6 ='
            # IF(D104*D105=1,B104+B105,MAX(B104,B105))
            ws['D104'] = 1 if rat5 > 0.2 else 0
            ws['D105'] = 1 if rat4 > 0.2 else 0

            if ws['D104'].value * ws['D105'].value == 1:
                rat6 = rat5 + rat4
            else:
                rat6 = max(rat5, rat4)
            ws['B106'] = rat6
            
            # Row 107: Control text
            #ws['A107'] = 'IF(D104*D105=1,"Interaction control",IF(B106=B104,"Shear control","Tension control"))'
            if ws['D104'].value * ws['D105'].value == 1:
                ws['B107'] = "Interaction control"
            elif rat6 == rat5:
                ws['B107'] = "Shear control"
            else:
                ws['B107'] = "Tension control"
            ws['B107'].font = Font(italic=True, color='0000FF')
            
            # ==================== UPDATE S4: Rat 6 ====================
            ws['S4'] = rat6
            ws['S4'].alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"Anchor Bolt calculation added: Rat4={rat4:.3f}, Rat5={rat5:.3f}, Rat6={rat6:.3f}")

            # ==================== ROW 110-162: PIER REINFORCEMENT CHECK ====================
            ws['A110'] = '3. Pier Reinforcement Check'
            ws['A110'].font = Font(bold=True, size=11)
            
            # ==================== SECTION 3.1: Transfer to Vertical Rebars ====================
            ws['A111'] = '3.1 Transfer of Anchor Load to Vertical Rebars'
            ws['A111'].font = Font(bold=True, italic=True, size=10)
            
            # Row 112: Control Case
            ws['A112'] = 'Control Case'
            ws['B112'] = ws['O3'].value if ws['O3'].value else ''  # Same as Tensile
            
            # Row 113: Vertical Load
            ws['A113'] = 'Vertical Load'
            vertical_load_pier = safe_float(ws['R3'].value)  # F3 from Tensile
            ws['B113'] = vertical_load_pier
            ws['C113'] = 'kN'
            
            # Row 114: Resistant factor
            ws['A114'] = 'Resitant factor φ ='
            ws['B114'] = 0.75
            
            # Parse rebar sizes
            main_size_str = str(ws['G11'].value) if ws['G11'].value else ''
            tie_size_str = str(ws['H11'].value) if ws['H11'].value else ''
            d_mainbar = int(main_size_str.replace('D', '').strip()) if main_size_str else 0
            d_tiebar = int(tie_size_str.replace('D', '').strip()) if tie_size_str else 0
            
            # Row 115: dmax
            ws['A115'] = 'dmax='
            hef_pier = safe_float(ws['L7'].value)  # heff
            w_pier = safe_float(bolt_data.get('W', 0)) if bolt_data else 0
            dmax = hef_pier / 3 + w_pier / 2
            ws['B115'] = dmax
            
            # Row 116: 0.5hef
            ws['A116'] = '0.5hef='
            val_0_5hef = 0.5 * hef_pier
            ws['B116'] = val_0_5hef
            # Check
            if dmax > val_0_5hef:
                ws['C116'] = 'NG'
            else:
                ws['C116'] = 'OK'
            ws['C116'].font = Font(color='0000FF')
            
            # Row 117: dact
            ws['A117'] = 'dact='
            cover_c = safe_float(ws['D11'].value)  # Cover c from Pier Detail
            c_bp = safe_float(ws['I3'].value)  # C from Base Plate Detail
            a_bp = safe_float(ws['J3'].value)  # A from Base Plate Detail
            
            ca1_pier = min((bp_val - a_bp) / 2, (np_pier - c_bp) / 2)
            ca2_pier = max((bp_val - a_bp) / 2, (np_pier - c_bp) / 2)
            
            dact = ((ca1_pier - cover_c - d_mainbar) ** 2 + (ca2_pier - cover_c - d_mainbar) ** 2) ** 0.5 - w_pier / 2
            ws['B117'] = dact
            # Check
            if dact > dmax:
                ws['C117'] = 'NG'
            else:
                ws['C117'] = 'OK'
            ws['C117'].font = Font(color='0000FF')
            
            # Row 118: Ca1-c-db/2
            ws['A118'] = 'Ca1-c-db/2='
            val_ca1_calc = (np_pier - c_bp) / 2 - cover_c - d_mainbar / 2
            ws['B118'] = val_ca1_calc
            # Check
            if val_ca1_calc < c_bp / 2:
                ws['C118'] = 'NG'
            else:
                ws['C118'] = 'OK'
            ws['C118'].font = Font(color='0000FF')
            
            # Row 119: Ca1-c-db/2
            ws['A119'] = 'Ca1-c-db/2='
            val_ca1_calc2 = (bp_val - a_bp) / 2 - cover_c - d_mainbar / 2
            ws['B119'] = val_ca1_calc2
            # Check
            if val_ca1_calc2 < a_bp / 2:
                ws['C119'] = 'NG'
            else:
                ws['C119'] = 'OK'
            ws['C119'].font = Font(color='0000FF')
            
            # Row 120: nreq'd
            ws['A120'] = "nreq'd ="
            fy_mainbar = safe_float(ws['C7'].value)  # fy(Mainbar)
            fy_tiebar = safe_float(ws['D7'].value)  # fy(Tiebar)
            as_mainbar = math.pi * (d_mainbar ** 2) / 4  # Area of one main bar
            
            if vertical_load_pier < 0:
                tension_force = abs(vertical_load_pier)
            else:
                tension_force = 0
            
            nreqd = math.ceil(tension_force * 1000 / (0.75 * fy_mainbar * as_mainbar)) if as_mainbar != 0 else 0
            ws['B120'] = nreqd
            #ws['C120'] = 'As là diện tích 1 thanh Rebar'
            #ws['C120'].font = Font(italic=True, color='FF0000')
            
            # Row 121: Rat 7
            ws['A121'] = 'Rat 7 ='
            qty_mainbar = safe_float(ws['F11'].value)  # Qty from Main Bar
            rat7 = nreqd / qty_mainbar if qty_mainbar != 0 else 0
            ws['B121'] = rat7
            #ws['C121'] = 'Qty là số thép dọc'
            #ws['C121'].font = Font(italic=True, color='FF0000')
            
            # Row 122: ld_prov
            ws['A122'] = 'ld_prov ='
            e_bp = safe_float(ws['K3'].value)  # E from Base Plate Detail
            f_bp = safe_float(ws['L3'].value)  # F from Base Plate Detail
            layer1 = safe_float(ws['K11'].value)  # Layer 1
            
            ld_prov = hef_pier - cover_c - (d_mainbar + dact) * math.tan(35 * math.pi / 180)
            ws['B122'] = ld_prov
            ws['C122'] = 'mm'
            
            # Row 123: ld
            ws['A123'] = 'ld='
            # Lookup from Rebar Development Length sheet
            ld_value = self.lookup_rebar_dev_length(d_mainbar, 'Ld')
            ws['B123'] = ld_value
            ws['C123'] = 'mm'
            
            # Row 124: ld_req'd
            ws['A124'] = "ld_req'd ="
            ld_reqd = ld_value * rat7 if ld_value else 0
            ws['B124'] = ld_reqd
            ws['C124'] = 'mm'
            
            # Row 125: Rat 8
            ws['A125'] = 'Rat 8 ='
            rat8 = ld_reqd / ld_prov if ld_prov != 0 else 0
            ws['B125'] = rat8
            
            # ==================== SECTION 3.2: Transfer to Horizontal Rebars ====================
            ws['A127'] = '3.2 Transfer of Anchor Load to Horizontal Rebars'
            ws['A127'].font = Font(bold=True, italic=True, size=10)
            
            # Row 128: Ψe (from Other Coefficient H15)
            ws['A128'] = 'Ψe ='
            ws['B128'] = f'=H15'
            
            # Row 129: λ (from Other Coefficient I15)
            ws['A129'] = 'λ ='
            ws['B129'] = f'=I15'
            
            # Row 130: ldh (using fy(Tiebar))
            ws['A130'] = 'ldh'
            # Lookup from Rebar Development Length sheet
            #ldh_value = self.lookup_rebar_dev_length(d_tiebar, 'Ldh')
            ldh_value = 0.24 * fy_tiebar * d_tiebar / (fc_val ** 0.5) if fc_val > 0 else 0
            ws['B130'] = ldh_value
            ws['C130'] = 'mm'

            # ==================== CONSIDER BETA ====================
            beta_val = node_data.get('beta', 90)  # Default 90 (current behavior)
            if beta_val == 0:
                # Web ∥ X → X-shear uses Np/C, Y-shear uses Bp/A
                pier_x  = np_pier
                space_x = c_val
                pier_y  = bp_val
                space_y = a_val
            else:
                # Web ⊥ X (beta=90) → X-shear uses Bp/A, Y-shear uses Np/C
                pier_x  = bp_val
                space_x = a_val
                pier_y  = np_pier
                space_y = c_val
                
            # ==================== FOR X - DIRECTION ====================
            ws['A132'] = '* For X - Direction'
            ws['A132'].font = Font(italic=True, bold=True)

            ws['A133'] = 'Control Case'
            ws['B133'] = ws['O5'].value if ws['O5'].value else ''

            ws['A134'] = 'Vertical Load'
            vertical_load_x = safe_float(ws['R5'].value)
            ws['B134'] = vertical_load_x
            ws['C134'] = 'kN'

            ws['A135'] = 'Horizontal Load'
            h_load_x_pier = safe_float(ws['P5'].value)
            ws['B135'] = h_load_x_pier
            ws['C135'] = 'kN'

            ws['A136'] = 'Vx='
            if vertical_load_x < 0:
                pu_x = 0
            else:
                pu_x = vertical_load_x
            vx_calc = abs(h_load_x_pier) - friction_mu * pu_x
            vx = max(vx_calc, 0)
            ws['B136'] = vx
            ws['C136'] = 'kN'

            ws['A137'] = 'Atie='
            ws['B137'] = math.pi * (d_tiebar ** 2) / 4
            ws['C137'] = 'mm2'

            # ✅ lda dùng pier_x, space_x
            ws['A138'] = 'lda_A_L(x) ='
            tan35 = math.tan(35 * math.pi / 180)
            layer1_val = safe_float(ws['K11'].value)
            layer2_val = safe_float(ws['L11'].value)
            lda_a_lx = (pier_x + space_x) / 2 - db_num - cover_c - layer1_val * tan35
            ws['B138'] = lda_a_lx
            ws['C138'] = 'mm'

            ws['A139'] = 'lda_A_R(x) ='
            lda_a_rx = (pier_x - 2 * cover_c) - lda_a_lx
            ws['B139'] = lda_a_rx
            ws['C139'] = 'mm'

            ws['A140'] = 'lda_B_L(x) ='
            lda_b_lx = (pier_x + space_x) / 2 - db_num - cover_c - (layer1_val + layer2_val) * tan35
            ws['B140'] = lda_b_lx
            ws['C140'] = 'mm'

            ws['A141'] = 'lda_B_R(x) ='
            lda_b_rx = (pier_x - 2 * cover_c) - lda_b_lx
            ws['B141'] = lda_b_rx
            ws['C141'] = 'mm'

            ws['A142'] = 'fsx='
            fsx = min(lda_a_lx, lda_a_rx, lda_b_lx, lda_b_rx) / ldh_value * fy_tiebar if ldh_value != 0 else 0
            ws['B142'] = fsx
            ws['C142'] = 'Mpa'

            ws['A143'] = 'Resitant factor φ ='
            ws['B143'] = 0.75

            ws['A144'] = "Atie_req'd(x)="
            x_legs = safe_float(ws['I11'].value)
            atie_reqd_x = vx / 2 * 1000 / (0.75 * fsx) if fsx != 0 else 0
            ws['B144'] = atie_reqd_x
            ws['C144'] = 'mm2'

            ws['A145'] = 'Atie_prov(x)='
            atie_prov_x = x_legs * (math.pi * (d_tiebar ** 2) / 4)
            ws['B145'] = atie_prov_x
            ws['C145'] = 'mm2'

            ws['A146'] = 'Rat 9 ='
            rat9 = atie_reqd_x / atie_prov_x if atie_prov_x != 0 else 0
            ws['B146'] = rat9
            
            # ==================== FOR Y - DIRECTION ====================
            ws['A148'] = '* For Y - Direction'
            ws['A148'].font = Font(italic=True, bold=True)

            ws['A149'] = 'Control Case'
            ws['B149'] = ws['O6'].value if ws['O6'].value else ''

            ws['A150'] = 'Vertical Load'
            vertical_load_y = safe_float(ws['R6'].value)
            ws['B150'] = vertical_load_y
            ws['C150'] = 'kN'

            ws['A151'] = 'Horizontal Load'
            h_load_y_pier = safe_float(ws['Q6'].value)
            ws['B151'] = h_load_y_pier
            ws['C151'] = 'kN'

            ws['A152'] = 'Vy='
            if vertical_load_y < 0:
                pu_y = 0
            else:
                pu_y = vertical_load_y
            vy_calc = abs(h_load_y_pier) - friction_mu * pu_y
            vy = max(vy_calc, 0)
            ws['B152'] = vy
            ws['C152'] = 'kN'

            ws['A153'] = 'Atie='
            ws['B153'] = math.pi * (d_tiebar ** 2) / 4
            ws['C153'] = 'mm2'

            # ✅ lda dùng pier_y, space_y
            ws['A154'] = 'lda_A_L(y) ='
            lda_a_ly = (pier_y + space_y) / 2 - db_num - cover_c - layer1_val * tan35
            ws['B154'] = lda_a_ly
            ws['C154'] = 'mm'

            ws['A155'] = 'lda_A_R(y) ='
            lda_a_ry = (pier_y - 2 * cover_c) - lda_a_ly
            ws['B155'] = lda_a_ry
            ws['C155'] = 'mm'

            ws['A156'] = 'lda_B_L(y) ='
            lda_b_ly = (pier_y + space_y) / 2 - db_num - cover_c - (layer1_val + layer2_val) * tan35
            ws['B156'] = lda_b_ly
            ws['C156'] = 'mm'

            ws['A157'] = 'lda_B_R(y) ='
            lda_b_ry = (pier_y - 2 * cover_c) - lda_b_ly
            ws['B157'] = lda_b_ry
            ws['C157'] = 'mm'

            ws['A158'] = 'fsy='
            fsy = min(lda_a_ly, lda_a_ry, lda_b_ly, lda_b_ry) / ldh_value * fy_tiebar if ldh_value != 0 else 0
            ws['B158'] = fsy
            ws['C158'] = 'Mpa'

            ws['A159'] = 'Resitant factor φ ='
            ws['B159'] = 0.75

            ws['A160'] = "Atie_req'd(y)="
            y_legs = safe_float(ws['J11'].value)
            atie_reqd_y = vy / 2 * 1000 / (0.75 * fsy) if fsy != 0 else 0
            ws['B160'] = atie_reqd_y
            ws['C160'] = 'mm2'

            ws['A161'] = 'Atie_prov(y)='
            atie_prov_y = y_legs * (math.pi * (d_tiebar ** 2) / 4)
            ws['B161'] = atie_prov_y
            ws['C161'] = 'mm2'

            ws['A162'] = 'Rat 10 ='
            rat10 = atie_reqd_y / atie_prov_y if atie_prov_y != 0 else 0
            ws['B162'] = rat10
            
            # ==================== UPDATE S5 and S6 ====================
            ws['S5'] = rat9
            ws['S5'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws['S6'] = rat10
            ws['S6'].alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"Pier Reinforcement calculation added: Rat7={rat7:.3f}, Rat8={rat8:.3f}, Rat9={rat9:.3f}, Rat10={rat10:.3f}")            
            
            # Add borders to calculation area
            for row in range(13, 163):  # Extend to row 162
                for col in ['A', 'B', 'C', 'D']:
                    if ws[f'{col}{row}'].value is not None:
                        ws[f'{col}{row}'].border = thin_border
            
            print(f"Compression calculation added: Rat1={rat1:.3f}, Rat2={rat2:.3f}, Final={final_ratio:.3f}")
            
        except Exception as e:
            print(f"Error in compression calculation: {e}")
            import traceback
            traceback.print_exc()
        
        # Save workbook
        wb.save(filepath)
        wb.close()
        
        print(f"Sheet '{sheet_name}' created/updated in {filename}")
        return filepath
    
    def edit_node_detail(self):
            """Open dialog to edit node detail from Hinge Type.xlsx"""
            if not self.selected_node:
                return
            
            node_data = self.base_plate_nodes[self.selected_node]
            
            # Check if Hinge Type.xlsx exists
            hinge_fixed = node_data.get('hinge_fixed_type', 'Hinge Type')
            if 'Fixed' in hinge_fixed:
                filename = 'Fixed Type.xlsx'
            else:
                filename = 'Hinge Type.xlsx'
            
            filepath = os.path.join(self.bpl_folder, filename)
            
            if not os.path.exists(filepath):
                messagebox.showwarning("Warning", 
                    f"{filename} not found!\n\n"
                    f"Please click 'Apply to Node' first to create the file.")
                return
            
            sheet_name = f"Node {self.selected_node}"
            
            # Check if sheet exists
            try:
                wb = openpyxl.load_workbook(filepath)
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    messagebox.showwarning("Warning", 
                        f"Sheet '{sheet_name}' not found!\n\n"
                        f"Please click 'Apply to Node' first.")
                    return
                
                ws = wb[sheet_name]
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to open {filename}:\n{str(e)}")
                return
            
            # Create dialog
            dialog = tk.Toplevel(self.root)
            dialog.title(f"Edit Node Detail - {self.selected_node}")
            dialog.geometry("1100x700")
            dialog.transient(self.root)
            dialog.grab_set()
            
            # Center dialog
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
            y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
            dialog.geometry(f"+{x}+{y}")
            
            # Title
            title_label = tk.Label(
                dialog,
                text=f"📝 Edit Detail for Node {self.selected_node}",
                font=('Arial', 14, 'bold'),
                bg='#1a472a',
                fg='white',
                pady=15
            )
            title_label.pack(fill='x')
            
            # Main container with scrollbar
            main_container = tk.Frame(dialog, bg='white')
            main_container.pack(fill='both', expand=True, padx=20, pady=(15, 0))
            
            canvas = tk.Canvas(main_container, bg='white', highlightthickness=0)
            scrollbar = ttk.Scrollbar(main_container, orient='vertical', command=canvas.yview)
            scrollable_frame = tk.Frame(canvas, bg='white')
            
            scrollable_frame.bind(
                '<Configure>',
                lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
            )
            
            canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            
            # Store entry widgets
            entries = {}
            
            # Helper to create editable grid
            def create_editable_grid(parent, title, header_row, value_row, columns, fill_width=True):
                """Create a grid with headers and editable values"""
                frame = tk.LabelFrame(
                    parent,
                    text=f" {title} ",
                    font=('Arial', 11, 'bold'),
                    bg='white',
                    fg='#1a472a',
                    relief='groove',
                    bd=2
                )
                if fill_width:
                    frame.pack(fill='x', pady=(0, 15))
                else:
                    frame.pack(anchor='w', pady=(0, 15))
                
                grid_frame = tk.Frame(frame, bg='white')
                grid_frame.pack(padx=15, pady=10)
                
                for col_idx, col_letter in enumerate(columns):
                    # Header
                    header_cell = ws[f'{col_letter}{header_row}']
                    header_text = header_cell.value if header_cell.value else ''
                    
                    tk.Label(
                        grid_frame,
                        text=str(header_text),
                        font=('Arial', 9, 'bold'),
                        bg='#E8E8E8',
                        relief='solid',
                        bd=1,
                        width=10,
                        height=2
                    ).grid(row=0, column=col_idx, padx=2, pady=2)
                    
                    # Value entry
                    value_cell = ws[f'{col_letter}{value_row}']
                    value_text = value_cell.value if value_cell.value is not None else ''
                    
                    entry = tk.Entry(
                        grid_frame,
                        font=('Arial', 10),
                        justify='center',
                        width=10,
                        relief='solid',
                        bd=1
                    )
                    entry.insert(0, str(value_text))
                    entry.grid(row=1, column=col_idx, padx=2, pady=2)
                    
                    # Store for saving
                    entries[f'{col_letter}{value_row}'] = entry
            
            # SECTION 1: Base Plate (rows 1-3)
            section1_container = tk.Frame(scrollable_frame, bg='white')
            section1_container.pack(fill='x', pady=(10, 0))
            
            left1 = tk.Frame(section1_container, bg='white')
            left1.pack(side='left', fill='both', expand=True, padx=(0, 10))
            
            section1_title = ws['A1'].value if ws['A1'].value else 'Section'
            create_editable_grid(left1, section1_title, 2, 3, ['A', 'B', 'C', 'D'])
            
            right1 = tk.Frame(section1_container, bg='white')
            right1.pack(side='left', fill='both', expand=True, padx=(10, 0))
            create_editable_grid(right1, 'Base Plate Detail', 2, 3, ['F', 'G', 'H', 'I', 'J', 'K', 'L'])
            
            # SECTION 2: Material (rows 5-7)
            section2_container = tk.Frame(scrollable_frame, bg='white')
            section2_container.pack(fill='x', pady=0)
            
            left2 = tk.Frame(section2_container, bg='white')
            left2.pack(side='left', fill='both', expand=True, padx=(0, 10))
            create_editable_grid(left2, 'Material', 6, 7, ['A', 'B', 'C', 'D'])
            
            right2 = tk.Frame(section2_container, bg='white')
            right2.pack(side='left', fill='both', expand=True, padx=(10, 0))
            create_editable_grid(right2, 'Anchor Bolt', 6, 7, ['F', 'G', 'H', 'I', 'J', 'K', 'L'])
            
            # SECTION 3: Pier Detail (rows 9-11)
            section3_container = tk.Frame(scrollable_frame, bg='white')
            section3_container.pack(fill='x', pady=0)
            
            left3 = tk.Frame(section3_container, bg='white')
            left3.pack(side='left', fill='both', expand=True, padx=(0, 10))
            create_editable_grid(left3, 'Pier Detail', 10, 11, ['A', 'B', 'C', 'D'])
            
            right3 = tk.Frame(section3_container, bg='white')
            right3.pack(side='left', fill='both', expand=True, padx=(10, 0))
            create_editable_grid(right3, 'Main Bar & Tie Bar', 10, 11, ['F', 'G', 'H', 'I', 'J', 'K', 'L'])
            
            # SECTION 4: Other Coefficient (rows 13-15)
            create_editable_grid(scrollable_frame, 'Other Coefficient', 14, 15, ['F', 'G', 'H', 'I'], fill_width=False)
            
            wb.close()
            
            # Button frame
            btn_frame = tk.Frame(dialog, bg='#f5f5f5')
            btn_frame.pack(side='bottom', fill='x', pady=15)
            
            def save_changes():
                try:
                    # Step 1: Update node_data['edited_*']
                    node_data = self.base_plate_nodes[self.selected_node]
                    
                    node_data['edited_section'] = {
                        'd':  entries['A3'].get().strip() if 'A3' in entries else '',
                        'bf': entries['B3'].get().strip() if 'B3' in entries else '',
                        'tw': entries['C3'].get().strip() if 'C3' in entries else '',
                        'tf': entries['D3'].get().strip() if 'D3' in entries else '',
                    }
                    node_data['edited_base_plate_detail'] = {
                        'N':  entries['F3'].get().strip() if 'F3' in entries else '',
                        'B':  entries['G3'].get().strip() if 'G3' in entries else '',
                        'P1': entries['H3'].get().strip() if 'H3' in entries else '',
                        'C':  entries['I3'].get().strip() if 'I3' in entries else '',
                        'A':  entries['J3'].get().strip() if 'J3' in entries else '',
                        'E':  entries['K3'].get().strip() if 'K3' in entries else '',
                        'F':  entries['L3'].get().strip() if 'L3' in entries else '',
                    }
                    node_data['edited_material'] = {
                        'fy_steel':   entries['A7'].get().strip() if 'A7' in entries else '',
                        'fc':         entries['B7'].get().strip() if 'B7' in entries else '',
                        'fy_mainbar': entries['C7'].get().strip() if 'C7' in entries else '',
                        'fy_tiebar':  entries['D7'].get().strip() if 'D7' in entries else '',
                    }
                    node_data['edited_anchor_bolt'] = {
                        'nb':   entries['F7'].get().strip() if 'F7' in entries else '',
                        'db':   entries['G7'].get().strip() if 'G7' in entries else '',
                        'Ase':  entries['H7'].get().strip() if 'H7' in entries else '',
                        'futa': entries['I7'].get().strip() if 'I7' in entries else '',
                        'A1':   entries['J7'].get().strip() if 'J7' in entries else '',
                        'Proj': entries['K7'].get().strip() if 'K7' in entries else '',
                        'heff': entries['L7'].get().strip() if 'L7' in entries else '',
                    }
                    node_data['edited_pier_detail'] = {
                        'Np': entries['A11'].get().strip() if 'A11' in entries else '',
                        'Bp': entries['B11'].get().strip() if 'B11' in entries else '',
                        'TG': entries['C11'].get().strip() if 'C11' in entries else '',
                        'c':  entries['D11'].get().strip() if 'D11' in entries else '',
                    }
                    node_data['edited_main_bar'] = {
                        'Qty':  entries['F11'].get().strip() if 'F11' in entries else '',
                        'Size': entries['G11'].get().strip() if 'G11' in entries else '',
                    }
                    node_data['edited_other_coeff'] = {
                        'friction_mu': entries['F15'].get().strip() if 'F15' in entries else '',
                        'psi_c_p':     entries['G15'].get().strip() if 'G15' in entries else '',
                        'psi_e':       entries['H15'].get().strip() if 'H15' in entries else '',
                        'lambda':      entries['I15'].get().strip() if 'I15' in entries else '',
                    }
                    
                    # Step 2: Regenerate full sheet using correct explicit calculation
                    result_path = self.create_or_update_hinge_fixed_xlsx(
                        self.selected_node, node_data
                    )
                    
                    messagebox.showinfo("✅ Success",
                        f"Changes saved & recalculated!\\n\\n"
                        f"File: {os.path.basename(result_path)}\\n"
                        f"Sheet: {sheet_name}")
                    
                    self.status_label.config(
                        text=f"● Node {self.selected_node} detail updated & recalculated",
                        fg='#90ee90'
                    )
                    dialog.destroy()
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to save:\\n{str(e)}")
                    import traceback
                    traceback.print_exc()           
            
            tk.Button(
                btn_frame,
                text="💾 Save Changes",
                command=save_changes,
                bg='#4CAF50',
                fg='white',
                font=('Arial', 11, 'bold'),
                cursor='hand2',
                relief='flat',
                padx=30,
                pady=10
            ).pack(side='left', padx=(30, 10))
            
            tk.Button(
                btn_frame,
                text="❌ Cancel",
                command=dialog.destroy,
                bg='#f44336',
                fg='white',
                font=('Arial', 11, 'bold'),
                cursor='hand2',
                relief='flat',
                padx=30,
                pady=10
            ).pack(side='left', padx=10)
# Main Execution
if __name__ == "__main__":
    root = tk.Tk()
    app = BasePlateApp(root)
    root.mainloop()