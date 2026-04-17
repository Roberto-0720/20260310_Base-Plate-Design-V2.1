"""
logic/design_check_mixin.py
=============================
Design check engine: reads Hinge Type.xlsx sheets, calculates control
ratios for every load combination, and colours nodes on the plan view.

App state used: self.base_plate_nodes, self.selected_node,
                self.bpl_folder, self.status_label
"""

import os
import math
import tkinter as tk
from tkinter import messagebox
import openpyxl


class DesignCheckMixin:

    def run_design_check(self):
        """Run design check for all assigned nodes"""
        if not self.base_plate_nodes:
            return

        # Count assigned nodes
        assigned_count = sum(1 for data in self.base_plate_nodes.values()
                           if data['bolt_size'] is not None)

        if assigned_count == 0:
            messagebox.showwarning("Warning", "No nodes have been assigned!\n\nPlease assign materials and bolt sizes first.")
            return

        # Confirm
        confirm = messagebox.askyesno("Confirm Design Check",
                                     f"Run design check for {assigned_count} assigned node(s)?")
        if not confirm:
            return

        # Run design check - Read ratios from Hinge Type.xlsx sheets
        checked_count = 0
        for node_name, data in self.base_plate_nodes.items():
            if data['bolt_size'] is not None:
                # Try to read max ratio from node's Hinge Type sheet
                max_ratio = self._get_max_ratio_from_sheet(node_name)

                if max_ratio is not None:
                    data['max_ratio'] = max_ratio
                    # Determine status based on ratio
                    if max_ratio <= 1.0:
                        data['design_status'] = 'OK'
                    else:
                        data['design_status'] = 'NG'
                else:
                    # If can't read sheet, set to "Not Checked"
                    data['design_status'] = 'Not Checked'
                    data['max_ratio'] = None

                checked_count += 1

        # Refresh plot
        self.plot_base_plate_plan()
        if self.selected_node:
            self.highlight_selected_node()

        messagebox.showinfo("Design Check Complete",
                          f"Design check completed for {checked_count} node(s)!\n\n"
                          f"Check the plan view for results:\n"
                          f"🟢 Green = Design OK (Ratio ≤ 1.0)\n"
                          f"🔴 Red = Design NG (Ratio > 1.0)\n"
                          f"⚪ Gray = Not Checked")

        self.status_label.config(text=f"● Design check complete: {checked_count} nodes", fg='#90ee90')

    def _get_max_ratio_from_sheet(self, node_name):
        """Get max ratio (from S2-S6) from node's Hinge Type sheet"""
        try:
            if not self.bpl_folder:
                return None

            # Try to open Hinge Type.xlsx
            filepath = os.path.join(self.bpl_folder, "Hinge Type.xlsx")
            if not os.path.exists(filepath):
                return None

            wb = openpyxl.load_workbook(filepath)
            sheet_name = f"Node {node_name}"

            if sheet_name not in wb.sheetnames:
                wb.close()
                return None

            ws = wb[sheet_name]

            # Get ratios from S2-S6
            ratios = []
            for row in range(2, 7):  # Rows 2-6
                try:
                    ratio_val = ws[f'S{row}'].value
                    if ratio_val is not None:
                        ratios.append(float(ratio_val))
                except:
                    pass

            wb.close()

            # Return max ratio
            if ratios:
                return max(ratios)
            else:
                return None

        except Exception as e:
            print(f"Error reading ratio from {node_name}: {e}")
            return None

    def _calculate_control_ratios(self, f1, f2, f3, ws, node_data):
        """
        Calculate 5 control ratios for a given load combination (f1, f2, f3)

        Returns: dict with keys 'compression', 'tensile', 'anchor_bolt', 'x_shear', 'y_shear'
        """
        import math

        def safe_float(value, default=0):
            """Safely convert value to float"""
            try:
                return float(value) if value else default
            except:
                return default

        # Get values from sheet (same as original code)
        n_val = safe_float(ws['F3'].value)
        b_val = safe_float(ws['G3'].value)
        p1_val = safe_float(ws['H3'].value)
        d_val = safe_float(ws['A3'].value)
        bf_val = safe_float(ws['B3'].value)

        np_val = safe_float(ws['A11'].value)
        bp_val = safe_float(ws['B11'].value)

        fy_steel = safe_float(ws['A7'].value)
        fc_val = safe_float(ws['B7'].value)
        fy_mainbar = safe_float(ws['C7'].value)  # fy(Mainbar)
        fy_tiebar = safe_float(ws['D7'].value)  # fy(Tiebar)
        friction_mu = safe_float(ws['F15'].value)  # Friction μ from Other Coefficient

        # Main bar
        main_qty_raw = ws['F11'].value
        main_size_raw = ws['G11'].value

        # Calculate Ast
        try:
            main_qty_num = int(safe_float(main_qty_raw))
            main_size = str(main_size_raw) if main_size_raw else ''
            size_num = int(main_size.replace('D', '').strip()) if main_size else 0
            area_one_bar = math.pi * (size_num ** 2) / 4
            ast = main_qty_num * area_one_bar
        except:
            ast = 0

        # Tie bar
        tie_size_raw = ws['H11'].value
        try:
            tie_size_str = str(tie_size_raw) if tie_size_raw else ''
            d_tiebar = int(tie_size_str.replace('D', '').strip()) if tie_size_str else 0
        except:
            d_tiebar = 0

        ratios = {}

        # ==================== 1. COMPRESSION RATIO ====================
        # Vertical load (F3)
        vertical_load = f3
        if vertical_load < 0:
            pu = 0
        else:
            pu = vertical_load

        # A1 and A2
        a1 = n_val * b_val
        a2 = np_val * bp_val

        # Pp (AISC)
        pp = 0.85 * fc_val * a1 * math.sqrt(a2 / a1) * 0.001 if a1 != 0 else 0
        val_1_7 = 1.7 * fc_val * a1 * 0.001

        # Pp (ACI) - using fy(Mainbar)
        pp_aci = 0.8 * (0.85 * fc_val * (a2 - ast) + fy_mainbar * ast) / 1000

        # Pn and φPn
        pn = min(min(pp, val_1_7), pp_aci)
        phi_pn = 0.65 * pn

        # Rat1 (Compression)
        rat1 = abs(vertical_load) / phi_pn if phi_pn != 0 else 0

        # m, n, n' for plate yielding
        m_val = (n_val - 0.95 * d_val) / 2
        n_calc = (b_val - 0.8 * bf_val) / 2
        n_prime = (d_val * bf_val) ** 0.5 / 4
        l_val = max(m_val, n_calc, n_prime)

        # tmin
        if fy_steel != 0 and b_val != 0 and n_val != 0:
            tmin = l_val * math.sqrt(2 * pu * 1000 / (0.9 * fy_steel * b_val * n_val))
        else:
            tmin = 0

        # Rat2 (Plate yielding)
        rat2 = tmin / p1_val if p1_val != 0 else 0

        # Max compression ratio
        ratios['compression'] = max(rat1, rat2)

        # ==================== 2. TENSILE RATIO ====================
        vertical_load_tensile = f3

        # Nua
        nb_val = safe_float(ws['F7'].value)
        if vertical_load_tensile < 0:
            nua = abs(vertical_load_tensile / nb_val) if nb_val != 0 else 0
        else:
            nua = 0

        # Mu
        a_val = safe_float(ws['J3'].value)
        tw_val = safe_float(ws['C3'].value)
        mu = nua * (a_val / 2 - tw_val / 2)

        # beff
        beff = (b_val / 2 - tw_val / 2) * 2

        # t_reqd
        if fy_steel != 0 and beff != 0:
            t_reqd = math.sqrt(4 * mu * 1000 / (0.9 * fy_steel * beff))
        else:
            t_reqd = 0

        # Rat3 (Tensile)
        rat3 = t_reqd / p1_val if p1_val != 0 else 0

        ratios['tensile'] = rat3

        # ==================== 3. ANCHOR BOLT RATIO ====================
        vertical_load_ab = f3
        h_load_x = f1
        h_load_y = f2

        # Steel tension
        ase_val = safe_float(ws['H7'].value)
        futa_val = safe_float(ws['I7'].value)
        nsa = ase_val * futa_val / 1000

        # Pullout resistance
        db_num = safe_float(ws['G7'].value)
        bolt_data = self.get_anchor_bolt_data(f"M{int(db_num)}")
        w_val = safe_float(bolt_data.get('W', 0)) if bolt_data else 0
        nutw_val = safe_float(bolt_data.get('NutW', 0)) if bolt_data else 0

        if w_val > 0:
            abrg = w_val ** 2 - math.pi / 4 * (db_num ** 2)
        else:
            abrg = 0.866 * (nutw_val ** 2)

        nap = 8 * abrg * fc_val / 1000
        npn = 1.4 * nap

        # Side-face blowout
        c_val = safe_float(ws['I3'].value)
        a_val = safe_float(ws['J3'].value)
        ca1 = min((bp_val - a_val) / 2, (np_val - c_val) / 2)
        ca2 = max((bp_val - a_val) / 2, (np_val - c_val) / 2)
        hef_val = safe_float(ws['L7'].value)

        ca2_ca1_ratio = ca2 / ca1 if ca1 != 0 else 0

        if ca2_ca1_ratio < 1:
            f1_factor = 1
        elif ca2_ca1_ratio > 3:
            f1_factor = 1 + (ca1 / ca2) / 4
        else:
            f1_factor = (1 + ca2_ca1_ratio) / 4

        if hef_val > 2.5 * ca1:
            nsb = f1_factor * npn
        else:
            nsb = npn

        # Pryout
        np_pier = safe_float(ws['A11'].value)
        bp_pier = safe_float(ws['B11'].value)

        ca1_p = min((bp_pier - a_val) / 2, (np_pier - c_val) / 2)
        np_effect = min(2 * ca1_p, hef_val / 2)
        ncp = 16 * np_effect * fc_val / 1000

        # Tension capacity
        nna = min(nsa, nsb, ncp)
        phi_nn = 0.65 * nna

        # Rat4 (Tension)
        if vertical_load_ab < 0:
            nua_ab = abs(vertical_load_ab)
        else:
            nua_ab = 0

        rat4 = nua_ab / phi_nn if phi_nn != 0 else 0

        # Shear
        vsa = (f1_factor * nsa if hef_val > 2.5 * ca1 else nsa) if abs(h_load_x) > 0 or abs(h_load_y) > 0 else nsa
        phi_vn = 0.65 * vsa

        # Vua
        h_resultant = math.sqrt(h_load_x ** 2 + h_load_y ** 2)
        if vertical_load_ab < 0:
            pu_shear = 0
        else:
            pu_shear = vertical_load_ab

        nb_val = safe_float(ws['F7'].value)
        vua_calc = (h_resultant - friction_mu * pu_shear) / nb_val if nb_val != 0 else 0
        vua = max(vua_calc, 0)

        # Rat5 (Shear)
        rat5 = vua / phi_vn if phi_vn != 0 else 0

        # Rat6 (Interaction)
        d104_val = 1 if rat5 > 0.2 else 0
        d105_val = 1 if rat4 > 0.2 else 0

        if d104_val * d105_val == 1:
            rat6 = rat5 + rat4
        else:
            rat6 = max(rat5, rat4)

        ratios['anchor_bolt'] = rat6

        # ==================== SETUP PIER/SPACING BASED ON BETA ====================
        beta_val = node_data.get('beta', 90)

        if beta_val == 0:
            pier_x  = np_val
            space_x = safe_float(ws['I3'].value)   # C
            pier_y  = bp_val
            space_y = safe_float(ws['J3'].value)   # A
        else:
            pier_x  = bp_val
            space_x = safe_float(ws['J3'].value)   # A
            pier_y  = np_val
            space_y = safe_float(ws['I3'].value)   # C

        # ==================== 4. X-SHEAR CHECK ====================
        vertical_load_x = f3
        h_load_x_pier = f1

        if vertical_load_x < 0:
            pu_x = 0
        else:
            pu_x = vertical_load_x

        vx_calc = abs(h_load_x_pier) - friction_mu * pu_x
        vx = max(vx_calc, 0)

        atie = math.pi * (d_tiebar ** 2) / 4 if d_tiebar > 0 else 1

        e_bp = safe_float(ws['K3'].value)
        f_bp = safe_float(ws['L3'].value)
        layer1 = safe_float(ws['K11'].value)
        layer2 = safe_float(ws['L11'].value)
        cover_c = safe_float(ws['D11'].value)

        ldh_value = 0.24 * fy_tiebar * d_tiebar / (fc_val ** 0.5) if fc_val > 0 and d_tiebar > 0 else 1

        # dùng pier_x, space_x
        tan35 = math.tan(35 * math.pi / 180)
        lda_a_lx = (pier_x + space_x) / 2 - db_num - cover_c - layer1 * tan35
        lda_a_rx = (pier_x - 2 * cover_c) - lda_a_lx
        lda_b_lx = (pier_x + space_x) / 2 - db_num - cover_c - (layer1 + layer2) * tan35
        lda_b_rx = (pier_x - 2 * cover_c) - lda_b_lx

        fsx = min(lda_a_lx, lda_a_rx, lda_b_lx, lda_b_rx) / ldh_value * fy_tiebar if ldh_value != 0 else 0

        x_legs = safe_float(ws['I11'].value)
        atie_reqd_x = vx / 2 * 1000 / (0.75 * fsx) if fsx != 0 else 0
        atie_prov_x = x_legs * atie

        rat9 = atie_reqd_x / atie_prov_x if atie_prov_x != 0 else 0
        ratios['x_shear'] = rat9

        # ==================== 5. Y-SHEAR CHECK ====================
        vertical_load_y = f3
        h_load_y_pier = f2

        if vertical_load_y < 0:
            pu_y = 0
        else:
            pu_y = vertical_load_y

        vy_calc = abs(h_load_y_pier) - friction_mu * pu_y
        vy = max(vy_calc, 0)

        # dùng pier_y, space_y
        lda_a_ly = (pier_y + space_y) / 2 - db_num - cover_c - layer1 * tan35
        lda_a_ry = (pier_y - 2 * cover_c) - lda_a_ly
        lda_b_ly = (pier_y + space_y) / 2 - db_num - cover_c - (layer1 + layer2) * tan35
        lda_b_ry = (pier_y - 2 * cover_c) - lda_b_ly

        fsy = min(lda_a_ly, lda_a_ry, lda_b_ly, lda_b_ry) / ldh_value * fy_tiebar if ldh_value != 0 else 0

        y_legs = safe_float(ws['J11'].value)
        atie_reqd_y = vy / 2 * 1000 / (0.75 * fsy) if fsy != 0 else 0
        atie_prov_y = y_legs * atie

        rat10 = atie_reqd_y / atie_prov_y if atie_prov_y != 0 else 0
        ratios['y_shear'] = rat10

        return ratios
