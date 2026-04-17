"""
logic/hinge_xlsx_mixin.py
==========================
Creates or updates the per-node sheet in Hinge Type.xlsx / Fixed Type.xlsx.
This is the core calculation output file: it writes all input tables, runs
every load combination through the design equations, and fills the Control
Load Table (N1:S6) with governing ratios.

Single public method:  create_or_update_hinge_fixed_xlsx(node_name, node_data)

App state used:
  self.bpl_folder, self.base_plate_nodes, self.status_label
  (also delegates to CalculationsMixin and DesignCheckMixin helpers)
"""

import os
import math
import tkinter as tk
from tkinter import messagebox

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class HingeXlsxMixin:

    def create_or_update_hinge_fixed_xlsx(self, node_name, node_data):
        """Create or update Hinge Type.xlsx / Fixed Type.xlsx with node sheet"""
        if not self.bpl_folder:
            messagebox.showerror("Error", "Base Plate Design folder not found!")
            return

        # Determine file name based on hinge/fixed type
        hinge_fixed = node_data.get('hinge_fixed_type', 'Hinge Type')
        if 'Fixed' in hinge_fixed:
            filename = 'Fixed Type.xlsx'
        else:
            filename = 'Hinge Type.xlsx'

        filepath = os.path.join(self.bpl_folder, filename)

        # Load or create workbook
        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

        # Create or get sheet for this node
        sheet_name = f"Node {node_name}"
        sheet_exists = sheet_name in wb.sheetnames

        # Always delete and recreate sheet to reflect latest dropdown selections
        if sheet_exists:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        is_update_mode = False  # Always write full input data

        # Get all required data
        section = node_data.get('section', '')
        section_parts = self.parse_section_name(section)
        hinge_data = self.get_hinge_type_row_data(section)

        # Materials
        fy_steel = self.get_material_strength(node_data.get('material_base_plate'), 'base_plate')
        fc = self.get_material_strength(node_data.get('material_concrete'), 'concrete')
        fy_mainbar = self.get_material_strength(node_data.get('material_mainbar'), 'rebar')
        fy_tiebar = self.get_material_strength(node_data.get('material_tiebar'), 'rebar')
        futa = self.get_material_strength(node_data.get('material_anchor_bolt'), 'anchor_bolt')

        # Anchor Bolt
        bolt_size = node_data.get('bolt_size', '')
        bolt_data = self.get_anchor_bolt_data(bolt_size)
        nb = hinge_data.get('No.AB', '') if hinge_data else ''
        db = bolt_size.replace('M', '') if bolt_size else ''
        ase_bolt = self.calculate_anchor_bolt_ase(bolt_size)
        a1, proj, heff = self.calculate_anchor_bolt_values(bolt_size, bolt_data, futa) if bolt_data else ('', '', '')

        # Main Bar
        main_qty = hinge_data.get('nrb', '') if hinge_data else ''
        main_size = hinge_data.get('drb', '') if hinge_data else ''
        main_ase = self.calculate_main_bar_ase(main_qty, main_size)

        # ==================== INPUT DATA SECTION (only for new sheets) ====================
        if not is_update_mode:
            # ==================== ROW 1: Base Plate Detail Title ====================
            ws.merge_cells('A1:D1')
            ws['A1'] = section
            ws['A1'].font = Font(bold=True, size=12)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            ws.merge_cells('F1:L1')
            ws['F1'] = 'Base Plate Detail'
            ws['F1'].font = Font(bold=True, size=12)
            ws['F1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            # ==================== ROW 2: Headers ====================
            ws['A2'] = 'd'
            ws['B2'] = 'bf'
            ws['C2'] = 'tw'
            ws['D2'] = 'tf'

            ws['F2'] = 'N'
            ws['G2'] = 'B'
            ws['H2'] = 'P1'
            ws['I2'] = 'C'
            ws['J2'] = 'A'
            ws['K2'] = 'E'
            ws['L2'] = 'F'

            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}2'].font = Font(bold=True)
                ws[f'{col}2'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}2'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')

            # ==================== ROW 3: Values ====================
            if 'edited_section' in node_data and node_data['edited_section']:
                ws['A3'] = node_data['edited_section'].get('d', section_parts.get('d', ''))
                ws['B3'] = node_data['edited_section'].get('bf', section_parts.get('bf', ''))
                ws['C3'] = node_data['edited_section'].get('tw', section_parts.get('tw', ''))
                ws['D3'] = node_data['edited_section'].get('tf', section_parts.get('tf', ''))
            else:
                ws['A3'] = section_parts.get('d', '')
                ws['B3'] = section_parts.get('bf', '')
                ws['C3'] = section_parts.get('tw', '')
                ws['D3'] = section_parts.get('tf', '')

            # Check if edited values exist for Base Plate Detail
            if 'edited_base_plate_detail' in node_data and node_data['edited_base_plate_detail']:
                ws['F3'] = node_data['edited_base_plate_detail'].get('N', hinge_data.get('N', '') if hinge_data else '')
                ws['G3'] = node_data['edited_base_plate_detail'].get('B', hinge_data.get('B', '') if hinge_data else '')
                ws['H3'] = node_data['edited_base_plate_detail'].get('P1', hinge_data.get('P1', '') if hinge_data else '')
                ws['I3'] = node_data['edited_base_plate_detail'].get('C', hinge_data.get('C', '') if hinge_data else '')
                ws['J3'] = node_data['edited_base_plate_detail'].get('A', hinge_data.get('A', '') if hinge_data else '')
                ws['K3'] = node_data['edited_base_plate_detail'].get('E', hinge_data.get('E', '') if hinge_data else '')
                ws['L3'] = node_data['edited_base_plate_detail'].get('F', hinge_data.get('F', '') if hinge_data else '')
            else:
                if hinge_data:
                    ws['F3'] = hinge_data.get('N', '')
                    ws['G3'] = hinge_data.get('B', '')
                    ws['H3'] = hinge_data.get('P1', '')
                    ws['I3'] = hinge_data.get('C', '')
                    ws['J3'] = hinge_data.get('A', '')
                    ws['K3'] = hinge_data.get('E', '')
                    ws['L3'] = hinge_data.get('F', '')

            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')

            # ==================== ROW 5: Material Title ====================
            ws.merge_cells('A5:D5')
            ws['A5'] = 'Material'
            ws['A5'].font = Font(bold=True, size=12)
            ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A5'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            ws.merge_cells('F5:L5')
            ws['F5'] = 'Anchor Bolt'
            ws['F5'].font = Font(bold=True, size=12)
            ws['F5'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F5'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            # ==================== ROW 6: Headers ====================
            ws['A6'] = 'fy(Steel)'
            ws['B6'] = "f'c"
            ws['C6'] = 'fy(Mainbar)'
            ws['D6'] = 'fy(Tiebar)'

            ws['F6'] = 'nb'
            ws['G6'] = 'db'
            ws['H6'] = 'Ase'
            ws['I6'] = 'futa'
            ws['J6'] = 'A1'
            ws['K6'] = 'Proj'
            ws['L6'] = 'heff'

            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}6'].font = Font(bold=True)
                ws[f'{col}6'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}6'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')

            # ==================== ROW 7: Values ====================
            # Check if edited material values exist
            if 'edited_material' in node_data and node_data['edited_material']:
                ws['A7'] = node_data['edited_material'].get('fy_steel', fy_steel)
                ws['B7'] = node_data['edited_material'].get('fc', fc)
                ws['C7'] = node_data['edited_material'].get('fy_mainbar', fy_mainbar)
                ws['D7'] = node_data['edited_material'].get('fy_tiebar', fy_tiebar)
            else:
                ws['A7'] = fy_steel
                ws['B7'] = fc
                ws['C7'] = fy_mainbar
                ws['D7'] = fy_tiebar

            # Check if edited anchor bolt values exist
            if 'edited_anchor_bolt' in node_data and node_data['edited_anchor_bolt']:
                ws['F7'] = node_data['edited_anchor_bolt'].get('nb', nb)
                ws['G7'] = node_data['edited_anchor_bolt'].get('db', db)
                ws['H7'] = node_data['edited_anchor_bolt'].get('Ase', ase_bolt)
                ws['I7'] = node_data['edited_anchor_bolt'].get('futa', futa)
                ws['J7'] = node_data['edited_anchor_bolt'].get('A1', a1)
                ws['K7'] = node_data['edited_anchor_bolt'].get('Proj', proj)
                ws['L7'] = node_data['edited_anchor_bolt'].get('heff', heff)
            else:
                ws['F7'] = nb
                ws['G7'] = db
                ws['H7'] = ase_bolt
                ws['I7'] = futa
                ws['J7'] = a1
                ws['K7'] = proj
                ws['L7'] = heff

            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}7'].alignment = Alignment(horizontal='center', vertical='center')

            # ==================== ROW 9: Pier Detail Title ====================
            ws.merge_cells('A9:D9')
            ws['A9'] = 'Pier Detail'
            ws['A9'].font = Font(bold=True, size=12)
            ws['A9'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A9'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            ws.merge_cells('F9:L9')
            ws['F9'] = 'Main Bar                                Tie Bar'
            ws['F9'].font = Font(bold=True, size=12)
            ws['F9'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F9'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            # ==================== ROW 10: Headers ====================
            ws['A10'] = 'Np'
            ws['B10'] = 'Bp'
            ws['C10'] = 'TG'
            ws['D10'] = 'Cover c'

            ws['F10'] = 'Qty'
            ws['G10'] = 'Size'
            ws['H10'] = 'Size'
            ws['I10'] = 'X-legs'
            ws['J10'] = 'Y-legs'
            ws['K10'] = 'Layer 1'
            ws['L10'] = 'Layer 2'

            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}10'].font = Font(bold=True)
                ws[f'{col}10'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}10'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')

            # ==================== ROW 11: Values ====================
            # Check if edited values exist, use them; otherwise use hinge_data
            if 'edited_pier_detail' in node_data and node_data['edited_pier_detail']:
                ws['A11'] = node_data['edited_pier_detail'].get('Np', hinge_data.get('Np', '') if hinge_data else '')
                ws['B11'] = node_data['edited_pier_detail'].get('Bp', hinge_data.get('Bp', '') if hinge_data else '')
                ws['C11'] = node_data['edited_pier_detail'].get('TG', 25)  # Default 25 if not set
                ws['D11'] = node_data['edited_pier_detail'].get('c', hinge_data.get('c', '') if hinge_data else '')
            else:
                if hinge_data:
                    ws['A11'] = hinge_data.get('Np', '')
                    ws['B11'] = hinge_data.get('Bp', '')
                ws['C11'] = 25  # TG fixed value
                if hinge_data:
                    ws['D11'] = hinge_data.get('c', '')

            # Check if edited main bar values exist
            if 'edited_main_bar' in node_data and node_data['edited_main_bar']:
                ws['F11'] = node_data['edited_main_bar'].get('Qty', main_qty)
                ws['G11'] = node_data['edited_main_bar'].get('Size', main_size)
            else:
                ws['F11'] = main_qty
                ws['G11'] = main_size

            if hinge_data:
                ws['H11'] = hinge_data.get('dtb', '')
                ws['I11'] = hinge_data.get('X-leg', '')
                ws['J11'] = hinge_data.get('Y-leg', '')
                ws['K11'] = hinge_data.get('Layer 1', '')
                ws['L11'] = hinge_data.get('Layer 2', '')

            for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}11'].alignment = Alignment(horizontal='center', vertical='center')

            # ==================== ROW 13: Other Coefficient Title ====================
            ws.merge_cells('F13:I13')
            ws['F13'] = 'Other Coefficient'
            ws['F13'].font = Font(bold=True, size=11)
            ws['F13'].alignment = Alignment(horizontal='center', vertical='center')
            ws['F13'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

            # ==================== ROW 14: Headers ====================
            ws['F14'] = 'Friction μ'
            ws['G14'] = 'Ψc,P'
            ws['H14'] = 'Ψe'
            ws['I14'] = 'λ'

            for col in ['F', 'G', 'H', 'I']:
                ws[f'{col}14'].font = Font(bold=True)
                ws[f'{col}14'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}14'].fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')

            # ==================== ROW 15: Values ====================
            # Check if edited other coefficient values exist
            if 'edited_other_coeff' in node_data and node_data['edited_other_coeff']:
                ws['F15'] = node_data['edited_other_coeff'].get('friction_mu', 0.4)
                ws['G15'] = node_data['edited_other_coeff'].get('psi_c_p', 1.4)
                ws['H15'] = node_data['edited_other_coeff'].get('psi_e', 1)
                ws['I15'] = node_data['edited_other_coeff'].get('lambda', 1)
            else:
                ws['F15'] = 0.4  # Friction μ default value
                ws['G15'] = 1.4  # Ψc,P default value
                ws['H15'] = 1    # Ψe default value
                ws['I15'] = 1    # λ default value

            for col in ['F', 'G', 'H', 'I']:
                ws[f'{col}15'].alignment = Alignment(horizontal='center', vertical='center')

            # Set column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 3
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 10
            ws.column_dimensions['K'].width = 10
            ws.column_dimensions['L'].width = 10

            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in range(1, 16):
                for col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                    if row <= 11 or (row >= 13 and row <= 15):
                        if (row <= 11 and col in ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L']) or \
                           (row >= 13 and col in ['F', 'G', 'H', 'I']):
                            ws[f'{col}{row}'].border = thin_border

        # ==================== CONTROL LOAD TABLE (N1:S6) ====================

        # Read reaction data from CSV
        reaction_file = os.path.join(self.bpl_folder, 'reaction_data_sap2000.csv')

        if os.path.exists(reaction_file):
            try:
                # Read CSV and find data for this node
                with open(reaction_file, 'r', encoding='utf-8') as f:
                    lines = f.readlines()

                # Skip header lines (first 3 lines)
                node_reactions = []
                for line in lines[3:]:
                    parts = line.strip().split(',')
                    if len(parts) >= 9 and parts[0].strip() == node_name:
                        # Found reaction for this node
                        node_reactions.append({
                            'node': parts[0].strip(),
                            'loadcase': parts[1].strip(),
                            'casetype': parts[2].strip(),
                            'f1': parts[3].strip(),
                            'f2': parts[4].strip(),
                            'f3': parts[5].strip(),
                            'm1': parts[6].strip() if len(parts) > 6 else '',
                            'm2': parts[7].strip() if len(parts) > 7 else '',
                            'm3': parts[8].strip() if len(parts) > 8 else ''
                        })

                if node_reactions:
                    # ==================== INITIALIZE TRACKING FOR MAX RATIOS ====================
                    max_ratios = {
                        'compression': {'ratio': -1, 'lc': None, 'f1': 0, 'f2': 0, 'f3': 0},
                        'tensile': {'ratio': -1, 'lc': None, 'f1': 0, 'f2': 0, 'f3': 0},
                        'anchor_bolt': {'ratio': -1, 'lc': None, 'f1': 0, 'f2': 0, 'f3': 0},
                        'x_shear': {'ratio': -1, 'lc': None, 'f1': 0, 'f2': 0, 'f3': 0},
                        'y_shear': {'ratio': -1, 'lc': None, 'f1': 0, 'f2': 0, 'f3': 0}
                    }

                    # ==================== LOOP THROUGH ALL COMBOS ====================
                    print(f"Debug: Processing {len(node_reactions)} load combinations for node {node_name}")

                    for combo in node_reactions:
                        lc_name = combo['loadcase']
                        try:
                            f1_val = float(combo['f1'])
                            f2_val = float(combo['f2'])
                            f3_val = float(combo['f3'])
                        except:
                            continue

                        # Calculate 5 ratios for this combo
                        ratios = self._calculate_control_ratios(
                            f1=f1_val, f2=f2_val, f3=f3_val,
                            ws=ws, node_data=node_data
                        )

                        # Update max ratios if this combo is better
                        for check_type in ['compression', 'tensile', 'anchor_bolt', 'x_shear', 'y_shear']:
                            current_ratio = ratios.get(check_type, 0)
                            if current_ratio > max_ratios[check_type]['ratio']:
                                max_ratios[check_type]['ratio'] = current_ratio
                                max_ratios[check_type]['lc'] = lc_name
                                max_ratios[check_type]['f1'] = f1_val
                                max_ratios[check_type]['f2'] = f2_val
                                max_ratios[check_type]['f3'] = f3_val

                        print(f"Debug: LC {lc_name}: Compress={ratios.get('compression', 0):.3f}, "
                              f"Tensile={ratios.get('tensile', 0):.3f}, "
                              f"AB={ratios.get('anchor_bolt', 0):.3f}, "
                              f"X={ratios.get('x_shear', 0):.3f}, "
                              f"Y={ratios.get('y_shear', 0):.3f}")

                    # ==================== FILL CONTROL LOAD TABLE WITH MAX RATIOS ====================

                    # ROW 1: Headers
                    ws['N1'] = 'Control Load'
                    ws['O1'] = 'L/C'
                    ws['P1'] = 'F1'
                    ws['Q1'] = 'F2'
                    ws['R1'] = 'F3'
                    ws['S1'] = 'Ratio'

                    for col in ['N', 'O', 'P', 'Q', 'R', 'S']:
                        ws[f'{col}1'].font = Font(bold=True, size=11)
                        ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')
                        ws[f'{col}1'].fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

                    # ROW 2-6: Data with best combos
                    control_loads = [
                        ('N2', 'Compression', 'compression'),
                        ('N3', 'Tensile', 'tensile'),
                        ('N4', 'Anchor Bolt', 'anchor_bolt'),
                        ('N5', 'X-Shear check', 'x_shear'),
                        ('N6', 'Y-Shear check', 'y_shear')
                    ]

                    for cell_ref, label, check_type in control_loads:
                        row = cell_ref[1]  # Get row number
                        best_combo = max_ratios[check_type]

                        # Control Load label
                        ws[cell_ref] = label
                        ws[cell_ref].alignment = Alignment(horizontal='left', vertical='center')

                        # L/C (best combo for this check type)
                        ws[f'O{row}'] = best_combo['lc'] if best_combo['lc'] else ''
                        ws[f'O{row}'].alignment = Alignment(horizontal='center', vertical='center')

                        # F1, F2, F3
                        ws[f'P{row}'] = best_combo['f1']
                        ws[f'P{row}'].alignment = Alignment(horizontal='center', vertical='center')

                        ws[f'Q{row}'] = best_combo['f2']
                        ws[f'Q{row}'].alignment = Alignment(horizontal='center', vertical='center')

                        ws[f'R{row}'] = best_combo['f3']
                        ws[f'R{row}'].alignment = Alignment(horizontal='center', vertical='center')

                        # Ratio (will be filled during calculation)
                        ws[f'S{row}'] = ''
                        ws[f'S{row}'].alignment = Alignment(horizontal='center', vertical='center')

                    # Set column widths for Control Load table
                    ws.column_dimensions['N'].width = 15
                    ws.column_dimensions['O'].width = 12
                    ws.column_dimensions['P'].width = 12
                    ws.column_dimensions['Q'].width = 12
                    ws.column_dimensions['R'].width = 12
                    ws.column_dimensions['S'].width = 12

                    # Add borders to Control Load table
                    for row in range(1, 7):
                        for col in ['N', 'O', 'P', 'Q', 'R', 'S']:
                            ws[f'{col}{row}'].border = thin_border

                    print(f"Control Load table added for Node {node_name}")
                    print(f"Best LCs: Compress={max_ratios['compression']['lc']} ({max_ratios['compression']['ratio']:.3f}), "
                          f"Tensile={max_ratios['tensile']['lc']} ({max_ratios['tensile']['ratio']:.3f}), "
                          f"AB={max_ratios['anchor_bolt']['lc']} ({max_ratios['anchor_bolt']['ratio']:.3f}), "
                          f"X={max_ratios['x_shear']['lc']} ({max_ratios['x_shear']['ratio']:.3f}), "
                          f"Y={max_ratios['y_shear']['lc']} ({max_ratios['y_shear']['ratio']:.3f})")

                else:
                    print(f"No reaction data found for Node {node_name}")

            except Exception as e:
                print(f"Error reading reaction data: {e}")
        else:
            print(f"Reaction file not found: {reaction_file}")

        # ==================== COMPRESSION DESIGN CALCULATION (Row 13-39) ====================

        # Get values from sheet and convert to float
        try:
            def safe_float(value, default=0):
                """Safely convert value to float"""
                try:
                    return float(value) if value else default
                except:
                    return default

            n_val = safe_float(ws['F3'].value)
            b_val = safe_float(ws['G3'].value)
            p1_val = safe_float(ws['H3'].value)  # Base plate thickness
            d_val = safe_float(ws['A3'].value)
            bf_val = safe_float(ws['B3'].value)

            np_val = safe_float(ws['A11'].value)
            bp_val = safe_float(ws['B11'].value)

            fy_steel = safe_float(ws['A7'].value)
            fc_val = safe_float(ws['B7'].value)
            fy_mainbar = safe_float(ws['C7'].value)  # fy(Mainbar)
            fy_tiebar = safe_float(ws['D7'].value)  # fy(Tiebar)

            # Main bar
            main_qty_raw = ws['F11'].value
            main_size_raw = ws['G11'].value

            # Calculate Ast
            try:
                main_qty_num = int(safe_float(main_qty_raw))
                # Parse D19 -> 19
                main_size = str(main_size_raw) if main_size_raw else ''
                size_num = int(main_size.replace('D', '').strip()) if main_size else 0
                # Area = π × d² / 4
                import math
                area_one_bar = math.pi * (size_num ** 2) / 4
                ast = main_qty_num * area_one_bar
            except Exception as e:
                print(f"Error calculating Ast: {e}")
                ast = 0

            vertical_load = safe_float(ws['R2'].value)  # F3 from Control Load

            # Calculate Pu
            if vertical_load < 0:
                pu = 0
            else:
                pu = vertical_load

            print(f"Debug values: N={n_val}, B={b_val}, P1={p1_val}, d={d_val}, bf={bf_val}")
            print(f"Debug values: Np={np_val}, Bp={bp_val}, fy={fy_steel}, fc={fc_val}")
            print(f"Debug values: Ast={ast}, Vertical Load={vertical_load}")

            # ==================== ROW 13-15: Title & Control ====================
            ws['A13'] = '1. Base Plate Design'
            ws['A13'].font = Font(bold=True, size=11)

            ws['A14'] = 'Control Case'
            ws['B14'] = ws['O2'].value if ws['O2'].value else ''  # L/C from Control Load

            ws['A15'] = 'Vertical Load'
            ws['B15'] = vertical_load
            ws['C15'] = 'kN'

            # ==================== ROW 16-28: Concrete Bearing Limit ====================
            ws['A16'] = '* Concrete Bearing Limit'
            ws['A16'].font = Font(italic=True, bold=True)

            # Row 17: As per AISC 360-16
            ws['A17'] = 'As per AISC 360-16:'

            # Row 18: A1
            ws['A18'] = 'A1='
            a1 = n_val * b_val
            ws['B18'] = a1
            ws['C18'] = 'mm2'

            print(f"Debug A1 calculation: {n_val} * {b_val} = {a1}")

            # Row 19: A2
            ws['A19'] = 'A2='
            a2 = np_val * bp_val
            ws['B19'] = a2
            ws['C19'] = 'mm2'

            # Row 20: Pp
            ws['A20'] = 'Pp='
            pp = 0.85 * fc_val * a1 * math.sqrt(a2 / a1) * 0.001 if a1 != 0 else 0
            ws['B20'] = pp
            ws['C20'] = 'kN'

            # Row 21: 1.7f'cA1
            ws['A21'] = "1.7f'cA1="
            val_1_7 = 1.7 * fc_val * a1 * 0.001
            ws['B21'] = val_1_7
            ws['C21'] = 'kN'

            # Row 23: As per ACI 318M-14
            ws['A23'] = 'As per ACI 318M-14:'

            # Row 24: Pp (ACI) - using fy(Mainbar)
            ws['A24'] = 'Pp='
            pp_aci = 0.8 * (0.85 * fc_val * (a2 - ast) + fy_mainbar * ast) / 1000  # Convert to kN
            ws['B24'] = pp_aci
            ws['C24'] = 'kN'

            # Row 25: Resistant factor
            ws['A25'] = 'Resitant factor φ ='
            ws['B25'] = 0.65

            # Row 26: Pn
            ws['A26'] = 'Pn ='
            pn = min(min(pp, val_1_7), pp_aci)
            ws['B26'] = pn
            ws['C26'] = 'kN'

            # Row 27: φPn
            ws['A27'] = 'φPn'
            phi_pn = 0.65 * pn
            ws['B27'] = phi_pn
            ws['C27'] = 'kN'

            # Row 28: Rat1
            ws['A28'] = 'Rat 1 ='
            rat1 = abs(vertical_load) / phi_pn if phi_pn != 0 else 0
            ws['B28'] = rat1
            #ws['C28'] = '(NG or OK: so sánh B28 với 1)'

            # ==================== ROW 30-38: Base Plate Yielding Limit ====================
            ws['A30'] = '* Base Plate Yielding Limit ( For concrete bearing )'
            ws['A30'].font = Font(italic=True, bold=True)

            # Row 31
            ws['A31'] = 'For the critical base plate cantilever dimension, l'

            # Row 32: m
            ws['A32'] = 'm='
            m_val = (n_val - 0.95 * d_val) / 2
            ws['B32'] = m_val
            ws['C32'] = 'mm'

            # Row 33: n
            ws['A33'] = 'n='
            n_calc = (b_val - 0.8 * bf_val) / 2
            ws['B33'] = n_calc
            ws['C33'] = 'mm'

            # Row 34: n'
            ws['A34'] = "n'="
            n_prime = (d_val * bf_val) ** 0.5 / 4
            ws['B34'] = n_prime
            ws['C34'] = 'mm'

            # Row 35: l
            ws['A35'] = 'l='
            l_val = max(m_val, n_calc, n_prime)
            ws['B35'] = l_val
            ws['C35'] = 'mm'

            # Row 36: Resistant factor
            ws['A36'] = 'Resitant factor φ ='
            ws['B36'] = 0.9

            # Row 37: tmin
            ws['A37'] = 'tmin='
            if fy_steel != 0 and b_val != 0 and n_val != 0:
                tmin = l_val * math.sqrt(2 * pu * 1000 / (0.9 * fy_steel * b_val * n_val))
            else:
                tmin = 0
            ws['B37'] = tmin
            ws['C37'] = 'mm'

            # Row 38: Rat2
            ws['A38'] = 'Rat 2 ='
            rat2 = tmin / p1_val if p1_val != 0 else 0
            ws['B38'] = rat2
            #ws['C38'] = '(NG or OK: so sánh B38 với 1)'

            # ==================== UPDATE S2: Final Ratio ====================
            final_ratio = max(rat1, rat2)
            ws['S2'] = final_ratio
            ws['S2'].alignment = Alignment(horizontal='center', vertical='center')

            # ==================== ROW 41-51: TENSILE LOADING ====================
            ws['A41'] = '1.2 Design Requirements for tensile loading'
            ws['A41'].font = Font(bold=True, size=11)

            # Row 42: Control Case
            ws['A42'] = 'Control Case'
            ws['B42'] = ws['O3'].value if ws['O3'].value else ''  # L/C from Tensile row

            # Row 43: Vertical Load
            ws['A43'] = 'Vertical Load'
            vertical_load_tensile = safe_float(ws['R3'].value)  # F3 from Tensile row
            ws['B43'] = vertical_load_tensile
            ws['C43'] = 'kN'

            # Row 45: Title
            ws['A45'] = '* Base Plate Yielding Limit ( For bolt tension )'
            ws['A45'].font = Font(italic=True, bold=True)

            # Row 46: Nua
            ws['A46'] = 'Nua='
            nb_val = safe_float(ws['F7'].value)  # nb from Anchor Bolt section
            if vertical_load_tensile < 0:
                nua = abs(vertical_load_tensile / nb_val) if nb_val != 0 else 0
            else:
                nua = 0
            ws['B46'] = nua
            ws['C46'] = 'kN'

            # Row 47: Mu
            ws['A47'] = 'Mu='
            a_val = safe_float(ws['J3'].value)  # A from Base Plate Detail
            tw_val = safe_float(ws['C3'].value)  # tw
            mu = nua * (a_val / 2 - tw_val / 2)
            ws['B47'] = mu
            ws['C47'] = 'kN-mm'

            # Row 48: beff
            ws['A48'] = 'beff='
            beff = (b_val / 2 - tw_val / 2) * 2
            ws['B48'] = beff
            ws['C48'] = 'mm'

            # Row 49: Resistant factor
            ws['A49'] = 'Resitant factor φ ='
            ws['B49'] = 0.9

            # Row 50: t_reqd
            ws['A50'] = 't_reqd='
            if fy_steel != 0 and beff != 0:
                t_reqd = math.sqrt(4 * mu *1000 / (0.9 * fy_steel * beff))
            else:
                t_reqd = 0
            ws['B50'] = t_reqd
            ws['C50'] = 'mm'

            # Row 51: Rat 3
            ws['A51'] = 'Rat 3 ='
            rat3 = t_reqd / p1_val if p1_val != 0 else 0
            ws['B51'] = rat3

            # ==================== UPDATE S3: Rat3 ====================
            ws['S3'] = rat3
            ws['S3'].alignment = Alignment(horizontal='center', vertical='center')

            print(f"Tensile calculation added: Nua={nua:.3f}, Mu={mu:.3f}, t_reqd={t_reqd:.3f}, Rat3={rat3:.3f}")

            # ==================== ROW 54-58: ANCHOR BOLT DESIGN HEADER ====================
            ws['A54'] = '2. Anchor Bolt Design'
            ws['A54'].font = Font(bold=True, size=11)

            ws['A55'] = 'Control Case'
            ws['B55'] = ws['O4'].value if ws['O4'].value else ''  # L/C from Anchor Bolt row

            ws['A56'] = 'Vertical Load'
            vertical_load_ab = safe_float(ws['R4'].value)  # F3
            ws['B56'] = vertical_load_ab
            ws['C56'] = 'kN'

            ws['A57'] = 'Horizontal load'
            h_load_x = safe_float(ws['P4'].value)  # F1 (X-Dir)
            ws['B57'] = h_load_x
            ws['C57'] = 'kN (X-Dir)'

            ws['A58'] = 'Horizontal load'
            h_load_y = safe_float(ws['Q4'].value)  # F2 (Y-Dir)
            ws['B58'] = h_load_y
            ws['C58'] = 'kN (Y-Dir)'

            # ==================== ROW 60-91: TENSILE LOADING ====================
            ws['A60'] = '2.1 Design Requirements for tensile loading'
            ws['A60'].font = Font(bold=True, italic=True, size=10)

            # Row 61: Steel Strength
            ws['A61'] = '* Steel Strength of Anchor in Tension'
            ws['A61'].font = Font(italic=True, bold=True)

            # Row 62: Nsa
            ws['A62'] = 'Nsa='
            ase_val = safe_float(ws['H7'].value)  # Ase from Anchor Bolt section
            futa_val = safe_float(ws['I7'].value)  # futa
            nsa = ase_val * futa_val / 1000  # Convert to kN
            ws['B62'] = nsa
            ws['C62'] = 'kN'

            # Row 64: Pullout Resistance
            ws['A64'] = '* Pullout Resistance of Anchor in Tension'
            ws['A64'].font = Font(italic=True, bold=True)

            # Row 65: Abrg
            ws['A65'] = 'Abrg='
            db_num = safe_float(ws['G7'].value)  # db

            # Get W and NutW from Anchor Bolt Table
            bolt_data = self.get_anchor_bolt_data(f"M{int(db_num)}")
            w_val = safe_float(bolt_data.get('W', 0)) if bolt_data else 0
            nutw_val = safe_float(bolt_data.get('NutW', 0)) if bolt_data else 0

            if w_val > 0:
                abrg = w_val ** 2 - math.pi / 4 * (db_num ** 2)
            else:
                abrg = 0.866 * (nutw_val ** 2)
            ws['B65'] = abrg
            ws['C65'] = 'mm2'

            # Row 66: Nap
            ws['A66'] = 'Nap='
            nap = 8 * abrg * fc_val / 1000  # Convert to kN
            ws['B66'] = nap
            ws['C66'] = 'kN'

            # Row 67: Ψc,P (from Other Coefficient G15)
            ws['A67'] = 'Ψc,P ='
            ws['B67'] = f'=G15'

            # Row 68: Npn
            ws['A68'] = 'Npn'
            npn = 1.4 * nap
            ws['B68'] = npn
            ws['C68'] = 'kN'

            # Row 70: Side-face Blowout
            ws['A70'] = '* Side-face Blowout Resistance of Anchor in Tension'
            ws['A70'].font = Font(italic=True, bold=True)

            # Row 71: Ca1
            ws['A71'] = 'Ca1='
            bp_val = safe_float(ws['B11'].value)  # Bp from Pier Detail
            np_pier = safe_float(ws['A11'].value)  # Np from Pier Detail
            a_val = safe_float(ws['J3'].value)  # A from Base Plate Detail (khoảng cách bolt)
            c_val = safe_float(ws['I3'].value)  # C from Base Plate Detail (khoảng cách bolt)

            ca1 = min((bp_val - a_val) / 2, (np_pier - c_val) / 2)
            ws['B71'] = ca1
            ws['C71'] = 'mm'

            # Row 72: 2.5Ca1
            ws['A72'] = '2.5Ca1='
            val_2_5ca1 = 2.5 * ca1
            ws['B72'] = val_2_5ca1
            ws['C72'] = 'mm'

            # Row 73: hef
            ws['A73'] = 'hef='
            hef_val = safe_float(ws['L7'].value)  # heff from Anchor Bolt
            ws['B73'] = hef_val
            ws['C73'] = 'mm'

            # Row 74: Ca2
            ws['A74'] = 'Ca2='
            ca2 = max((bp_val - a_val) / 2, (np_pier - c_val) / 2)
            ws['B74'] = ca2
            ws['C74'] = 'mm'

            # Row 75: ca2/ca1
            ws['A75'] = 'ca2/Ca1='
            ca2_ca1_ratio = ca2 / ca1 if ca1 != 0 else 0
            ws['B75'] = ca2_ca1_ratio

            # Row 76: f1
            ws['A76'] = 'f1='
            if ca2_ca1_ratio < 1:
                f1 = 1
            elif ca2_ca1_ratio > 3:
                f1 = 1 + (ca1 / ca2) / 4
            else:
                f1 = (1 + ca2_ca1_ratio) / 4
            ws['B76'] = f1

            # Row 77: Check if hef > 2.5Ca1 (side-face blowout applies)
            if hef_val > val_2_5ca1:
                ws['B77'] = "Consider side-face blowout"
            else:
                ws['B77'] = "Side-face blowout strength does not apply"
            ws['B77'].font = Font(italic=True, color='0000FF')

            # Row 78: Check corner effect
            if ca2_ca1_ratio < 1 or ca2_ca1_ratio > 3:
                ws['B78'] = "Corner effect does not apply"
            else:
                ws['B78'] = "Consider corner effect"
            ws['B78'].font = Font(italic=True, color='0000FF')

            # Row 79: Nsb
            ws['A79'] = 'Nsb='
            if hef_val > val_2_5ca1:
                nsb = f1 * (13 * ca1 * abrg * 0.5) * (fc_val ** 0.5) / 1000
            else:
                nsb = 0  # NA
            ws['B79'] = nsb if nsb != 0 else 'NA'
            ws['C79'] = 'kN'

            # Row 80: S1
            ws['A80'] = 'S1'
            s1 = min(a_val, c_val)
            ws['B80'] = s1
            ws['C80'] = 'mm'

            # Row 81: 6 x Ca1
            ws['A81'] = '6 x Ca1'
            val_6ca1 = 6 * ca1
            ws['B81'] = val_6ca1
            ws['C81'] = 'mm'

            # Row 82: f2
            ws['A82'] = 'f2'
            if s1 < val_6ca1:
                f2 = 1 + s1 / (6 * ca1) if ca1 != 0 else 0
            else:
                f2 = 1
            ws['B82'] = f2

            # Row 83: Check close spacing effect
            if s1 < val_6ca1:
                ws['B83'] = "Consider close spacing effect"
            else:
                ws['B83'] = "Close spacing effect does not apply"
            ws['B83'].font = Font(italic=True, color='0000FF')

            # Row 84: NSbg
            ws['A84'] = 'NSbg='
            if hef_val > val_2_5ca1 and nsb != 'NA':
                nsbg = f2 * nsb / nb_val * 2 if nb_val != 0 else 0
            else:
                nsbg = 'NA'
            ws['B84'] = nsbg if nsbg != 'NA' else 'NA'
            ws['C84'] = 'kN'

            # Row 86: Tensile Design Summary
            ws['A86'] = '* Tensile Design Summary'
            ws['A86'].font = Font(italic=True, bold=True)

            # Row 87: Resistant factor
            ws['A87'] = 'Resitant factor φ ='
            ws['B87'] = 0.75

            # Row 88: Nn
            ws['A88'] = 'Nn='
            # IF(Nsbg="NA", MIN(Nsa,Npn), MIN(Nsa,Npn,Nsbg))
            if nsbg == 'NA':
                nn = min(nsa, npn)
            else:
                nn = min(nsa, npn, nsbg)
            ws['B88'] = nn
            ws['C88'] = 'kN'

            # Row 89: φNn
            ws['A89'] = 'φNn='
            phi_nn = 0.75 * nn
            ws['B89'] = phi_nn
            ws['C89'] = 'kN'

            # Row 90: Nua
            ws['A90'] = 'Nua='
            if vertical_load_ab < 0:
                nua_ab = abs(vertical_load_ab / nb_val) if nb_val != 0 else 0
            else:
                nua_ab = 0
            ws['B90'] = nua_ab
            ws['C90'] = 'kN'

            # Row 91: Rat 4
            ws['A91'] = 'Rat 4 ='
            rat4 = nua_ab / phi_nn if phi_nn != 0 else 0
            ws['B91'] = rat4

            # ==================== ROW 93-101: SHEAR LOADING ====================
            ws['A93'] = '2.2 Design Requirements for shear loading'
            ws['A93'].font = Font(bold=True, italic=True, size=10)

            # Row 94: Steel Strength
            ws['A94'] = '* Steel Strength of Anchor in Shear'
            ws['A94'].font = Font(italic=True, bold=True)

            # Row 95: Vsa
            ws['A95'] = 'Vsa='
            vsa = 0.6 * ase_val * futa_val * 0.8 / 1000
            ws['B95'] = vsa
            ws['C95'] = 'kN'

            # Row 97: Shear Design Summary
            ws['A97'] = '* Shear Design Summary'
            ws['A97'].font = Font(italic=True, bold=True)

            # Row 98: Resistant factor
            ws['A98'] = 'Resitant factor φ ='
            ws['B98'] = 0.65

            # Row 99: φVn
            ws['A99'] = 'φVn='
            phi_vn = 0.65 * vsa
            ws['B99'] = phi_vn
            ws['C99'] = 'kN'

            # Row 100: Vua
            ws['A100'] = 'Vua='
            friction_mu = safe_float(ws['C7'].value)  # Friction μ
            h_resultant = math.sqrt(h_load_x ** 2 + h_load_y ** 2)
            if vertical_load_ab < 0:
                pu_shear = 0
            else:
                pu_shear = vertical_load_ab

            vua_calc = (h_resultant - friction_mu * pu_shear) / nb_val
            vua = max(vua_calc, 0)
            ws['B100'] = vua
            ws['C100'] = 'kN'

            # Row 101: Rat 5
            ws['A101'] = 'Rat 5 ='
            rat5 = vua / phi_vn if phi_vn != 0 else 0
            ws['B101'] = rat5

            # ==================== ROW 103-107: INTERACTION ====================
            ws['A103'] = '2.3 Interaction of Tensile and Shear Forces'
            ws['A103'].font = Font(bold=True, italic=True, size=10)

            # Row 104: Vua / φVn
            ws['A104'] = 'Vua / φVn ='
            ws['B104'] = rat5
            ws['C104'] = '≤0.2' if rat5 <= 0.2 else '>0.2'

            # Row 105: Nua / φNn
            ws['A105'] = 'Nua / φNn ='
            ws['B105'] = rat4
            ws['C105'] = '≤0.2' if rat4 <= 0.2 else '>0.2'

            # Row 106: Rat 6
            ws['A106'] = 'Rat 6 ='
            # IF(D104*D105=1,B104+B105,MAX(B104,B105))
            ws['D104'] = 1 if rat5 > 0.2 else 0
            ws['D105'] = 1 if rat4 > 0.2 else 0

            if ws['D104'].value * ws['D105'].value == 1:
                rat6 = rat5 + rat4
            else:
                rat6 = max(rat5, rat4)
            ws['B106'] = rat6

            # Row 107: Control text
            if ws['D104'].value * ws['D105'].value == 1:
                ws['B107'] = "Interaction control"
            elif rat6 == rat5:
                ws['B107'] = "Shear control"
            else:
                ws['B107'] = "Tension control"
            ws['B107'].font = Font(italic=True, color='0000FF')

            # ==================== UPDATE S4: Rat 6 ====================
            ws['S4'] = rat6
            ws['S4'].alignment = Alignment(horizontal='center', vertical='center')

            print(f"Anchor Bolt calculation added: Rat4={rat4:.3f}, Rat5={rat5:.3f}, Rat6={rat6:.3f}")

            # ==================== ROW 110-162: PIER REINFORCEMENT CHECK ====================
            ws['A110'] = '3. Pier Reinforcement Check'
            ws['A110'].font = Font(bold=True, size=11)

            # ==================== SECTION 3.1: Transfer to Vertical Rebars ====================
            ws['A111'] = '3.1 Transfer of Anchor Load to Vertical Rebars'
            ws['A111'].font = Font(bold=True, italic=True, size=10)

            # Row 112: Control Case
            ws['A112'] = 'Control Case'
            ws['B112'] = ws['O3'].value if ws['O3'].value else ''  # Same as Tensile

            # Row 113: Vertical Load
            ws['A113'] = 'Vertical Load'
            vertical_load_pier = safe_float(ws['R3'].value)  # F3 from Tensile
            ws['B113'] = vertical_load_pier
            ws['C113'] = 'kN'

            # Row 114: Resistant factor
            ws['A114'] = 'Resitant factor φ ='
            ws['B114'] = 0.75

            # Parse rebar sizes
            main_size_str = str(ws['G11'].value) if ws['G11'].value else ''
            tie_size_str = str(ws['H11'].value) if ws['H11'].value else ''
            d_mainbar = int(main_size_str.replace('D', '').strip()) if main_size_str else 0
            d_tiebar = int(tie_size_str.replace('D', '').strip()) if tie_size_str else 0

            # Row 115: dmax
            ws['A115'] = 'dmax='
            hef_pier = safe_float(ws['L7'].value)  # heff
            w_pier = safe_float(bolt_data.get('W', 0)) if bolt_data else 0
            dmax = hef_pier / 3 + w_pier / 2
            ws['B115'] = dmax

            # Row 116: 0.5hef
            ws['A116'] = '0.5hef='
            val_0_5hef = 0.5 * hef_pier
            ws['B116'] = val_0_5hef
            # Check
            if dmax > val_0_5hef:
                ws['C116'] = 'NG'
            else:
                ws['C116'] = 'OK'
            ws['C116'].font = Font(color='0000FF')

            # Row 117: dact
            ws['A117'] = 'dact='
            cover_c = safe_float(ws['D11'].value)  # Cover c from Pier Detail
            c_bp = safe_float(ws['I3'].value)  # C from Base Plate Detail
            a_bp = safe_float(ws['J3'].value)  # A from Base Plate Detail

            ca1_pier = min((bp_val - a_bp) / 2, (np_pier - c_bp) / 2)
            ca2_pier = max((bp_val - a_bp) / 2, (np_pier - c_bp) / 2)

            dact = ((ca1_pier - cover_c - d_mainbar) ** 2 + (ca2_pier - cover_c - d_mainbar) ** 2) ** 0.5 - w_pier / 2
            ws['B117'] = dact
            # Check
            if dact > dmax:
                ws['C117'] = 'NG'
            else:
                ws['C117'] = 'OK'
            ws['C117'].font = Font(color='0000FF')

            # Row 118: Ca1-c-db/2
            ws['A118'] = 'Ca1-c-db/2='
            val_ca1_calc = (np_pier - c_bp) / 2 - cover_c - d_mainbar / 2
            ws['B118'] = val_ca1_calc
            # Check
            if val_ca1_calc < c_bp / 2:
                ws['C118'] = 'NG'
            else:
                ws['C118'] = 'OK'
            ws['C118'].font = Font(color='0000FF')

            # Row 119: Ca1-c-db/2
            ws['A119'] = 'Ca1-c-db/2='
            val_ca1_calc2 = (bp_val - a_bp) / 2 - cover_c - d_mainbar / 2
            ws['B119'] = val_ca1_calc2
            # Check
            if val_ca1_calc2 < a_bp / 2:
                ws['C119'] = 'NG'
            else:
                ws['C119'] = 'OK'
            ws['C119'].font = Font(color='0000FF')

            # Row 120: nreq'd
            ws['A120'] = "nreq'd ="
            fy_mainbar = safe_float(ws['C7'].value)  # fy(Mainbar)
            fy_tiebar = safe_float(ws['D7'].value)  # fy(Tiebar)
            as_mainbar = math.pi * (d_mainbar ** 2) / 4  # Area of one main bar

            if vertical_load_pier < 0:
                tension_force = abs(vertical_load_pier)
            else:
                tension_force = 0

            nreqd = math.ceil(tension_force * 1000 / (0.75 * fy_mainbar * as_mainbar)) if as_mainbar != 0 else 0
            ws['B120'] = nreqd

            # Row 121: Rat 7
            ws['A121'] = 'Rat 7 ='
            qty_mainbar = safe_float(ws['F11'].value)  # Qty from Main Bar
            rat7 = nreqd / qty_mainbar if qty_mainbar != 0 else 0
            ws['B121'] = rat7

            # Row 122: ld_prov
            ws['A122'] = 'ld_prov ='
            e_bp = safe_float(ws['K3'].value)  # E from Base Plate Detail
            f_bp = safe_float(ws['L3'].value)  # F from Base Plate Detail
            layer1 = safe_float(ws['K11'].value)  # Layer 1

            ld_prov = hef_pier - cover_c - (d_mainbar + dact) * math.tan(35 * math.pi / 180)
            ws['B122'] = ld_prov
            ws['C122'] = 'mm'

            # Row 123: ld
            ws['A123'] = 'ld='
            # Lookup from Rebar Development Length sheet
            ld_value = self.lookup_rebar_dev_length(d_mainbar, 'Ld')
            ws['B123'] = ld_value
            ws['C123'] = 'mm'

            # Row 124: ld_req'd
            ws['A124'] = "ld_req'd ="
            ld_reqd = ld_value * rat7 if ld_value else 0
            ws['B124'] = ld_reqd
            ws['C124'] = 'mm'

            # Row 125: Rat 8
            ws['A125'] = 'Rat 8 ='
            rat8 = ld_reqd / ld_prov if ld_prov != 0 else 0
            ws['B125'] = rat8

            # ==================== SECTION 3.2: Transfer to Horizontal Rebars ====================
            ws['A127'] = '3.2 Transfer of Anchor Load to Horizontal Rebars'
            ws['A127'].font = Font(bold=True, italic=True, size=10)

            # Row 128: Ψe (from Other Coefficient H15)
            ws['A128'] = 'Ψe ='
            ws['B128'] = f'=H15'

            # Row 129: λ (from Other Coefficient I15)
            ws['A129'] = 'λ ='
            ws['B129'] = f'=I15'

            # Row 130: ldh (using fy(Tiebar))
            ws['A130'] = 'ldh'
            ldh_value = 0.24 * fy_tiebar * d_tiebar / (fc_val ** 0.5) if fc_val > 0 else 0
            ws['B130'] = ldh_value
            ws['C130'] = 'mm'

            # ==================== CONSIDER BETA ====================
            beta_val = node_data.get('beta', 90)  # Default 90 (current behavior)
            if beta_val == 0:
                # Web ∥ X → X-shear uses Np/C, Y-shear uses Bp/A
                pier_x  = np_pier
                space_x = c_val
                pier_y  = bp_val
                space_y = a_val
            else:
                # Web ⊥ X (beta=90) → X-shear uses Bp/A, Y-shear uses Np/C
                pier_x  = bp_val
                space_x = a_val
                pier_y  = np_pier
                space_y = c_val

            # ==================== FOR X - DIRECTION ====================
            ws['A132'] = '* For X - Direction'
            ws['A132'].font = Font(italic=True, bold=True)

            ws['A133'] = 'Control Case'
            ws['B133'] = ws['O5'].value if ws['O5'].value else ''

            ws['A134'] = 'Vertical Load'
            vertical_load_x = safe_float(ws['R5'].value)
            ws['B134'] = vertical_load_x
            ws['C134'] = 'kN'

            ws['A135'] = 'Horizontal Load'
            h_load_x_pier = safe_float(ws['P5'].value)
            ws['B135'] = h_load_x_pier
            ws['C135'] = 'kN'

            ws['A136'] = 'Vx='
            if vertical_load_x < 0:
                pu_x = 0
            else:
                pu_x = vertical_load_x
            vx_calc = abs(h_load_x_pier) - friction_mu * pu_x
            vx = max(vx_calc, 0)
            ws['B136'] = vx
            ws['C136'] = 'kN'

            ws['A137'] = 'Atie='
            ws['B137'] = math.pi * (d_tiebar ** 2) / 4
            ws['C137'] = 'mm2'

            # lda dùng pier_x, space_x
            ws['A138'] = 'lda_A_L(x) ='
            tan35 = math.tan(35 * math.pi / 180)
            layer1_val = safe_float(ws['K11'].value)
            layer2_val = safe_float(ws['L11'].value)
            lda_a_lx = (pier_x + space_x) / 2 - db_num - cover_c - layer1_val * tan35
            ws['B138'] = lda_a_lx
            ws['C138'] = 'mm'

            ws['A139'] = 'lda_A_R(x) ='
            lda_a_rx = (pier_x - 2 * cover_c) - lda_a_lx
            ws['B139'] = lda_a_rx
            ws['C139'] = 'mm'

            ws['A140'] = 'lda_B_L(x) ='
            lda_b_lx = (pier_x + space_x) / 2 - db_num - cover_c - (layer1_val + layer2_val) * tan35
            ws['B140'] = lda_b_lx
            ws['C140'] = 'mm'

            ws['A141'] = 'lda_B_R(x) ='
            lda_b_rx = (pier_x - 2 * cover_c) - lda_b_lx
            ws['B141'] = lda_b_rx
            ws['C141'] = 'mm'

            ws['A142'] = 'fsx='
            fsx = min(lda_a_lx, lda_a_rx, lda_b_lx, lda_b_rx) / ldh_value * fy_tiebar if ldh_value != 0 else 0
            ws['B142'] = fsx
            ws['C142'] = 'Mpa'

            ws['A143'] = 'Resitant factor φ ='
            ws['B143'] = 0.75

            ws['A144'] = "Atie_req'd(x)="
            x_legs = safe_float(ws['I11'].value)
            atie_reqd_x = vx / 2 * 1000 / (0.75 * fsx) if fsx != 0 else 0
            ws['B144'] = atie_reqd_x
            ws['C144'] = 'mm2'

            ws['A145'] = 'Atie_prov(x)='
            atie_prov_x = x_legs * (math.pi * (d_tiebar ** 2) / 4)
            ws['B145'] = atie_prov_x
            ws['C145'] = 'mm2'

            ws['A146'] = 'Rat 9 ='
            rat9 = atie_reqd_x / atie_prov_x if atie_prov_x != 0 else 0
            ws['B146'] = rat9

            # ==================== FOR Y - DIRECTION ====================
            ws['A148'] = '* For Y - Direction'
            ws['A148'].font = Font(italic=True, bold=True)

            ws['A149'] = 'Control Case'
            ws['B149'] = ws['O6'].value if ws['O6'].value else ''

            ws['A150'] = 'Vertical Load'
            vertical_load_y = safe_float(ws['R6'].value)
            ws['B150'] = vertical_load_y
            ws['C150'] = 'kN'

            ws['A151'] = 'Horizontal Load'
            h_load_y_pier = safe_float(ws['Q6'].value)
            ws['B151'] = h_load_y_pier
            ws['C151'] = 'kN'

            ws['A152'] = 'Vy='
            if vertical_load_y < 0:
                pu_y = 0
            else:
                pu_y = vertical_load_y
            vy_calc = abs(h_load_y_pier) - friction_mu * pu_y
            vy = max(vy_calc, 0)
            ws['B152'] = vy
            ws['C152'] = 'kN'

            ws['A153'] = 'Atie='
            ws['B153'] = math.pi * (d_tiebar ** 2) / 4
            ws['C153'] = 'mm2'

            # lda dùng pier_y, space_y
            ws['A154'] = 'lda_A_L(y) ='
            lda_a_ly = (pier_y + space_y) / 2 - db_num - cover_c - layer1_val * tan35
            ws['B154'] = lda_a_ly
            ws['C154'] = 'mm'

            ws['A155'] = 'lda_A_R(y) ='
            lda_a_ry = (pier_y - 2 * cover_c) - lda_a_ly
            ws['B155'] = lda_a_ry
            ws['C155'] = 'mm'

            ws['A156'] = 'lda_B_L(y) ='
            lda_b_ly = (pier_y + space_y) / 2 - db_num - cover_c - (layer1_val + layer2_val) * tan35
            ws['B156'] = lda_b_ly
            ws['C156'] = 'mm'

            ws['A157'] = 'lda_B_R(y) ='
            lda_b_ry = (pier_y - 2 * cover_c) - lda_b_ly
            ws['B157'] = lda_b_ry
            ws['C157'] = 'mm'

            ws['A158'] = 'fsy='
            fsy = min(lda_a_ly, lda_a_ry, lda_b_ly, lda_b_ry) / ldh_value * fy_tiebar if ldh_value != 0 else 0
            ws['B158'] = fsy
            ws['C158'] = 'Mpa'

            ws['A159'] = 'Resitant factor φ ='
            ws['B159'] = 0.75

            ws['A160'] = "Atie_req'd(y)="
            y_legs = safe_float(ws['J11'].value)
            atie_reqd_y = vy / 2 * 1000 / (0.75 * fsy) if fsy != 0 else 0
            ws['B160'] = atie_reqd_y
            ws['C160'] = 'mm2'

            ws['A161'] = 'Atie_prov(y)='
            atie_prov_y = y_legs * (math.pi * (d_tiebar ** 2) / 4)
            ws['B161'] = atie_prov_y
            ws['C161'] = 'mm2'

            ws['A162'] = 'Rat 10 ='
            rat10 = atie_reqd_y / atie_prov_y if atie_prov_y != 0 else 0
            ws['B162'] = rat10

            # ==================== UPDATE S5 and S6 ====================
            ws['S5'] = rat9
            ws['S5'].alignment = Alignment(horizontal='center', vertical='center')

            ws['S6'] = rat10
            ws['S6'].alignment = Alignment(horizontal='center', vertical='center')

            print(f"Pier Reinforcement calculation added: Rat7={rat7:.3f}, Rat8={rat8:.3f}, Rat9={rat9:.3f}, Rat10={rat10:.3f}")

            # Add borders to calculation area
            for row in range(13, 163):  # Extend to row 162
                for col in ['A', 'B', 'C', 'D']:
                    if ws[f'{col}{row}'].value is not None:
                        ws[f'{col}{row}'].border = thin_border

            print(f"Compression calculation added: Rat1={rat1:.3f}, Rat2={rat2:.3f}, Final={final_ratio:.3f}")

        except Exception as e:
            print(f"Error in compression calculation: {e}")
            import traceback
            traceback.print_exc()

        # Save workbook
        wb.save(filepath)
        wb.close()

        print(f"Sheet '{sheet_name}' created/updated in {filename}")
        return filepath
