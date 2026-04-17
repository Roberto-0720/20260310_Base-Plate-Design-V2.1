"""
logic/node_manager_mixin.py
============================
Node selection, material dropdowns, apply/copy settings,
and control load table dialog.

App state used: self.base_plate_nodes, self.selected_node,
                self.material_trees, self.bpl_folder,
                self.dropdown_frame, self.node_info_label,
                self.apply_btn, self.edit_detail_btn, self.copy_multi_btn,
                self.anchor_bolt_var, self.base_plate_var, self.concrete_var,
                self.mainbar_var, self.tiebar_var, self.bolt_var,
                self.hinge_fixed_var, self.detail_var,
                self.status_label, self.root
"""

import os
import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl


class NodeManagerMixin:

    # ------------------------------------------------------------------
    # Node selection + dropdown population
    # ------------------------------------------------------------------

    def select_node(self, node_name):
        """Select a node and show its details"""
        self.selected_node = node_name
        node_data = self.base_plate_nodes[node_name]

        # Update info label
        info_text = (
            f"Node: {node_name}\n"
            f"Section: {node_data['section'] or 'N/A'}\n"
            f"Position: ({node_data['x']:.2f}, {node_data['y']:.2f})"
        )
        self.node_info_label.config(text=info_text, fg='#1a472a')

        # If design check has been run, show control load table
        if node_data.get('design_status') and node_data['design_status'] != 'Not Checked':
            self.show_control_load_table(node_name)

        # Clear and rebuild dropdown frame
        for widget in self.dropdown_frame.winfo_children():
            widget.destroy()

        # Get material lists from Data.xlsx
        anchor_bolt_list = self.get_material_list('anchor_bolt')
        base_plate_list = self.get_material_list('base_plate')
        concrete_list = self.get_material_list('concrete')
        rebar_list = self.get_material_list('rebar')
        bolt_list = self.get_bolt_size_list()
        detail_list = self.get_detail_type_list(node_data['section'])

        # Helper function to get default value (first non-empty item or stored value)
        def get_default(stored_value, option_list):
            if stored_value:
                return stored_value
            # Find first non-empty item in list
            for item in option_list:
                if item and item.strip():  # Skip empty strings
                    return item
            return ''

        # Anchor Bolt Material
        tk.Label(self.dropdown_frame, text="Anchor Bolt Material:", font=('Arial', 9, 'bold'),
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(5, 2))
        self.anchor_bolt_var = tk.StringVar(value=get_default(node_data.get('material_anchor_bolt', ''), anchor_bolt_list))
        anchor_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.anchor_bolt_var,
                                   values=anchor_bolt_list, state='readonly', width=25)
        anchor_menu.pack(fill='x', pady=(0, 8))

        # Base Plate Material
        tk.Label(self.dropdown_frame, text="Base Plate Material:", font=('Arial', 9, 'bold'),
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.base_plate_var = tk.StringVar(value=get_default(node_data.get('material_base_plate', ''), base_plate_list))
        baseplate_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.base_plate_var,
                                      values=base_plate_list, state='readonly', width=25)
        baseplate_menu.pack(fill='x', pady=(0, 8))

        # Concrete Material
        tk.Label(self.dropdown_frame, text="Concrete Material:", font=('Arial', 9, 'bold'),
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.concrete_var = tk.StringVar(value=get_default(node_data.get('material_concrete', ''), concrete_list))
        concrete_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.concrete_var,
                                     values=concrete_list, state='readonly', width=25)
        concrete_menu.pack(fill='x', pady=(0, 8))

        # Mainbar Material
        tk.Label(self.dropdown_frame, text="Mainbar Material:", font=('Arial', 9, 'bold'),
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.mainbar_var = tk.StringVar(value=get_default(node_data.get('material_mainbar', ''), rebar_list))
        mainbar_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.mainbar_var,
                                  values=rebar_list, state='readonly', width=25)
        mainbar_menu.pack(fill='x', pady=(0, 8))

        # Tiebar Material
        tk.Label(self.dropdown_frame, text="Tiebar Material:", font=('Arial', 9, 'bold'),
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.tiebar_var = tk.StringVar(value=get_default(node_data.get('material_tiebar', ''), rebar_list))
        tiebar_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.tiebar_var,
                                  values=rebar_list, state='readonly', width=25)
        tiebar_menu.pack(fill='x', pady=(0, 8))

        # Anchor Bolt Size
        tk.Label(self.dropdown_frame, text="Anchor Bolt Size:", font=('Arial', 9, 'bold'),
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.bolt_var = tk.StringVar(value=get_default(node_data.get('bolt_size', ''), bolt_list))
        bolt_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.bolt_var,
                                values=bolt_list, state='readonly', width=25)
        bolt_menu.pack(fill='x', pady=(0, 8))

        # Hinge/Fixed Type
        tk.Label(self.dropdown_frame, text="Hinge/Fixed Type:", font=('Arial', 9, 'bold'),
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        hinge_list = ['Hinge Type', 'Fixed Type']
        self.hinge_fixed_var = tk.StringVar(value=get_default(node_data.get('hinge_fixed_type', ''), hinge_list))
        hinge_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.hinge_fixed_var,
                                  values=hinge_list, state='readonly', width=25)
        hinge_menu.pack(fill='x', pady=(0, 8))

        # Detail Type
        tk.Label(self.dropdown_frame, text="Detail Type:", font=('Arial', 9, 'bold'),
                bg='#f8f9fa', anchor='w').pack(fill='x', pady=(0, 2))
        self.detail_var = tk.StringVar(value=get_default(node_data.get('detail_type', ''), detail_list))
        detail_menu = ttk.Combobox(self.dropdown_frame, textvariable=self.detail_var,
                                   values=detail_list, state='readonly', width=25)
        detail_menu.pack(fill='x', pady=(0, 8))

        # Block mouse wheel on all Comboboxes to prevent accidental value changes while scrolling
        def _block_mousewheel(event):
            return "break"
        for combo in [anchor_menu, baseplate_menu, concrete_menu, mainbar_menu,
                      tiebar_menu, bolt_menu, hinge_menu, detail_menu]:
            combo.bind("<MouseWheel>", _block_mousewheel)

        # Enable buttons
        self.apply_btn.config(state='normal')
        self.edit_detail_btn.config(state='normal')  # THÊM DÒNG NÀY
        self.copy_multi_btn.config(state='normal')

        # Highlight selected node in plot
        self.highlight_selected_node()

    # ------------------------------------------------------------------
    # Material lookup helpers
    # ------------------------------------------------------------------

    def get_material_list(self, material_type):
        """Get material list from Material Strength tree"""
        tree = self.material_trees.get(material_type)
        if not tree:
            return ['']

        materials = ['']  # Empty option
        for item in tree.get_children():
            values = tree.item(item, 'values')
            if values and values[0]:
                materials.append(values[0])

        return materials

    def get_bolt_size_list(self):
        """Get bolt sizes from Anchor Bolt Table"""
        tree = self.material_trees.get('Anchor Bolt Table')
        if not tree:
            return ['']

        bolts = ['']
        for item in tree.get_children():
            values = tree.item(item, 'values')
            if values and values[0]:  # db column
                bolts.append(f"M{values[0]}")

        return bolts

    def get_detail_type_list(self, section):
            """Get detail types from Hinge Type that match section"""
            tree = self.material_trees.get('Hinge Type')
            if not tree:
                return ['']

            details = ['']
            for item in tree.get_children():
                values = tree.item(item, 'values')
                if values and values[0]:
                    if section and section.upper() in values[0].upper():
                        detail_type = values[1] if len(values) > 1 and values[1] else 'Standard'
                        detail_str = f"{values[0]} - Type {detail_type}"
                        if detail_str not in details:
                            details.append(detail_str)

            if len(details) == 1:
                for item in tree.get_children():
                    values = tree.item(item, 'values')
                    if values and values[0]:
                        detail_type = values[1] if len(values) > 1 and values[1] else 'Standard'
                        details.append(f"{values[0]} - Type {detail_type}")

            return details

    # ------------------------------------------------------------------
    # Apply / Copy settings
    # ------------------------------------------------------------------

    def apply_node_settings(self):
            """Apply settings to selected node and create/update Hinge Type.xlsx"""
            if not self.selected_node:
                return

            node_data = self.base_plate_nodes[self.selected_node]

            # Validate required fields
            if not node_data.get('section'):
                messagebox.showwarning("Warning", "Section is required!")
                return

            # Update node data
            node_data['material_anchor_bolt'] = self.anchor_bolt_var.get() if self.anchor_bolt_var.get() else None
            node_data['material_base_plate'] = self.base_plate_var.get() if self.base_plate_var.get() else None
            node_data['material_concrete'] = self.concrete_var.get() if self.concrete_var.get() else None
            node_data['material_mainbar'] = self.mainbar_var.get() if self.mainbar_var.get() else None
            node_data['material_tiebar'] = self.tiebar_var.get() if self.tiebar_var.get() else None
            node_data['bolt_size'] = self.bolt_var.get() if self.bolt_var.get() else None
            node_data['hinge_fixed_type'] = self.hinge_fixed_var.get() if self.hinge_fixed_var.get() else None
            node_data['detail_type'] = self.detail_var.get() if self.detail_var.get() else None

            # Reset design status
            node_data['design_status'] = 'Not Checked'

            # ✅ THÊM: Clear tất cả edited_* khi Apply lại từ dropdown
            for key in ['edited_section','edited_pier_detail', 'edited_base_plate_detail',
                        'edited_material', 'edited_anchor_bolt',
                        'edited_main_bar', 'edited_other_coeff']:
                node_data.pop(key, None)

            # Check if Fixed Type is selected (not yet supported)
            if node_data.get('hinge_fixed_type') == 'Fixed Type':
                messagebox.showwarning("⚠️ Not Supported",
                    "Fixed Type is still under development.\n\n"
                    "Please select 'Hinge Type' instead.")
                return

            # Create/Update Hinge Type.xlsx
            try:
                filepath = self.create_or_update_hinge_fixed_xlsx(self.selected_node, node_data)

                # Refresh plot
                self.plot_base_plate_plan()
                self.highlight_selected_node()

                self.status_label.config(text=f"● Settings applied to {self.selected_node}, sheet created in {os.path.basename(filepath)}", fg='#90ee90')

                messagebox.showinfo("✅ Success",
                    f"Settings applied to node {self.selected_node}\n\n"
                    #f"Sheet created/updated:\n{filepath}\n\n"
                    f"Sheet name: Node {self.selected_node}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to create sheet:\n{str(e)}")
                print(f"Error details: {e}")

    def copy_to_multiple(self):
        """Copy current settings to multiple nodes"""
        if not self.selected_node:
            return

        # Get source data
        source_data = self.base_plate_nodes[self.selected_node]
        source_section = source_data.get('section', '')

        # Split nodes into same section vs different section
        same_section_nodes = []
        diff_section_nodes = []
        for node_name in sorted(self.base_plate_nodes.keys()):
            if node_name != self.selected_node:
                node_section = self.base_plate_nodes[node_name].get('section', '')
                if node_section and source_section and node_section.upper() == source_section.upper():
                    same_section_nodes.append(node_name)
                else:
                    diff_section_nodes.append(node_name)

        # Create selection dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Copy to Multiple Nodes")
        dialog.geometry("950x600")
        dialog.transient(self.root)
        dialog.grab_set()

        # Header
        header_frame = tk.Frame(dialog, bg='#f0f0f0')
        header_frame.pack(fill='x', padx=15, pady=(15, 5))
        tk.Label(header_frame, text=f"Copy settings from: {self.selected_node}",
                font=('Arial', 11, 'bold'), bg='#f0f0f0', fg='#1a472a').pack(side='left')
        tk.Label(header_frame, text=f"Section: {source_section or 'N/A'}",
                font=('Arial', 10, 'italic'), bg='#f0f0f0', fg='#666').pack(side='left', padx=(20, 0))

        # Two-column container
        columns_frame = tk.Frame(dialog, bg='white')
        columns_frame.pack(fill='both', expand=True, padx=15, pady=10)
        columns_frame.columnconfigure(0, weight=1)
        columns_frame.columnconfigure(1, weight=1)

        node_vars = {}

        def create_column(parent, col_idx, title, subtitle, nodes, title_bg, title_fg):
            """Create a scrollable column with checkboxes"""
            col_frame = tk.Frame(parent, bg='white', bd=1, relief='groove')
            col_frame.grid(row=0, column=col_idx, sticky='nsew', padx=(0 if col_idx == 0 else 5, 0))

            # Column header
            header = tk.Frame(col_frame, bg=title_bg)
            header.pack(fill='x')
            tk.Label(header, text=title, font=('Arial', 10, 'bold'),
                    bg=title_bg, fg=title_fg).pack(side='left', padx=10, pady=5)
            tk.Label(header, text=f"({len(nodes)})", font=('Arial', 9),
                    bg=title_bg, fg=title_fg).pack(side='left')

            # Subtitle
            tk.Label(col_frame, text=subtitle, font=('Arial', 8, 'italic'),
                    bg='white', fg='#999').pack(anchor='w', padx=10, pady=(3, 0))

            # Select All / None buttons
            btn_row = tk.Frame(col_frame, bg='white')
            btn_row.pack(fill='x', padx=5, pady=3)

            def select_all():
                for nn in nodes:
                    if nn in node_vars:
                        node_vars[nn].set(True)
            def select_none():
                for nn in nodes:
                    if nn in node_vars:
                        node_vars[nn].set(False)

            tk.Button(btn_row, text="✓ All", command=select_all, bg='#4CAF50', fg='white',
                     font=('Arial', 8, 'bold'), cursor='hand2', width=6).pack(side='left', padx=2)
            tk.Button(btn_row, text="✗ None", command=select_none, bg='#FF9800', fg='white',
                     font=('Arial', 8, 'bold'), cursor='hand2', width=6).pack(side='left', padx=2)

            # Scrollable checkbox area
            scroll_container = tk.Frame(col_frame, bg='white')
            scroll_container.pack(fill='both', expand=True, padx=2, pady=2)

            canvas = tk.Canvas(scroll_container, bg='white', highlightthickness=0)
            scrollbar = ttk.Scrollbar(scroll_container, orient='vertical', command=canvas.yview)
            scrollable = tk.Frame(canvas, bg='white')

            scrollable.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
            canvas.create_window((0, 0), window=scrollable, anchor='nw')
            canvas.configure(yscrollcommand=scrollbar.set)

            def _on_mousewheel(event):
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            canvas.bind("<MouseWheel>", _on_mousewheel)

            canvas.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')

            # Add checkboxes
            for nn in nodes:
                var = tk.BooleanVar(value=False)
                node_vars[nn] = var
                cb = tk.Checkbutton(scrollable, text=f" {nn}", variable=var,
                                   font=('Arial', 10), bg='white', anchor='w')
                cb.pack(fill='x', padx=5, pady=2)

            if not nodes:
                tk.Label(scrollable, text="(No nodes)", font=('Arial', 9, 'italic'),
                        bg='white', fg='#ccc').pack(pady=20)

        # Left column: same section
        create_column(columns_frame, 0,
                     f"✅ Same Section", f"{source_section}",
                     same_section_nodes, '#E8F5E9', '#2E7D32')

        # Right column: different section
        create_column(columns_frame, 1,
                     "⚠️ Different Section", "Other sections",
                     diff_section_nodes, '#FFF3E0', '#E65100')

        def apply_to_selected():
            count = 0
            failed_nodes = []

            for node_name, var in node_vars.items():
                if var.get():
                    try:
                        # Update node data (basic dropdowns/materials)
                        self.base_plate_nodes[node_name]['material_anchor_bolt'] = source_data.get('material_anchor_bolt')
                        self.base_plate_nodes[node_name]['material_base_plate'] = source_data.get('material_base_plate')
                        self.base_plate_nodes[node_name]['material_concrete'] = source_data.get('material_concrete')
                        self.base_plate_nodes[node_name]['material_mainbar'] = source_data.get('material_mainbar')
                        self.base_plate_nodes[node_name]['material_tiebar'] = source_data.get('material_tiebar')
                        self.base_plate_nodes[node_name]['bolt_size'] = source_data.get('bolt_size')
                        self.base_plate_nodes[node_name]['hinge_fixed_type'] = source_data.get('hinge_fixed_type')
                        self.base_plate_nodes[node_name]['detail_type'] = source_data.get('detail_type')
                        self.base_plate_nodes[node_name]['design_status'] = 'Not Checked'

                        # Copy edited values if source node has them (from Edit Node Detail)
                        if 'edited_pier_detail' in source_data:
                            self.base_plate_nodes[node_name]['edited_pier_detail'] = source_data.get('edited_pier_detail')
                        if 'edited_base_plate_detail' in source_data:
                            self.base_plate_nodes[node_name]['edited_base_plate_detail'] = source_data.get('edited_base_plate_detail')
                        if 'edited_material' in source_data:
                            self.base_plate_nodes[node_name]['edited_material'] = source_data.get('edited_material')
                        if 'edited_anchor_bolt' in source_data:
                            self.base_plate_nodes[node_name]['edited_anchor_bolt'] = source_data.get('edited_anchor_bolt')
                        if 'edited_main_bar' in source_data:
                            self.base_plate_nodes[node_name]['edited_main_bar'] = source_data.get('edited_main_bar')
                        if 'edited_other_coeff' in source_data:
                            self.base_plate_nodes[node_name]['edited_other_coeff'] = source_data.get('edited_other_coeff')

                        # Create/update Hinge Type.xlsx sheet for this node
                        self.create_or_update_hinge_fixed_xlsx(node_name, self.base_plate_nodes[node_name])
                        count += 1
                    except Exception as e:
                        failed_nodes.append(f"{node_name}: {str(e)}")
                        print(f"Error applying to {node_name}: {e}")

            if count > 0:
                self.plot_base_plate_plan()
                self.highlight_selected_node()

                msg = f"Settings copied to {count} node(s)"
                if failed_nodes:
                    msg += f"\n\nFailed ({len(failed_nodes)}):\n" + "\n".join(failed_nodes[:5])
                    if len(failed_nodes) > 5:
                        msg += f"\n... and {len(failed_nodes) - 5} more"
                    messagebox.showwarning("Partial Success", msg)
                else:
                    messagebox.showinfo("Success", msg)
            else:
                messagebox.showwarning("No Selection", "Please select at least one node")

            dialog.destroy()

        # Buttons
        btn_frame = tk.Frame(dialog, bg='white')
        btn_frame.pack(side='bottom', fill='x', pady=15, padx=15)

        tk.Button(btn_frame, text="✓ Apply", command=apply_to_selected,
                 bg='#4CAF50', fg='white', font=('Arial', 10, 'bold'),
                 width=15, cursor='hand2').pack(side='left', padx=5)
        tk.Button(btn_frame, text="✗ Cancel", command=dialog.destroy,
                 bg='#f44336', fg='white', font=('Arial', 10, 'bold'),
                 width=15, cursor='hand2').pack(side='left', padx=5)

    # ------------------------------------------------------------------
    # Control Load Table dialog
    # ------------------------------------------------------------------

    def show_control_load_table(self, node_name):
        """Display Control Load Table (N1:S6) from node's Hinge Type sheet in a dialog"""
        try:
            if not self.bpl_folder:
                return

            # Open Hinge Type.xlsx
            filepath = os.path.join(self.bpl_folder, "Hinge Type.xlsx")
            if not os.path.exists(filepath):
                return

            wb = openpyxl.load_workbook(filepath)
            sheet_name = f"Node {node_name}"

            if sheet_name not in wb.sheetnames:
                wb.close()
                return

            ws = wb[sheet_name]

            # Read Control Load Table (N1:S6)
            table_data = []
            headers = ['Control Load', 'L/C', 'F1', 'F2', 'F3', 'Ratio']

            # Row 1: Headers
            table_data.append(headers)

            # Rows 2-6: Data
            control_checks = ['Compression', 'Tensile', 'Anchor Bolt', 'X-Shear check', 'Y-Shear check']
            for row in range(2, 7):
                row_data = []
                row_data.append(ws[f'N{row}'].value or '')
                row_data.append(ws[f'O{row}'].value or '')

                # Format forces and ratio with 2 decimal places
                try:
                    f1 = ws[f'P{row}'].value
                    f1_str = f"{float(f1):.2f}" if f1 is not None else ''
                except:
                    f1_str = ws[f'P{row}'].value or ''
                row_data.append(f1_str)

                try:
                    f2 = ws[f'Q{row}'].value
                    f2_str = f"{float(f2):.2f}" if f2 is not None else ''
                except:
                    f2_str = ws[f'Q{row}'].value or ''
                row_data.append(f2_str)

                try:
                    f3 = ws[f'R{row}'].value
                    f3_str = f"{float(f3):.2f}" if f3 is not None else ''
                except:
                    f3_str = ws[f'R{row}'].value or ''
                row_data.append(f3_str)

                # Ratio - format with 2 decimal places
                try:
                    ratio = ws[f'S{row}'].value
                    ratio_str = f"{float(ratio):.2f}" if ratio is not None else ''
                except:
                    ratio_str = ws[f'S{row}'].value or ''
                row_data.append(ratio_str)

                table_data.append(row_data)

            wb.close()

            # Create dialog to display table
            dialog = tk.Toplevel(self.root)
            dialog.title(f"Control Load Table - Node {node_name}")
            dialog.geometry("700x300")
            dialog.transient(self.root)

            # Add title
            tk.Label(dialog, text=f"Control Load Table - Node {node_name}",
                    font=('Arial', 12, 'bold'), fg='#1a472a').pack(pady=10)

            # Create treeview for table
            tree_frame = tk.Frame(dialog)
            tree_frame.pack(fill='both', expand=True, padx=10, pady=10)

            # Columns
            columns = ('Control Load', 'L/C', 'F1', 'F2', 'F3', 'Ratio')
            tree = ttk.Treeview(tree_frame, columns=columns, height=6, show='headings')

            # Define column widths and headings
            widths = [120, 100, 80, 80, 80, 80]
            for i, (col, width) in enumerate(zip(columns, widths)):
                tree.column(col, width=width, anchor='center')
                tree.heading(col, text=col)

            # Add data rows
            for row in table_data[1:]:  # Skip header
                tree.insert('', 'end', values=row)

            # Add scrollbar
            scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
            tree.configure(yscroll=scrollbar.set)

            tree.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')

            # Add close button
            tk.Button(dialog, text="Close", command=dialog.destroy,
                     bg='#1a472a', fg='white', font=('Arial', 10, 'bold'),
                     padx=20, pady=8).pack(pady=10)

        except Exception as e:
            print(f"Error displaying control load table: {e}")
