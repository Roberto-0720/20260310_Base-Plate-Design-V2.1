"""
ui/dialogs_mixin.py
====================
Standalone dialog windows: User Guide, About, and the Edit Node Detail
dialog (which reads and writes back to Hinge Type.xlsx / Fixed Type.xlsx).

App state used:
  self.root, self.selected_node, self.base_plate_nodes,
  self.bpl_folder, self.status_label
"""

import os
import tkinter as tk
from tkinter import ttk, messagebox

import openpyxl

try:
    from PIL import Image, ImageTk
    _PIL_AVAILABLE = True
except ImportError:
    _PIL_AVAILABLE = False


class DialogsMixin:

    def show_guide(self):
        """Show user guide"""
        guide_window = tk.Toplevel(self.root)
        guide_window.title("Instructions & Workflow")
        guide_window.geometry("550x650")
        guide_window.configure(bg='white')
        guide_window.transient(self.root)

        # Title
        title_frame = tk.Frame(guide_window, bg='white')
        title_frame.pack(fill='x', pady=(15, 5))
        tk.Label(title_frame, text="📋 HOW TO USE THIS TOOL",
                font=('Arial', 14, 'bold'), bg='white').pack()

        # Scrollable text area
        text_frame = tk.Frame(guide_window, bg='white')
        text_frame.pack(fill='both', expand=True, padx=20, pady=10)

        text_widget = tk.Text(text_frame, wrap='word', font=('Consolas', 10),
                             bg='#fafafa', fg='#333', relief='groove', bd=1,
                             padx=15, pady=15, spacing2=3)
        scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        text_widget.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        guide_text = """🏗  BASE PLATE DESIGN TOOL

📋 PREPARATION STEPS:
  • Have your SAP2000 model file (.sdb) ready
  • Unit system: kN, m

⚡ WORKFLOW:
  1  □ "Get Model" to connect to SAP2000
  2  □ Select frame elements in SAP2000 model
  3  □ "Load (Auto)" to extract coordinates & reactions
  4  □ Select load cases/combinations for analysis
  5  □ Define Material Properties
  6  □ Assign Base Plate details for each node
  7  □ "Run Analysis" to calculate all ratios
  8  □ "Export to Excel" to generate Calculation Report

📤 EXPORT & RESULTS:
  • Calculation Report.xlsx - Detailed calculation sheets
  • Summary sheet with all node ratios
  • Color-coded pass/fail indication

⚠ IMPORTANT NOTES:
  • Select ELEMENTS only for Load (Auto), DON'T select any nodes
  • "Fixed Type" is still under development
  • "STAAD Pro" software option is under development

💡 TIPS:
  • Add/Modify Material to suit your project
  • Click Reload Data to update Base Plate Plan View
  • Click a Node on Base Plate Plan View to assign Parameters
  • Use "Copy to Multiple Nodes" for same-section nodes

📞 SUPPORT: Contact Roberto for technical assistance"""

        text_widget.insert('1.0', guide_text)
        text_widget.config(state='disabled')

    def show_about(self):
        """Show about information"""
        about_window = tk.Toplevel(self.root)
        about_window.title("About")
        about_window.geometry("450x350")
        about_window.configure(bg='white')
        about_window.transient(self.root)
        about_window.resizable(False, False)

        # Icon
        tk.Label(about_window, text="ℹ️", font=('Arial', 28), bg='white').pack(pady=(20, 5))

        # Tool name
        tk.Label(about_window, text="Base Plate Design Tool",
                font=('Arial', 13, 'bold'), bg='white', fg='#1a472a').pack()

        # Version info
        info_frame = tk.Frame(about_window, bg='white')
        info_frame.pack(pady=10)
        tk.Label(info_frame, text="Version: 2.0\nAnalysis Type: Base Plate Design",
                font=('Arial', 10), bg='white', fg='#555', justify='center').pack()

        # Features
        features_frame = tk.Frame(about_window, bg='white')
        features_frame.pack(pady=5)
        tk.Label(features_frame, text="Features:\n"
                "- Load combination analysis\n"
                "- Compression & Tensile checks\n"
                "- Shear & Interaction checks\n"
                "- Visual capacity ratio diagrams\n"
                "- Color-coded failure indication",
                font=('Arial', 9), bg='white', fg='#666', justify='left').pack(anchor='w', padx=40)

        # Credits
        tk.Label(about_window, text="Civil & Structural Engineering Analysis\nRoberto Inhouse",
                font=('Arial', 9, 'italic'), bg='white', fg='#888').pack(pady=(10, 2))
        tk.Label(about_window, text="© 2026 All rights reserved",
                font=('Arial', 9), bg='white', fg='#aaa').pack()

        # OK button
        tk.Button(about_window, text="OK", command=about_window.destroy,
                 width=10, font=('Arial', 10)).pack(pady=(10, 15))

    def edit_node_detail(self):
        """Open dialog to edit node detail from Hinge Type.xlsx"""
        if not self.selected_node:
            return

        node_data = self.base_plate_nodes[self.selected_node]

        # Check if Hinge Type.xlsx exists
        hinge_fixed = node_data.get('hinge_fixed_type', 'Hinge Type')
        if 'Fixed' in hinge_fixed:
            filename = 'Fixed Type.xlsx'
        else:
            filename = 'Hinge Type.xlsx'

        filepath = os.path.join(self.bpl_folder, filename)

        if not os.path.exists(filepath):
            messagebox.showwarning("Warning",
                f"{filename} not found!\n\n"
                f"Please click 'Apply to Node' first to create the file.")
            return

        sheet_name = f"Node {self.selected_node}"

        # Check if sheet exists
        try:
            wb = openpyxl.load_workbook(filepath)
            if sheet_name not in wb.sheetnames:
                wb.close()
                messagebox.showwarning("Warning",
                    f"Sheet '{sheet_name}' not found!\n\n"
                    f"Please click 'Apply to Node' first.")
                return

            ws = wb[sheet_name]

        except Exception as e:
            messagebox.showerror("Error", f"Failed to open {filename}:\n{str(e)}")
            return

        # Xác định Type để chọn hình minh họa
        section = node_data.get('section', '')
        hinge_data = self.get_hinge_type_row_data(section)
        plate_type = hinge_data.get('Type', '-') if hinge_data else '-'
        is_type_B = str(plate_type).strip().upper() == 'B'

        # Create dialog — mở rộng để chứa panel ảnh
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Edit Node Detail - {self.selected_node}")
        dialog.geometry("1420x780")
        dialog.transient(self.root)
        dialog.grab_set()

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f"+{x}+{y}")

        # Title
        title_label = tk.Label(
            dialog,
            text=f"📝 Edit Detail for Node {self.selected_node}",
            font=('Arial', 14, 'bold'),
            bg='#1a472a',
            fg='white',
            pady=15
        )
        title_label.pack(fill='x')

        # Outer frame chia 2 cột: data (trái) + ảnh (phải)
        outer_frame = tk.Frame(dialog, bg='white')
        outer_frame.pack(fill='both', expand=True, padx=10, pady=(10, 0))

        # ---- CỘT TRÁI: data tables với scrollbar ----
        left_panel = tk.Frame(outer_frame, bg='white')
        left_panel.pack(side='left', fill='both', expand=True, padx=(10, 5))

        main_container = tk.Frame(left_panel, bg='white')
        main_container.pack(fill='both', expand=True)

        canvas = tk.Canvas(main_container, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')

        scrollable_frame.bind(
            '<Configure>',
            lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # ---- CỘT PHẢI: hình minh họa ----
        right_panel = tk.Frame(outer_frame, bg='white', width=400)
        right_panel.pack(side='right', fill='y', padx=(5, 10))
        right_panel.pack_propagate(False)

        # Load và hiển thị ảnh
        resources_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'resources')
        img_filename = 'baseplate_type_B.png' if is_type_B else 'baseplate_type_minus.png'
        img_path = os.path.normpath(os.path.join(resources_dir, img_filename))

        img_label = tk.Label(right_panel, bg='white')
        img_label.pack(fill='both', expand=True, padx=5, pady=5)

        if _PIL_AVAILABLE and os.path.exists(img_path):
            try:
                pil_img = Image.open(img_path)
                # Resize vừa panel (350 x 480)
                pil_img.thumbnail((350, 500), Image.LANCZOS)
                tk_img = ImageTk.PhotoImage(pil_img)
                img_label.config(image=tk_img)
                img_label.image = tk_img  # giữ reference
            except Exception as e:
                print(f"Warning: Could not load image: {e}")
                img_label.config(
                    text=f"[Image load error]\n{os.path.basename(img_path)}",
                    font=('Arial', 9), fg='#aaa'
                )
        elif not _PIL_AVAILABLE:
            img_label.config(
                text="Install Pillow to show\nillustration image:\n\npip install Pillow",
                font=('Arial', 9, 'italic'), fg='#888', justify='center'
            )
        else:
            img_label.config(
                text=f"Image not found:\n{img_filename}",
                font=('Arial', 9, 'italic'), fg='#aaa', justify='center'
            )

        # Store entry widgets
        entries = {}

        # Helper to create editable grid
        def create_editable_grid(parent, title, header_row, value_row, columns, fill_width=True):
            """Create a grid with headers and editable values"""
            frame = tk.LabelFrame(
                parent,
                text=f" {title} ",
                font=('Arial', 11, 'bold'),
                bg='white',
                fg='#1a472a',
                relief='groove',
                bd=2
            )
            if fill_width:
                frame.pack(fill='x', pady=(0, 15))
            else:
                frame.pack(anchor='w', pady=(0, 15))

            grid_frame = tk.Frame(frame, bg='white')
            grid_frame.pack(padx=15, pady=10)

            for col_idx, col_letter in enumerate(columns):
                # Header
                header_cell = ws[f'{col_letter}{header_row}']
                header_text = header_cell.value if header_cell.value else ''

                tk.Label(
                    grid_frame,
                    text=str(header_text),
                    font=('Arial', 9, 'bold'),
                    bg='#E8E8E8',
                    relief='solid',
                    bd=1,
                    width=10,
                    height=2
                ).grid(row=0, column=col_idx, padx=2, pady=2)

                # Value entry
                value_cell = ws[f'{col_letter}{value_row}']
                value_text = value_cell.value if value_cell.value is not None else ''

                entry = tk.Entry(
                    grid_frame,
                    font=('Arial', 10),
                    justify='center',
                    width=10,
                    relief='solid',
                    bd=1
                )
                entry.insert(0, str(value_text))
                entry.grid(row=1, column=col_idx, padx=2, pady=2)

                # Store for saving
                entries[f'{col_letter}{value_row}'] = entry

        # SECTION 1: Base Plate (rows 1-3)
        section1_container = tk.Frame(scrollable_frame, bg='white')
        section1_container.pack(fill='x', pady=(10, 0))

        left1 = tk.Frame(section1_container, bg='white')
        left1.pack(side='left', fill='both', expand=True, padx=(0, 10))

        section1_title = ws['A1'].value if ws['A1'].value else 'Section'
        create_editable_grid(left1, section1_title, 2, 3, ['A', 'B', 'C', 'D'])

        right1 = tk.Frame(section1_container, bg='white')
        right1.pack(side='left', fill='both', expand=True, padx=(10, 0))
        create_editable_grid(right1, 'Base Plate Detail', 2, 3, ['F', 'G', 'H', 'I', 'J', 'K', 'L'])

        # SECTION 2: Material (rows 5-7)
        section2_container = tk.Frame(scrollable_frame, bg='white')
        section2_container.pack(fill='x', pady=0)

        left2 = tk.Frame(section2_container, bg='white')
        left2.pack(side='left', fill='both', expand=True, padx=(0, 10))
        create_editable_grid(left2, 'Material', 6, 7, ['A', 'B', 'C', 'D'])

        right2 = tk.Frame(section2_container, bg='white')
        right2.pack(side='left', fill='both', expand=True, padx=(10, 0))
        create_editable_grid(right2, 'Anchor Bolt', 6, 7, ['F', 'G', 'H', 'I', 'J', 'K', 'L'])

        # SECTION 3: Pier Detail (rows 9-11)
        section3_container = tk.Frame(scrollable_frame, bg='white')
        section3_container.pack(fill='x', pady=0)

        left3 = tk.Frame(section3_container, bg='white')
        left3.pack(side='left', fill='both', expand=True, padx=(0, 10))
        create_editable_grid(left3, 'Pier Detail', 10, 11, ['A', 'B', 'C', 'D'])

        right3 = tk.Frame(section3_container, bg='white')
        right3.pack(side='left', fill='both', expand=True, padx=(10, 0))
        create_editable_grid(right3, 'Main Bar & Tie Bar', 10, 11, ['F', 'G', 'H', 'I', 'J', 'K', 'L'])

        # SECTION 4: Other Coefficient (rows 13-15)
        create_editable_grid(scrollable_frame, 'Other Coefficient', 14, 15, ['F', 'G', 'H', 'I'], fill_width=False)

        wb.close()

        # Button frame
        btn_frame = tk.Frame(dialog, bg='#f5f5f5')
        btn_frame.pack(side='bottom', fill='x', pady=12)

        def save_changes():
            try:
                # Step 1: Update node_data['edited_*']
                node_data = self.base_plate_nodes[self.selected_node]

                node_data['edited_section'] = {
                    'd':  entries['A3'].get().strip() if 'A3' in entries else '',
                    'bf': entries['B3'].get().strip() if 'B3' in entries else '',
                    'tw': entries['C3'].get().strip() if 'C3' in entries else '',
                    'tf': entries['D3'].get().strip() if 'D3' in entries else '',
                }
                node_data['edited_base_plate_detail'] = {
                    'N':  entries['F3'].get().strip() if 'F3' in entries else '',
                    'B':  entries['G3'].get().strip() if 'G3' in entries else '',
                    'P1': entries['H3'].get().strip() if 'H3' in entries else '',
                    'C':  entries['I3'].get().strip() if 'I3' in entries else '',
                    'A':  entries['J3'].get().strip() if 'J3' in entries else '',
                    'E':  entries['K3'].get().strip() if 'K3' in entries else '',
                    'F':  entries['L3'].get().strip() if 'L3' in entries else '',
                }
                node_data['edited_material'] = {
                    'fy_steel':   entries['A7'].get().strip() if 'A7' in entries else '',
                    'fc':         entries['B7'].get().strip() if 'B7' in entries else '',
                    'fy_mainbar': entries['C7'].get().strip() if 'C7' in entries else '',
                    'fy_tiebar':  entries['D7'].get().strip() if 'D7' in entries else '',
                }
                node_data['edited_anchor_bolt'] = {
                    'nb':   entries['F7'].get().strip() if 'F7' in entries else '',
                    'db':   entries['G7'].get().strip() if 'G7' in entries else '',
                    'Ase':  entries['H7'].get().strip() if 'H7' in entries else '',
                    'futa': entries['I7'].get().strip() if 'I7' in entries else '',
                    'A1':   entries['J7'].get().strip() if 'J7' in entries else '',
                    'Proj': entries['K7'].get().strip() if 'K7' in entries else '',
                    'heff': entries['L7'].get().strip() if 'L7' in entries else '',
                }
                node_data['edited_pier_detail'] = {
                    'Np': entries['A11'].get().strip() if 'A11' in entries else '',
                    'Bp': entries['B11'].get().strip() if 'B11' in entries else '',
                    'TG': entries['C11'].get().strip() if 'C11' in entries else '',
                    'c':  entries['D11'].get().strip() if 'D11' in entries else '',
                }
                node_data['edited_main_bar'] = {
                    'Qty':  entries['F11'].get().strip() if 'F11' in entries else '',
                    'Size': entries['G11'].get().strip() if 'G11' in entries else '',
                }
                node_data['edited_other_coeff'] = {
                    'friction_mu': entries['F15'].get().strip() if 'F15' in entries else '',
                    'psi_c_p':     entries['G15'].get().strip() if 'G15' in entries else '',
                    'psi_e':       entries['H15'].get().strip() if 'H15' in entries else '',
                    'lambda':      entries['I15'].get().strip() if 'I15' in entries else '',
                }

                # Step 2: Regenerate full sheet using correct explicit calculation
                result_path = self.create_or_update_hinge_fixed_xlsx(
                    self.selected_node, node_data
                )

                messagebox.showinfo("✅ Success",
                    f"Changes saved & recalculated!\n\n"
                    f"File: {os.path.basename(result_path)}\n"
                    f"Sheet: {sheet_name}")

                self.status_label.config(
                    text=f"● Node {self.selected_node} detail updated & recalculated",
                    fg='#90ee90'
                )
                dialog.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Failed to save:\n{str(e)}")
                import traceback
                traceback.print_exc()

        tk.Button(
            btn_frame,
            text="💾 Save Changes",
            command=save_changes,
            bg='#4CAF50',
            fg='white',
            font=('Arial', 11, 'bold'),
            cursor='hand2',
            relief='flat',
            padx=30,
            pady=10
        ).pack(side='left', padx=(30, 10))

        tk.Button(
            btn_frame,
            text="❌ Cancel",
            command=dialog.destroy,
            bg='#f44336',
            fg='white',
            font=('Arial', 11, 'bold'),
            cursor='hand2',
            relief='flat',
            padx=30,
            pady=10
        ).pack(side='left', padx=10)
