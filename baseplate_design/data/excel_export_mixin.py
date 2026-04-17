"""
data/excel_export_mixin.py
===========================
Exports per-node calculation reports to Excel using Template.xlsx as the
base and mapping calculated values from Hinge Type.xlsx into each page.

Public methods:
  export_to_excel()           — entry point; validates state, shows dialog
  _show_export_dialog()       — node-selection dialog
  _create_calculation_report()— copies template, fills pages, saves file
  _map_page1_data()           — Page 1 (rows 1-44): summary + ratios
  _map_page2_data()           — Page 2 (rows 46-85): bearing & yielding
  _map_page3_data()           — Page 3 (rows 87-134): tensile + AB tension
  _map_page4_data()           — Page 4 (rows 135-167): summaries + interaction
  _map_page5_data()           — Page 5 (rows 168-212): pier vertical rebars
  _map_page6_data()           — Page 6 (rows 213-247): pier horizontal rebars
  _create_summary_sheet()     — Summary sheet at index 0

App state used:
  self.base_plate_nodes, self.bpl_folder, self.status_label, self.root
"""

import os
import math
import shutil
import platform
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelExportMixin:

    def export_to_excel(self):
        """Export calculation reports for checked nodes to Calculation Report.xlsx"""
        # Check if any nodes have been checked
        checked_nodes = [name for name, data in self.base_plate_nodes.items()
                        if data['design_status'] in ['OK', 'NG']]

        if not checked_nodes:
            messagebox.showwarning("No Data",
                "Please run design check first!\n\n"
                "Steps:\n"
                "1. Select nodes in Base Plate Detail\n"
                "2. Click 'RUN DESIGN CHECK'\n"
                "3. Then export")
            return

        # Create selection dialog for nodes
        self._show_export_dialog(checked_nodes)

    def _show_export_dialog(self, checked_nodes):
        """Show dialog to select nodes for export"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Export to Excel")
        dialog.geometry("500x650")
        dialog.transient(self.root)
        dialog.grab_set()

        # Header
        header_frame = tk.Frame(dialog, bg='#f0f0f0')
        header_frame.pack(fill='x', padx=15, pady=(15, 10))
        tk.Label(header_frame, text="Select nodes to export:",
                font=('Arial', 12, 'bold'), bg='#f0f0f0', fg='#1a472a').pack(side='left')
        tk.Label(header_frame, text=f"({len(checked_nodes)} nodes available)",
                font=('Arial', 10), bg='#f0f0f0', fg='#666').pack(side='left', padx=(15, 0))

        # Scrollable frame for checkboxes
        scroll_container = tk.Frame(dialog, bg='white')
        scroll_container.pack(fill='both', expand=True, padx=15, pady=10)

        canvas = tk.Canvas(scroll_container, bg='white', highlightthickness=0, height=350)
        scrollbar = ttk.Scrollbar(scroll_container, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')

        scrollable_frame.bind('<Configure>',
                             lambda e: canvas.configure(scrollregion=canvas.bbox('all')))

        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        # Mouse wheel scroll
        def _on_mousewheel(event):
            if canvas.winfo_containing(event.x_root, event.y_root) == canvas:
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        canvas.bind("<MouseWheel>", _on_mousewheel)

        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        # Checkboxes
        node_vars = {}
        for node_name in sorted(checked_nodes):
            var = tk.BooleanVar(value=True)  # Default: all selected
            node_vars[node_name] = var

            # Node info
            node_data = self.base_plate_nodes[node_name]
            status_color = '#4CAF50' if node_data['design_status'] == 'OK' else '#f44336'
            status_text = f"[{node_data['design_status']}]"
            ratio_text = f"(R: {node_data.get('max_ratio', '?'):.2f})" if node_data.get('max_ratio') else ""

            cb_frame = tk.Frame(scrollable_frame, bg='white')
            cb_frame.pack(fill='x', padx=10, pady=4)

            cb = tk.Checkbutton(cb_frame, text=node_name, variable=var,
                               font=('Arial', 10), bg='white', anchor='w')
            cb.pack(side='left')

            status_label = tk.Label(cb_frame, text=status_text,
                                   font=('Arial', 9, 'bold'), bg='white', fg=status_color)
            status_label.pack(side='left', padx=(10, 5))

            ratio_label = tk.Label(cb_frame, text=ratio_text,
                                  font=('Arial', 9), bg='white', fg='#666')
            ratio_label.pack(side='left')

        # Buttons
        def select_all():
            for var in node_vars.values():
                var.set(True)

        def deselect_all():
            for var in node_vars.values():
                var.set(False)

        def export_selected():
            selected = [name for name, var in node_vars.items() if var.get()]

            if not selected:
                messagebox.showwarning("No Selection", "Please select at least one node")
                return

            dialog.destroy()
            self._create_calculation_report(selected)

        # Button frame
        btn_frame = tk.Frame(dialog, bg='white')
        btn_frame.pack(fill='x', pady=(10, 15), padx=15)

        # Top buttons
        select_frame = tk.Frame(btn_frame, bg='white')
        select_frame.pack(fill='x', pady=(0, 10))

        tk.Button(select_frame, text="✓ All", command=select_all,
                 bg='#2196F3', fg='white', font=('Arial', 9, 'bold'),
                 width=12, cursor='hand2').pack(side='left', padx=5)
        tk.Button(select_frame, text="✗ None", command=deselect_all,
                 bg='#FF9800', fg='white', font=('Arial', 9, 'bold'),
                 width=12, cursor='hand2').pack(side='left', padx=5)

        # Bottom buttons
        bottom_frame = tk.Frame(btn_frame, bg='white')
        bottom_frame.pack(fill='x')

        tk.Button(bottom_frame, text="✓ Export", command=export_selected,
                 bg='#4CAF50', fg='white', font=('Arial', 10, 'bold'),
                 width=15, cursor='hand2').pack(side='left', padx=5)
        tk.Button(bottom_frame, text="✗ Cancel", command=dialog.destroy,
                 bg='#f44336', fg='white', font=('Arial', 10, 'bold'),
                 width=15, cursor='hand2').pack(side='left', padx=5)

    def _create_calculation_report(self, selected_nodes):
        """Create Calculation Report.xlsx using Template.xlsx"""
        try:
            # 1. Find Template.xlsx in bpl_folder or script directory
            template_path = os.path.join(self.bpl_folder, 'Template.xlsx')
            if not os.path.exists(template_path):
                script_dir = os.path.dirname(os.path.abspath(__file__))
                template_src = os.path.join(script_dir, 'Template.xlsx')
                if os.path.exists(template_src):
                    shutil.copy2(template_src, template_path)
                else:
                    messagebox.showerror("Error",
                        "Template.xlsx not found!\n\n"
                        "Please place Template.xlsx in:\n"
                        f"1. {self.bpl_folder}\n"
                        f"2. {script_dir}")
                    return

            # 2. Copy Template.xlsx → Calculation Report.xlsx
            report_file = os.path.join(self.bpl_folder, 'Calculation Report.xlsx')
            shutil.copy2(template_path, report_file)

            # 3. Open workbooks
            report_wb = openpyxl.load_workbook(report_file)
            hinge_file = os.path.join(self.bpl_folder, 'Hinge Type.xlsx')
            if not os.path.exists(hinge_file):
                messagebox.showerror("Error", "Hinge Type.xlsx not found!")
                report_wb.close()
                return
            hinge_wb = openpyxl.load_workbook(hinge_file)

            # 4. Get template sheet (first sheet)
            template_ws = report_wb.worksheets[0]

            # 5. Create copies of template sheet for additional nodes
            for i in range(1, len(selected_nodes)):
                new_ws = report_wb.copy_worksheet(template_ws)

                # Copy drawing (images) from template sheet to new sheet
                try:
                   if hasattr(template_ws, '_drawing') and template_ws._drawing:
                       # Deep copy the drawing object
                       import copy
                       new_ws._drawing = copy.deepcopy(template_ws._drawing)
                       if new_ws._drawing:
                           new_ws._drawing.anchor = new_ws
                except Exception as e:
                   print(f"Warning: Could not copy images: {e}")

            # 6. Rename sheets to node names
            for i, node_name in enumerate(selected_nodes):
                report_wb.worksheets[i].title = node_name

            # 7. Map data for each node from Hinge Type.xlsx
            for i, node_name in enumerate(selected_nodes):
                ws = report_wb.worksheets[i]
                hinge_sheet_name = f"Node {node_name}"

                if hinge_sheet_name not in hinge_wb.sheetnames:
                    print(f"Warning: Sheet '{hinge_sheet_name}' not found in Hinge Type.xlsx")
                    continue

                hs = hinge_wb[hinge_sheet_name]

                # Map Page 1 data (rows 1-44)
                self._map_page1_data(ws, hs)

                # Map Page 2 data (rows 46-85)
                self._map_page2_data(ws, hs)

                # Map Page 3 data (rows 87-134)
                self._map_page3_data(ws, hs)

                # Map Page 4 data (rows 135-167)
                self._map_page4_data(ws, hs)

                # Map Page 5 data (rows 168-212)
                self._map_page5_data(ws, hs)

                # Map Page 6 data (rows 213-247)
                self._map_page6_data(ws, hs)

                # Apply formatting: Page Break Preview + Hide Gridlines
                ws.sheet_view.view = 'pageBreakPreview'  # Enable Page Break Preview
                ws.sheet_view.showGridLines = False  # Hide gridlines

                print(f"Exported: Node {node_name}")

            hinge_wb.close()

            # 8. Create Summary sheet at position 0
            self._create_summary_sheet(report_wb, selected_nodes)

            # 9. Save
            report_wb.save(report_file)
            report_wb.close()

            # Show success message
            messagebox.showinfo("Export Success!",
                f"Calculation Report created!\n\n"
                f"File: Calculation Report.xlsx\n"
                f"Location: {self.bpl_folder}\n\n"
                f"Nodes exported: {len(selected_nodes)}\n"
                f"Total sheets: {len(selected_nodes) + 1} (Summary + {len(selected_nodes)} nodes)")

            # Open file
            try:
                if platform.system() == 'Windows':
                    os.startfile(report_file)
                elif platform.system() == 'Darwin':
                    subprocess.call(['open', report_file])
                else:
                    subprocess.call(['xdg-open', report_file])
            except:
                pass

            self.status_label.config(text=f"● Export complete: {len(selected_nodes)} nodes", fg='#90ee90')

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to create report:\n{str(e)}")
            print(f"Export error: {e}")
            import traceback
            traceback.print_exc()

    def _map_page1_data(self, ws, hs):
        """Map Page 1 data (rows 1-44) from Hinge Type sheet to Template"""

        def sv(value):
            """Safe value - convert to number if possible, return '' for None"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, 6)
            except:
                return value

        # ===== 1. Column Section (Row 6-10) =====
        ws['J6'] = hs['A1'].value or ''      # Section name e.g. "H340X250X9X14"
        ws['K7'] = sv(hs['A3'].value)         # d (depth)
        ws['K8'] = sv(hs['B3'].value)         # bf (width)
        ws['K9'] = sv(hs['C3'].value)         # tw (web)
        ws['K10'] = sv(hs['D3'].value)        # tf (flange)

        # ===== 2. Base Plate Detail (Row 12-19) =====
        n_val = sv(hs['F3'].value)
        b_val = sv(hs['G3'].value)
        p1_val = sv(hs['H3'].value)
        ws['J12'] = f"PL-{n_val}X{b_val}X{p1_val}" if all([n_val != '', b_val != '', p1_val != '']) else ''
        ws['K13'] = n_val                     # N (length)
        ws['K14'] = b_val                     # B (width)
        ws['K15'] = p1_val                    # P1 (thickness)
        ws['K16'] = sv(hs['I3'].value)        # C (bolt spacing)
        ws['K17'] = sv(hs['J3'].value)        # A (bolt spacing)
        ws['K18'] = sv(hs['L3'].value)        # F (edge distance)
        ws['K19'] = sv(hs['K3'].value)        # E (edge distance)

        # ===== 3. Anchor Bolt Detail (Row 21-28) =====
        nb = sv(hs['F7'].value)
        db = sv(hs['G7'].value)
        lo = sv(hs['J7'].value)
        ws['J21'] = f"{nb}-M{db}, Lo={lo} mm" if all([nb != '', db != '', lo != '']) else ''
        ws['K22'] = nb                        # nb
        ws['K23'] = db                        # db
        ws['K24'] = sv(hs['H7'].value)        # Ase
        ws['K25'] = sv(hs['I7'].value)        # futa
        ws['K26'] = lo                        # Lo (A1)
        ws['K27'] = sv(hs['K7'].value)        # proj
        ws['K28'] = sv(hs['L7'].value)        # hef

        # ===== 4. Material (Row 31-35) =====
        ws['K31'] = sv(hs['A7'].value)        # fy (steel)
        ws['K32'] = sv(hs['B7'].value)        # f'c
        ws['K33'] = sv(hs['F15'].value)       # μ (friction) from Other Coefficient
        ws['K34'] = sv(hs['C7'].value)        # fy (mainbar)
        ws['K35'] = sv(hs['D7'].value)        # fy (tiebar)

        # ===== 5. Pier Detail (Row 31-34 right side) =====
        ws['AB31'] = sv(hs['A11'].value)      # Np
        ws['AB32'] = sv(hs['B11'].value)      # Bp
        ws['AB33'] = sv(hs['C11'].value)      # TG
        ws['AB34'] = sv(hs['D11'].value)      # Cover c

        # ===== 6. Control Ratios - Base Plate Design =====
        rat1 = sv(hs['B28'].value)
        rat2 = sv(hs['B38'].value)
        rat3 = sv(hs['B51'].value)
        rat6 = sv(hs['B106'].value)

        ws['J40'] = 'NA' if rat1 == 0 else hs['O2'].value or ''       # L/C Compression
        ws['P40'] = rat1        # rat1 (Concrete Bearing)
        ws['J41'] = 'NA' if rat2 == 0 else hs['O2'].value or ''       # L/C Compression (same case)
        ws['P41'] = rat2        # rat2 (Plate Yielding C)
        ws['J42'] = 'NA' if rat3 == 0 else hs['O3'].value or ''       # L/C Tensile
        ws['P42'] = rat3        # rat3 (Plate Yielding T)
        ws['C44'] = hs['B107'].value or ''     # Control type text
        ws['J44'] = 'NA' if rat6 == 0 else hs['O4'].value or ''       # L/C Anchor Bolt
        ws['P44'] = rat6        # rat6 (Anchor Bolt interaction)

        # ===== 7. Control Ratios - Pier Reinforcement =====
        rat7 = sv(hs['B121'].value)
        rat8 = sv(hs['B125'].value)
        rat9 = sv(hs['B146'].value)
        rat10 = sv(hs['B162'].value)

        ws['Y40'] = 'NA' if rat7 == 0 else hs['O3'].value or ''       # L/C Main Bar (Tensile case)
        ws['AE40'] = rat7       # rat7 (Main Bar Area)
        ws['Y41'] = 'NA' if rat8 == 0 else hs['O3'].value or ''       # L/C Dev Length (Tensile case)
        ws['AE41'] = rat8       # rat8 (Development Length)
        ws['Y43'] = 'NA' if rat9 == 0 else hs['O5'].value or ''       # L/C Tie Bar X-Dir
        ws['AE43'] = rat9       # rat9 (Tie Bar Area X)
        ws['Y44'] = 'NA' if rat10 == 0 else hs['O6'].value or ''      # L/C Tie Bar Y-Dir
        ws['AE44'] = rat10      # rat10 (Tie Bar Area Y)

    def _map_page2_data(self, ws, hs):
        """Map Page 2 data (rows 46-85): Base Plate Design - Bearing & Yielding"""

        def sv(value, dp=4):
            """Safe value - convert to number if possible"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, dp)
            except:
                return value

        # ===== 1.1 Design Requirements for Bearing =====
        # Row 49: Control case
        ws['I49'] = hs['O2'].value or ''         # L/C (Compression control case)

        # Row 50: Vertical load
        ws['I50'] = sv(hs['B15'].value)          # Vertical load (from Hinge row 15)

        # Row 57-58: A1, A2
        ws['M57'] = sv(hs['B18'].value)          # A1 = N x B (M57)
        ws['M58'] = sv(hs['B19'].value)          # A2 = Np x Bp (M58)

        # Row 60-61: Pp (AISC)
        ws['M60'] = sv(hs['B20'].value)          # Pp (AISC) (M60)
        ws['M61'] = sv(hs['B21'].value)          # 1.7f'cA1 (M61)
        ws['D61'] = '=IF(M60<=M61,"≤",">")'     # Comparison: Pp <= 1.7f'cA1

        # Row 65: Pp (ACI)
        ws['E65'] = sv(hs['B24'].value)          # Pp (ACI) (E65)

        # Row 70-71: Pn, PhiPn
        ws['K70'] = sv(hs['B26'].value)          # Pn
        ws['K71'] = sv(hs['B27'].value)          # PhiPn
        ws['O71'] = '=IF(K71<P71,"<",">")'      # Comparison: PhiPn < Pu
        ws['P71'] = sv(hs['B15'].value)          # Pu (P71 for comparison)

        # Row 71: OK/NG check - based on Ratio 1
        try:
            rat1 = float(hs['B28'].value) if hs['B28'].value else 0
            ws['U71'] = 'OK' if rat1 < 1.0 else 'NG'
        except:
            ws['U71'] = ''

        # Row 71: Ratio1
        ws['AE71'] = 'Ratio1 ='
        ws['AF71'] = sv(hs['B28'].value)         # rat1

        # ===== Base Plate Yielding Limit =====
        # Row 75-77: m, n, n' - set to 2 decimal places
        ws['K75'] = sv(hs['B32'].value, dp=2)    # m
        ws['K76'] = sv(hs['B33'].value, dp=2)    # n
        ws['K77'] = sv(hs['B34'].value, dp=2)    # n'

        # Row 79: l = max(m, n, n')
        ws['K79'] = sv(hs['B35'].value, dp=2)    # l

        # Row 84: tmin, P1, OK/NG, Ratio2
        ws['K84'] = sv(hs['B37'].value, dp=2)    # tmin
        ws['S84'] = sv(hs['H3'].value)           # P1 (plate thickness)
        ws['P84'] = '=IF(K84<=S84,"≤",">")'     # Comparison: tmin <= P1

        # OK/NG check - based on Ratio 2
        try:
            rat2 = float(hs['B38'].value) if hs['B38'].value else 0
            ws['X84'] = 'OK' if rat2 < 1.0 else 'NG'
        except:
            ws['X84'] = ''

        ws['AE84'] = 'Ratio2 ='
        ws['AF84'] = sv(hs['B38'].value)         # rat2

    def _map_page3_data(self, ws, hs):
        """Map Page 3 data (rows 87-134): Tensile Loading & AB Design (Tension)"""

        def sv(value, dp=4):
            """Safe value - convert to number if possible"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, dp)
            except:
                return value

        # ===== 1.2 Design Requirements for tensile loading =====
        ws['I88'] = hs['O3'].value or ''         # Control Case
        ws['I89'] = sv(hs['B43'].value)          # Vertical Load (F3)
        ws['L93'] = sv(hs['B46'].value)          # Nua = Vert.Load / nb
        ws['L94'] = sv(hs['B47'].value)          # Mu
        ws['L97'] = sv(hs['B48'].value)          # beff
        ws['J99'] = 0.9                         # Resistance Factor (J99-K99)
        ws['K102'] = sv(hs['B50'].value, dp=2)   # treq'd
        ws['S102'] = sv(hs['H3'].value)          # P1
        ws['P102'] = '=IF(K102<=S102,"≤",">")'  # Comparison: treq'd <= P1

        # OK/NG check for Ratio 3
        try:
            rat3 = float(hs['B51'].value) if hs['B51'].value else 0
            ws['X102'] = 'OK' if rat3 < 1.0 else 'NG'
        except:
            ws['X102'] = ''
        ws['AF102'] = sv(hs['B51'].value)        # Ratio3

        # ===== 2. Anchor Bolt Design (Tensile resistance) =====
        ws['I106'] = hs['O4'].value or ''        # Control case
        ws['I107'] = sv(hs['B56'].value)         # Vertical load
        ws['I108'] = sv(hs['B57'].value)         # Horizontal load (X Dir.)
        ws['I109'] = sv(hs['B58'].value)         # Horizontal load (Y Dir.)

        # 2.1 Design Requirements for tensile loading
        ws['I114'] = sv(hs['B62'].value)         # Nsa (I114-K114)
        ws['F118'] = sv(hs['B65'].value)         # Abrg (F118-H118)
        ws['M119'] = sv(hs['B66'].value)         # Nap (M119-O119)
        ws['E121'] = 1.4                         # Psi_c,P (E121-F121)
        ws['M122'] = sv(hs['B68'].value)         # Npn (M122-O122)

        # Side-face Blowout
        ws['E125'] = sv(hs['B71'].value)         # Ca1 (E125-G125)
        ws['K125'] = sv(hs['B72'].value)         # 2.5Ca1 (K125-M125)
        ws['Q125'] = sv(hs['B73'].value, dp=1)   # hef (Q125-S125)
        ws['T125'] = hs['B77'].value or ''       # Blowout message

        ws['E126'] = sv(hs['B74'].value)         # Ca2 (E126-G126)
        ws['K126'] = sv(hs['B75'].value)         # Ca2/Ca1 (K126-M126)
        ws['Q126'] = sv(hs['B76'].value)         # f1 (Q126-S126)
        ws['T126'] = hs['B78'].value or ''       # Corner message

        ws['Q128'] = hs['B79'].value or ''       # Nsb (Q128-S128)

        ws['E130'] = sv(hs['B80'].value)         # S1 (E130-G130)
        ws['L130'] = sv(hs['B81'].value)         # 6 x Ca1 (L130-N130)
        ws['Q130'] = sv(hs['B82'].value)         # f2 (Q130-S130)
        ws['T130'] = hs['B83'].value or ''       # Spacing message

        ws['Q133'] = hs['B84'].value or ''       # Nsbg (Q133-S133)

    def _map_page4_data(self, ws, hs):
        """Map Page 4 data (rows 135-167): Tensile/Shear Summaries & Interaction"""

        def sv(value, dp=4):
            """Safe value - convert to number if possible"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, dp)
            except:
                return value

        # ===== Tensile Design Summary =====
        ws['J137'] = sv(hs['B87'].value)         # Phi = 0.75 (J137-K137)
        ws['M138'] = sv(hs['B88'].value)         # Nn (M138-O138)
        ws['M139'] = sv(hs['B89'].value)         # PhiNn (M139-O139)

        ws['F142'] = sv(hs['B90'].value, dp=2)   # Nua (F142-H142)
        ws['M142'] = sv(hs['B89'].value)         # PhiNn (M142-O142)

        try:
            rat4 = float(hs['B91'].value) if hs['B91'].value else 0
            ws['R142'] = 'OK' if rat4 < 1.0 else 'NG'
        except:
            ws['R142'] = ''

        ws['AF143'] = sv(hs['B91'].value)        # Ratio4 (AF143-AH143)

        # ===== 2.2 Design Requirements for shear loading =====
        ws['M147'] = sv(hs['B95'].value)         # Vsa (M147-O147)

        # ===== Shear Design Summary =====
        ws['J150'] = sv(hs['B98'].value)         # Phi = 0.65 (J150-K150)
        ws['J151'] = sv(hs['B99'].value)         # PhiVn (J151-L151)

        ws['E154'] = sv(hs['B100'].value, dp=2)  # Vua (E154-G154)
        ws['L154'] = sv(hs['B99'].value)         # PhiVn (L154-N154)

        try:
            rat5 = float(hs['B101'].value) if hs['B101'].value else 0
            ws['Q154'] = 'OK' if rat5 < 1.0 else 'NG'
        except:
            ws['Q154'] = ''

        ws['AF155'] = sv(hs['B101'].value)       # Ratio5 (AF155-AH155)

        # ===== 2.3 Interaction of Tensile and Shear Forces =====
        ws['I157'] = sv(hs['B104'].value, dp=2)  # Vua / PhiVn (I157-K157)
        ws['I158'] = sv(hs['B105'].value, dp=2)  # Nua / PhiNn (I158-K158)

        # Get values from Hinge Type C104 & C105 for L157 & L158
        c104_value = sv(hs['C104'].value, dp=2) if hs['C104'].value else ''
        c105_value = sv(hs['C105'].value, dp=2) if hs['C105'].value else ''
        ws['L157'] = c104_value    # Result from Hinge Type C104
        ws['L158'] = c105_value    # Result from Hinge Type C105

        ws['C159'] = hs['B107'].value or ''      # Interaction control text (moved to C159)
        try:
            rat6 = float(hs['B106'].value) if hs['B106'].value else 0
            ws['I159'] = 'OK' if rat6 < 1.0 else 'NG'
        except:
            ws['I159'] = ''

        ws['AF159'] = sv(hs['B106'].value)       # Ratio6 (AF159-AH159)

    def _map_page5_data(self, ws, hs):
        """Map Page 5 data (rows 168-212): Pier Reinforcement - Vertical Rebars"""

        def sv(value, dp=4):
            """Safe value - convert to number if possible"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, dp)
            except:
                return value

        # ===== 3.1 Transfer of Anchor Load to Vertical Rebars =====
        ws['I171'] = hs['O3'].value or ''         # Control case (Tensile)
        ws['I172'] = sv(hs['B43'].value)          # Vertical load (F3)

        # Row 175: Rebar text e.g. "For 14 - D22 bars,"
        main_qty = sv(hs['F11'].value)
        main_size = hs['G11'].value or ''
        ws['C175'] = f"For {main_qty} - {main_size} bars,"
        import math
        try:
            size_num = int(str(main_size).replace('D', '').strip())
            ws['K175'] = sv(math.pi * (size_num ** 2) / 4, dp=0)
        except:
            ws['K175'] = ''

        # Row 179: dmax
        ws['K179'] = sv(hs['B115'].value, dp=0)   # dmax
        ws['T179'] = sv(hs['B116'].value, dp=0)   # 0.5hef
        ws['Y179'] = hs['C116'].value or ''       # OK/NG (Y179-AA179)

        # Row 182: dactual
        ws['S182'] = sv(hs['B117'].value, dp=0)   # dact (S182-U182)
        ws['AA182'] = sv(hs['B115'].value, dp=0)  # dmax (AA182-AC182)
        ws['AF182'] = hs['C117'].value or ''      # OK/NG

        # Row 185-186: Ca checks
        ws['G185'] = sv(hs['B118'].value, dp=0)   # Ca1-c-db/2 (G185-I185)
        ws['O185'] = sv(float(hs['I3'].value)/2 if hs['I3'].value else 0, dp=0) # S1/2 (O185-Q185)
        ws['T185'] = hs['C118'].value or ''       # OK/NG

        ws['G186'] = sv(hs['B119'].value, dp=0)   # Ca2-c-db/2 (G186-I186)
        ws['O186'] = sv(float(hs['J3'].value)/2 if hs['J3'].value else 0, dp=0) # S2/2 (O186-Q186)
        ws['T186'] = hs['C119'].value or ''       # OK/NG

        # Check for development length
        ws['J189'] = 0.75                         # Resistance Factor (J189-K189)
        ws['L190'] = sv(hs['B120'].value, dp=2)   # nreqd (L190-M190)
        ws['O190'] = sv(hs['F11'].value, dp=0)    # nprov (O190-P190)
        try:
            rat7 = float(hs['B121'].value) if hs['B121'].value else 0
            ws['Q190'] = 'OK' if rat7 < 1.0 else 'NG'
        except:
            ws['Q190'] = ''
        ws['AF191'] = sv(hs['B121'].value, dp=2)  # Ratio7 (AF191-AH191)

        ws['F193'] = sv(hs['B122'].value, dp=0)   # ldprov (F193-H193)
        ws['N195'] = hs['G11'].value or ''        # Main bar size from G11 (e.g. D22)
        ws['R195'] = sv(hs['B123'].value, dp=2)   # ld min (R195-T195)
        ws['R196'] = sv(hs['B124'].value, dp=2)   # ldreqd (R196-T196)
        ws['R197'] = sv(hs['B122'].value, dp=2)   # ldprov (R197-T197)
        try:
            rat8 = float(hs['B125'].value) if hs['B125'].value else 0
            ws['W197'] = 'OK' if rat8 < 1.0 else 'NG'
        except:
            ws['W197'] = ''
        ws['AF198'] = sv(hs['B125'].value, dp=2)  # Ratio8

        # 3.2 Horizontal Rebars
        ws['K202'] = sv(hs['B130'].value, dp=0)   # ldh
        ws['H204'] = 1                            # Psi_e
        ws['L204'] = 1                            # Lambda

        # X-Direction
        ws['I207'] = hs['O5'].value or ''         # Control Case X
        ws['I208'] = sv(hs['B134'].value)         # Vertical Load X
        ws['I209'] = sv(hs['B135'].value)         # Horizontal Load X
        ws['N211'] = sv(hs['B136'].value, dp=0)   # Vx

    def _map_page6_data(self, ws, hs):
        """Map Page 6 data (rows 213-247): Pier Reinforcement - Horizontal Rebars (X & Z)"""

        def sv(value, dp=4):
            """Safe value - convert to number if possible"""
            if value is None or value == '' or value == '-':
                return ''
            try:
                f = float(value)
                if f == int(f) and '.' not in str(value):
                    return int(f)
                return round(f, dp)
            except:
                return value

        # ===== X-Direction Continued =====
        # Extract diameter number from H11 (e.g. D10 -> 10)
        h11_val = hs['H11'].value or ''
        tie_diameter = str(h11_val).replace('D', '').replace('d', '') if h11_val else ''
        ws['H214'] = tie_diameter  # Tie diameter number only
        ws['N214'] = sv(hs['B137'].value, dp=1)   # Atie (N214-P214)
        ws['F215'] = sv(hs['B138'].value, dp=1)   # lda_A_L (F215-H215)
        ws['N215'] = sv(hs['B139'].value, dp=1)   # lda_A_R (N215-P215)
        ws['F216'] = sv(hs['B140'].value, dp=1)   # lda_B_L (F216-H216)
        ws['N216'] = sv(hs['B141'].value, dp=1)   # lda_B_R (N216-P216)
        ws['M218'] = sv(hs['B142'].value, dp=1)   # fs (M218-O218)
        ws['J220'] = 0.75                         # Resistance Factor (J220-K220)
        ws['K222'] = sv(hs['B144'].value, dp=3)  # Atie_reqd (K222-M222)
        ws['F224'] = sv(hs['I11'].value, dp=3)   # Legs
        ws['J224'] = hs['H11'].value or ''        # Tie bar size from H11 (e.g. D10)
        ws['F225'] = sv(hs['B145'].value, dp=3)   # Atie_prov (F225-H225)
        ws['P225'] = sv(hs['B144'].value, dp=3)   # Atie_reqd (P225-R225)
        try:
            rat9 = float(hs['B146'].value) if hs['B146'].value else 0
            ws['U225'] = 'OK' if rat9 < 1.0 else 'NG'
        except:
            ws['U225'] = ''
        ws['AF226'] = sv(hs['B146'].value, dp=2)  # Ratio9 (AF226-AH226)

        # ===== Y-Direction =====
        ws['I229'] = hs['O6'].value or ''         # Control Case Y (I229-K229)
        ws['I230'] = sv(hs['B150'].value)         # Vertical Load Y (I230-K230)

        # Check if vertical load is tension
        try:
            vy_load = float(hs['B150'].value) if hs['B150'].value else 0
            if vy_load < 0: ws['M230'] = '( Tension )'
        except: pass

        ws['I231'] = sv(hs['B151'].value)         # Horizontal Load Y (I231-K231)
        ws['N233'] = sv(hs['B152'].value, dp=2)   # Vy (N233-P233)

        # Extract diameter number from H11 (e.g. D10 -> 10)
        ws['H235'] = tie_diameter  # Tie diameter number only
        ws['N235'] = sv(hs['B153'].value, dp=1)   # Atie (N235-P235)
        ws['F236'] = sv(hs['B154'].value, dp=1)   # lda_A_L (F236-H236)
        ws['N236'] = sv(hs['B155'].value, dp=1)   # lda_A_R (N236-P236)
        ws['F237'] = sv(hs['B156'].value, dp=1)   # lda_B_L (F237-H237)
        ws['N237'] = sv(hs['B157'].value, dp=1)   # lda_B_R (N237-P237)

        # fs for Y
        ws['K239'] = sv(hs['B158'].value, dp=1)   # fs (K239-M239)

        ws['J241'] = 0.75                         # Resistance Factor (J241-K241)
        ws['K243'] = sv(hs['B160'].value, dp=3)   # Atie_reqd (K243-M243)
        ws['F245'] = sv(hs['J11'].value, dp=3)    # Legs
        ws['J245'] = hs['H11'].value or ''        # Tie bar size from H11 (e.g. D10)
        ws['F246'] = sv(hs['B161'].value, dp=3)   # Atie_prov (F246-H246)
        ws['P246'] = sv(hs['B160'].value, dp=3)   # Atie_reqd (P246-R246)
        try:
            rat10 = float(hs['B162'].value) if hs['B162'].value else 0
            ws['U246'] = 'OK' if rat10 < 1.0 else 'NG'
        except:
            ws['U246'] = ''
        ws['AF247'] = sv(hs['B162'].value, dp=2)  # Ratio10 (AF247-AH247)

    def _create_summary_sheet(self, workbook, selected_nodes):
        """Create summary sheet with node list and status"""
        ws = workbook.create_sheet("Summary", 0)

        # Header
        ws['A1'] = "Calculation Report Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25

        # Info
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"Total Nodes: {len(selected_nodes)}"

        # Table headers
        headers = ['Node', 'Status', 'Max Ratio', 'Design', 'L/C Control']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1a472a', end_color='1a472a', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row_idx, node_name in enumerate(sorted(selected_nodes), 7):
            node_data = self.base_plate_nodes[node_name]

            # Node name
            ws.cell(row=row_idx, column=1).value = node_name

            # Status
            status = node_data.get('design_status', 'N/A')
            status_cell = ws.cell(row=row_idx, column=2)
            status_cell.value = status
            status_cell.fill = PatternFill(
                start_color='4CAF50' if status == 'OK' else 'f44336',
                end_color='4CAF50' if status == 'OK' else 'f44336',
                fill_type='solid'
            )
            status_cell.font = Font(color='FFFFFF', bold=True)
            status_cell.alignment = Alignment(horizontal='center')

            # Max Ratio
            max_ratio = node_data.get('max_ratio')
            ratio_cell = ws.cell(row=row_idx, column=3)
            ratio_cell.value = f"{max_ratio:.2f}" if max_ratio else "N/A"
            ratio_cell.alignment = Alignment(horizontal='center')

            # Design check
            design_cell = ws.cell(row=row_idx, column=4)
            design_cell.value = "✓ Checked" if status != 'Not Checked' else "✗ Not Checked"
            design_cell.alignment = Alignment(horizontal='center')

            # L/C Control (placeholder - would need to read from sheet)
            lc_cell = ws.cell(row=row_idx, column=5)
            lc_cell.value = "See detail"
            lc_cell.alignment = Alignment(horizontal='center')

            # Apply borders
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
