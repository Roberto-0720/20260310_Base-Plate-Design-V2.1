"""
data/material_data_mixin.py
============================
Material data I/O methods — load/save all material tables
from/to Data.xlsx (Material Strength, Anchor Bolt, Hinge Type, etc.).

App state used: self.bpl_folder, self.sap_model_file,
                self.material_trees, self.material_status_label,
                self.status_label,
                self.anchor_bolt_material_tree, self.concrete_tree,
                self.rebar_tree, self.base_plate_material_tree
"""

import os
import shutil
import openpyxl
from tkinter import messagebox


class MaterialDataMixin:

    # ------------------------------------------------------------------
    # Reload all
    # ------------------------------------------------------------------

    def reload_all_material_data(self):
            """Reload tất cả dữ liệu từ Data.xlsx nếu bpl_folder hợp lệ"""
            if not self.bpl_folder or not os.path.exists(os.path.join(self.bpl_folder, 'Data.xlsx')):
                if self.material_status_label:
                    self.material_status_label.config(
                        text="⚠️ Please connect to SAP2000 first",
                        fg='#FF9800'
                    )
                for tree in self.material_trees.values():
                    if tree:
                        for item in tree.get_children():
                            tree.delete(item)
                return False

            # Load dữ liệu
            self.load_anchor_bolt_material_data(self.material_trees.get('anchor_bolt'))
            self.load_concrete_data(self.material_trees.get('concrete'))
            self.load_rebar_data(self.material_trees.get('rebar'))
            self.load_base_plate_material_data(self.material_trees.get('base_plate'))
            self.load_anchor_bolt_data(self.material_trees.get('Anchor Bolt Table'))
            self.load_hinge_type_data(self.material_trees.get('Hinge Type'))
            self.load_rebar_dev_length_data(self.material_trees.get('Rebar Development Length'))

            if self.material_status_label:
                sap_model_name = os.path.basename(self.sap_model_file) if self.sap_model_file else "Unknown"
                self.material_status_label.config(
                    text=f"📁 Base Plate Design folder (from {sap_model_name})",
                    fg='#4CAF50'
                )

            self.status_label.config(text="● Material data reloaded successfully", fg='#90ee90')
            return True

    # ------------------------------------------------------------------
    # Load methods — Material Strength sheet
    # ------------------------------------------------------------------

    def load_steel_data(self, tree):
        """Load Steel data from Data.xlsx (columns A-B)"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return

        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return

        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Material Strength']

            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)

            # Read ALL Steel data from columns A-B (starting from row 2)
            for row_idx in range(2, ws.max_row + 1):
                mat_type = ws[f'A{row_idx}'].value
                strength = ws[f'B{row_idx}'].value

                # Stop when reaching empty row
                if not mat_type or mat_type == 'Material Type':
                    break

                tree.insert('', 'end', values=(mat_type, strength))

            wb.close()
            print(f"Debug: Loaded {len(tree.get_children())} steel materials from Data.xlsx")
        except Exception as e:
            print(f"Error loading Steel data: {e}")

    def load_concrete_data(self, tree):
        """Load Concrete data from Data.xlsx (columns D-E)"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return

        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return

        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Material Strength']

            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)

            # Read ALL Concrete data from columns D-E (starting from row 2)
            for row_idx in range(2, ws.max_row + 1):
                mat_type = ws[f'D{row_idx}'].value
                strength = ws[f'E{row_idx}'].value

                # Stop when reaching empty row
                if not mat_type or mat_type == 'Material Type':
                    break

                tree.insert('', 'end', values=(mat_type, strength))

            wb.close()
            print(f"Debug: Loaded {len(tree.get_children())} concrete materials from Data.xlsx")
        except Exception as e:
            print(f"Error loading Concrete data: {e}")

    def load_rebar_data(self, tree):
        """Load Rebar data from Data.xlsx (columns G-H)"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return
        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return
        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Material Strength']
            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)
            # Read ALL Rebar data from columns G-H (starting from row 2)
            for row_idx in range(2, ws.max_row + 1):
                mat_type = ws[f'G{row_idx}'].value
                strength = ws[f'H{row_idx}'].value
                # Stop when reaching empty row
                if not mat_type or mat_type == 'Material Type':
                    break

                tree.insert('', 'end', values=(mat_type, strength))
            wb.close()
            print(f"Debug: Loaded {len(tree.get_children())} rebar materials from Data.xlsx")
        except Exception as e:
            print(f"Error loading Rebar data: {e}")

    def load_anchor_bolt_material_data(self, tree):
            """Load Anchor Bolt Material data from Data.xlsx (columns A-B)"""
            if not self.bpl_folder or not os.path.exists(self.bpl_folder):
                return
            data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
            if not os.path.exists(data_file):
                return
            try:
                wb = openpyxl.load_workbook(data_file)
                ws = wb['Material Strength']
                for item in tree.get_children():
                    tree.delete(item)
                for row_idx in range(2, ws.max_row + 1):
                    mat_type = ws[f'A{row_idx}'].value
                    strength = ws[f'B{row_idx}'].value
                    if not mat_type or mat_type == 'Material Type':
                        break
                    tree.insert('', 'end', values=(mat_type, strength))
                wb.close()
            except Exception as e:
                print(f"Error loading Anchor Bolt Material data: {e}")

    def load_base_plate_material_data(self, tree):
        """Load Base Plate Material data from Data.xlsx (columns J-K)"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return
        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return
        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Material Strength']
            for item in tree.get_children():
                tree.delete(item)
            for row_idx in range(2, ws.max_row + 1):
                mat_type = ws[f'J{row_idx}'].value
                strength = ws[f'K{row_idx}'].value
                if not mat_type or mat_type == 'Material Type':
                    break
                tree.insert('', 'end', values=(mat_type, strength))
            wb.close()
        except Exception as e:
            print(f"Error loading Base Plate Material data: {e}")

    # ------------------------------------------------------------------
    # Load methods — individual sheets
    # ------------------------------------------------------------------

    def load_anchor_bolt_data(self, tree):
        """Load Anchor Bolt Table data from Data.xlsx"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return

        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return

        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Anchor Bolt Table']

            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)

            # Read data starting from row 2 (skip header)
            for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
                if row[0] and row[0] != 'db':
                    values = tuple('' if v is None else str(v) for v in row)
                    tree.insert('', 'end', values=values)

            wb.close()
        except Exception as e:
            print(f"Error loading Anchor Bolt data: {e}")

    def load_base_plate_dimension_data(self, tree):
        """Load Base Plate Dimension data from Data.xlsx"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return

        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return

        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Base Plate Dimension']

            # Clear existing data
            for item in tree.get_children():
                tree.delete(item)

            # Read data starting from row 2 (skip header)
            for row in ws.iter_rows(min_row=2, max_col=14, values_only=True):
                if row[0] and row[0] != 'Column size':
                    values = tuple('' if v is None else str(v) for v in row)
                    tree.insert('', 'end', values=values)

            wb.close()
        except Exception as e:
            print(f"Error loading Base Plate Dimension data: {e}")

    def load_hinge_type_data(self, tree):
            """Load Hinge Type data from Data.xlsx"""
            if not self.bpl_folder or not os.path.exists(self.bpl_folder):
                return
            data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
            if not os.path.exists(data_file):
                return
            try:
                wb = openpyxl.load_workbook(data_file)
                ws = wb['Hinge Type']
                for item in tree.get_children():
                    tree.delete(item)
                # Đọc 22 cột
                for row in ws.iter_rows(min_row=2, max_col=22, values_only=True):
                    if row[0] and row[0] != 'Column size':
                        values = tuple('' if v is None else str(v) for v in row)
                        tree.insert('', 'end', values=values)
                wb.close()
            except Exception as e:
                print(f"Error loading Hinge Type data: {e}")

    def load_rebar_dev_length_data(self, tree):
        """Load Rebar Development Length data from Data.xlsx"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            return
        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            return
        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb['Rebar Development Length']
            for item in tree.get_children():
                tree.delete(item)
            # Read data starting from row 2 (skip header)
            for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                if row[0] and row[0] != 'Bars':
                    values = tuple('' if v is None else str(v) for v in row)
                    tree.insert('', 'end', values=values)
            wb.close()
        except Exception as e:
            print(f"Error loading Rebar Development Length data: {e}")

    def lookup_rebar_dev_length(self, bar_size, column_name):
            """Lookup Ld or Ldh from Rebar Development Length sheet

            Args:
                bar_size: int (e.g., 19 for D19)
                column_name: 'Ld' or 'Ldh'

            Returns:
                float value or 0
            """
            if not self.bpl_folder:
                return 0

            data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
            if not os.path.exists(data_file):
                return 0

            try:
                wb = openpyxl.load_workbook(data_file)
                ws = wb['Rebar Development Length']

                # Find column index
                col_idx = 2 if column_name == 'Ld' else 3  # Ld = col B, Ldh = col C

                # Search for bar size
                for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                    if row[0] and int(float(row[0])) == bar_size:
                        result = row[col_idx - 1] if len(row) >= col_idx else 0
                        wb.close()
                        return float(result) if result else 0

                wb.close()
                return 0
            except Exception as e:
                print(f"Error looking up rebar dev length: {e}")
                return 0

    # ------------------------------------------------------------------
    # Save methods
    # ------------------------------------------------------------------

    def save_material_strength_data(self):
            """Save all material strength data back to Data.xlsx"""
            if not self.bpl_folder or not os.path.exists(self.bpl_folder):
                messagebox.showwarning("Warning", "Base Plate Design folder not found!")
                return

            data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
            if not os.path.exists(data_file):
                messagebox.showwarning("Warning", "Data.xlsx not found!")
                return

            try:
                wb = openpyxl.load_workbook(data_file)
                ws = wb['Material Strength']

                # Clear và write Anchor Bolt (A-B)
                for row_idx in range(2, ws.max_row + 1):
                    ws[f'A{row_idx}'] = None
                    ws[f'B{row_idx}'] = None
                anchor_items = self.anchor_bolt_material_tree.get_children()
                for idx, item in enumerate(anchor_items, start=2):
                    values = self.anchor_bolt_material_tree.item(item, 'values')
                    ws[f'A{idx}'] = values[0] if values and values[0] else None
                    ws[f'B{idx}'] = values[1] if values and len(values) > 1 and values[1] else None

                # Clear và write Concrete (D-E)
                for row_idx in range(2, ws.max_row + 1):
                    ws[f'D{row_idx}'] = None
                    ws[f'E{row_idx}'] = None
                concrete_items = self.concrete_tree.get_children()
                for idx, item in enumerate(concrete_items, start=2):
                    values = self.concrete_tree.item(item, 'values')
                    ws[f'D{idx}'] = values[0] if values and values[0] else None
                    ws[f'E{idx}'] = values[1] if values and len(values) > 1 and values[1] else None

                # Clear và write Rebar (G-H)
                for row_idx in range(2, ws.max_row + 1):
                    ws[f'G{row_idx}'] = None
                    ws[f'H{row_idx}'] = None
                rebar_items = self.rebar_tree.get_children()
                for idx, item in enumerate(rebar_items, start=2):
                    values = self.rebar_tree.item(item, 'values')
                    ws[f'G{idx}'] = values[0] if values and values[0] else None
                    ws[f'H{idx}'] = values[1] if values and len(values) > 1 and values[1] else None

                # Clear và write Base Plate (J-K)
                for row_idx in range(2, ws.max_row + 1):
                    ws[f'J{row_idx}'] = None
                    ws[f'K{row_idx}'] = None
                baseplate_items = self.base_plate_material_tree.get_children()
                for idx, item in enumerate(baseplate_items, start=2):
                    values = self.base_plate_material_tree.item(item, 'values')
                    ws[f'J{idx}'] = values[0] if values and values[0] else None
                    ws[f'K{idx}'] = values[1] if values and len(values) > 1 and values[1] else None

                wb.save(data_file)
                wb.close()

                messagebox.showinfo("✅ Success",
                        f"Material Strength data saved successfully!\n\n"
                        f"Anchor Bolt: {len(anchor_items)} items\n"
                        f"Concrete: {len(concrete_items)} items\n"
                        f"Rebar: {len(rebar_items)} items\n"
                        f"Base Plate: {len(baseplate_items)} items")

                self.status_label.config(text="● Material Strength saved", fg='#90ee90')

            except Exception as e:
                messagebox.showerror("Error", f"Failed to save Material Strength:\n{str(e)}")

    def save_material_data(self, tree, sheet_name):
        """Save treeview data back to Data.xlsx"""
        if not self.bpl_folder or not os.path.exists(self.bpl_folder):
            messagebox.showwarning("Warning", "Base Plate Design folder not found!")
            return

        data_file = os.path.join(self.bpl_folder, 'Data.xlsx')
        if not os.path.exists(data_file):
            messagebox.showwarning("Warning", "Data.xlsx not found!")
            return

        try:
            wb = openpyxl.load_workbook(data_file)
            ws = wb[sheet_name]

            # Clear old data (keep header)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None

            # Write new data
            for idx, item in enumerate(tree.get_children(), start=2):
                values = tree.item(item, 'values')
                for col_idx, value in enumerate(values, start=1):
                    ws.cell(row=idx, column=col_idx, value=value if value else None)

            wb.save(data_file)
            wb.close()

            messagebox.showinfo("✅ Success", f"{sheet_name} data saved successfully!")
            self.status_label.config(text=f"● {sheet_name} saved", fg='#90ee90')

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save {sheet_name}:\n{str(e)}")

    # ------------------------------------------------------------------
    # Data file initialisation (software-agnostic)
    # ------------------------------------------------------------------

    def create_or_check_data_file(self):
        """Create Data.xlsx if it doesn't exist in Base Plate Design folder"""
        try:
            data_file_path = os.path.join(self.bpl_folder, "Data.xlsx")

            # If Data.xlsx already exists, don't create it again
            if os.path.exists(data_file_path):
                print(f"Debug: Data.xlsx already exists at {data_file_path}")
                return

            # Create new Data.xlsx with 3 sheets
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # ==================== SHEET 1: Material Strength ====================
            ws1 = wb.create_sheet("Material Strength")

            # Anchor Bolt Table (Columns A-B) - ĐỔI TÊN TỪ STEEL
            ws1['A1'] = "Material Type"
            ws1['B1'] = "Tensile Strength futa (MPa)"
            ws1['A2'] = "ASTM F1554 Grade 36"
            ws1['B2'] = 400
            ws1['A3'] = "ASTM A193 Grade B7"
            ws1['B3'] = 800

            # Concrete Table (Columns D-E)
            ws1['D1'] = "Material Type"
            ws1['E1'] = "Compressive Strength f'c (MPa)"
            ws1['D2'] = "F'c = 280"
            ws1['E2'] = 28
            ws1['D3'] = "F'c = 350"
            ws1['E3'] = 35

            # Rebar Table (Columns G-H)
            ws1['G1'] = "Material Type"
            ws1['H1'] = "Reinforcement yield stress fy (MPa)"
            ws1['G2'] = "SD280W"
            ws1['H2'] = 280
            ws1['G3'] = "SD420W"
            ws1['H3'] = 420

            # Base Plate Table (Columns J-K)
            ws1['J1'] = "Material Type"
            ws1['K1'] = "Yield Strength fy (MPa)"
            ws1['J2'] = "ASTM A36"
            ws1['K2'] = 245
            ws1['J3'] = "SM400"
            ws1['K3'] = 235
            ws1['J4'] = "SM490"
            ws1['K4'] = 325

            # ==================== SHEET 2: Anchor Bolt Table ====================
            ws2 = wb.create_sheet("Anchor Bolt Table")

            headers = ['db', 'Rmin', 'a', 'W', 'T', 'S', 'NutW', 'nt', 'Nut Allowance', 'Edge Min', 'Leng A1', 'Leng A2']
            for col, header in enumerate(headers, 1):
                ws2.cell(row=1, column=col, value=header)

            bolt_data = [
                [16, 319, 18, 0, 0, 25, 26.16, 11, 70, 150, 420, '-'],
                [20, 340, 22, 0, 0, 30, 33, 10, 90, 150, 460, '-'],
                [24, 400, 26, 70, 22, 48, 40, 9, 130, 150, 560, '-'],
                [30, 480, 33, 80, 24, 62, 49, 7, 170, 180, 680, '-'],
                [36, 570, 39, 100, 30, 73, 58.8, 6, 190, 220, 790, '-'],
                [42, 730, 45, 110, 32, 83, 67.9, 5, 210, 255, 970, '-'],
                [48, 860, 51, 120, 34, 93, 77.6, 4.5, 240, 290, 1130, '-'],
                [56, 860, 59, 140, 40, 106, 87.2, 4.5, 270, 340, 1160, '-'],
                [64, 970, 67, 150, 42, 120, 96.8, 4, 290, 385, 1290, '-'],
            ]

            for row_idx, row_data in enumerate(bolt_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws2.cell(row=row_idx, column=col_idx, value=value)

            # ==================== SHEET 3: Hinge Type ====================
            ws3 = wb.create_sheet("Hinge Type")

            headers = ['Column size', 'Type', 'No.AB', 'P1', 'N', 'A', 'B', 'C', 'E', 'F', 'P2', 'Y',
                      'Np', 'Bp', 'c', 'nrb', 'drb', 'dtb', 'X-leg', 'Y-leg', 'Layer 1', 'Layer 2']
            for col, header in enumerate(headers, 1):
                ws3.cell(row=1, column=col, value=header)

            hinge_data = [
                ['H194X150X6X9', '-', 2, 19, 250, 120, 200, '-', 40, 125, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H244X175X7X11', '-', 2, 19, 300, 120, 250, '-', 65, 150, '-', '-', 500, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H125X125X6.5X9', '-', 2, 16, 200, 120, 200, '-', 40, 100, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H150X150X7X10', '-', 2, 19, 200, 120, 200, '-', 40, 100, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H175X175X7.5X11', '-', 2, 19, 250, 120, 250, '-', 65, 125, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H200X200X8X12', '-', 2, 25, 250, 120, 250, '-', 65, 125, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H250X250X9X14', 'B', 4, 25, 300, 120, 300, 120, 90, 90, '-', '-', 450, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H294X200X8X12', 'B', 4, 25, 350, 120, 250, 120, 65, 115, '-', '-', 500, 450, 50, 8, 'D19', 'D10', 2, 2, 50, 75],
                ['H300X300X10X15', 'B', 4, 32, 350, 160, 350, 160, 95, 95, '-', '-', 550, 550, 50, 12, 'D19', 'D10', 3, 3, 50, 75],
                ['H340X250X9X14', 'B', 4, 32, 400, 160, 300, 160, 70, 120, '-', '-', 600, 550, 50, 12, 'D19', 'D10', 3, 3, 50, 75],
                ['H350X350X12X19', 'B', 4, 36, 400, 200, 400, 200, 100, 100, '-', '-', 650, 650, 50, 16, 'D19', 'D13', 3, 3, 50, 75],
                ['H390X300X10X16', 'B', 4, 36, 450, 160, 350, 180, 95, 135, '-', '-', 600, 550, 50, 12, 'D19', 'D13', 3, 3, 50, 75],
                ['H400X400X13X21', 'B', 4, 36, 450, 200, 450, 200, 125, 125, '-', '-', 700, 650, 50, 16, 'D19', 'D13', 3, 3, 50, 75],
                ['H488X300X11X18', 'B', 4, 32, 550, 200, 350, 240, 75, 155, '-', '-', 750, 650, 50, 18, 'D19', 'D13', 4, 3, 50, 75],
                ['H588X300X12X20', 'B', 4, 36, 650, 200, 350, 280, 75, 185, '-', '-', 750, 700, 50, 20, 'D19', 'D13', 4, 4, 50, 75],
                ['BH400X340X16X25', 'B', 4, 36, 450, 200, 400, 200, 100, 125, '-', '-', 650, 600, 50, 16, 'D19', 'D13', 3, 3, 50, 75],
                ['BH500X340X16X25', 'B', 4, 40, 550, 240, 400, 240, 80, 155, '-', '-', 750, 700, 50, 20, 'D19', 'D13', 4, 4, 50, 75],
                ['BH514X340X19X32', 'B', 4, 40, 600, 240, 400, 240, 80, 180, '-', '-', 750, 700, 50, 20, 'D19', 'D13', 4, 4, 50, 75],
            ]

            for row_idx, row_data in enumerate(hinge_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws3.cell(row=row_idx, column=col_idx, value=value)

            # ==================== SHEET 4: Rebar Development Length ====================
            ws4 = wb.create_sheet("Rebar Development Length")

            headers = ['Bars', 'Ld', 'Ldh']
            for col, header in enumerate(headers, 1):
                ws4.cell(row=1, column=col, value=header)

            rebar_dev_data = [
                [10, 490, 150],
                [13, 490, 150],
                [16, 520, 160],
                [19, 950, 290],
                [22, 1110, 340],
                [25, 1270, 390],
                [32, 1600, 490],
                [36, 1760, 550],
            ]

            for row_idx, row_data in enumerate(rebar_dev_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws4.cell(row=row_idx, column=col_idx, value=value)

            wb.save(data_file_path)
            print(f"Debug: Data.xlsx created successfully at {data_file_path}")
            messagebox.showinfo("✅ Success", f"Data.xlsx created successfully in Base Plate Design folder")

        except Exception as e:
            messagebox.showerror("❌ Error", f"Failed to create Data.xlsx:\n{str(e)}")
            print(f"Error creating Data.xlsx: {str(e)}")

    def _copy_template_to_bpl_folder(self):
        """Copy Template.xlsx from script directory to Base Plate Design folder"""
        if not self.bpl_folder:
            return

        template_dest = os.path.join(self.bpl_folder, 'Template.xlsx')
        if os.path.exists(template_dest):
            return  # Already exists

        # Find Template.xlsx in script directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_src = os.path.join(script_dir, 'Template.xlsx')

        if os.path.exists(template_src):
            try:
                shutil.copy2(template_src, template_dest)
                print(f"Debug: Template.xlsx copied to {template_dest}")
            except Exception as e:
                print(f"Error copying Template.xlsx: {e}")
        else:
            print(f"Warning: Template.xlsx not found at {script_dir}")
